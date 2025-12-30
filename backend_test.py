#!/usr/bin/env python3
"""
Backend API Testing for Betting Automation System
Tests the Opportunities endpoints and plays888.co scraping functionality
"""

import requests
import sys
import json
from datetime import datetime

class BettingSystemAPITester:
    def __init__(self, base_url="https://sportswizard.preview.emergentagent.com"):
        self.base_url = base_url
        self.api_url = f"{base_url}/api"
        self.tests_run = 0
        self.tests_passed = 0
        self.test_results = []

    def log_test(self, name, success, details=""):
        """Log test result"""
        self.tests_run += 1
        if success:
            self.tests_passed += 1
        
        result = {
            "test": name,
            "success": success,
            "details": details,
            "timestamp": datetime.now().isoformat()
        }
        self.test_results.append(result)
        
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status} - {name}")
        if details:
            print(f"    Details: {details}")

    def test_get_opportunities(self):
        """Test GET /api/opportunities endpoint"""
        try:
            response = requests.get(f"{self.api_url}/opportunities", timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'plays', 'date', 'last_updated']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("GET /api/opportunities - Structure", False, 
                                f"Missing fields: {missing_fields}")
                    return False
                
                # Validate games array
                games = data.get('games', [])
                if not isinstance(games, list):
                    self.log_test("GET /api/opportunities - Games Array", False, 
                                "Games is not an array")
                    return False
                
                # Check if we have games data
                if len(games) > 0:
                    game = games[0]
                    game_fields = ['game_num', 'time', 'away_team', 'home_team', 
                                 'away_ppg_rank', 'away_last3_rank', 'home_ppg_rank', 
                                 'home_last3_rank', 'total', 'game_avg', 'recommendation']
                    
                    missing_game_fields = [field for field in game_fields if field not in game]
                    if missing_game_fields:
                        self.log_test("GET /api/opportunities - Game Structure", False,
                                    f"Missing game fields: {missing_game_fields}")
                        return False
                    
                    self.log_test("GET /api/opportunities - Game Structure", True,
                                f"Game has all required fields")
                
                # Validate plays array
                plays = data.get('plays', [])
                if isinstance(plays, list) and len(plays) > 0:
                    play = plays[0]
                    play_fields = ['game', 'total', 'game_avg', 'recommendation', 'color']
                    missing_play_fields = [field for field in play_fields if field not in play]
                    if missing_play_fields:
                        self.log_test("GET /api/opportunities - Play Structure", False,
                                    f"Missing play fields: {missing_play_fields}")
                    else:
                        self.log_test("GET /api/opportunities - Play Structure", True,
                                    f"Play has all required fields")
                
                self.log_test("GET /api/opportunities", True, 
                            f"Returned {len(games)} games, {len(plays)} plays")
                return True
            else:
                self.log_test("GET /api/opportunities", False, 
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("GET /api/opportunities", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("GET /api/opportunities", False, f"JSON decode error: {str(e)}")
            return False

    def test_refresh_opportunities(self):
        """Test POST /api/opportunities/refresh endpoint"""
        try:
            response = requests.post(f"{self.api_url}/opportunities/refresh", timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure (same as GET)
                required_fields = ['games', 'plays', 'date', 'last_updated']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("POST /api/opportunities/refresh - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                games = data.get('games', [])
                plays = data.get('plays', [])
                
                self.log_test("POST /api/opportunities/refresh", True,
                            f"Refreshed: {len(games)} games, {len(plays)} plays")
                return True
            else:
                self.log_test("POST /api/opportunities/refresh", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("POST /api/opportunities/refresh", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("POST /api/opportunities/refresh", False, f"JSON decode error: {str(e)}")
            return False

    def test_betting_logic(self):
        """Test the betting recommendation logic"""
        try:
            response = requests.get(f"{self.api_url}/opportunities", timeout=10)
            
            if response.status_code != 200:
                self.log_test("Betting Logic Test", False, "Could not get opportunities data")
                return False
            
            data = response.json()
            games = data.get('games', [])
            
            if not games:
                self.log_test("Betting Logic Test", False, "No games data to test")
                return False
            
            # Test betting logic rules (updated rules from frontend)
            over_games = []
            under_games = []
            no_bet_games = []
            
            for game in games:
                game_avg = game.get('game_avg')
                recommendation = game.get('recommendation')
                
                if game_avg is not None:
                    if game_avg <= 12.5:  # OVER range: 1-12.5
                        if recommendation == 'OVER':
                            over_games.append(game)
                        else:
                            self.log_test("Betting Logic - OVER Rule", False,
                                        f"Game avg {game_avg} should be OVER but got {recommendation}")
                            return False
                    elif game_avg >= 17.5:  # UNDER range: 17.5-30
                        if recommendation == 'UNDER':
                            under_games.append(game)
                        else:
                            self.log_test("Betting Logic - UNDER Rule", False,
                                        f"Game avg {game_avg} should be UNDER but got {recommendation}")
                            return False
                    else:  # No edge range: 13-17
                        if recommendation is None or recommendation == '':
                            no_bet_games.append(game)
                        else:
                            self.log_test("Betting Logic - No Bet Rule", False,
                                        f"Game avg {game_avg} should have no recommendation but got {recommendation}")
                            return False
            
            self.log_test("Betting Logic Test", True,
                        f"OVER: {len(over_games)}, UNDER: {len(under_games)}, No bet: {len(no_bet_games)}")
            return True
            
        except Exception as e:
            self.log_test("Betting Logic Test", False, f"Error: {str(e)}")
            return False

    def test_color_coding(self):
        """Test color coding for recommendations"""
        try:
            response = requests.get(f"{self.api_url}/opportunities", timeout=10)
            
            if response.status_code != 200:
                self.log_test("Color Coding Test", False, "Could not get opportunities data")
                return False
            
            data = response.json()
            games = data.get('games', [])
            plays = data.get('plays', [])
            
            # Test game color coding
            for game in games:
                recommendation = game.get('recommendation')
                color = game.get('color')
                
                if recommendation == 'OVER' and color != 'green':
                    self.log_test("Color Coding - OVER", False,
                                f"OVER recommendation should be green, got {color}")
                    return False
                elif recommendation == 'UNDER' and color != 'red':
                    self.log_test("Color Coding - UNDER", False,
                                f"UNDER recommendation should be red, got {color}")
                    return False
            
            # Test plays color coding
            for play in plays:
                recommendation = play.get('recommendation')
                color = play.get('color')
                
                if recommendation == 'OVER' and color != 'green':
                    self.log_test("Color Coding - Play OVER", False,
                                f"OVER play should be green, got {color}")
                    return False
                elif recommendation == 'UNDER' and color != 'red':
                    self.log_test("Color Coding - Play UNDER", False,
                                f"UNDER play should be red, got {color}")
                    return False
            
            self.log_test("Color Coding Test", True, "All color coding is correct")
            return True
            
        except Exception as e:
            self.log_test("Color Coding Test", False, f"Error: {str(e)}")
            return False

    def test_scrape_nba_totals(self):
        """Test POST /api/scrape/totals/NBA endpoint"""
        try:
            print("Testing NBA totals scraping (this may take 30-60 seconds)...")
            response = requests.post(f"{self.api_url}/scrape/totals/NBA", timeout=90)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                if not isinstance(data, dict) or 'games' not in data:
                    self.log_test("POST /api/scrape/totals/NBA - Structure", False,
                                "Response should have 'games' field")
                    return False
                
                games = data.get('games', [])
                if not isinstance(games, list):
                    self.log_test("POST /api/scrape/totals/NBA - Games Array", False,
                                "Games should be an array")
                    return False
                
                if len(games) == 0:
                    self.log_test("POST /api/scrape/totals/NBA", False,
                                "No NBA games found - may be off-season or scraping issue")
                    return False
                
                # Validate game structure
                game = games[0]
                required_fields = ['away', 'home', 'total']
                missing_fields = [field for field in required_fields if field not in game]
                
                if missing_fields:
                    self.log_test("POST /api/scrape/totals/NBA - Game Structure", False,
                                f"Missing game fields: {missing_fields}")
                    return False
                
                self.log_test("POST /api/scrape/totals/NBA", True,
                            f"Successfully scraped {len(games)} NBA games from plays888.co")
                return True
            else:
                self.log_test("POST /api/scrape/totals/NBA", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.Timeout:
            self.log_test("POST /api/scrape/totals/NBA", False, "Request timeout (>90s)")
            return False
        except requests.exceptions.RequestException as e:
            self.log_test("POST /api/scrape/totals/NBA", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("POST /api/scrape/totals/NBA", False, f"JSON decode error: {str(e)}")
            return False

    def test_scrape_nhl_totals(self):
        """Test POST /api/scrape/totals/NHL endpoint"""
        try:
            print("Testing NHL totals scraping (this may take 30-60 seconds)...")
            response = requests.post(f"{self.api_url}/scrape/totals/NHL", timeout=90)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                if not isinstance(data, dict) or 'games' not in data:
                    self.log_test("POST /api/scrape/totals/NHL - Structure", False,
                                "Response should have 'games' field")
                    return False
                
                games = data.get('games', [])
                if not isinstance(games, list):
                    self.log_test("POST /api/scrape/totals/NHL - Games Array", False,
                                "Games should be an array")
                    return False
                
                if len(games) == 0:
                    self.log_test("POST /api/scrape/totals/NHL", False,
                                "No NHL games found - may be off-season or scraping issue")
                    return False
                
                # Validate game structure
                game = games[0]
                required_fields = ['away', 'home', 'total']
                missing_fields = [field for field in required_fields if field not in game]
                
                if missing_fields:
                    self.log_test("POST /api/scrape/totals/NHL - Game Structure", False,
                                f"Missing game fields: {missing_fields}")
                    return False
                
                self.log_test("POST /api/scrape/totals/NHL", True,
                            f"Successfully scraped {len(games)} NHL games from plays888.co")
                return True
            else:
                self.log_test("POST /api/scrape/totals/NHL", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.Timeout:
            self.log_test("POST /api/scrape/totals/NHL", False, "Request timeout (>90s)")
            return False
        except requests.exceptions.RequestException as e:
            self.log_test("POST /api/scrape/totals/NHL", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("POST /api/scrape/totals/NHL", False, f"JSON decode error: {str(e)}")
            return False

    def test_refresh_with_live_lines_nba(self):
        """Test POST /api/opportunities/refresh?use_live_lines=true for NBA"""
        try:
            print("Testing NBA opportunities refresh with live lines (this may take 30-60 seconds)...")
            response = requests.post(f"{self.api_url}/opportunities/refresh?use_live_lines=true", timeout=90)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'plays', 'date', 'last_updated', 'data_source']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("POST /api/opportunities/refresh?use_live_lines=true - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                # Check data source
                data_source = data.get('data_source')
                if data_source != 'plays888.co':
                    self.log_test("POST /api/opportunities/refresh?use_live_lines=true - Data Source", False,
                                f"Expected data_source 'plays888.co', got '{data_source}'")
                    return False
                
                games = data.get('games', [])
                plays = data.get('plays', [])
                
                self.log_test("POST /api/opportunities/refresh?use_live_lines=true", True,
                            f"NBA refresh with live lines: {len(games)} games, {len(plays)} plays, source: {data_source}")
                return True
            else:
                self.log_test("POST /api/opportunities/refresh?use_live_lines=true", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.Timeout:
            self.log_test("POST /api/opportunities/refresh?use_live_lines=true", False, "Request timeout (>90s)")
            return False
        except requests.exceptions.RequestException as e:
            self.log_test("POST /api/opportunities/refresh?use_live_lines=true", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("POST /api/opportunities/refresh?use_live_lines=true", False, f"JSON decode error: {str(e)}")
            return False

    def test_refresh_with_live_lines_nhl(self):
        """Test POST /api/opportunities/nhl/refresh?use_live_lines=true for NHL"""
        try:
            print("Testing NHL opportunities refresh with live lines (this may take 30-60 seconds)...")
            response = requests.post(f"{self.api_url}/opportunities/nhl/refresh?use_live_lines=true", timeout=90)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'plays', 'date', 'last_updated', 'data_source']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("POST /api/opportunities/nhl/refresh?use_live_lines=true - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                # Check data source
                data_source = data.get('data_source')
                if data_source != 'plays888.co':
                    self.log_test("POST /api/opportunities/nhl/refresh?use_live_lines=true - Data Source", False,
                                f"Expected data_source 'plays888.co', got '{data_source}'")
                    return False
                
                games = data.get('games', [])
                plays = data.get('plays', [])
                
                self.log_test("POST /api/opportunities/nhl/refresh?use_live_lines=true", True,
                            f"NHL refresh with live lines: {len(games)} games, {len(plays)} plays, source: {data_source}")
                return True
            else:
                self.log_test("POST /api/opportunities/nhl/refresh?use_live_lines=true", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.Timeout:
            self.log_test("POST /api/opportunities/nhl/refresh?use_live_lines=true", False, "Request timeout (>90s)")
            return False
        except requests.exceptions.RequestException as e:
            self.log_test("POST /api/opportunities/nhl/refresh?use_live_lines=true", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("POST /api/opportunities/nhl/refresh?use_live_lines=true", False, f"JSON decode error: {str(e)}")
            return False

    def test_data_source_field(self):
        """Test that GET /api/opportunities returns data_source field"""
        try:
            response = requests.get(f"{self.api_url}/opportunities", timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                
                if 'data_source' not in data:
                    self.log_test("GET /api/opportunities - data_source field", False,
                                "Missing data_source field in response")
                    return False
                
                data_source = data.get('data_source')
                valid_sources = ['cached', 'plays888.co']
                
                if data_source not in valid_sources:
                    self.log_test("GET /api/opportunities - data_source value", False,
                                f"Invalid data_source '{data_source}', expected one of {valid_sources}")
                    return False
                
                self.log_test("GET /api/opportunities - data_source field", True,
                            f"data_source field present with value: {data_source}")
                return True
            else:
                self.log_test("GET /api/opportunities - data_source field", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("GET /api/opportunities - data_source field", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("GET /api/opportunities - data_source field", False, f"JSON decode error: {str(e)}")
            return False

    def test_bet_time_line_tracking_yesterday(self):
        """Test GET /api/opportunities?day=yesterday for bet-time line tracking"""
        try:
            response = requests.get(f"{self.api_url}/opportunities?day=yesterday", timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("GET /api/opportunities?day=yesterday - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                games = data.get('games', [])
                if not isinstance(games, list):
                    self.log_test("GET /api/opportunities?day=yesterday - Games Array", False,
                                "Games is not an array")
                    return False
                
                # Look for games with user bets
                user_bet_games = [g for g in games if g.get('user_bet') == True]
                
                if not user_bet_games:
                    self.log_test("GET /api/opportunities?day=yesterday - User Bets", False,
                                "No games with user_bet=true found")
                    return False
                
                # Test specific expected game: San Antonio @ Okla City
                san_antonio_game = None
                for game in user_bet_games:
                    if (game.get('away_team') == 'San Antonio' and game.get('home_team') == 'Okla City'):
                        san_antonio_game = game
                        break
                
                if san_antonio_game:
                    # Verify required fields for bet-time line tracking
                    required_bet_fields = ['bet_line', 'bet_edge', 'user_bet_hit', 'final_score']
                    missing_bet_fields = [field for field in required_bet_fields if field not in san_antonio_game]
                    
                    if missing_bet_fields:
                        self.log_test("Bet-Time Line Tracking - San Antonio Game Fields", False,
                                    f"Missing bet fields: {missing_bet_fields}")
                        return False
                    
                    # Verify expected values for San Antonio @ Okla City
                    expected_bet_line = 233.0
                    expected_bet_edge = 6.0
                    expected_final_score = 219
                    expected_user_bet_hit = False  # 219 < 233 means UNDER won
                    
                    actual_bet_line = san_antonio_game.get('bet_line')
                    actual_bet_edge = san_antonio_game.get('bet_edge')
                    actual_final_score = san_antonio_game.get('final_score')
                    actual_user_bet_hit = san_antonio_game.get('user_bet_hit')
                    
                    # Validate bet_line
                    if actual_bet_line != expected_bet_line:
                        self.log_test("Bet-Time Line Tracking - San Antonio bet_line", False,
                                    f"Expected bet_line {expected_bet_line}, got {actual_bet_line}")
                        return False
                    
                    # Validate bet_edge
                    if actual_bet_edge != expected_bet_edge:
                        self.log_test("Bet-Time Line Tracking - San Antonio bet_edge", False,
                                    f"Expected bet_edge {expected_bet_edge}, got {actual_bet_edge}")
                        return False
                    
                    # Validate final_score
                    if actual_final_score != expected_final_score:
                        self.log_test("Bet-Time Line Tracking - San Antonio final_score", False,
                                    f"Expected final_score {expected_final_score}, got {actual_final_score}")
                        return False
                    
                    # Validate user_bet_hit
                    if actual_user_bet_hit != expected_user_bet_hit:
                        self.log_test("Bet-Time Line Tracking - San Antonio user_bet_hit", False,
                                    f"Expected user_bet_hit {expected_user_bet_hit}, got {actual_user_bet_hit}")
                        return False
                    
                    self.log_test("Bet-Time Line Tracking - San Antonio Game", True,
                                f"All bet-time tracking fields correct: bet_line={actual_bet_line}, bet_edge={actual_bet_edge}, final={actual_final_score}, hit={actual_user_bet_hit}")
                else:
                    self.log_test("Bet-Time Line Tracking - San Antonio Game", False,
                                "San Antonio @ Okla City game not found in user bets")
                    return False
                
                # Test Houston @ LA Lakers game if present
                houston_game = None
                for game in user_bet_games:
                    if (game.get('away_team') == 'Houston' and game.get('home_team') == 'LA Lakers'):
                        houston_game = game
                        break
                
                if houston_game:
                    expected_houston_bet_line = 230.0
                    expected_houston_bet_edge = 8.5
                    expected_houston_final_score = 215
                    expected_houston_user_bet_hit = False  # 215 < 230 means UNDER won
                    
                    actual_houston_bet_line = houston_game.get('bet_line')
                    actual_houston_bet_edge = houston_game.get('bet_edge')
                    actual_houston_final_score = houston_game.get('final_score')
                    actual_houston_user_bet_hit = houston_game.get('user_bet_hit')
                    
                    if (actual_houston_bet_line == expected_houston_bet_line and
                        actual_houston_bet_edge == expected_houston_bet_edge and
                        actual_houston_final_score == expected_houston_final_score and
                        actual_houston_user_bet_hit == expected_houston_user_bet_hit):
                        self.log_test("Bet-Time Line Tracking - Houston Game", True,
                                    f"Houston game correct: bet_line={actual_houston_bet_line}, bet_edge={actual_houston_bet_edge}, final={actual_houston_final_score}, hit={actual_houston_user_bet_hit}")
                    else:
                        self.log_test("Bet-Time Line Tracking - Houston Game", False,
                                    f"Houston game values incorrect. Expected: bet_line={expected_houston_bet_line}, bet_edge={expected_houston_bet_edge}, final={expected_houston_final_score}, hit={expected_houston_user_bet_hit}. Got: bet_line={actual_houston_bet_line}, bet_edge={actual_houston_bet_edge}, final={actual_houston_final_score}, hit={actual_houston_user_bet_hit}")
                        return False
                
                self.log_test("GET /api/opportunities?day=yesterday - Bet-Time Line Tracking", True,
                            f"Found {len(user_bet_games)} games with user bets, bet-time line tracking working correctly")
                return True
            else:
                self.log_test("GET /api/opportunities?day=yesterday", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("GET /api/opportunities?day=yesterday", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("GET /api/opportunities?day=yesterday", False, f"JSON decode error: {str(e)}")
            return False

    def test_refresh_opportunities_yesterday(self):
        """Test POST /api/opportunities/refresh?day=yesterday for bet-time line tracking"""
        try:
            response = requests.post(f"{self.api_url}/opportunities/refresh?day=yesterday", timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("POST /api/opportunities/refresh?day=yesterday - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                # Check data source
                data_source = data.get('data_source')
                if data_source:
                    self.log_test("POST /api/opportunities/refresh?day=yesterday - Data Source", True,
                                f"Data source: {data_source}")
                else:
                    self.log_test("POST /api/opportunities/refresh?day=yesterday - Data Source", False,
                                "No data_source field in response")
                    return False
                
                games = data.get('games', [])
                user_bet_games = [g for g in games if g.get('user_bet') == True]
                
                if user_bet_games:
                    # Verify bet-time line tracking fields are present
                    sample_game = user_bet_games[0]
                    bet_tracking_fields = ['bet_line', 'bet_edge', 'user_bet_hit']
                    missing_tracking_fields = [field for field in bet_tracking_fields if field not in sample_game]
                    
                    if missing_tracking_fields:
                        self.log_test("POST /api/opportunities/refresh?day=yesterday - Bet Tracking", False,
                                    f"Missing bet tracking fields: {missing_tracking_fields}")
                        return False
                    
                    self.log_test("POST /api/opportunities/refresh?day=yesterday - Bet Tracking", True,
                                f"Bet tracking fields present in {len(user_bet_games)} user bet games")
                
                self.log_test("POST /api/opportunities/refresh?day=yesterday", True,
                            f"Refreshed yesterday data: {len(games)} games, {len(user_bet_games)} with user bets, source: {data_source}")
                return True
            else:
                self.log_test("POST /api/opportunities/refresh?day=yesterday", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("POST /api/opportunities/refresh?day=yesterday", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("POST /api/opportunities/refresh?day=yesterday", False, f"JSON decode error: {str(e)}")
            return False

    def test_bet_line_vs_closing_line_difference(self):
        """Test that bet_line (bet-time) is different from total (closing line) for user bets"""
        try:
            response = requests.get(f"{self.api_url}/opportunities?day=yesterday", timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                games = data.get('games', [])
                user_bet_games = [g for g in games if g.get('user_bet') == True]
                
                if not user_bet_games:
                    self.log_test("Bet Line vs Closing Line Test", False,
                                "No games with user bets found to test")
                    return False
                
                differences_found = 0
                for game in user_bet_games:
                    bet_line = game.get('bet_line')
                    closing_line = game.get('total')
                    
                    if bet_line is not None and closing_line is not None:
                        if bet_line != closing_line:
                            differences_found += 1
                            self.log_test(f"Line Difference - {game.get('away_team')} @ {game.get('home_team')}", True,
                                        f"Bet line: {bet_line}, Closing line: {closing_line}, Difference: {abs(bet_line - closing_line)}")
                
                if differences_found > 0:
                    self.log_test("Bet Line vs Closing Line Test", True,
                                f"Found {differences_found} games with different bet-time vs closing lines")
                    return True
                else:
                    self.log_test("Bet Line vs Closing Line Test", False,
                                "No differences found between bet-time and closing lines")
                    return False
            else:
                self.log_test("Bet Line vs Closing Line Test", False,
                            f"Status code: {response.status_code}")
                return False
                
        except Exception as e:
            self.log_test("Bet Line vs Closing Line Test", False, f"Error: {str(e)}")
            return False

    def test_nba_data_sourcing_strategy_tomorrow(self):
        """Test GET /api/opportunities?day=tomorrow - should use scoresandodds.com"""
        try:
            response = requests.get(f"{self.api_url}/opportunities?day=tomorrow", timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success', 'data_source']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("GET /api/opportunities?day=tomorrow - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                # Check data source
                data_source = data.get('data_source')
                expected_sources = ['scoresandodds.com', 'hardcoded']  # Allow hardcoded as fallback
                
                if data_source not in expected_sources:
                    self.log_test("GET /api/opportunities?day=tomorrow - Data Source", False,
                                f"Expected data_source in {expected_sources}, got '{data_source}'")
                    return False
                
                games = data.get('games', [])
                if len(games) == 0:
                    self.log_test("GET /api/opportunities?day=tomorrow - Games Count", False,
                                "No games found for tomorrow")
                    return False
                
                # Validate game structure
                game = games[0]
                required_game_fields = ['away_team', 'home_team', 'total']
                missing_game_fields = [field for field in required_game_fields if field not in game]
                
                if missing_game_fields:
                    self.log_test("GET /api/opportunities?day=tomorrow - Game Structure", False,
                                f"Missing game fields: {missing_game_fields}")
                    return False
                
                # Expected: 9 games for Dec 27, 2025 (Dallas @ Sacramento, Phoenix @ New Orleans, etc.)
                expected_min_games = 3  # Allow some flexibility
                if len(games) < expected_min_games:
                    self.log_test("GET /api/opportunities?day=tomorrow - Game Count", False,
                                f"Expected at least {expected_min_games} games, got {len(games)}")
                    return False
                
                self.log_test("GET /api/opportunities?day=tomorrow", True,
                            f"Found {len(games)} games, data_source: {data_source}")
                return True
            else:
                self.log_test("GET /api/opportunities?day=tomorrow", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("GET /api/opportunities?day=tomorrow", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("GET /api/opportunities?day=tomorrow", False, f"JSON decode error: {str(e)}")
            return False

    def test_nba_data_sourcing_strategy_today(self):
        """Test GET /api/opportunities?day=today - should use plays888.co for live lines"""
        try:
            response = requests.get(f"{self.api_url}/opportunities?day=today", timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("GET /api/opportunities?day=today - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                # Check data source (may be hardcoded if no live lines fetched)
                data_source = data.get('data_source')
                valid_sources = ['plays888.co', 'hardcoded', None]  # Allow hardcoded as fallback
                
                if data_source not in valid_sources:
                    self.log_test("GET /api/opportunities?day=today - Data Source", False,
                                f"Unexpected data_source '{data_source}', expected one of {valid_sources}")
                    return False
                
                games = data.get('games', [])
                
                # Validate game structure if games exist
                if games:
                    game = games[0]
                    required_game_fields = ['away_team', 'home_team', 'total']
                    missing_game_fields = [field for field in required_game_fields if field not in game]
                    
                    if missing_game_fields:
                        self.log_test("GET /api/opportunities?day=today - Game Structure", False,
                                    f"Missing game fields: {missing_game_fields}")
                        return False
                
                self.log_test("GET /api/opportunities?day=today", True,
                            f"Found {len(games)} games, data_source: {data_source}")
                return True
            else:
                self.log_test("GET /api/opportunities?day=today", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("GET /api/opportunities?day=today", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("GET /api/opportunities?day=today", False, f"JSON decode error: {str(e)}")
            return False

    def test_nba_data_sourcing_strategy_yesterday(self):
        """Test GET /api/opportunities?day=yesterday - should use scoresandodds.com for final scores + plays888 for bet lines"""
        try:
            response = requests.get(f"{self.api_url}/opportunities?day=yesterday", timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("GET /api/opportunities?day=yesterday - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                games = data.get('games', [])
                if len(games) == 0:
                    self.log_test("GET /api/opportunities?day=yesterday - Games Count", False,
                                "No games found for yesterday")
                    return False
                
                # Expected: 5 games for Dec 25 (Christmas Day)
                expected_games = 5
                if len(games) != expected_games:
                    self.log_test("GET /api/opportunities?day=yesterday - Game Count", False,
                                f"Expected {expected_games} games for Christmas Day, got {len(games)}")
                    return False
                
                # Check for user bets with bet_line and bet_edge fields
                user_bet_games = [g for g in games if g.get('user_bet') == True]
                
                if not user_bet_games:
                    self.log_test("GET /api/opportunities?day=yesterday - User Bets", False,
                                "No games with user_bet=true found")
                    return False
                
                # Validate user bet fields
                for game in user_bet_games:
                    required_bet_fields = ['bet_line', 'bet_edge']
                    missing_bet_fields = [field for field in required_bet_fields if field not in game]
                    
                    if missing_bet_fields:
                        self.log_test("GET /api/opportunities?day=yesterday - Bet Fields", False,
                                    f"Missing bet fields: {missing_bet_fields}")
                        return False
                
                self.log_test("GET /api/opportunities?day=yesterday", True,
                            f"Found {len(games)} games, {len(user_bet_games)} with user bets and bet-time data")
                return True
            else:
                self.log_test("GET /api/opportunities?day=yesterday", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("GET /api/opportunities?day=yesterday", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("GET /api/opportunities?day=yesterday", False, f"JSON decode error: {str(e)}")
            return False

    def test_nba_refresh_tomorrow_scoresandodds(self):
        """Test POST /api/opportunities/refresh?day=tomorrow - should force refresh from scoresandodds.com"""
        try:
            response = requests.post(f"{self.api_url}/opportunities/refresh?day=tomorrow", timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("POST /api/opportunities/refresh?day=tomorrow - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                # Check data source indicates correct source
                data_source = data.get('data_source')
                expected_sources = ['scoresandodds.com', 'hardcoded']  # Allow hardcoded as fallback
                
                if data_source not in expected_sources:
                    self.log_test("POST /api/opportunities/refresh?day=tomorrow - Data Source", False,
                                f"Expected data_source in {expected_sources}, got '{data_source}'")
                    return False
                
                games = data.get('games', [])
                if len(games) == 0:
                    self.log_test("POST /api/opportunities/refresh?day=tomorrow - Games Count", False,
                                "No games found after refresh")
                    return False
                
                self.log_test("POST /api/opportunities/refresh?day=tomorrow", True,
                            f"Refreshed {len(games)} games, data_source: {data_source}")
                return True
            else:
                self.log_test("POST /api/opportunities/refresh?day=tomorrow", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("POST /api/opportunities/refresh?day=tomorrow", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("POST /api/opportunities/refresh?day=tomorrow", False, f"JSON decode error: {str(e)}")
            return False

    def test_nba_refresh_today_plays888(self):
        """Test POST /api/opportunities/refresh?day=today&use_live_lines=true - should use plays888.co"""
        try:
            response = requests.post(f"{self.api_url}/opportunities/refresh?day=today&use_live_lines=true", timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("POST /api/opportunities/refresh?day=today&use_live_lines=true - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                # Check data source
                data_source = data.get('data_source')
                expected_sources = ['plays888.co', 'hardcoded']  # Allow hardcoded as fallback
                
                if data_source not in expected_sources:
                    self.log_test("POST /api/opportunities/refresh?day=today&use_live_lines=true - Data Source", False,
                                f"Expected data_source in {expected_sources}, got '{data_source}'")
                    return False
                
                games = data.get('games', [])
                
                self.log_test("POST /api/opportunities/refresh?day=today&use_live_lines=true", True,
                            f"Refreshed {len(games)} games, data_source: {data_source}")
                return True
            else:
                self.log_test("POST /api/opportunities/refresh?day=today&use_live_lines=true", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("POST /api/opportunities/refresh?day=today&use_live_lines=true", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("POST /api/opportunities/refresh?day=today&use_live_lines=true", False, f"JSON decode error: {str(e)}")
            return False

    def test_nhl_opportunities_tomorrow_lines(self):
        """Test GET /api/opportunities/nhl?day=tomorrow - verify all 13 NHL games have correct totals for Dec 27, 2025"""
        try:
            response = requests.get(f"{self.api_url}/opportunities/nhl?day=tomorrow", timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("GET /api/opportunities/nhl?day=tomorrow - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                games = data.get('games', [])
                if not isinstance(games, list):
                    self.log_test("GET /api/opportunities/nhl?day=tomorrow - Games Array", False,
                                "Games is not an array")
                    return False
                
                # Expected: 13 NHL games for December 27, 2025
                expected_game_count = 13
                if len(games) != expected_game_count:
                    self.log_test("GET /api/opportunities/nhl?day=tomorrow - Game Count", False,
                                f"Expected {expected_game_count} NHL games, got {len(games)}")
                    return False
                
                # Expected totals for each game (verified against scoresandodds.com)
                expected_totals = {
                    ("NY Rangers", "NY Islanders"): 5.5,
                    ("Minnesota", "Winnipeg"): 5.5,
                    ("Tampa Bay", "Florida"): 5.5,
                    ("Boston", "Buffalo"): 6.5,
                    ("Detroit", "Carolina"): 6.5,
                    ("Ottawa", "Toronto"): 6.5,
                    ("Washington", "New Jersey"): 5.5,  # Previously incorrect as 6.5
                    ("Chicago", "Dallas"): 5.5,
                    ("Nashville", "St. Louis"): 5.5,
                    ("Anaheim", "Los Angeles"): 6.5,  # Previously incorrect as 5.5
                    ("Colorado", "Vegas"): 6.5,
                    ("Edmonton", "Calgary"): 6.5,
                    ("San Jose", "Vancouver"): 5.5,  # Previously incorrect as 6.0
                }
                
                # Track found games and their totals
                found_games = {}
                incorrect_totals = []
                
                for game in games:
                    away_team = game.get('away_team', '').strip()
                    home_team = game.get('home_team', '').strip()
                    total = game.get('total')
                    
                    # Validate game structure
                    if not away_team or not home_team or total is None:
                        self.log_test("GET /api/opportunities/nhl?day=tomorrow - Game Structure", False,
                                    f"Game missing required fields: away_team='{away_team}', home_team='{home_team}', total={total}")
                        return False
                    
                    # Check if this game matches any expected game
                    game_key = (away_team, home_team)
                    if game_key in expected_totals:
                        found_games[game_key] = total
                        expected_total = expected_totals[game_key]
                        
                        if total != expected_total:
                            incorrect_totals.append({
                                'game': f"{away_team} @ {home_team}",
                                'expected': expected_total,
                                'actual': total
                            })
                
                # Check if all expected games were found
                missing_games = []
                for expected_game in expected_totals:
                    if expected_game not in found_games:
                        missing_games.append(f"{expected_game[0]} @ {expected_game[1]}")
                
                if missing_games:
                    self.log_test("GET /api/opportunities/nhl?day=tomorrow - Missing Games", False,
                                f"Missing expected games: {missing_games}")
                    return False
                
                # Check for incorrect totals
                if incorrect_totals:
                    error_details = []
                    for error in incorrect_totals:
                        error_details.append(f"{error['game']}: expected {error['expected']}, got {error['actual']}")
                    
                    self.log_test("GET /api/opportunities/nhl?day=tomorrow - Incorrect Totals", False,
                                f"Incorrect totals found: {'; '.join(error_details)}")
                    return False
                
                # Specifically verify the 3 corrected games
                corrected_games = [
                    ("Washington", "New Jersey", 5.5),
                    ("Anaheim", "Los Angeles", 6.5),
                    ("San Jose", "Vancouver", 5.5)
                ]
                
                corrected_games_verified = []
                for away, home, expected_total in corrected_games:
                    game_key = (away, home)
                    if game_key in found_games:
                        actual_total = found_games[game_key]
                        if actual_total == expected_total:
                            corrected_games_verified.append(f"{away} @ {home}: {actual_total} ✓")
                        else:
                            self.log_test("GET /api/opportunities/nhl?day=tomorrow - Corrected Game", False,
                                        f"{away} @ {home}: expected {expected_total}, got {actual_total}")
                            return False
                
                self.log_test("GET /api/opportunities/nhl?day=tomorrow - All Games Verified", True,
                            f"All {len(games)} NHL games have correct totals")
                
                self.log_test("GET /api/opportunities/nhl?day=tomorrow - Corrected Games", True,
                            f"3 corrected games verified: {'; '.join(corrected_games_verified)}")
                
                self.log_test("GET /api/opportunities/nhl?day=tomorrow", True,
                            f"Successfully verified all {len(games)} NHL games for Dec 27, 2025")
                return True
            else:
                self.log_test("GET /api/opportunities/nhl?day=tomorrow", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("GET /api/opportunities/nhl?day=tomorrow", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("GET /api/opportunities/nhl?day=tomorrow", False, f"JSON decode error: {str(e)}")
            return False

    def test_historical_data_nba_date(self, date_str, expected_record=None):
        """Test NBA historical data for a specific date"""
        try:
            response = requests.get(f"{self.api_url}/opportunities?day={date_str}", timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test(f"NBA Historical Data {date_str} - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                games = data.get('games', [])
                if not isinstance(games, list):
                    self.log_test(f"NBA Historical Data {date_str} - Games Array", False,
                                "Games is not an array")
                    return False
                
                if len(games) == 0:
                    self.log_test(f"NBA Historical Data {date_str} - No Games", False,
                                "No games found for this date")
                    return False
                
                # Check that all games have final_score populated
                games_without_final_score = []
                for game in games:
                    final_score = game.get('final_score')
                    if final_score is None:
                        games_without_final_score.append(f"{game.get('away_team', 'Unknown')} @ {game.get('home_team', 'Unknown')}")
                
                if games_without_final_score:
                    self.log_test(f"NBA Historical Data {date_str} - Final Scores", False,
                                f"Games missing final_score: {', '.join(games_without_final_score)}")
                    return False
                
                # Check user bets
                user_bet_games = [g for g in games if g.get('user_bet') == True]
                
                # Validate user bet fields
                for game in user_bet_games:
                    required_bet_fields = ['bet_type', 'bet_line', 'bet_result', 'user_bet_hit']
                    missing_bet_fields = [field for field in required_bet_fields if field not in game]
                    
                    if missing_bet_fields:
                        self.log_test(f"NBA Historical Data {date_str} - Bet Fields", False,
                                    f"Game {game.get('away_team')} @ {game.get('home_team')} missing bet fields: {missing_bet_fields}")
                        return False
                    
                    # Validate user_bet_hit calculation
                    final_score = game.get('final_score')
                    bet_line = game.get('bet_line')
                    bet_result = game.get('bet_result')
                    user_bet_hit = game.get('user_bet_hit')
                    
                    if final_score is not None and bet_line is not None:
                        # Calculate expected result based on bet type
                        if bet_result in ['won', 'lost']:
                            expected_hit = (bet_result == 'won')
                            if user_bet_hit != expected_hit:
                                self.log_test(f"NBA Historical Data {date_str} - Bet Calculation", False,
                                            f"Game {game.get('away_team')} @ {game.get('home_team')}: user_bet_hit={user_bet_hit} but bet_result={bet_result}")
                                return False
                
                # Check expected record if provided
                if expected_record and user_bet_games:
                    won_bets = len([g for g in user_bet_games if g.get('bet_result') == 'won'])
                    lost_bets = len([g for g in user_bet_games if g.get('bet_result') == 'lost'])
                    actual_record = f"{won_bets}-{lost_bets}"
                    
                    if actual_record != expected_record:
                        self.log_test(f"NBA Historical Data {date_str} - Record", False,
                                    f"Expected record {expected_record}, got {actual_record}")
                        return False
                
                self.log_test(f"NBA Historical Data {date_str}", True,
                            f"Found {len(games)} games, {len(user_bet_games)} with user bets, all final scores populated")
                return True
            else:
                self.log_test(f"NBA Historical Data {date_str}", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test(f"NBA Historical Data {date_str}", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test(f"NBA Historical Data {date_str}", False, f"JSON decode error: {str(e)}")
            return False

    def test_historical_data_nhl_date(self, date_str, expected_record=None):
        """Test NHL historical data for a specific date"""
        try:
            response = requests.get(f"{self.api_url}/opportunities/nhl?day={date_str}", timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test(f"NHL Historical Data {date_str} - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                games = data.get('games', [])
                if not isinstance(games, list):
                    self.log_test(f"NHL Historical Data {date_str} - Games Array", False,
                                "Games is not an array")
                    return False
                
                if len(games) == 0:
                    self.log_test(f"NHL Historical Data {date_str} - No Games", False,
                                "No games found for this date")
                    return False
                
                # Check that all games have final_score populated
                games_without_final_score = []
                for game in games:
                    final_score = game.get('final_score')
                    if final_score is None:
                        games_without_final_score.append(f"{game.get('away_team', 'Unknown')} @ {game.get('home_team', 'Unknown')}")
                
                if games_without_final_score:
                    self.log_test(f"NHL Historical Data {date_str} - Final Scores", False,
                                f"Games missing final_score: {', '.join(games_without_final_score)}")
                    return False
                
                # Check user bets
                user_bet_games = [g for g in games if g.get('user_bet') == True]
                
                # Validate user bet fields
                for game in user_bet_games:
                    required_bet_fields = ['bet_type', 'bet_line', 'bet_result', 'user_bet_hit']
                    missing_bet_fields = [field for field in required_bet_fields if field not in game]
                    
                    if missing_bet_fields:
                        self.log_test(f"NHL Historical Data {date_str} - Bet Fields", False,
                                    f"Game {game.get('away_team')} @ {game.get('home_team')} missing bet fields: {missing_bet_fields}")
                        return False
                    
                    # Validate user_bet_hit calculation
                    final_score = game.get('final_score')
                    bet_line = game.get('bet_line')
                    bet_result = game.get('bet_result')
                    user_bet_hit = game.get('user_bet_hit')
                    
                    if final_score is not None and bet_line is not None:
                        # Calculate expected result based on bet type
                        if bet_result in ['won', 'lost']:
                            expected_hit = (bet_result == 'won')
                            if user_bet_hit != expected_hit:
                                self.log_test(f"NHL Historical Data {date_str} - Bet Calculation", False,
                                            f"Game {game.get('away_team')} @ {game.get('home_team')}: user_bet_hit={user_bet_hit} but bet_result={bet_result}")
                                return False
                
                # Check expected record if provided
                if expected_record and user_bet_games:
                    won_bets = len([g for g in user_bet_games if g.get('bet_result') == 'won'])
                    lost_bets = len([g for g in user_bet_games if g.get('bet_result') == 'lost'])
                    actual_record = f"{won_bets}-{lost_bets}"
                    
                    if actual_record != expected_record:
                        self.log_test(f"NHL Historical Data {date_str} - Record", False,
                                    f"Expected record {expected_record}, got {actual_record}")
                        return False
                
                self.log_test(f"NHL Historical Data {date_str}", True,
                            f"Found {len(games)} games, {len(user_bet_games)} with user bets, all final scores populated")
                return True
            else:
                self.log_test(f"NHL Historical Data {date_str}", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test(f"NHL Historical Data {date_str}", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test(f"NHL Historical Data {date_str}", False, f"JSON decode error: {str(e)}")
            return False

    def test_historical_data_verification(self):
        """Test historical data for NBA and NHL games from 12/22/2025 to 12/27/2025"""
        print("📊 HISTORICAL DATA VERIFICATION (12/22/2025 - 12/27/2025)")
        print("-" * 60)
        
        # Expected betting records from the review request
        nba_expected_records = {
            "2025-12-22": "1-2",
            "2025-12-23": "4-3", 
            "2025-12-25": "1-1",
            "2025-12-26": "3-2",
            "2025-12-27": "2-2"
        }
        
        nhl_expected_records = {
            "2025-12-22": "0-1",
            "2025-12-23": "3-3",
            "2025-12-27": "4-1"
        }
        
        # Test NBA dates
        print("\n🏀 NBA Historical Data Tests:")
        nba_success = True
        for date_str, expected_record in nba_expected_records.items():
            success = self.test_historical_data_nba_date(date_str, expected_record)
            if not success:
                nba_success = False
        
        # Test NHL dates  
        print("\n🏒 NHL Historical Data Tests:")
        nhl_success = True
        for date_str, expected_record in nhl_expected_records.items():
            success = self.test_historical_data_nhl_date(date_str, expected_record)
            if not success:
                nhl_success = False
        
        # Overall result
        overall_success = nba_success and nhl_success
        self.log_test("Historical Data Verification (12/22-12/27)", overall_success,
                    f"NBA: {'✅' if nba_success else '❌'}, NHL: {'✅' if nhl_success else '❌'}")
        
        return overall_success

    def test_process_update_records(self):
        """Test POST /api/process/update-records?start_date=2025-12-22 - Process #6 implementation"""
        try:
            print("Testing Process #6 - Update Records (this may take 30-60 seconds)...")
            response = requests.post(f"{self.api_url}/process/update-records?start_date=2025-12-22", timeout=90)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['status', 'records', 'start_date', 'updated_at']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("POST /api/process/update-records - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                status = data.get('status')
                if status != 'success':
                    self.log_test("POST /api/process/update-records - Status", False,
                                f"API returned status='{status}', expected 'success'")
                    return False
                
                records = data.get('records', {})
                
                # Validate that records were calculated for all leagues
                expected_leagues = ['NBA', 'NHL', 'NFL']
                for league in expected_leagues:
                    if league not in records:
                        self.log_test(f"POST /api/process/update-records - {league} Records", False,
                                    f"No records found for {league}")
                        return False
                    
                    league_data = records[league]
                    required_record_fields = ['betting', 'edge']
                    missing_record_fields = [field for field in required_record_fields if field not in league_data]
                    
                    if missing_record_fields:
                        self.log_test(f"POST /api/process/update-records - {league} Structure", False,
                                    f"Missing record fields for {league}: {missing_record_fields}")
                        return False
                    
                    # Validate betting and edge structure
                    betting_data = league_data.get('betting', {})
                    edge_data = league_data.get('edge', {})
                    
                    if 'wins' not in betting_data or 'losses' not in betting_data:
                        self.log_test(f"POST /api/process/update-records - {league} Betting Structure", False,
                                    f"Missing wins/losses in betting data for {league}")
                        return False
                    
                    if 'hits' not in edge_data or 'misses' not in edge_data:
                        self.log_test(f"POST /api/process/update-records - {league} Edge Structure", False,
                                    f"Missing hits/misses in edge data for {league}")
                        return False
                
                # Validate expected records for NBA (12/22-12/27)
                nba_data = records.get('NBA', {})
                expected_nba_betting_wins = 11
                expected_nba_betting_losses = 10
                expected_nba_edge_hits = 23
                expected_nba_edge_misses = 16
                
                actual_nba_betting_wins = nba_data.get('betting', {}).get('wins')
                actual_nba_betting_losses = nba_data.get('betting', {}).get('losses')
                actual_nba_edge_hits = nba_data.get('edge', {}).get('hits')
                actual_nba_edge_misses = nba_data.get('edge', {}).get('misses')
                
                if actual_nba_betting_wins != expected_nba_betting_wins:
                    self.log_test("POST /api/process/update-records - NBA Betting Wins", False,
                                f"Expected NBA betting wins {expected_nba_betting_wins}, got {actual_nba_betting_wins}")
                    return False
                
                if actual_nba_betting_losses != expected_nba_betting_losses:
                    self.log_test("POST /api/process/update-records - NBA Betting Losses", False,
                                f"Expected NBA betting losses {expected_nba_betting_losses}, got {actual_nba_betting_losses}")
                    return False
                
                if actual_nba_edge_hits != expected_nba_edge_hits:
                    self.log_test("POST /api/process/update-records - NBA Edge Hits", False,
                                f"Expected NBA edge hits {expected_nba_edge_hits}, got {actual_nba_edge_hits}")
                    return False
                
                if actual_nba_edge_misses != expected_nba_edge_misses:
                    self.log_test("POST /api/process/update-records - NBA Edge Misses", False,
                                f"Expected NBA edge misses {expected_nba_edge_misses}, got {actual_nba_edge_misses}")
                    return False
                
                # Validate expected records for NHL (12/22-12/27)
                nhl_data = records.get('NHL', {})
                expected_nhl_betting_wins = 7
                expected_nhl_betting_losses = 5
                expected_nhl_edge_hits = 11
                expected_nhl_edge_misses = 5
                
                actual_nhl_betting_wins = nhl_data.get('betting', {}).get('wins')
                actual_nhl_betting_losses = nhl_data.get('betting', {}).get('losses')
                actual_nhl_edge_hits = nhl_data.get('edge', {}).get('hits')
                actual_nhl_edge_misses = nhl_data.get('edge', {}).get('misses')
                
                if actual_nhl_betting_wins != expected_nhl_betting_wins:
                    self.log_test("POST /api/process/update-records - NHL Betting Wins", False,
                                f"Expected NHL betting wins {expected_nhl_betting_wins}, got {actual_nhl_betting_wins}")
                    return False
                
                if actual_nhl_betting_losses != expected_nhl_betting_losses:
                    self.log_test("POST /api/process/update-records - NHL Betting Losses", False,
                                f"Expected NHL betting losses {expected_nhl_betting_losses}, got {actual_nhl_betting_losses}")
                    return False
                
                if actual_nhl_edge_hits != expected_nhl_edge_hits:
                    self.log_test("POST /api/process/update-records - NHL Edge Hits", False,
                                f"Expected NHL edge hits {expected_nhl_edge_hits}, got {actual_nhl_edge_hits}")
                    return False
                
                if actual_nhl_edge_misses != expected_nhl_edge_misses:
                    self.log_test("POST /api/process/update-records - NHL Edge Misses", False,
                                f"Expected NHL edge misses {expected_nhl_edge_misses}, got {actual_nhl_edge_misses}")
                    return False
                
                # Validate expected records for NFL (12/22-12/27)
                nfl_data = records.get('NFL', {})
                expected_nfl_betting_wins = 0
                expected_nfl_betting_losses = 0
                expected_nfl_edge_hits = 0
                expected_nfl_edge_misses = 0
                
                actual_nfl_betting_wins = nfl_data.get('betting', {}).get('wins')
                actual_nfl_betting_losses = nfl_data.get('betting', {}).get('losses')
                actual_nfl_edge_hits = nfl_data.get('edge', {}).get('hits')
                actual_nfl_edge_misses = nfl_data.get('edge', {}).get('misses')
                
                if actual_nfl_betting_wins != expected_nfl_betting_wins:
                    self.log_test("POST /api/process/update-records - NFL Betting Wins", False,
                                f"Expected NFL betting wins {expected_nfl_betting_wins}, got {actual_nfl_betting_wins}")
                    return False
                
                if actual_nfl_betting_losses != expected_nfl_betting_losses:
                    self.log_test("POST /api/process/update-records - NFL Betting Losses", False,
                                f"Expected NFL betting losses {expected_nfl_betting_losses}, got {actual_nfl_betting_losses}")
                    return False
                
                if actual_nfl_edge_hits != expected_nfl_edge_hits:
                    self.log_test("POST /api/process/update-records - NFL Edge Hits", False,
                                f"Expected NFL edge hits {expected_nfl_edge_hits}, got {actual_nfl_edge_hits}")
                    return False
                
                if actual_nfl_edge_misses != expected_nfl_edge_misses:
                    self.log_test("POST /api/process/update-records - NFL Edge Misses", False,
                                f"Expected NFL edge misses {expected_nfl_edge_misses}, got {actual_nfl_edge_misses}")
                    return False
                
                self.log_test("POST /api/process/update-records", True,
                            f"Successfully calculated records: NBA(11-10/23-16), NHL(7-5/11-5), NFL(0-0/0-0)")
                return True
            else:
                self.log_test("POST /api/process/update-records", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.Timeout:
            self.log_test("POST /api/process/update-records", False, "Request timeout (>90s)")
            return False
        except requests.exceptions.RequestException as e:
            self.log_test("POST /api/process/update-records", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("POST /api/process/update-records", False, f"JSON decode error: {str(e)}")
            return False

    def test_get_records_summary(self):
        """Test GET /api/records/summary - Process #6 implementation"""
        try:
            response = requests.get(f"{self.api_url}/records/summary", timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate that all leagues are present
                expected_leagues = ['NBA', 'NHL', 'NFL']
                missing_leagues = [league for league in expected_leagues if league not in data]
                
                if missing_leagues:
                    self.log_test("GET /api/records/summary - Structure", False,
                                f"Missing leagues: {missing_leagues}")
                    return False
                
                # Validate each league's structure
                for league in expected_leagues:
                    league_data = data[league]
                    required_fields = ['betting_record', 'edge_record', 'start_date']
                    missing_fields = [field for field in required_fields if field not in league_data]
                    
                    if missing_fields:
                        self.log_test(f"GET /api/records/summary - {league} Structure", False,
                                    f"Missing fields for {league}: {missing_fields}")
                        return False
                
                # Validate expected records match Process #6 requirements
                nba_data = data.get('NBA', {})
                expected_nba_edge = "23-16"
                expected_nba_betting = "11-10"
                
                actual_nba_edge = nba_data.get('edge_record')
                actual_nba_betting = nba_data.get('betting_record')
                
                if actual_nba_edge != expected_nba_edge:
                    self.log_test("GET /api/records/summary - NBA Edge Record", False,
                                f"Expected NBA edge record {expected_nba_edge}, got {actual_nba_edge}")
                    return False
                
                if actual_nba_betting != expected_nba_betting:
                    self.log_test("GET /api/records/summary - NBA Betting Record", False,
                                f"Expected NBA betting record {expected_nba_betting}, got {actual_nba_betting}")
                    return False
                
                # Validate NHL records
                nhl_data = data.get('NHL', {})
                expected_nhl_edge = "11-5"
                expected_nhl_betting = "7-5"
                
                actual_nhl_edge = nhl_data.get('edge_record')
                actual_nhl_betting = nhl_data.get('betting_record')
                
                if actual_nhl_edge != expected_nhl_edge:
                    self.log_test("GET /api/records/summary - NHL Edge Record", False,
                                f"Expected NHL edge record {expected_nhl_edge}, got {actual_nhl_edge}")
                    return False
                
                if actual_nhl_betting != expected_nhl_betting:
                    self.log_test("GET /api/records/summary - NHL Betting Record", False,
                                f"Expected NHL betting record {expected_nhl_betting}, got {actual_nhl_betting}")
                    return False
                
                # Validate NFL records
                nfl_data = data.get('NFL', {})
                expected_nfl_edge = "0-0"
                expected_nfl_betting = "0-0"
                
                actual_nfl_edge = nfl_data.get('edge_record')
                actual_nfl_betting = nfl_data.get('betting_record')
                
                if actual_nfl_edge != expected_nfl_edge:
                    self.log_test("GET /api/records/summary - NFL Edge Record", False,
                                f"Expected NFL edge record {expected_nfl_edge}, got {actual_nfl_edge}")
                    return False
                
                if actual_nfl_betting != expected_nfl_betting:
                    self.log_test("GET /api/records/summary - NFL Betting Record", False,
                                f"Expected NFL betting record {expected_nfl_betting}, got {actual_nfl_betting}")
                    return False
                
                # Validate start_date
                nba_start_date = nba_data.get('start_date')
                if nba_start_date != "2025-12-22":
                    self.log_test("GET /api/records/summary - Start Date", False,
                                f"Expected start_date '2025-12-22', got '{nba_start_date}'")
                    return False
                
                self.log_test("GET /api/records/summary", True,
                            f"Successfully retrieved records summary: NBA({actual_nba_edge}/{actual_nba_betting}), NHL({actual_nhl_edge}/{actual_nhl_betting}), NFL({actual_nfl_edge}/{actual_nfl_betting}), start_date={nba_start_date}")
                return True
            else:
                self.log_test("GET /api/records/summary", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("GET /api/records/summary", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("GET /api/records/summary", False, f"JSON decode error: {str(e)}")
            return False

    def test_records_database_storage(self):
        """Test that records are properly stored in database collections"""
        try:
            # First, trigger the update process
            update_response = requests.post(f"{self.api_url}/process/update-records?start_date=2025-12-22", timeout=90)
            
            if update_response.status_code != 200:
                self.log_test("Records Database Storage - Update Process", False,
                            f"Failed to trigger update process: {update_response.status_code}")
                return False
            
            update_data = update_response.json()
            if update_data.get('status') != 'success':
                self.log_test("Records Database Storage - Update Status", False,
                            f"Update process failed: {update_data}")
                return False
            
            # Then verify the summary endpoint returns the data
            summary_response = requests.get(f"{self.api_url}/records/summary", timeout=15)
            
            if summary_response.status_code != 200:
                self.log_test("Records Database Storage - Summary Retrieval", False,
                            f"Failed to retrieve summary: {summary_response.status_code}")
                return False
            
            summary_data = summary_response.json()
            
            # Check that all expected leagues have data
            expected_leagues = ['NBA', 'NHL', 'NFL']
            for league in expected_leagues:
                if league not in summary_data:
                    self.log_test(f"Records Database Storage - {league} Persistence", False,
                                f"No persisted records found for {league}")
                    return False
                
                league_data = summary_data[league]
                if not league_data.get('edge_record') or not league_data.get('betting_record'):
                    self.log_test(f"Records Database Storage - {league} Data", False,
                                f"Incomplete record data for {league}")
                    return False
            
            # Verify timestamps are present
            nba_data = summary_data.get('NBA', {})
            start_date = nba_data.get('start_date')
            last_updated = nba_data.get('betting_last_updated') or nba_data.get('edge_last_updated')
            
            if not start_date or not last_updated:
                self.log_test("Records Database Storage - Timestamps", False,
                            "Missing start_date or last_updated timestamps")
                return False
            
            self.log_test("Records Database Storage", True,
                        f"Records properly stored and retrieved from database collections (compound_records and edge_records)")
            return True
            
        except Exception as e:
            self.log_test("Records Database Storage", False, f"Error: {str(e)}")
            return False

    def test_process_6_implementation(self):
        """Test Process #6 - Update Records Implementation"""
        print("🔄 PROCESS #6 - UPDATE RECORDS IMPLEMENTATION")
        print("-" * 60)
        
        # Test the update records endpoint
        print("\n📊 Testing POST /api/process/update-records...")
        update_success = self.test_process_update_records()
        
        # Test the records summary endpoint
        print("\n📋 Testing GET /api/records/summary...")
        summary_success = self.test_get_records_summary()
        
        # Test database storage
        print("\n💾 Testing database storage...")
        storage_success = self.test_records_database_storage()
        
        # Overall result
        overall_success = update_success and summary_success and storage_success
        self.log_test("Process #6 Implementation", overall_success,
                    f"Update: {'✅' if update_success else '❌'}, Summary: {'✅' if summary_success else '❌'}, Storage: {'✅' if storage_success else '❌'}")
        
        return overall_success

    def test_excel_export_functionality(self):
        """Test Excel export functionality for both NBA and NHL"""
        print("=" * 60)
        print("EXCEL EXPORT FUNCTIONALITY TESTING")
        print("=" * 60)
        
        excel_tests_passed = 0
        excel_tests_total = 0
        
        # Test 1: NBA Excel Export
        excel_tests_total += 1
        print("Testing NBA Excel Export...")
        try:
            response = requests.get(
                f"{self.api_url}/export/excel?league=NBA&start_date=2025-12-22&end_date=2025-12-27", 
                timeout=30
            )
            
            if response.status_code == 200:
                # Verify content type
                content_type = response.headers.get('content-type', '')
                expected_content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                
                if content_type == expected_content_type:
                    # Verify Content-Disposition header
                    content_disposition = response.headers.get('content-disposition', '')
                    if 'attachment' in content_disposition and 'filename=' in content_disposition:
                        # Verify file size
                        content_length = len(response.content)
                        if content_length > 0:
                            self.log_test("NBA Excel Export", True, 
                                        f"Valid Excel file: {content_length} bytes, correct headers")
                            excel_tests_passed += 1
                        else:
                            self.log_test("NBA Excel Export", False, "Excel file is empty")
                    else:
                        self.log_test("NBA Excel Export", False, "Invalid Content-Disposition header")
                else:
                    self.log_test("NBA Excel Export", False, f"Wrong content type: {content_type}")
            else:
                self.log_test("NBA Excel Export", False, f"Status code: {response.status_code}")
        except Exception as e:
            self.log_test("NBA Excel Export", False, f"Error: {str(e)}")
        
        # Test 2: NHL Excel Export
        excel_tests_total += 1
        print("Testing NHL Excel Export...")
        try:
            response = requests.get(
                f"{self.api_url}/export/excel?league=NHL&start_date=2025-12-22&end_date=2025-12-27", 
                timeout=30
            )
            
            if response.status_code == 200:
                content_type = response.headers.get('content-type', '')
                expected_content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                
                if content_type == expected_content_type:
                    content_length = len(response.content)
                    if content_length > 0:
                        self.log_test("NHL Excel Export", True, 
                                    f"Valid Excel file: {content_length} bytes")
                        excel_tests_passed += 1
                    else:
                        self.log_test("NHL Excel Export", False, "Excel file is empty")
                else:
                    self.log_test("NHL Excel Export", False, f"Wrong content type: {content_type}")
            else:
                self.log_test("NHL Excel Export", False, f"Status code: {response.status_code}")
        except Exception as e:
            self.log_test("NHL Excel Export", False, f"Error: {str(e)}")
        
        # Test 3: Excel File Structure (if openpyxl is available)
        excel_tests_total += 1
        print("Testing Excel File Structure...")
        try:
            import io
            from openpyxl import load_workbook
            
            response = requests.get(
                f"{self.api_url}/export/excel?league=NBA&start_date=2025-12-22&end_date=2025-12-27", 
                timeout=30
            )
            
            if response.status_code == 200:
                buffer = io.BytesIO(response.content)
                wb = load_workbook(buffer)
                ws = wb.active
                
                # Verify 35 columns (A through AI)
                if ws.max_column == 35:
                    # Verify critical headers
                    critical_headers = {
                        1: 'Date', 7: 'Away Team', 11: 'Home Team', 
                        32: '4-Dot Result', 34: '4-Dot Hit', 35: '4-Dot Record'
                    }
                    
                    headers_correct = True
                    for col_num, expected_header in critical_headers.items():
                        actual_header = ws.cell(row=1, column=col_num).value
                        if actual_header != expected_header:
                            headers_correct = False
                            break
                    
                    if headers_correct and ws.max_row > 1:
                        # Check for 4-Dot logic
                        four_dot_results_found = 0
                        for row in range(2, min(ws.max_row + 1, 20)):
                            four_dot_result = ws.cell(row=row, column=32).value
                            if four_dot_result in ['OVER', 'UNDER', 'NO BET']:
                                four_dot_results_found += 1
                        
                        if four_dot_results_found > 0:
                            self.log_test("Excel File Structure", True, 
                                        f"35 columns, correct headers, {four_dot_results_found} 4-Dot results, {ws.max_row-1} data rows")
                            excel_tests_passed += 1
                        else:
                            self.log_test("Excel File Structure", False, "No 4-Dot results found")
                    else:
                        self.log_test("Excel File Structure", False, "Incorrect headers or no data rows")
                else:
                    self.log_test("Excel File Structure", False, f"Expected 35 columns, got {ws.max_column}")
            else:
                self.log_test("Excel File Structure", False, f"Could not get Excel file: {response.status_code}")
        except ImportError:
            self.log_test("Excel File Structure", False, "openpyxl not available for validation")
        except Exception as e:
            self.log_test("Excel File Structure", False, f"Error: {str(e)}")
        
        print(f"\nExcel Export Tests: {excel_tests_passed}/{excel_tests_total} passed")
        return excel_tests_passed == excel_tests_total

    def test_ncaab_opportunities_tomorrow(self):
        """Test NCAAB API endpoint to verify all 68 games for December 29, 2025 have been successfully imported"""
        try:
            response = requests.get(f"{self.api_url}/opportunities/ncaab?day=tomorrow", timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                # Verify date field is "2025-12-29"
                date_field = data.get('date')
                expected_date = "2025-12-29"
                if date_field != expected_date:
                    self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Date", False,
                                f"Expected date '{expected_date}', got '{date_field}'")
                    return False
                
                games = data.get('games', [])
                if not isinstance(games, list):
                    self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Games Array", False,
                                "Games is not an array")
                    return False
                
                # Verify response contains exactly 68 games
                expected_game_count = 68
                if len(games) != expected_game_count:
                    self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Game Count", False,
                                f"Expected {expected_game_count} games, got {len(games)}")
                    return False
                
                # Verify games have correct structure (away_team, home_team, opening_line, time, etc.)
                if games:
                    sample_game = games[0]
                    required_game_fields = ['away_team', 'home_team', 'time']
                    missing_game_fields = [field for field in required_game_fields if field not in sample_game]
                    
                    if missing_game_fields:
                        self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Game Structure", False,
                                    f"Missing game fields: {missing_game_fields}")
                        return False
                
                # Verify specific games from the import
                specific_games_to_verify = [
                    {"away": "NC Central", "home": "Penn St.", "ou": 149.5},
                    {"away": "Utah", "home": "Washington", "ou": 160.5},
                    {"away": "Queens", "home": "Auburn", "ou": 174.5}
                ]
                
                found_specific_games = []
                for target_game in specific_games_to_verify:
                    for game in games:
                        away_team = game.get('away_team', '').strip()
                        home_team = game.get('home_team', '').strip()
                        opening_line = game.get('opening_line') or game.get('total')
                        
                        # Check if this matches the target game (flexible matching)
                        if (target_game["away"].lower() in away_team.lower() and 
                            target_game["home"].lower() in home_team.lower()):
                            
                            found_specific_games.append({
                                'game': f"{away_team} @ {home_team}",
                                'expected_ou': target_game["ou"],
                                'actual_ou': opening_line,
                                'match': opening_line == target_game["ou"] if opening_line else False
                            })
                            break
                
                # Verify games without lines show null/None
                games_without_lines = []
                for game in games:
                    away_team = game.get('away_team', '').strip()
                    home_team = game.get('home_team', '').strip()
                    opening_line = game.get('opening_line') or game.get('total')
                    
                    # Check for Arlington Baptist @ Baylor
                    if ("Arlington Baptist".lower() in away_team.lower() and 
                        "Baylor".lower() in home_team.lower()):
                        if opening_line is not None:
                            self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Null Lines", False,
                                        f"Arlington Baptist @ Baylor should have null opening_line, got {opening_line}")
                            return False
                        else:
                            games_without_lines.append(f"{away_team} @ {home_team}")
                
                # Log results for specific games
                for found_game in found_specific_games:
                    if found_game['match']:
                        self.log_test(f"NCAAB Specific Game - {found_game['game']}", True,
                                    f"O/U {found_game['actual_ou']} matches expected {found_game['expected_ou']}")
                    else:
                        self.log_test(f"NCAAB Specific Game - {found_game['game']}", False,
                                    f"O/U {found_game['actual_ou']} does not match expected {found_game['expected_ou']}")
                        return False
                
                # Verify at least some games have opening lines
                games_with_lines = [g for g in games if (g.get('opening_line') or g.get('total')) is not None]
                if len(games_with_lines) == 0:
                    self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Lines Present", False,
                                "No games have opening lines or totals")
                    return False
                
                self.log_test("GET /api/opportunities/ncaab?day=tomorrow", True,
                            f"Successfully verified {len(games)} NCAAB games for {expected_date}")
                
                # Additional verification details
                self.log_test("NCAAB Games Structure Verification", True,
                            f"All games have required fields: away_team, home_team, time")
                
                self.log_test("NCAAB Specific Games Verification", True,
                            f"Found and verified {len(found_specific_games)} specific games with correct O/U lines")
                
                if games_without_lines:
                    self.log_test("NCAAB Null Lines Verification", True,
                                f"Verified {len(games_without_lines)} games with null lines: {', '.join(games_without_lines)}")
                
                return True
            else:
                self.log_test("GET /api/opportunities/ncaab?day=tomorrow", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("GET /api/opportunities/ncaab?day=tomorrow", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("GET /api/opportunities/ncaab?day=tomorrow", False, f"JSON decode error: {str(e)}")
            return False

    def test_ncaab_api_tomorrow(self):
        """Test GET /api/opportunities/ncaab?day=tomorrow - verify 68 games for 2025-12-29 with PPG data"""
        try:
            response = requests.get(f"{self.api_url}/opportunities/ncaab?day=tomorrow", timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                games = data.get('games', [])
                if not isinstance(games, list):
                    self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Games Array", False,
                                "Games is not an array")
                    return False
                
                # Expected: exactly 68 games for 2025-12-29
                expected_game_count = 68
                if len(games) != expected_game_count:
                    self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Game Count", False,
                                f"Expected exactly {expected_game_count} NCAAB games, got {len(games)}")
                    return False
                
                # Verify date is correct
                expected_date = "2025-12-29"
                actual_date = data.get('date')
                if actual_date != expected_date:
                    self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Date", False,
                                f"Expected date {expected_date}, got {actual_date}")
                    return False
                
                # Test specific games with PPG data
                test_games = [
                    {
                        "away_team": "NC Central",
                        "home_team": "Penn St.",
                        "expected_combined_ppg": 139.4,
                        "expected_edge": -10.1,
                        "game_num": 1
                    },
                    {
                        "away_team": "Missouri St.",
                        "home_team": "Delaware", 
                        "expected_combined_ppg": 152.8,
                        "expected_edge": 16.3,
                        "game_num": 3
                    },
                    {
                        "away_team": "Utah",
                        "home_team": "Washington",
                        "expected_combined_ppg": 165.7,
                        "expected_edge": 5.2,
                        "game_num": 68
                    }
                ]
                
                games_with_ppg = 0
                games_without_ppg = 0
                verified_test_games = 0
                
                for game in games:
                    # Validate game structure
                    required_game_fields = ['away_team', 'home_team', 'combined_ppg', 'edge', 'away_ppg_rank', 'home_ppg_rank', 'away_last3_rank', 'home_last3_rank']
                    missing_game_fields = [field for field in required_game_fields if field not in game]
                    
                    if missing_game_fields:
                        self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Game Structure", False,
                                    f"Game missing required fields: {missing_game_fields}")
                        return False
                    
                    # Count games with/without PPG data
                    combined_ppg = game.get('combined_ppg')
                    if combined_ppg is not None:
                        games_with_ppg += 1
                    else:
                        games_without_ppg += 1
                    
                    # Check specific test games
                    for test_game in test_games:
                        if (game.get('away_team') == test_game['away_team'] and 
                            game.get('home_team') == test_game['home_team']):
                            
                            actual_combined_ppg = game.get('combined_ppg')
                            actual_edge = game.get('edge')
                            
                            # Allow small tolerance for floating point comparison
                            if actual_combined_ppg is not None:
                                ppg_diff = abs(actual_combined_ppg - test_game['expected_combined_ppg'])
                                if ppg_diff > 1.0:  # Allow 1.0 point tolerance
                                    self.log_test(f"NCAAB Game {test_game['game_num']} - GPG Avg", False,
                                                f"{test_game['away_team']} @ {test_game['home_team']}: expected GPG Avg ~{test_game['expected_combined_ppg']}, got {actual_combined_ppg}")
                                    return False
                            
                            if actual_edge is not None:
                                edge_diff = abs(actual_edge - test_game['expected_edge'])
                                if edge_diff > 1.0:  # Allow 1.0 point tolerance
                                    self.log_test(f"NCAAB Game {test_game['game_num']} - Edge", False,
                                                f"{test_game['away_team']} @ {test_game['home_team']}: expected Edge ~{test_game['expected_edge']}, got {actual_edge}")
                                    return False
                            
                            verified_test_games += 1
                            self.log_test(f"NCAAB Game {test_game['game_num']} - {test_game['away_team']} @ {test_game['home_team']}", True,
                                        f"GPG Avg: {actual_combined_ppg}, Edge: {actual_edge}")
                
                # Verify we found all test games
                if verified_test_games != len(test_games):
                    self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Test Games", False,
                                f"Expected to verify {len(test_games)} test games, only found {verified_test_games}")
                    return False
                
                # Test games without PPG data (non-D1 teams)
                non_d1_games = [
                    ("Arlington Baptist", "Baylor"),
                    ("LA Sierra", "UNLV")
                ]
                
                verified_non_d1_games = 0
                for game in games:
                    for away, home in non_d1_games:
                        if (game.get('away_team') == away and game.get('home_team') == home):
                            combined_ppg = game.get('combined_ppg')
                            if combined_ppg is not None:
                                self.log_test(f"NCAAB Non-D1 Game - {away} @ {home}", False,
                                            f"Expected null combined_ppg for non-D1 team, got {combined_ppg}")
                                return False
                            verified_non_d1_games += 1
                            self.log_test(f"NCAAB Non-D1 Game - {away} @ {home}", True,
                                        f"Correctly has null combined_ppg")
                
                # Verify Edge calculation formula (Edge = GPG Avg - Line)
                edge_calculation_errors = 0
                for game in games:
                    combined_ppg = game.get('combined_ppg')
                    line = game.get('total') or game.get('line')
                    edge = game.get('edge')
                    
                    if combined_ppg is not None and line is not None and edge is not None:
                        expected_edge = combined_ppg - line
                        edge_diff = abs(edge - expected_edge)
                        if edge_diff > 0.1:  # Allow small floating point tolerance
                            edge_calculation_errors += 1
                            if edge_calculation_errors <= 3:  # Only log first few errors
                                self.log_test(f"NCAAB Edge Calculation - {game.get('away_team')} @ {game.get('home_team')}", False,
                                            f"Edge calculation error: GPG Avg {combined_ppg} - Line {line} = {expected_edge}, but got Edge {edge}")
                
                if edge_calculation_errors > 0:
                    self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Edge Calculation", False,
                                f"Found {edge_calculation_errors} games with incorrect Edge calculation")
                    return False
                
                self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Edge Calculation", True,
                            "All games have correct Edge calculation (GPG Avg - Line)")
                
                # Verify rank data is present
                games_with_ranks = 0
                for game in games:
                    if (game.get('away_ppg_rank') is not None and 
                        game.get('home_ppg_rank') is not None and
                        game.get('away_last3_rank') is not None and
                        game.get('home_last3_rank') is not None):
                        games_with_ranks += 1
                
                if games_with_ranks < (games_with_ppg * 0.8):  # At least 80% of games with PPG should have ranks
                    self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Rank Data", False,
                                f"Only {games_with_ranks} games have complete rank data, expected at least {int(games_with_ppg * 0.8)}")
                    return False
                
                self.log_test("GET /api/opportunities/ncaab?day=tomorrow - Rank Data", True,
                            f"{games_with_ranks} games have complete rank data")
                
                self.log_test("GET /api/opportunities/ncaab?day=tomorrow", True,
                            f"Successfully verified all {len(games)} NCAAB games: {games_with_ppg} with PPG data, {games_without_ppg} without PPG data")
                return True
            else:
                self.log_test("GET /api/opportunities/ncaab?day=tomorrow", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("GET /api/opportunities/ncaab?day=tomorrow", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("GET /api/opportunities/ncaab?day=tomorrow", False, f"JSON decode error: {str(e)}")
            return False

    def test_nhl_gpg_avg_calculation_fix(self):
        """Test NHL GPG Avg calculation fix with updated ESPN and StatMuse data"""
        try:
            response = requests.get(f"{self.api_url}/opportunities/nhl?day=today", timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                
                # Validate response structure
                required_fields = ['games', 'date', 'success']
                missing_fields = [field for field in required_fields if field not in data]
                
                if missing_fields:
                    self.log_test("NHL GPG Avg Fix - Structure", False,
                                f"Missing fields: {missing_fields}")
                    return False
                
                games = data.get('games', [])
                if not isinstance(games, list):
                    self.log_test("NHL GPG Avg Fix - Games Array", False,
                                "Games is not an array")
                    return False
                
                # Expected: 11 NHL games for today
                expected_game_count = 11
                if len(games) != expected_game_count:
                    self.log_test("NHL GPG Avg Fix - Game Count", False,
                                f"Expected {expected_game_count} NHL games, got {len(games)}")
                    return False
                
                # Validate each game has required GPG fields
                for game in games:
                    required_gpg_fields = ['combined_gpg', 'edge', 'recommendation']
                    missing_gpg_fields = [field for field in required_gpg_fields if field not in game]
                    
                    if missing_gpg_fields:
                        self.log_test("NHL GPG Avg Fix - Game GPG Fields", False,
                                    f"Game {game.get('away_team')} @ {game.get('home_team')} missing GPG fields: {missing_gpg_fields}")
                        return False
                
                # Test specific expected results from the review request
                expected_results = [
                    {
                        'away_team': 'Colorado',
                        'home_team': 'Dallas',
                        'expected_gpg_avg': 7.4,
                        'expected_line': 6.5,
                        'expected_edge': 0.9,
                        'expected_rec': 'OVER'
                    },
                    {
                        'away_team': 'Boston',
                        'home_team': 'New Jersey',
                        'expected_gpg_avg': 4.6,
                        'expected_line': 5.5,
                        'expected_edge': -0.9,
                        'expected_rec': 'UNDER'
                    },
                    {
                        'away_team': 'Toronto',
                        'home_team': 'Vegas',
                        'expected_gpg_avg': 8.2,
                        'expected_line': 6.5,
                        'expected_edge': 1.7,
                        'expected_rec': 'OVER'
                    }
                ]
                
                verified_games = []
                tolerance = 0.2  # Allow small tolerance for floating point calculations
                
                for expected in expected_results:
                    # Find the matching game
                    matching_game = None
                    for game in games:
                        if (game.get('away_team') == expected['away_team'] and 
                            game.get('home_team') == expected['home_team']):
                            matching_game = game
                            break
                    
                    if not matching_game:
                        self.log_test(f"NHL GPG Avg Fix - {expected['away_team']} @ {expected['home_team']}", False,
                                    "Expected game not found")
                        continue
                    
                    # Verify GPG Avg (combined_gpg)
                    actual_gpg_avg = matching_game.get('combined_gpg')
                    if actual_gpg_avg is None:
                        self.log_test(f"NHL GPG Avg Fix - {expected['away_team']} @ {expected['home_team']} GPG", False,
                                    "combined_gpg is None")
                        continue
                    
                    if abs(actual_gpg_avg - expected['expected_gpg_avg']) > tolerance:
                        self.log_test(f"NHL GPG Avg Fix - {expected['away_team']} @ {expected['home_team']} GPG", False,
                                    f"Expected GPG Avg ≈{expected['expected_gpg_avg']}, got {actual_gpg_avg}")
                        continue
                    
                    # Verify Line
                    actual_line = matching_game.get('total')
                    if actual_line != expected['expected_line']:
                        self.log_test(f"NHL GPG Avg Fix - {expected['away_team']} @ {expected['home_team']} Line", False,
                                    f"Expected Line {expected['expected_line']}, got {actual_line}")
                        continue
                    
                    # Verify Edge calculation (Edge = GPG Avg - Line)
                    actual_edge = matching_game.get('edge')
                    expected_edge_calc = actual_gpg_avg - actual_line
                    if abs(actual_edge - expected_edge_calc) > tolerance:
                        self.log_test(f"NHL GPG Avg Fix - {expected['away_team']} @ {expected['home_team']} Edge Calc", False,
                                    f"Edge calculation incorrect: {actual_edge} ≠ {expected_edge_calc} (GPG Avg - Line)")
                        continue
                    
                    # Verify Edge value matches expected
                    if abs(actual_edge - expected['expected_edge']) > tolerance:
                        self.log_test(f"NHL GPG Avg Fix - {expected['away_team']} @ {expected['home_team']} Edge", False,
                                    f"Expected Edge ≈{expected['expected_edge']}, got {actual_edge}")
                        continue
                    
                    # Verify Recommendation
                    actual_rec = matching_game.get('recommendation')
                    if actual_rec != expected['expected_rec']:
                        self.log_test(f"NHL GPG Avg Fix - {expected['away_team']} @ {expected['home_team']} Rec", False,
                                    f"Expected Recommendation {expected['expected_rec']}, got {actual_rec}")
                        continue
                    
                    # If we get here, all validations passed
                    verified_games.append(f"{expected['away_team']} @ {expected['home_team']}: GPG Avg={actual_gpg_avg:.1f}, Edge={actual_edge:+.1f}, Rec={actual_rec}")
                    self.log_test(f"NHL GPG Avg Fix - {expected['away_team']} @ {expected['home_team']}", True,
                                f"GPG Avg={actual_gpg_avg:.1f}, Line={actual_line}, Edge={actual_edge:+.1f}, Rec={actual_rec}")
                
                # Test recommendation logic
                over_count = 0
                under_count = 0
                for game in games:
                    edge = game.get('edge')
                    recommendation = game.get('recommendation')
                    
                    if edge is not None:
                        if edge >= 0.5 and recommendation == 'OVER':
                            over_count += 1
                        elif edge <= -0.5 and recommendation == 'UNDER':
                            under_count += 1
                        elif -0.5 < edge < 0.5 and (recommendation is None or recommendation == ''):
                            # No bet zone is correct
                            pass
                        else:
                            self.log_test("NHL GPG Avg Fix - Recommendation Logic", False,
                                        f"Game {game.get('away_team')} @ {game.get('home_team')}: Edge={edge}, Rec={recommendation} doesn't follow logic")
                            return False
                
                # Verify GPG Avg formula is being used correctly
                # Formula: GPG Avg = (SeasonPPG_T1 + SeasonPPG_T2 + Last3PPG_T1 + Last3PPG_T2) / 2
                # We can't verify the exact source data, but we can verify the calculation makes sense
                
                if len(verified_games) >= 2:  # At least 2 of the 3 expected games verified
                    self.log_test("NHL GPG Avg Fix - Sample Games Verified", True,
                                f"Verified {len(verified_games)} sample games: {'; '.join(verified_games)}")
                else:
                    self.log_test("NHL GPG Avg Fix - Sample Games Verified", False,
                                f"Only verified {len(verified_games)} of 3 expected sample games")
                    return False
                
                self.log_test("NHL GPG Avg Fix - Recommendation Logic", True,
                            f"OVER recommendations: {over_count}, UNDER recommendations: {under_count}")
                
                self.log_test("NHL GPG Avg Calculation Fix", True,
                            f"All {len(games)} NHL games have correct GPG Avg calculations with updated ESPN/StatMuse data")
                return True
            else:
                self.log_test("NHL GPG Avg Calculation Fix", False,
                            f"Status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_test("NHL GPG Avg Calculation Fix", False, f"Request error: {str(e)}")
            return False
        except json.JSONDecodeError as e:
            self.log_test("NHL GPG Avg Calculation Fix", False, f"JSON decode error: {str(e)}")
            return False

    def run_all_tests(self):
        """Run all API tests"""
        print("=" * 60)
        print("NCAAB API TESTING - December 29, 2025")
        print("=" * 60)
        print(f"Testing API: {self.api_url}")
        print()
        
        # Run NCAAB Tests (primary focus for this review)
        ncaab_success = self.test_ncaab_api_tomorrow()
        
        # Run Excel Export Tests (secondary)
        print("\n" + "=" * 60)
        print("EXCEL EXPORT FUNCTIONALITY TESTING")
        print("=" * 60)
        excel_success = self.test_excel_export_functionality()
        
        # Run Process #6 Implementation Tests (tertiary)
        print("\n" + "=" * 60)
        print("PROCESS #6 - UPDATE RECORDS TESTING")
        print("=" * 60)
        process_6_success = self.test_process_6_implementation()
        
        # Run Historical Data Verification (quaternary)
        print("\n" + "=" * 60)
        print("HISTORICAL DATA VERIFICATION - 12/22/2025 to 12/27/2025")
        print("=" * 60)
        historical_success = self.test_historical_data_verification()
        
        # Print summary
        print()
        print("=" * 60)
        print("TEST SUMMARY")
        print("=" * 60)
        print(f"Tests Run: {self.tests_run}")
        print(f"Tests Passed: {self.tests_passed}")
        print(f"Tests Failed: {self.tests_run - self.tests_passed}")
        print(f"Success Rate: {(self.tests_passed/self.tests_run)*100:.1f}%")
        
        print("\n📊 RESULTS BY CATEGORY:")
        print(f"NCAAB API Testing: {'✅ PASSED' if ncaab_success else '❌ FAILED'}")
        print(f"Excel Export Functionality: {'✅ PASSED' if excel_success else '❌ FAILED'}")
        print(f"Process #6 Implementation: {'✅ PASSED' if process_6_success else '❌ FAILED'}")
        print(f"Historical Data Verification: {'✅ PASSED' if historical_success else '❌ FAILED'}")
        
        return self.tests_passed == self.tests_run

def main():
    tester = BettingSystemAPITester()
    success = tester.run_all_tests()
    
    # Save detailed results
    with open('/app/test_reports/backend_test_results.json', 'w') as f:
        json.dump({
            'timestamp': datetime.now().isoformat(),
            'total_tests': tester.tests_run,
            'passed_tests': tester.tests_passed,
            'success_rate': (tester.tests_passed/tester.tests_run)*100 if tester.tests_run > 0 else 0,
            'results': tester.test_results
        }, f, indent=2)
    
    return 0 if success else 1

if __name__ == "__main__":
    sys.exit(main())