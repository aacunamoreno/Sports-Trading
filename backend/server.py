from fastapi import FastAPI, APIRouter, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any, Tuple
import uuid
import random
from datetime import datetime, timezone, timedelta
from cryptography.fernet import Fernet
import base64
import asyncio
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
import httpx
from telegram import Bot
from telegram.constants import ParseMode
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.interval import IntervalTrigger
from apscheduler.triggers.cron import CronTrigger

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Encryption key for storing credentials
encryption_key = os.environ.get('ENCRYPTION_KEY')
if not encryption_key:
    encryption_key = Fernet.generate_key().decode()
    print(f"Generated new encryption key. Add to .env: ENCRYPTION_KEY={encryption_key}")
cipher_suite = Fernet(encryption_key.encode())

# Telegram Bot setup (optional - will be configured later)
telegram_bot = None
telegram_chat_id = None

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Telegram after logger is set up
async def init_telegram_from_db():
    """Initialize Telegram bot from database configuration"""
    global telegram_bot, telegram_chat_id
    try:
        # Try to load from database first
        config = await db.telegram_config.find_one({}, {"_id": 0})
        if config:
            telegram_bot = Bot(token=config["bot_token"])
            telegram_chat_id = int(config["chat_id"])
            logger.info(f"Telegram initialized from database (Chat ID: {telegram_chat_id})")
            return
    except Exception as e:
        logger.error(f"Error loading Telegram from database: {e}")
    
    # Fall back to environment variables
    token = os.environ.get('TELEGRAM_BOT_TOKEN')
    chat_id = os.environ.get('TELEGRAM_CHAT_ID')
    
    if token and chat_id:
        telegram_bot = Bot(token=token)
        telegram_chat_id = int(chat_id)
        logger.info("Telegram initialized from environment variables")
    else:
        logger.info("Telegram not configured (optional feature)")

# Startup event to initialize Telegram and auto-start monitoring
@app.on_event("startup")
async def startup_event():
    await init_telegram_from_db()
    await auto_start_monitoring()
    schedule_daily_summary()
    await startup_recovery()  # Check if we missed anything overnight
    await process_pending_deletions()  # Clean up any scheduled deletions from before restart

async def schedule_message_deletion(chat_id: int, message_id: int, delay_minutes: int = 15):
    """Schedule a message for deletion by storing in database (survives server restarts)"""
    try:
        from zoneinfo import ZoneInfo
        delete_at = datetime.now(timezone.utc) + timedelta(minutes=delay_minutes)
        await db.scheduled_deletions.insert_one({
            "chat_id": chat_id,
            "message_id": message_id,
            "delete_at": delete_at,
            "created_at": datetime.now(timezone.utc)
        })
        logger.debug(f"Scheduled message {message_id} for deletion at {delete_at}")
    except Exception as e:
        logger.error(f"Error scheduling message deletion: {e}")

async def process_pending_deletions():
    """Process any messages that are past their deletion time"""
    try:
        now = datetime.now(timezone.utc)
        pending = await db.scheduled_deletions.find({"delete_at": {"$lte": now}}).to_list(100)
        
        if not pending:
            return
        
        logger.info(f"Processing {len(pending)} pending message deletions...")
        
        telegram_config = await db.telegram_config.find_one({}, {"_id": 0})
        if not telegram_config or not telegram_config.get("bot_token"):
            logger.warning("No Telegram config found for deletion processing")
            return
        
        bot = Bot(token=telegram_config["bot_token"])
        
        deleted_count = 0
        for item in pending:
            try:
                await bot.delete_message(chat_id=item["chat_id"], message_id=item["message_id"])
                deleted_count += 1
            except Exception as e:
                # Message may already be deleted or too old
                logger.debug(f"Could not delete message {item['message_id']}: {e}")
            
            # Remove from scheduled deletions regardless of success
            await db.scheduled_deletions.delete_one({"_id": item["_id"]})
        
        if deleted_count > 0:
            logger.info(f"Auto-deleted {deleted_count} status messages")
            
    except Exception as e:
        logger.error(f"Error processing pending deletions: {e}")

async def delete_message_later(bot, chat_id, message_id, delay_minutes=15):
    """Schedule a Telegram message for deletion - now uses database for reliability"""
    # Use database-backed scheduling instead of asyncio task
    await schedule_message_deletion(chat_id, message_id, delay_minutes)

def schedule_daily_summary():
    """Schedule the daily summary to run at 11 PM Arizona time"""
    global scheduler
    
    # Arizona is UTC-7 (no daylight saving)
    # 11 PM Arizona = 6 AM UTC next day (23:00 - 7 = 16:00? No wait...)
    # Actually: Arizona UTC-7, so 11 PM Arizona = 11 PM + 7 hours = 6 AM UTC next day
    # Let's use cron with timezone
    
    try:
        # Activity summary at 10:59 PM Arizona
        scheduler.add_job(
            send_activity_summary,
            trigger=CronTrigger(hour=22, minute=45, timezone='America/Phoenix'),  # 10:45 PM Arizona
            id='activity_summary',
            replace_existing=True
        )
        logger.info("Activity summary scheduled for 10:45 PM Arizona time")
        
        # Betting summary at 11:00 PM Arizona
        scheduler.add_job(
            send_daily_summary,
            trigger=CronTrigger(hour=22, minute=45, timezone='America/Phoenix'),  # 10:45 PM Arizona
            id='daily_summary',
            replace_existing=True
        )
        logger.info("Daily summary scheduled for 10:45 PM Arizona time")
        
        # Daily cleanup at 9:00 AM Arizona
        scheduler.add_job(
            daily_cleanup,
            trigger=CronTrigger(hour=9, minute=0, timezone='America/Phoenix'),  # 9 AM Arizona
            id='daily_cleanup',
            replace_existing=True
        )
        logger.info("Daily cleanup scheduled for 9:00 AM Arizona time")
        
        # NBA Opportunities refresh at 10:30 PM Arizona (before sleep mode at 10:45 PM)
        scheduler.add_job(
            refresh_nba_opportunities_scheduled,
            trigger=CronTrigger(hour=22, minute=30, timezone='America/Phoenix'),  # 10:30 PM Arizona
            id='nba_opportunities_refresh',
            replace_existing=True
        )
        logger.info("NBA opportunities refresh scheduled for 10:30 PM Arizona time")
        
        # ============================================
        # PROCESS NOTEBOOK SCHEDULED JOBS
        # ============================================
        
        # #1 - 8:00 PM Arizona: Scrape tomorrow's games (opening lines) from ScoresAndOdds
        scheduler.add_job(
            scrape_tomorrows_opening_lines,
            trigger=CronTrigger(hour=20, minute=0, timezone='America/Phoenix'),  # 8:00 PM Arizona
            id='scrape_tomorrows_opening_lines',
            replace_existing=True
        )
        logger.info("Tomorrow's opening lines scrape scheduled for 8:00 PM Arizona time")
        
        # #4, #5, #6 - 5:00 AM Arizona: Morning data refresh
        scheduler.add_job(
            morning_data_refresh,
            trigger=CronTrigger(hour=5, minute=0, timezone='America/Phoenix'),  # 5:00 AM Arizona
            id='morning_data_refresh',
            replace_existing=True
        )
        logger.info("Morning data refresh scheduled for 5:00 AM Arizona time")
        
    except Exception as e:
        logger.error(f"Error scheduling daily tasks: {str(e)}")

# ============================================================================
# OPENING LINES TRACKING - Store first seen lines for each game
# ============================================================================

async def store_opening_line(league: str, date: str, away_team: str, home_team: str, line: float):
    """
    Store the opening line for a game. Only stores if no opening line exists yet.
    This captures the FIRST line we see for a game.
    """
    try:
        game_key = f"{league}_{date}_{away_team}_{home_team}".lower().replace(" ", "_")
        
        # Check if opening line already exists
        existing = await db.opening_lines.find_one({"game_key": game_key}, {"_id": 0})
        if existing:
            logger.debug(f"Opening line already exists for {away_team} @ {home_team}: {existing.get('opening_line')}")
            return existing.get('opening_line')
        
        # Store new opening line
        await db.opening_lines.insert_one({
            "game_key": game_key,
            "league": league,
            "date": date,
            "away_team": away_team,
            "home_team": home_team,
            "opening_line": line,
            "created_at": datetime.now(timezone.utc)
        })
        logger.info(f"Stored opening line for {away_team} @ {home_team}: {line}")
        return line
    except Exception as e:
        logger.error(f"Error storing opening line: {e}")
        return None

async def get_opening_line(league: str, date: str, away_team: str, home_team: str) -> Optional[float]:
    """
    Get the stored opening line for a game.
    """
    try:
        game_key = f"{league}_{date}_{away_team}_{home_team}".lower().replace(" ", "_")
        existing = await db.opening_lines.find_one({"game_key": game_key}, {"_id": 0})
        if existing:
            return existing.get('opening_line')
        return None
    except Exception as e:
        logger.error(f"Error getting opening line: {e}")
        return None

async def store_opening_lines_batch(league: str, date: str, games: List[Dict]):
    """
    Store opening lines for multiple games at once.
    Only stores lines that don't already exist.
    """
    stored_count = 0
    for game in games:
        away = game.get('away_team') or game.get('away')
        home = game.get('home_team') or game.get('home')
        line = game.get('total')
        
        if away and home and line:
            result = await store_opening_line(league, date, away, home, line)
            if result:
                stored_count += 1
    
    logger.info(f"Stored {stored_count} opening lines for {league} on {date}")
    return stored_count

async def get_opening_lines_batch(league: str, date: str, games: List[Dict]) -> Dict[str, float]:
    """
    Get opening lines for multiple games. Returns a dict mapping game_key to opening_line.
    """
    opening_lines = {}
    for game in games:
        away = game.get('away_team') or game.get('away')
        home = game.get('home_team') or game.get('home')
        
        if away and home:
            opening = await get_opening_line(league, date, away, home)
            if opening:
                game_key = f"{away}_{home}".lower().replace(" ", "_")
                opening_lines[game_key] = opening
    
    return opening_lines

# ============================================================================


async def auto_start_monitoring():
    """Auto-start bet monitoring if it was previously enabled"""
    global monitoring_enabled
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    
    try:
        # Check if there's an active connection
        conn = await db.connections.find_one({}, {"_id": 0}, sort=[("created_at", -1)])
        if conn and conn.get("is_connected"):
            # Check if monitoring was previously enabled (stored in DB)
            monitor_config = await db.monitor_config.find_one({}, {"_id": 0})
            if monitor_config and monitor_config.get("auto_start", True):
                monitoring_enabled = True
                
                # Check if we missed a scheduled check (server restart detection)
                state = await db.monitor_state.find_one({"_id": "main"})
                if state and state.get("next_check_utc"):
                    next_check_utc = state["next_check_utc"]
                    if isinstance(next_check_utc, str):
                        next_check_utc = datetime.fromisoformat(next_check_utc.replace('Z', '+00:00'))
                    if next_check_utc.tzinfo is None:
                        next_check_utc = next_check_utc.replace(tzinfo=timezone.utc)
                    
                    now_utc = datetime.now(timezone.utc)
                    if next_check_utc < now_utc:
                        minutes_overdue = (now_utc - next_check_utc).total_seconds() / 60
                        if minutes_overdue > 2:  # More than 2 minutes overdue
                            now_arizona = datetime.now(arizona_tz)
                            logger.warning(f"SERVER RESTART DETECTED! Missed check was {minutes_overdue:.0f} min overdue.")
                            
                            # Send alert to Telegram about the restart - auto-delete after 30 min
                            try:
                                telegram_config = await db.telegram_config.find_one({}, {"_id": 0})
                                if telegram_config and telegram_config.get("bot_token"):
                                    bot = Bot(token=telegram_config["bot_token"])
                                    msg = f"⚠️ *SERVER RESTART*\n\nMonitoring was interrupted.\nMissed check by ~{minutes_overdue:.0f} min.\nRunning immediate catch-up check.\n\nTime: {now_arizona.strftime('%I:%M %p')} Arizona"
                                    sent_msg = await bot.send_message(
                                        chat_id=telegram_config["chat_id"],
                                        text=msg,
                                        parse_mode=ParseMode.MARKDOWN
                                    )
                                    # Schedule auto-deletion after 30 minutes
                                    asyncio.create_task(delete_message_later(bot, telegram_config["chat_id"], sent_msg.message_id, 15))
                            except Exception as e:
                                logger.error(f"Failed to send restart alert: {e}")
                
                # Start background monitoring loop
                asyncio.create_task(monitoring_loop())
                
                logger.info("Bet monitoring auto-started with background loop (5 min intervals, paused 10:00 PM - 7:00 AM Arizona)")
            else:
                logger.info("Bet monitoring not auto-started (disabled in config)")
        else:
            logger.info("Bet monitoring not auto-started (no active connection)")
    except Exception as e:
        logger.error(f"Error auto-starting monitoring: {str(e)}")


async def monitoring_loop():
    """Background loop for bet monitoring - designed to be resilient to all errors and server restarts"""
    global monitoring_enabled
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    
    logger.info("=" * 60)
    logger.info("MONITORING LOOP STARTED - This loop should NEVER stop")
    logger.info("=" * 60)
    
    loop_iteration = 0
    
    # On startup, check if we missed a scheduled check due to server restart
    try:
        state = await db.monitor_state.find_one({"_id": "main"})
        if state and state.get("next_check_utc"):
            next_check_utc = state["next_check_utc"]
            if isinstance(next_check_utc, str):
                next_check_utc = datetime.fromisoformat(next_check_utc.replace('Z', '+00:00'))
            if next_check_utc.tzinfo is None:
                next_check_utc = next_check_utc.replace(tzinfo=timezone.utc)
            
            now_utc = datetime.now(timezone.utc)
            if next_check_utc < now_utc:
                # We missed a check! Calculate how overdue
                minutes_overdue = (now_utc - next_check_utc).total_seconds() / 60
                logger.warning(f"MISSED CHECK DETECTED! Was scheduled for {next_check_utc.isoformat()}, now {minutes_overdue:.1f} min overdue. Running immediately!")
                # Run immediately by not sleeping on first iteration
    except Exception as e:
        logger.error(f"Error checking for missed checks: {e}")
    
    while True:
        loop_iteration += 1
        try:
            if monitoring_enabled:
                # Check if we're in sleep hours
                now_arizona = datetime.now(arizona_tz)
                current_hour = now_arizona.hour
                current_minute = now_arizona.minute
                current_time_minutes = current_hour * 60 + current_minute
                
                sleep_start = 22 * 60 + 0   # 10:00 PM
                sleep_end = 7 * 60 + 0       # 7:00 AM
                
                if current_time_minutes >= sleep_start or current_time_minutes < sleep_end:
                    logger.info(f"[Loop #{loop_iteration}] Sleep hours ({now_arizona.strftime('%I:%M %p')} Arizona) - waiting 5 min...")
                    await asyncio.sleep(300)  # Check again in 5 minutes during sleep
                    continue
                
                # Run monitoring check - wrapped in its own try/except
                logger.info(f"[Loop #{loop_iteration}] Starting monitoring cycle...")
                try:
                    await run_monitoring_cycle()
                    logger.info(f"[Loop #{loop_iteration}] Monitoring cycle completed successfully")
                except Exception as e:
                    logger.error(f"[Loop #{loop_iteration}] Monitoring cycle error (loop continues): {str(e)}", exc_info=True)
                
                # Fixed 5 minute interval
                next_interval = random.randint(MIN_INTERVAL, MAX_INTERVAL)
                next_check_time = now_arizona + timedelta(minutes=next_interval)
                next_check_utc = datetime.now(timezone.utc) + timedelta(minutes=next_interval)
                
                # CRITICAL: Store next check time in database so we can detect missed checks after restart
                try:
                    await db.monitor_state.update_one(
                        {"_id": "main"},
                        {"$set": {
                            "next_check_utc": next_check_utc.isoformat(),
                            "next_check_arizona": next_check_time.strftime('%I:%M %p'),
                            "last_check_utc": datetime.now(timezone.utc).isoformat(),
                            "loop_iteration": loop_iteration
                        }},
                        upsert=True
                    )
                except Exception as e:
                    logger.error(f"Failed to save monitor state: {e}")
                
                logger.info(f"[Loop #{loop_iteration}] Sleeping for {next_interval} minutes. Next check at ~{next_check_time.strftime('%I:%M %p')} Arizona")
                await asyncio.sleep(next_interval * 60)
                
            else:
                # Monitoring disabled, check status again in 1 minute
                logger.info(f"[Loop #{loop_iteration}] Monitoring disabled, checking status in 60s...")
                await asyncio.sleep(60)
                
        except asyncio.CancelledError:
            logger.warning("Monitoring loop received CancelledError - this should only happen on shutdown")
            break
        except Exception as e:
            # This catch-all ensures the loop NEVER dies
            logger.error(f"[Loop #{loop_iteration}] UNEXPECTED ERROR in monitoring loop (will retry in 60s): {str(e)}", exc_info=True)
            await asyncio.sleep(60)  # Wait 1 minute before retrying
            continue  # Explicitly continue the loop


async def run_monitoring_cycle():
    """Run a single monitoring cycle with notifications"""
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    check_time = datetime.now(arizona_tz)
    new_bets_found = {"jac075": 0, "jac083": 0}  # Default value
    
    try:
        # Process any pending message deletions first
        await process_pending_deletions()
        
        # Log to database
        await db.activity_log.insert_one({
            "type": "bet_check",
            "account": "jac083",
            "timestamp": datetime.now(timezone.utc),
            "timestamp_arizona": check_time.strftime('%I:%M %p'),
            "date": check_time.strftime('%Y-%m-%d')
        })
        
        # Run monitoring
        logger.info(f"Running monitoring cycle at {check_time.strftime('%I:%M %p')} Arizona")
        
        try:
            result = await monitor_open_bets()
            if result is not None:
                new_bets_found = result
        except Exception as e:
            logger.error(f"Error in monitor_open_bets: {str(e)}", exc_info=True)
        
        # Check for settled bets
        try:
            await check_bet_results()
        except Exception as e:
            logger.error(f"Error in check_bet_results: {str(e)}", exc_info=True)
        
        # DISABLED: Status notifications to keep chat clean
        # User only wants compilation messages (daily bet summaries)
        # await send_check_notification(check_time, new_bets_found)
        logger.info(f"Check complete - ENANO: {new_bets_found.get('jac075', 0)}, TIPSTER: {new_bets_found.get('jac083', 0)} new bets")
        
    except Exception as e:
        logger.error(f"Monitoring cycle error: {str(e)}", exc_info=True)


async def startup_recovery():
    """Check if we missed overnight period and send catch-up notifications"""
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        now_arizona = datetime.now(arizona_tz)
        
        # Check when the last monitoring check happened
        last_activity = await db.activity_log.find_one({}, sort=[("timestamp", -1)])
        
        if last_activity:
            last_check = last_activity.get("timestamp")
            if last_check:
                # Handle both datetime objects and ISO strings
                if isinstance(last_check, str):
                    last_check = datetime.fromisoformat(last_check.replace('Z', '+00:00'))
                
                # If last_check is naive, assume UTC
                if last_check.tzinfo is None:
                    last_check = last_check.replace(tzinfo=timezone.utc)
                
                hours_since_last = (datetime.now(timezone.utc) - last_check).total_seconds() / 3600
                
                # Also check if a scheduled check was missed
                next_check = await get_next_check_time()
                check_was_missed = False
                if next_check:
                    if isinstance(next_check, str):
                        next_check = datetime.fromisoformat(next_check.replace('Z', '+00:00'))
                    if next_check.tzinfo is None:
                        next_check = next_check.replace(tzinfo=timezone.utc)
                    
                    # If next_check time has passed, we missed a check
                    if next_check < datetime.now(timezone.utc):
                        minutes_overdue = (datetime.now(timezone.utc) - next_check).total_seconds() / 60
                        if minutes_overdue > 2:  # More than 2 minutes overdue
                            check_was_missed = True
                            logger.warning(f"Startup recovery: Scheduled check was {minutes_overdue:.0f} minutes overdue!")
                
                if hours_since_last > 1 or check_was_missed:  # More than 1 hour gap OR missed scheduled check
                    logger.warning(f"Startup recovery: {hours_since_last:.1f} hours since last check. Sending catch-up notification...")
                    
                    # DISABLED: SYSTEM RESTART notification to keep chat clean
                    # User only wants compilation messages
                    logger.info(f"System restart detected after {hours_since_last:.1f} hours - running catch-up check (no notification sent)")
                    
                    # Trigger immediate bet check (use the full cycle with notification)
                    logger.info("Triggering immediate catch-up bet check...")
                    asyncio.create_task(monitor_and_reschedule())
                else:
                    logger.info(f"Startup recovery: Last check was {hours_since_last:.1f} hours ago, no recovery needed")
        else:
            logger.info("Startup recovery: No previous activity found, skipping recovery")
            
    except Exception as e:
        logger.error(f"Startup recovery error: {str(e)}")


# Models
class ConnectionConfig(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    password_encrypted: str
    website: str = "plays888.co"
    is_connected: bool = False
    last_connection: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ConnectionCreate(BaseModel):
    username: str
    password: str

class BettingRule(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    enabled: bool = True
    min_odds: Optional[int] = None  # American odds (e.g., -110, +150)
    max_odds: Optional[int] = None  # American odds (e.g., -110, +150)
    wager_amount: float
    auto_place: bool = False
    sport: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BettingRuleCreate(BaseModel):
    name: str
    enabled: bool = True
    min_odds: Optional[int] = None  # American odds (e.g., -110, +150)
    max_odds: Optional[int] = None  # American odds (e.g., -110, +150)
    wager_amount: float
    auto_place: bool = False
    sport: Optional[str] = None

class BetOpportunity(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_name: str
    odds: int  # American odds (e.g., -110, +150)
    sport: str
    bet_type: str
    available: bool = True
    matched_rule_id: Optional[str] = None
    discovered_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BetHistory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    opportunity_id: str
    rule_id: str
    wager_amount: float
    odds: int  # American odds (e.g., -110, +150)
    status: str  # placed, won, lost, pending
    placed_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    result: Optional[str] = None

class PlaceBetRequest(BaseModel):
    opportunity_id: str
    wager_amount: float

class AccountStatus(BaseModel):
    is_connected: bool
    balance: Optional[float] = None
    username: Optional[str] = None
    last_sync: Optional[datetime] = None


# Helper functions
def encrypt_password(password: str) -> str:
    return cipher_suite.encrypt(password.encode()).decode()

def decrypt_password(encrypted: str) -> str:
    return cipher_suite.decrypt(encrypted.encode()).decode()

def calculate_american_odds_payout(wager: float, odds: int) -> float:
    """Calculate potential payout from American odds"""
    if odds > 0:
        # Positive odds: shows profit on $100 bet
        return wager * (odds / 100)
    else:
        # Negative odds: shows how much to bet to win $100
        return wager * (100 / abs(odds))

def format_american_odds(odds: int) -> str:
    """Format American odds with + or - sign"""
    return f"+{odds}" if odds > 0 else str(odds)

# Account username to display name mapping
ACCOUNT_LABELS = {
    "jac075": "ENANO",
    "jac083": "TIPSTER"
}

def format_amount_short(amount: float) -> str:
    """Format amount as short string like $2.2K"""
    if amount >= 1000:
        return f"${amount/1000:.1f}K".replace('.0K', 'K')
    return f"${amount:.0f}"

def extract_short_game_name(game: str, description: str = "") -> str:
    """Extract VERY short team names - max 3-4 chars per team"""
    import re
    
    text = (game or description or "").upper()
    
    # Common team abbreviations dictionary
    TEAM_ABBREVS = {
        # NFL
        'JACKSONVILLE': 'JAX', 'JAGUARS': 'JAX', 'DENVER': 'DEN', 'BRONCOS': 'DEN',
        'ATLANTA': 'ATL', 'FALCONS': 'ATL', 'ARIZONA': 'ARI', 'CARDINALS': 'ARI',
        'PITTSBURGH': 'PIT', 'STEELERS': 'PIT', 'PENGUINS': 'PIT',
        'DETROIT': 'DET', 'LIONS': 'DET', 'PATRIOTS': 'NE', 'NEW ENGLAND': 'NE',
        'BALTIMORE': 'BAL', 'RAVENS': 'BAL', 'LAS VEGAS': 'LV', 'RAIDERS': 'LV',
        'HOUSTON': 'HOU', 'TEXANS': 'HOU',
        # NHL
        'WINNIPEG': 'WPG', 'JETS': 'WPG', 'UTAH': 'UTA', 'MAMMOTH': 'UTA',
        'COLORADO': 'COL', 'AVALANCHE': 'COL', 'MINNESOTA': 'MIN', 'WILD': 'MIN',
        'MONTREAL': 'MTL', 'CANADIENS': 'MTL', 'NEW YORK': 'NY', 'RANGERS': 'NYR',
        'NASHVILLE': 'NSH', 'PREDATORS': 'NSH',
        # College Basketball common
        'SACRAMENTO': 'SAC', 'STATE': '', 'BETHUNE': 'BETH', 'COOKMAN': 'COOK',
        'CENTRAL': 'CEN', 'MICHIGAN': 'MICH', 'FAIRLEIGH': 'FDU', 'DICKINSON': 'FDU',
        'INDIANAPOLIS': 'INDY', 'IRVINE': 'UCI', 'NORFOLK': 'NORF',
        'CREIGHTON': 'CRE', 'EDWARDSVILLE': 'SIU', 'TECH': 'TCH',
        'ALABAMA': 'ALA', 'KENTUCKY': 'KEN', 'DUKE': 'DUKE', 'CAROLINA': 'CAR',
        'GONZAGA': 'GONZ', 'VILLANOVA': 'NOVA', 'KANSAS': 'KAN', 'TEXAS': 'TEX',
        'FLORIDA': 'FLA', 'OHIO': 'OHIO', 'OREGON': 'ORE', 'UCLA': 'UCLA',
        'MEMPHIS': 'MEM', 'TENNESSEE': 'TENN', 'VIRGINIA': 'UVA', 'LOUISVILLE': 'LOU',
        'XAVIER': 'XAV', 'PURDUE': 'PUR', 'ILLINOIS': 'ILL', 'IOWA': 'IOWA',
        'WISCONSIN': 'WISC', 'MARYLAND': 'MD', 'INDIANA': 'IND', 'PENN': 'PENN',
        'SYRACUSE': 'SYR', 'CLEMSON': 'CLEM', 'MIAMI': 'MIA', 'AUBURN': 'AUB',
        'ARKANSAS': 'ARK', 'MISSISSIPPI': 'MISS', 'MISSOURI': 'MIZZ', 'BAYLOR': 'BAY',
        'STANFORD': 'STAN', 'WASHINGTON': 'WASH', 'ARIZONA STATE': 'ASU',
        'BOSTON': 'BOS', 'COLLEGE': 'BC', 'CONNECTICUT': 'UCON', 'TEMPLE': 'TEM',
        'CINCINNATI': 'CIN', 'HOUSTON': 'HOU', 'SMU': 'SMU', 'TULANE': 'TUL',
        'RICE': 'RICE', 'PEPPERDINE': 'PEP', 'NEW MEXICO': 'NM', 'MEXICO': 'NM',
        # Soccer
        'ABHA': 'ABHA', 'CLUB': '',
        # Generic
        'UNIVERSITY': '', 'OF': '', 'THE': '', 'VRS': '', 'VS': '', 'REG.TIME': '', 'REG': '',
    }
    
    def get_abbrev(team_name):
        """Get abbreviation for a team name"""
        team_name = team_name.strip().upper()
        
        # Direct lookup
        if team_name in TEAM_ABBREVS:
            return TEAM_ABBREVS[team_name]
        
        # Try each word
        words = team_name.split()
        abbrev_parts = []
        for word in words:
            if word in TEAM_ABBREVS:
                if TEAM_ABBREVS[word]:  # Skip empty abbreviations
                    abbrev_parts.append(TEAM_ABBREVS[word])
            elif len(word) >= 3:
                abbrev_parts.append(word[:3])
        
        if abbrev_parts:
            return abbrev_parts[0]  # Return first significant part
        
        # Fallback: first 4 chars
        return team_name[:4] if team_name else "TM"
    
    # Remove common noise
    text = re.sub(r'REG\.?TIME|OT INCLUDED|\[.*?\]|\d{1,3}\s*', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    
    # Try to find matchup pattern
    match = re.search(r'(.+?)\s*(?:VRS|VS\.?|@|\/)\s*(.+)', text, re.IGNORECASE)
    if match:
        team1 = get_abbrev(match.group(1).strip())
        team2 = get_abbrev(match.group(2).strip())
        return f"{team1}/{team2}"
    
    # Single team (like spreads)
    return get_abbrev(text)

def adjust_time_for_arizona(time_str: str) -> str:
    """Add 1 hour to game time to convert from Plays888 time to Arizona time"""
    if not time_str:
        return time_str
    
    try:
        time_str = time_str.strip()
        # Parse the time
        parts = time_str.upper().replace('AM', ' AM').replace('PM', ' PM').replace('  ', ' ').split()
        time_part = parts[0]
        period = parts[1] if len(parts) > 1 else 'AM'
        
        time_parts = time_part.split(':')
        hour = int(time_parts[0])
        minute = int(time_parts[1]) if len(time_parts) > 1 else 0
        
        # Convert to 24-hour
        if period == 'PM' and hour != 12:
            hour += 12
        elif period == 'AM' and hour == 12:
            hour = 0
        
        # Add 1 hour for Arizona adjustment
        hour += 1
        if hour >= 24:
            hour = hour - 24
        
        # Convert back to 12-hour format
        if hour == 0:
            new_hour = 12
            new_period = 'AM'
        elif hour < 12:
            new_hour = hour
            new_period = 'AM'
        elif hour == 12:
            new_hour = 12
            new_period = 'PM'
        else:
            new_hour = hour - 12
            new_period = 'PM'
        
        return f"{new_hour:02d}:{minute:02d} {new_period}"
    except Exception:
        return time_str

def extract_bet_type_short(bet_type: str) -> str:
    """Extract short bet type like 'u48.5' or 'o47' from bet description"""
    import re
    
    if not bet_type:
        return ""
    
    # First, convert ½ to .5 for consistent handling
    bet_type_normalized = bet_type.replace('½', '.5')
    
    # Look for over/under patterns
    over_match = re.search(r'(?:over|o)\s*(\d+(?:\.\d+)?)', bet_type_normalized, re.IGNORECASE)
    if over_match:
        return f"o{over_match.group(1)}"
    
    under_match = re.search(r'(?:under|u)\s*(\d+(?:\.\d+)?)', bet_type_normalized, re.IGNORECASE)
    if under_match:
        return f"u{under_match.group(1)}"
    
    # Look for spread patterns (like +3.5 or -7)
    spread_match = re.search(r'([+-]\d+(?:\.\d+)?)', bet_type_normalized)
    if spread_match:
        return spread_match.group(1)
    
    # Return truncated original
    return bet_type[:15] if len(bet_type) > 15 else bet_type


# Team name mapping from plays888 full names to short names for opportunities
PLAYS888_TO_SHORT_TEAM_NAME = {
    # NBA Teams
    'ATLANTA HAWKS': 'Atlanta',
    'BOSTON CELTICS': 'Boston',
    'BROOKLYN NETS': 'Brooklyn',
    'CHARLOTTE HORNETS': 'Charlotte',
    'CHICAGO BULLS': 'Chicago',
    'CLEVELAND CAVALIERS': 'Cleveland',
    'DALLAS MAVERICKS': 'Dallas',
    'DENVER NUGGETS': 'Denver',
    'DETROIT PISTONS': 'Detroit',
    'GOLDEN STATE WARRIORS': 'Golden State',
    'HOUSTON ROCKETS': 'Houston',
    'INDIANA PACERS': 'Indiana',
    'LOS ANGELES CLIPPERS': 'LA Clippers',
    'LOS ANGELES LAKERS': 'LA Lakers',
    'MEMPHIS GRIZZLIES': 'Memphis',
    'MIAMI HEAT': 'Miami',
    'MILWAUKEE BUCKS': 'Milwaukee',
    'MINNESOTA TIMBERWOLVES': 'Minnesota',
    'NEW ORLEANS PELICANS': 'New Orleans',
    'NEW YORK KNICKS': 'New York',
    'OKLAHOMA CITY THUNDER': 'Okla City',
    'ORLANDO MAGIC': 'Orlando',
    'PHILADELPHIA 76ERS': 'Philadelphia',
    'PHOENIX SUNS': 'Phoenix',
    'PORTLAND TRAIL BLAZERS': 'Portland',
    'SACRAMENTO KINGS': 'Sacramento',
    'SAN ANTONIO SPURS': 'San Antonio',
    'TORONTO RAPTORS': 'Toronto',
    'UTAH JAZZ': 'Utah',
    'WASHINGTON WIZARDS': 'Washington',
    # NHL Teams
    'ANAHEIM DUCKS': 'Anaheim',
    'ARIZONA COYOTES': 'Arizona',
    'BOSTON BRUINS': 'Boston',
    'BUFFALO SABRES': 'Buffalo',
    'CALGARY FLAMES': 'Calgary',
    'CAROLINA HURRICANES': 'Carolina',
    'CHICAGO BLACKHAWKS': 'Chicago',
    'COLORADO AVALANCHE': 'Colorado',
    'COLUMBUS BLUE JACKETS': 'Columbus',
    'DALLAS STARS': 'Dallas',
    'DETROIT RED WINGS': 'Detroit',
    'EDMONTON OILERS': 'Edmonton',
    'FLORIDA PANTHERS': 'Florida',
    'LOS ANGELES KINGS': 'Los Angeles',
    'MINNESOTA WILD': 'Minnesota',
    'MONTREAL CANADIENS': 'Montreal',
    'NASHVILLE PREDATORS': 'Nashville',
    'NEW JERSEY DEVILS': 'New Jersey',
    'NEW YORK ISLANDERS': 'NY Islanders',
    'NEW YORK RANGERS': 'NY Rangers',
    'OTTAWA SENATORS': 'Ottawa',
    'PHILADELPHIA FLYERS': 'Philadelphia',
    'PITTSBURGH PENGUINS': 'Pittsburgh',
    'SAN JOSE SHARKS': 'San Jose',
    'SEATTLE KRAKEN': 'Seattle',
    'ST. LOUIS BLUES': 'St. Louis',
    'TAMPA BAY LIGHTNING': 'Tampa Bay',
    'TORONTO MAPLE LEAFS': 'Toronto',
    'UTAH MAMMOTH': 'Utah',  # New team (formerly Arizona Coyotes)
    'VANCOUVER CANUCKS': 'Vancouver',
    'VEGAS GOLDEN KNIGHTS': 'Vegas',
    'WASHINGTON CAPITALS': 'Washington',
    'WINNIPEG JETS': 'Winnipeg',
    # NFL Teams
    'ARIZONA CARDINALS': 'Arizona',
    'ATLANTA FALCONS': 'Atlanta',
    'BALTIMORE RAVENS': 'Baltimore',
    'BUFFALO BILLS': 'Buffalo',
    'CAROLINA PANTHERS': 'Carolina',
    'CHICAGO BEARS': 'Chicago',
    'CINCINNATI BENGALS': 'Cincinnati',
    'CLEVELAND BROWNS': 'Cleveland',
    'DALLAS COWBOYS': 'Dallas',
    'DENVER BRONCOS': 'Denver',
    'DETROIT LIONS': 'Detroit',
    'GREEN BAY PACKERS': 'Green Bay',
    'HOUSTON TEXANS': 'Houston',
    'INDIANAPOLIS COLTS': 'Indianapolis',
    'JACKSONVILLE JAGUARS': 'Jacksonville',
    'KANSAS CITY CHIEFS': 'Kansas City',
    'LAS VEGAS RAIDERS': 'Las Vegas',
    'LOS ANGELES CHARGERS': 'LA Chargers',
    'LOS ANGELES RAMS': 'LA Rams',
    'MIAMI DOLPHINS': 'Miami',
    'MINNESOTA VIKINGS': 'Minnesota',
    'NEW ENGLAND PATRIOTS': 'New England',
    'NEW ORLEANS SAINTS': 'New Orleans',
    'NEW YORK GIANTS': 'NY Giants',
    'NEW YORK JETS': 'NY Jets',
    'PHILADELPHIA EAGLES': 'Philadelphia',
    'PITTSBURGH STEELERS': 'Pittsburgh',
    'SAN FRANCISCO 49ERS': 'San Francisco',
    'SEATTLE SEAHAWKS': 'Seattle',
    'TAMPA BAY BUCCANEERS': 'Tampa Bay',
    'TENNESSEE TITANS': 'Tennessee',
    'WASHINGTON COMMANDERS': 'Washington',
}

def convert_plays888_team_name(full_name: str) -> str:
    """Convert plays888 full team name to short name for opportunities"""
    return PLAYS888_TO_SHORT_TEAM_NAME.get(full_name.upper(), full_name)


async def get_or_create_daily_compilation(account: str) -> dict:
    """Get or create the daily bet compilation for an account"""
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    today = datetime.now(arizona_tz).strftime('%Y-%m-%d')
    
    compilation = await db.daily_compilations.find_one({
        "account": account,
        "date": today
    })
    
    if not compilation:
        compilation = {
            "account": account,
            "date": today,
            "message_id": None,
            "bets": [],
            "total_result": 0,
            "created_at": datetime.now(timezone.utc)
        }
        await db.daily_compilations.insert_one(compilation)
    
    return compilation

async def build_compilation_message(account: str, detailed: bool = False) -> str:
    """Build the compilation message for an account
    
    Args:
        account: The account username
        detailed: If True, use full team names. If False, use short abbreviations.
    """
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    today = datetime.now(arizona_tz).strftime('%Y-%m-%d')
    
    compilation = await db.daily_compilations.find_one({
        "account": account,
        "date": today
    })
    
    if not compilation or not compilation.get('bets'):
        return None
    
    account_label = ACCOUNT_LABELS.get(account, account)
    bets = compilation['bets']
    total_result = compilation.get('total_result', 0)
    
    # Sort bets by game time (earliest first, games without time at the end)
    def parse_time_for_sort(bet):
        """Convert time string to sortable value"""
        time_str = bet.get('game_time', '')
        date_str = bet.get('game_date', '')
        if not time_str:
            return (9999, 0)  # Put games without time at the end
        
        try:
            # Parse time like "3:15 PM" or "10:00 AM"
            time_str = time_str.upper().strip()
            # Handle formats like "3:15PM" or "3:15 PM"
            time_str = time_str.replace('AM', ' AM').replace('PM', ' PM').replace('  ', ' ')
            
            parts = time_str.replace(':', ' ').split()
            hour = int(parts[0])
            minute = int(parts[1]) if len(parts) > 1 else 0
            is_pm = 'PM' in time_str.upper()
            
            # Convert to 24-hour for sorting
            if is_pm and hour != 12:
                hour += 12
            elif not is_pm and hour == 12:
                hour = 0
            
            # If game_date is provided, parse it for multi-day sorting
            day_offset = 0
            if date_str:
                # date_str is like "Jan 19"
                try:
                    now = datetime.now(arizona_tz)
                    month_map = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                                 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}
                    parts = date_str.lower().split()
                    if len(parts) >= 2:
                        month = month_map.get(parts[0][:3], now.month)
                        day = int(parts[1])
                        if month < now.month or (month == now.month and day < now.day):
                            # Past date or today
                            day_offset = 0
                        else:
                            # Calculate days from now
                            from datetime import date
                            game_date = date(now.year, month, day)
                            today_date = now.date()
                            day_offset = (game_date - today_date).days
                except:
                    pass
            
            return (day_offset, hour * 60 + minute)
        except:
            return (9998, 0)  # Put unparseable times before "no time" items
    
    # Sort bets by time
    bets_sorted = sorted(bets, key=parse_time_for_sort)
    
    # For ENANO (jac075), get TIPSTER's bets to compare losses
    tipster_bet_keys = set()
    if account == "jac075":
        tipster_compilation = await db.daily_compilations.find_one({
            "account": "jac083",
            "date": today
        })
        if tipster_compilation and tipster_compilation.get('bets'):
            for tb in tipster_compilation['bets']:
                # Create a unique key for each bet (game + bet_type)
                key = f"{tb.get('game_short', '')}-{tb.get('bet_type_short', '')}"
                tipster_bet_keys.add(key)
    
    # Header differs for short vs detailed
    if detailed:
        lines = [f"📋 *{account_label}* (Detail)", ""]
    else:
        lines = [f"👤 *{account_label}*", ""]
    
    for i, bet in enumerate(bets_sorted, 1):
        # Use full game name for detailed, short for compact
        if detailed:
            game_name = bet.get('game', bet.get('game_short', 'GAME')).upper()
            # Clean up the game name
            game_name = game_name.replace('REG.TIME', '').strip()
        else:
            game_name = bet.get('game_short', 'GAME')
        
        bet_type_short = bet.get('bet_type_short', '')
        wager_short = bet.get('wager_short', '$0')
        to_win_short = bet.get('to_win_short', '$0')
        result = bet.get('result')
        game_time = bet.get('game_time', '')
        country = bet.get('country', '')
        is_new = bet.get('is_new', False)
        
        # Build line with game time if available
        # Add 🔵 prefix for NEW bets
        new_prefix = "🔵" if is_new else ""
        if game_time:
            bet_line = f"{new_prefix}#{i} {game_time} {game_name}"
        else:
            bet_line = f"{new_prefix}#{i} {game_name}"
        if bet_type_short:
            bet_line += f" {bet_type_short}"
        # Add country if available
        if country:
            bet_line += f" ({country})"
        bet_line += f" ({wager_short}/{to_win_short})"
        
        # Add result emoji
        if result == 'won':
            bet_line += "🟢"
        elif result == 'lost':
            # For ENANO: color-code losses by bet amount
            if account == "jac075":
                wager_short = bet.get('wager_short', '')
                if wager_short.startswith('$5') or wager_short.startswith('$6'):
                    bet_line += "🟤"  # Brown: Loss for $.5K bets
                elif wager_short.startswith('$1'):
                    bet_line += "🟣"  # Purple: Loss for $1K bets
                else:
                    bet_line += "🔴"  # Red: Loss for $2K+ bets
            else:
                bet_line += "🔴"
        elif result == 'push':
            bet_line += "🔵"
        else:
            bet_line += "🟡"
        
        lines.append(bet_line)
    
    # Add result total if any bets are settled
    settled_bets = [b for b in bets if b.get('result') in ['won', 'lost', 'push']]
    if settled_bets:
        lines.append("")
        result_sign = "+" if total_result >= 0 else ""
        lines.append(f"*Result: {result_sign}{format_amount_short(total_result)}*")
        
        # Add records
        if account == "jac075":
            # ENANO: Show 3 records (overall, $2K bets, $1K bets)
            overall_wins = len([b for b in bets if b.get('result') == 'won'])
            overall_losses = len([b for b in bets if b.get('result') == 'lost'])
            
            # $2K bets (includes $2K, $2.2K, etc.)
            bets_2k = [b for b in bets if b.get('wager_short', '').startswith('$2')]
            wins_2k = len([b for b in bets_2k if b.get('result') == 'won'])
            losses_2k = len([b for b in bets_2k if b.get('result') == 'lost'])
            
            # $1K bets (includes $1K, $1.3K, etc.)
            bets_1k = [b for b in bets if b.get('wager_short', '').startswith('$1')]
            wins_1k = len([b for b in bets_1k if b.get('result') == 'won'])
            losses_1k = len([b for b in bets_1k if b.get('result') == 'lost'])
            
            # $.5K bets ($500-$999 range, typically $500-$600)
            bets_05k = [b for b in bets if b.get('wager_short', '').startswith('$5') or b.get('wager_short', '').startswith('$6')]
            wins_05k = len([b for b in bets_05k if b.get('result') == 'won'])
            losses_05k = len([b for b in bets_05k if b.get('result') == 'lost'])
            
            lines.append(f"*Record: {overall_wins}-{overall_losses}*")
            lines.append(f"*$2K: {wins_2k}-{losses_2k}*")
            lines.append(f"*$1K: {wins_1k}-{losses_1k}*")
            lines.append(f"*$.5K: {wins_05k}-{losses_05k}*")
        else:
            # TIPSTER: Single overall record
            wins = len([b for b in bets if b.get('result') == 'won'])
            losses = len([b for b in bets if b.get('result') == 'lost'])
            lines.append(f"*Record: {wins}-{losses}*")
    
    return "\n".join(lines)


async def build_enano_comparison_message() -> str:
    """Build ENANO message comparing against TIPSTER bets
    
    Shows TIPSTER bets with indicators:
    - 🟢 = Placed and WON
    - 🔴 = Placed and LOST
    - 🟠 = MISSED (game started/ended, we didn't bet)
    - 🔵 = Placed (game not finished yet)
    - 🟡 = Pending (can still place)
    - (line)🟣 = Line difference indicator
    
    Then shows ENANO-only bets at the bottom (separated)
    Shows ENANO's own Result and Record at the bottom
    """
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    now = datetime.now(arizona_tz)
    today = now.strftime('%Y-%m-%d')
    today_day = now.day
    today_month = now.month
    current_time_minutes = now.hour * 60 + now.minute
    
    # Get TIPSTER bets
    tipster_comp = await db.daily_compilations.find_one({
        "account": "jac083",
        "date": today
    })
    
    # Get ENANO bets
    enano_comp = await db.daily_compilations.find_one({
        "account": "jac075",
        "date": today
    })
    
    tipster_bets = tipster_comp.get('bets', []) if tipster_comp else []
    enano_bets = enano_comp.get('bets', []) if enano_comp else []
    
    if not tipster_bets and not enano_bets:
        return None
    
    # Create a lookup for ENANO bets by game_short and bet_type_short
    enano_bet_keys = set()
    for bet in enano_bets:
        game = bet.get('game_short', bet.get('game', '')).upper()
        bet_type = bet.get('bet_type_short', bet.get('bet_type', '')).upper()
        enano_bet_keys.add(f"{game}|{bet_type}")
        enano_bet_keys.add(game)
    
    def parse_time_for_sort(bet):
        """Sort by: today by time first, then tomorrow by time"""
        time_str = bet.get('game_time', '')
        date_str = bet.get('game_date', '')
        
        if not time_str:
            return (9999, 9999)
        
        try:
            time_str = time_str.upper().strip()
            time_str = time_str.replace('AM', ' AM').replace('PM', ' PM').replace('  ', ' ')
            parts = time_str.replace(':', ' ').split()
            hour = int(parts[0])
            minute = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
            is_pm = 'PM' in time_str.upper()
            if is_pm and hour != 12:
                hour += 12
            elif not is_pm and hour == 12:
                hour = 0
            time_minutes = hour * 60 + minute
            
            day_offset = 0
            if date_str:
                month_map = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                             'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}
                parts = date_str.lower().split()
                if len(parts) >= 2:
                    month = month_map.get(parts[0][:3], today_month)
                    day = int(parts[1])
                    if day > today_day or month > today_month:
                        day_offset = 1
            
            return (day_offset, time_minutes)
        except:
            return (9998, 9998)
    
    def get_game_time_minutes(bet):
        """Get game time in minutes for comparison with current time"""
        time_str = bet.get('game_time', '')
        if not time_str:
            return None
        try:
            time_str = time_str.upper().strip()
            time_str = time_str.replace('AM', ' AM').replace('PM', ' PM').replace('  ', ' ')
            parts = time_str.replace(':', ' ').split()
            hour = int(parts[0])
            minute = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
            is_pm = 'PM' in time_str.upper()
            if is_pm and hour != 12:
                hour += 12
            elif not is_pm and hour == 12:
                hour = 0
            return hour * 60 + minute
        except:
            return None
    
    def is_bet_placed_by_enano(tipster_bet):
        """Check if ENANO has placed this TIPSTER bet
        Returns: (is_placed: bool, enano_line: str or None, enano_bet: dict or None)
        """
        import re
        
        game = tipster_bet.get('game_short', tipster_bet.get('game', '')).upper()
        game_full = tipster_bet.get('game', '').upper()
        bet_type = tipster_bet.get('bet_type_short', tipster_bet.get('bet_type', '')).upper()
        
        def extract_number(s):
            match = re.search(r'([+-]?\d+\.?\d*)', s)
            return float(match.group(1)) if match else None
        
        def get_direction(s):
            s = s.upper()
            if s.startswith('O') or 'OVER' in s:
                return 'O'
            if s.startswith('U') or 'UNDER' in s:
                return 'U'
            if '+' in s:
                return '+'
            if '-' in s:
                return '-'
            return None
        
        # Just return first match for backward compatibility
        all_matches = get_all_matching_enano_bets(tipster_bet)
        if all_matches:
            enano_bet, enano_line = all_matches[0]
            return True, enano_line, enano_bet
        return False, None, None
    
    def get_all_matching_enano_bets(tipster_bet):
        """Get ALL ENANO bets that match this TIPSTER bet (for duplicates)
        Returns: list of (enano_bet, enano_line) tuples
        """
        import re
        
        game = tipster_bet.get('game_short', tipster_bet.get('game', '')).upper()
        game_full = tipster_bet.get('game', '').upper()
        bet_type = tipster_bet.get('bet_type_short', tipster_bet.get('bet_type', '')).upper()
        
        def extract_number(s):
            match = re.search(r'([+-]?\d+\.?\d*)', s)
            return float(match.group(1)) if match else None
        
        def get_direction(s):
            s = s.upper()
            if s.startswith('O') or 'OVER' in s:
                return 'O'
            if s.startswith('U') or 'UNDER' in s:
                return 'U'
            if '+' in s:
                return '+'
            if '-' in s:
                return '-'
            return None
        
        matches = []
        # Check exact match first
        if f"{game}|{bet_type}" in enano_bet_keys:
            # Find the matching ENANO bet
            for enano_bet in enano_bets:
                if enano_bet.get('game_short', '').upper() == game and enano_bet.get('bet_type_short', '').upper() == bet_type:
                    return True, None, enano_bet
            return True, None, None
        
        # Check partial match
        for enano_bet in enano_bets:
            enano_game = enano_bet.get('game_short', enano_bet.get('game', '')).upper()
            enano_game_full = enano_bet.get('game', '').upper()
            enano_type = enano_bet.get('bet_type_short', enano_bet.get('bet_type', '')).upper()
            
            game_match = game in enano_game or enano_game in game
            
            if not game_match and game_full and enano_game_full:
                tipster_words = set(game_full.replace('VS', ' ').replace('/', ' ').split())
                enano_words = set(enano_game_full.replace('VS', ' ').replace('/', ' ').split())
                common_words = tipster_words & enano_words
                significant_common = [w for w in common_words if len(w) > 3]
                if len(significant_common) >= 1:
                    game_match = True
            
            if game_match:
                if bet_type == enano_type:
                    return True, None, enano_bet
                
                tipster_dir = get_direction(bet_type)
                enano_dir = get_direction(enano_type)
                tipster_num = extract_number(bet_type)
                enano_num = extract_number(enano_type)
                
                if tipster_dir and enano_dir and tipster_dir == enano_dir:
                    if tipster_num is not None and enano_num is not None:
                        if tipster_num != enano_num:
                            return True, enano_type, enano_bet
                        return True, None, enano_bet
                    return True, enano_type if enano_type != bet_type else None, enano_bet
                
                if enano_type in ['STRAIGHT', 'STRAIGHT BET', '']:
                    return True, None, enano_bet
                if bet_type in ['STRAIGHT', 'STRAIGHT BET', '']:
                    return True, enano_type, enano_bet
        
        return False, None, None
    
    def is_game_started_or_ended(bet):
        """Check if game has started or ended based on time"""
        result = bet.get('result')
        if result in ['won', 'lost', 'push']:
            return True
        
        date_str = bet.get('game_date', '')
        if date_str:
            month_map = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                         'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}
            parts = date_str.lower().split()
            if len(parts) >= 2:
                month = month_map.get(parts[0][:3], today_month)
                day = int(parts[1])
                if day > today_day or month > today_month:
                    return False
        
        game_time = get_game_time_minutes(bet)
        if game_time is None:
            return False
        
        return current_time_minutes >= game_time
    
    def is_tomorrow(bet):
        """Check if bet is for tomorrow"""
        date_str = bet.get('game_date', '')
        if date_str:
            month_map = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                         'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}
            parts = date_str.lower().split()
            if len(parts) >= 2:
                month = month_map.get(parts[0][:3], today_month)
                day = int(parts[1])
                if day > today_day or month > today_month:
                    return True
        return False
    
    # Sort TIPSTER bets by time
    tipster_sorted = sorted(tipster_bets, key=parse_time_for_sort)
    
    # Build ENANO comparison message
    lines = ["📋 *ENANO* (vs TIPSTER)", ""]
    
    # Split into today and tomorrow only (no separate completed section)
    today_bets = []
    tomorrow_bets = []
    
    for bet in tipster_sorted:
        if is_tomorrow(bet):
            tomorrow_bets.append(bet)
        else:
            today_bets.append(bet)
    
    # Find ENANO-only bets (not in TIPSTER) - track matched bets to handle duplicates
    enano_only_bets = []
    matched_enano_indices = set()  # Track which ENANO bets were matched
    
    # First, identify which ENANO bets match TIPSTER bets
    for tipster_bet in tipster_bets:
        tipster_game = tipster_bet.get('game_short', tipster_bet.get('game', '')).upper()
        tipster_game_full = tipster_bet.get('game', '').upper()
        tipster_type = tipster_bet.get('bet_type_short', tipster_bet.get('bet_type', '')).upper()
        
        for i, enano_bet in enumerate(enano_bets):
            if i in matched_enano_indices:
                continue  # Already matched to another TIPSTER bet
            
            enano_game = enano_bet.get('game_short', enano_bet.get('game', '')).upper()
            enano_game_full = enano_bet.get('game', '').upper()
            enano_type = enano_bet.get('bet_type_short', enano_bet.get('bet_type', '')).upper()
            
            game_match = enano_game in tipster_game or tipster_game in enano_game
            
            if not game_match and enano_game_full and tipster_game_full:
                enano_words = set(enano_game_full.replace('VS', ' ').replace('/', ' ').split())
                tipster_words = set(tipster_game_full.replace('VS', ' ').replace('/', ' ').split())
                common_words = enano_words & tipster_words
                significant_common = [w for w in common_words if len(w) > 3]
                if len(significant_common) >= 1:
                    game_match = True
            
            if game_match:
                type_match = False
                if enano_type in tipster_type or tipster_type in enano_type:
                    type_match = True
                elif ('O' in enano_type and 'O' in tipster_type) or ('U' in enano_type and 'U' in tipster_type):
                    type_match = True
                elif ('+' in enano_type and '+' in tipster_type) or ('-' in enano_type and '-' in tipster_type):
                    type_match = True
                elif enano_type in ['STRAIGHT', 'STRAIGHT BET', ''] or tipster_type in ['STRAIGHT', 'STRAIGHT BET', '']:
                    type_match = True
                
                if type_match:
                    matched_enano_indices.add(i)
                    break  # Only match one ENANO bet per TIPSTER bet
    
    # Unmatched ENANO bets go to ENANO Only
    for i, enano_bet in enumerate(enano_bets):
        if i not in matched_enano_indices:
            enano_only_bets.append(enano_bet)
    
    bet_num = 1
    
    # Track ENANO's results
    enano_wins = 0
    enano_losses = 0
    enano_result_amount = 0.0
    
    # Process TODAY's bets (all sorted by time, including completed)
    for bet in today_bets:
        game_name = bet.get('game', bet.get('game_short', 'GAME')).upper()
        game_name = game_name.replace('REG.TIME', '').strip()
        bet_type_short = bet.get('bet_type_short', '')
        wager_short = bet.get('wager_short', '$0')
        to_win_short = bet.get('to_win_short', '$0')
        tipster_result = bet.get('result')
        game_time = bet.get('game_time', '')
        country = bet.get('country', '')
        
        is_placed, enano_line, enano_bet = is_bet_placed_by_enano(bet)
        game_started = is_game_started_or_ended(bet)
        
        # Determine emoji based on status
        if is_placed and enano_bet:
            # ENANO placed this bet
            enano_result = enano_bet.get('result')
            enano_wager = enano_bet.get('wager', 0)
            enano_to_win = enano_bet.get('to_win', 0)
            
            if enano_result == 'won' or (not enano_result and tipster_result == 'won'):
                emoji = "🟢"
                if enano_line:
                    emoji += f"({enano_line})🟣"
                enano_wins += 1
                enano_result_amount += enano_to_win
            elif enano_result == 'lost' or (not enano_result and tipster_result == 'lost'):
                emoji = "🔴"
                if enano_line:
                    emoji += f"({enano_line})🟣"
                enano_losses += 1
                enano_result_amount -= enano_wager
            elif tipster_result in ['won', 'lost', 'push'] or game_started:
                # Game completed/started but ENANO's bet not graded yet
                if enano_line:
                    emoji = f"🔵({enano_line})🟣"
                else:
                    emoji = "🔵"
            else:
                # Game not started yet
                if enano_line:
                    emoji = f"🔵({enano_line})🟣"
                else:
                    emoji = "🔵"
        elif game_started:
            # ENANO missed (game started, not placed)
            emoji = "🟠"
        else:
            # Pending (can still place)
            emoji = "🟡"
        
        # Build line
        if game_time:
            bet_line = f"#{bet_num} {game_time} {game_name}"
        else:
            bet_line = f"#{bet_num} {game_name}"
        
        if bet_type_short and 'Straight' not in bet_type_short:
            bet_line += f" {bet_type_short}"
        if country:
            bet_line += f" ({country})"
        bet_line += f" ({wager_short}/{to_win_short}){emoji}"
        
        lines.append(bet_line)
        bet_num += 1
    
    # Add separator for tomorrow's bets
    if tomorrow_bets:
        lines.append("")
        
        for bet in tomorrow_bets:
            game_name = bet.get('game', bet.get('game_short', 'GAME')).upper()
            game_name = game_name.replace('REG.TIME', '').strip()
            bet_type_short = bet.get('bet_type_short', '')
            wager_short = bet.get('wager_short', '$0')
            to_win_short = bet.get('to_win_short', '$0')
            game_time = bet.get('game_time', '')
            country = bet.get('country', '')
            
            is_placed, enano_line, _ = is_bet_placed_by_enano(bet)
            if is_placed:
                if enano_line:
                    emoji = f"🔵({enano_line})🟣"
                else:
                    emoji = "🔵"
            else:
                emoji = "🟡"
            
            if game_time:
                bet_line = f"#{bet_num} {game_time} {game_name}"
            else:
                bet_line = f"#{bet_num} {game_name}"
            
            if bet_type_short and 'Straight' not in bet_type_short:
                bet_line += f" {bet_type_short}"
            if country:
                bet_line += f" ({country})"
            bet_line += f" ({wager_short}/{to_win_short}){emoji}"
            
            lines.append(bet_line)
            bet_num += 1
    
    # Add ENANO-only bets at the bottom
    if enano_only_bets:
        lines.append("")
        lines.append("*ENANO Only:*")
        
        enano_only_sorted = sorted(enano_only_bets, key=parse_time_for_sort)
        for bet in enano_only_sorted:
            game_name = bet.get('game', bet.get('game_short', 'GAME')).upper()
            game_name = game_name.replace('REG.TIME', '').strip()
            bet_type_short = bet.get('bet_type_short', '')
            wager_short = bet.get('wager_short', '$0')
            to_win_short = bet.get('to_win_short', '$0')
            result = bet.get('result')
            game_time = bet.get('game_time', '')
            country = bet.get('country', '')
            wager = bet.get('wager', 0)
            to_win = bet.get('to_win', 0)
            
            if result == 'won':
                emoji = "🟢"
                enano_wins += 1
                enano_result_amount += to_win
            elif result == 'lost':
                emoji = "🔴"
                enano_losses += 1
                enano_result_amount -= wager
            elif result == 'push':
                emoji = "🔵"
            else:
                emoji = "🟡"
            
            if game_time:
                bet_line = f"#{bet_num} {game_time} {game_name}"
            else:
                bet_line = f"#{bet_num} {game_name}"
            
            if bet_type_short and 'Straight' not in bet_type_short:
                bet_line += f" {bet_type_short}"
            if country:
                bet_line += f" ({country})"
            bet_line += f" ({wager_short}/{to_win_short}){emoji}"
            
            lines.append(bet_line)
            bet_num += 1
    else:
        lines.append("")
        lines.append("*ENANO Only:*")
        lines.append("-")
    
    # Add summary at the bottom
    total_tipster = len(tipster_bets)
    total_placed = sum(1 for b in tipster_bets if is_bet_placed_by_enano(b)[0])
    
    total_missed = 0
    for b in tipster_bets:
        is_placed, _, _ = is_bet_placed_by_enano(b)
        if is_placed:
            continue
        result = b.get('result')
        if result in ['won', 'lost', 'push'] or is_game_started_or_ended(b):
            total_missed += 1
    
    total_pending = total_tipster - total_placed - total_missed
    
    lines.append("")
    lines.append(f"*Copied: {total_placed}/{total_tipster}* | 🟠 Missed: {total_missed} | 🟡 Pending: {total_pending}")
    
    # Calculate ENANO's result from ALL ENANO bets (not just matched ones)
    total_enano_wins = 0
    total_enano_losses = 0
    total_enano_result = 0.0
    
    for eb in enano_bets:
        r = eb.get('result')
        w = eb.get('wager', 0)
        tw = eb.get('to_win', 0)
        if r == 'won':
            total_enano_wins += 1
            total_enano_result += tw
        elif r == 'lost':
            total_enano_losses += 1
            total_enano_result -= w
    
    # Add ENANO's own result and record if we have completed bets
    if total_enano_wins > 0 or total_enano_losses > 0:
        lines.append("")
        if total_enano_result >= 0:
            lines.append(f"*Result: +${total_enano_result:,.2f}*")
        else:
            lines.append(f"*Result: -${abs(total_enano_result):,.2f}*")
        lines.append(f"*Record: {total_enano_wins}-{total_enano_losses}*")
    
    return "\n".join(lines)


async def update_compilation_message(account: str):
    """Update the Telegram message - Both accounts get detailed messages only
    TIPSTER: Shows all bets with standard format
    ENANO: Shows comparison view against TIPSTER bets
    """
    if not telegram_bot or not telegram_chat_id:
        logger.info("Telegram not configured, skipping compilation update")
        return
    
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        today = datetime.now(arizona_tz).strftime('%Y-%m-%d')
        now = datetime.now(arizona_tz)
        
        compilation = await db.daily_compilations.find_one({
            "account": account,
            "date": today
        })
        
        if not compilation or not compilation.get('bets'):
            return
        
        # Generate detailed message based on account type
        if account == "jac083":  # TIPSTER
            detailed_message = await build_compilation_message(account, detailed=True)
        else:  # ENANO (jac075)
            detailed_message = await build_enano_comparison_message()
        
        if not detailed_message:
            return
        
        # Delete ALL old messages for this account
        all_compilations = await db.daily_compilations.find({"account": account}).to_list(100)
        for old_comp in all_compilations:
            for field in ['message_id_short', 'message_id_detailed', 'message_id']:
                old_id = old_comp.get(field)
                if old_id:
                    try:
                        await telegram_bot.delete_message(chat_id=telegram_chat_id, message_id=old_id)
                    except Exception:
                        pass
        
        # Send DETAILED message only (no more short messages)
        detailed_sent = await telegram_bot.send_message(
            chat_id=telegram_chat_id,
            text=detailed_message,
            parse_mode=ParseMode.MARKDOWN
        )
        
        # Store message ID
        await db.daily_compilations.update_one(
            {"account": account, "date": today},
            {"$set": {
                "message_id_detailed": detailed_sent.message_id,
                "message_id_short": None,
                "message_id": None
            }}
        )
        logger.info(f"Sent detailed compilation for {account}: {detailed_sent.message_id}")
        
        # Clear the is_new flag for all bets after message is sent
        bets = compilation.get('bets', [])
        for bet in bets:
            bet['is_new'] = False
        await db.daily_compilations.update_one(
            {"account": account, "date": today},
            {"$set": {"bets": bets}}
        )
        
        # Clean up old compilations from database (keep only last 7 days)
        seven_days_ago = (datetime.now(arizona_tz) - timedelta(days=7)).strftime('%Y-%m-%d')
        await db.daily_compilations.delete_many({
            "account": account,
            "date": {"$lt": seven_days_ago}
        })
    
    except Exception as e:
        logger.error(f"Failed to update compilation message: {str(e)}")

async def add_bet_to_compilation(account: str, bet_details: dict):
    """Add a new bet to the daily compilation and update Telegram"""
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    today = datetime.now(arizona_tz).strftime('%Y-%m-%d')
    
    # Prepare bet entry
    game = bet_details.get('game', '')
    description = bet_details.get('description', '')
    bet_type = bet_details.get('bet_type', '')
    wager = bet_details.get('wager', 0)
    to_win = bet_details.get('potential_win', wager)
    ticket = bet_details.get('ticket_number', '')
    game_time = bet_details.get('game_time', '')
    game_date = bet_details.get('game_date', '')
    country = bet_details.get('country', '')
    
    bet_entry = {
        "ticket": ticket,
        "game": game,
        "game_short": extract_short_game_name(game, description),
        "bet_type": bet_type,
        "bet_type_short": extract_bet_type_short(bet_type),
        "game_time": game_time,
        "game_date": game_date,
        "country": country,
        "wager": wager,
        "wager_short": format_amount_short(wager),
        "to_win": to_win,
        "to_win_short": format_amount_short(to_win),
        "result": None,
        "is_new": True,  # Flag to highlight new bets with 🔵
        "added_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Ensure compilation exists
    compilation = await db.daily_compilations.find_one({
        "account": account,
        "date": today
    })
    
    if not compilation:
        compilation = {
            "account": account,
            "date": today,
            "message_id": None,
            "bets": [],
            "total_result": 0,
            "created_at": datetime.now(timezone.utc)
        }
        await db.daily_compilations.insert_one(compilation)
    
    # Add bet to compilation
    await db.daily_compilations.update_one(
        {"account": account, "date": today},
        {"$push": {"bets": bet_entry}}
    )
    
    # Update Telegram message
    await update_compilation_message(account)
    logger.info(f"Added bet to compilation for {account}: {bet_entry['game_short']}")

async def update_bet_result_in_compilation(account: str, ticket: str, result: str, win_amount: float = 0):
    """Update a bet's result in the compilation and update Telegram"""
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    today = datetime.now(arizona_tz).strftime('%Y-%m-%d')
    
    compilation = await db.daily_compilations.find_one({
        "account": account,
        "date": today
    })
    
    if not compilation:
        logger.warning(f"No compilation found for {account} on {today}")
        return
    
    bets = compilation.get('bets', [])
    total_result = compilation.get('total_result', 0)
    
    # Find and update the bet
    for bet in bets:
        if bet.get('ticket') == ticket and bet.get('result') is None:
            bet['result'] = result
            if result == 'won':
                total_result += win_amount
            elif result == 'lost':
                total_result -= bet.get('wager', 0)
            # push doesn't change total
            break
    
    # Update in database
    await db.daily_compilations.update_one(
        {"account": account, "date": today},
        {"$set": {"bets": bets, "total_result": total_result}}
    )
    
    # Update Telegram message
    await update_compilation_message(account)
    logger.info(f"Updated result for ticket {ticket} in {account}'s compilation: {result}")

async def send_telegram_notification(bet_details: dict, account: str = None):
    """Send Telegram notification when a bet is placed - now uses compilation system"""
    if not telegram_bot or not telegram_chat_id:
        logger.info("Telegram not configured, skipping notification")
        return
    
    try:
        # Add to daily compilation instead of sending individual message
        await add_bet_to_compilation(account, bet_details)
        logger.info(f"Telegram compilation updated for Ticket#{bet_details.get('ticket_number')}")
        
    except Exception as e:
        logger.error(f"Failed to send Telegram notification: {str(e)}")


async def get_plays888_daily_totals(username: str, password: str) -> dict:
    """Scrape daily totals directly from plays888.co History page using Win/Loss row"""
    service = None
    try:
        service = Plays888Service()
        await service.initialize()
        
        login_result = await service.login(username, password)
        if not login_result["success"]:
            logger.error(f"Login failed for {username}: {login_result['message']}")
            return None
        
        # Go to History page
        await service.page.goto('https://www.plays888.co/wager/History.aspx', timeout=30000)
        await service.page.wait_for_timeout(4000)
        
        # Extract the daily summary table
        # The Win/Loss row has the actual daily profits directly - much more reliable
        # Supports both English (Mon, Tue, etc.) and Spanish (lun, mar, etc.) headers
        totals = await service.page.evaluate('''() => {
            const result = {
                daily_profits: [],
                week_total: null,
                win_loss_row: [],
                detected_language: null,
                debug_header: [],
                error: null
            };
            
            // Standard day names (we'll normalize everything to these)
            const standardDays = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun'];
            
            // Mapping from Spanish to English
            const spanishToEnglish = {
                'lun': 'mon',
                'mar': 'tue',
                'mié': 'wed',
                'mi\u00e9': 'wed',
                'mie': 'wed',
                'jue': 'thu',
                'vie': 'fri',
                'sáb': 'sat',
                's\u00e1b': 'sat',
                'sab': 'sat',
                'dom': 'sun'
            };
            
            try {
                const tables = document.querySelectorAll('table');
                
                for (const table of tables) {
                    const text = table.textContent.trim().toLowerCase();
                    
                    // Look for summary table - check for both English and Spanish patterns
                    const hasSpanishDays = text.includes('lun') && text.includes('mar') && text.includes('jue');
                    const hasEnglishDays = text.includes('mon') && text.includes('tue') && text.includes('thu');
                    const hasBeginning = text.includes('beginning');
                    
                    if (hasBeginning || hasSpanishDays || hasEnglishDays) {
                        const rows = table.querySelectorAll('tr');
                        
                        // Build header-to-day mapping from FIRST row
                        // Header: [empty, Beginning, Mon/lun, Tue/mar, Wed/mié, Thu/jue, Fri/vie, Sat/sáb, Sun/dom, Total]
                        // Index:   0      1          2        3        4        5        6        7        8        9
                        let headerDayMap = {}; // Maps cell index to normalized day name
                        
                        const firstRow = rows[0];
                        if (firstRow) {
                            const headerCells = firstRow.querySelectorAll('td, th');
                            result.debug_header = Array.from(headerCells).map(c => c.textContent.trim());
                            
                            for (let i = 0; i < headerCells.length; i++) {
                                let dayText = headerCells[i].textContent.trim().toLowerCase().replace(/\\xa0/g, ' ').replace(/\\s+/g, '');
                                
                                // Check Spanish days first
                                if (spanishToEnglish[dayText]) {
                                    headerDayMap[i] = spanishToEnglish[dayText];
                                    result.detected_language = 'spanish';
                                }
                                // Check English days
                                else if (standardDays.includes(dayText.substring(0, 3))) {
                                    headerDayMap[i] = dayText.substring(0, 3);
                                    result.detected_language = result.detected_language || 'english';
                                }
                                // Check for 'total'
                                else if (dayText.includes('total')) {
                                    headerDayMap[i] = 'total';
                                }
                                // Check for 'beginning'
                                else if (dayText.includes('beginning')) {
                                    headerDayMap[i] = 'beginning';
                                }
                            }
                        }
                        
                        // Find Win/Loss row and extract values using headerDayMap
                        for (const row of rows) {
                            const cells = row.querySelectorAll('td');
                            if (cells.length < 2) continue;
                            
                            const firstCell = cells[0].textContent.trim().toLowerCase();
                            
                            // Win/Loss row has the daily profits directly
                            if (firstCell.includes('win') || firstCell.includes('loss')) {
                                for (let i = 1; i < cells.length; i++) {
                                    const cellText = cells[i].textContent.trim();
                                    const match = cellText.match(/(-?[\\d,]+\\.\\d+)/);
                                    if (match) {
                                        const profit = parseFloat(match[1].replace(/,/g, ''));
                                        // Use headerDayMap to get the day name for this column
                                        const dayName = headerDayMap[i] || ('col' + i);
                                        result.win_loss_row.push({
                                            day: dayName,
                                            profit: profit
                                        });
                                    }
                                }
                                break;
                            }
                        }
                        
                        // Use Win/Loss row for daily profits
                        if (result.win_loss_row.length > 0) {
                            // Filter out Total, Beginning, and unknown columns for daily_profits
                            result.daily_profits = result.win_loss_row.filter(d => 
                                d.day !== 'total' && d.day !== 'beginning' && !d.day.startsWith('col')
                            );
                            // Get week total
                            const total = result.win_loss_row.find(d => d.day === 'total');
                            if (total) {
                                result.week_total = total.profit;
                            } else {
                                // Sum up all days if Total not found
                                result.week_total = result.daily_profits.reduce((sum, d) => sum + d.profit, 0);
                            }
                        }
                        
                        break;
                    }
                }
                
                // Count bet rows from the bet history table (rows with dates and dollar amounts)
                // Exclude ACCRUAL ADJUSTMENT rows which aren't actual bets
                result.total_bets = 0;
                const allTables = document.querySelectorAll('table');
                for (const table of allTables) {
                    const rows = table.querySelectorAll('tr');
                    for (const row of rows) {
                        const text = row.textContent;
                        // Bet rows have date format (MM/DD/YYYY) and Ticket #
                        // Exclude ACCRUAL ADJUSTMENT rows
                        if (text.match(/\\d{1,2}\\/\\d{1,2}\\/\\d{4}/) && 
                            text.match(/Ticket #:/i) && 
                            !text.toUpperCase().includes('ACCRUAL')) {
                            result.total_bets++;
                        }
                    }
                }
                
            } catch (e) {
                result.error = e.toString();
            }
            
            return result;
        }''')
        
        await service.close()
        logger.info(f"plays888 totals for {username}: {totals}")
        return totals
        
    except Exception as e:
        logger.error(f"Error getting plays888 totals: {str(e)}")
        if service:
            try:
                await service.close()
            except:
                pass
        return None


async def send_daily_summary():
    """Send daily betting summary to Telegram at 10:45 PM Arizona time"""
    if not telegram_bot or not telegram_chat_id:
        logger.info("Telegram not configured, skipping daily summary")
        return
    
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        now_arizona = datetime.now(arizona_tz)
        
        # Get day of week for matching plays888 data
        day_names_es = ['lun', 'mar', 'mié', 'jue', 'vie', 'sáb', 'dom']
        day_names_en = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']
        today_dow = now_arizona.weekday()  # 0=Monday
        today_day_es = day_names_es[today_dow]
        today_day_en = day_names_en[today_dow]
        
        # Get accounts
        connections = await db.connections.find({"is_connected": True}, {"_id": 0}).to_list(100)
        
        for conn in connections:
            username = conn["username"]
            password = decrypt_password(conn["password_encrypted"])
            label = ACCOUNT_LABELS.get(username, username)
            
            # Get plays888 daily totals
            totals = await get_plays888_daily_totals(username, password)
            
            if totals and totals.get('daily_profits'):
                # Day names are now normalized to English lowercase (mon, tue, wed, thu, fri, sat, sun)
                day_display_names = {
                    'mon': 'Monday', 'tue': 'Tuesday', 'wed': 'Wednesday',
                    'thu': 'Thursday', 'fri': 'Friday', 'sat': 'Saturday', 'sun': 'Sunday'
                }
                
                # Find today's profit using normalized day name
                today_profit = None
                for day_data in totals['daily_profits']:
                    day = day_data['day'].lower()
                    if day == today_day_en or day.startswith(today_day_en):
                        today_profit = day_data['profit']
                        break
                
                # Build week summary
                week_lines = []
                for day_data in totals['daily_profits']:
                    amt = day_data['profit']
                    emoji = "📈" if amt >= 0 else "📉"
                    day_name = day_display_names.get(day_data['day'], day_data['day'].capitalize())
                    week_lines.append(f"{emoji} {day_name}: ${amt:+,.2f}")
                
                week_text = "\n".join(week_lines) if week_lines else "No data"
                
                # Week total
                week_total = totals.get('week_total', 0)
                week_emoji = "📈" if week_total >= 0 else "📉"
                
                if today_profit is not None:
                    profit_emoji = "📈" if today_profit >= 0 else "📉"
                    profit_text = f"{profit_emoji} *Today's Profit:* ${today_profit:+,.2f} MXN"
                else:
                    profit_text = "⚠️ Could not get today's profit"
                
                message = f"""
📊 *{label} - DAILY SUMMARY*
📅 {now_arizona.strftime('%B %d, %Y')}

{profit_text}

📆 *This Week:*
{week_text}

{week_emoji} *Week Total:* ${week_total:+,.2f} MXN

_Data from plays888.co_
_Have a good night! 🌙_
                """
            else:
                message = f"""
📊 *{label} - DAILY SUMMARY*
📅 {now_arizona.strftime('%B %d, %Y')}

⚠️ Could not retrieve data from plays888.co

_Have a good night! 🌙_
                """
            
            await telegram_bot.send_message(
                chat_id=telegram_chat_id,
                text=message.strip(),
                parse_mode=ParseMode.MARKDOWN
            )
            logger.info(f"Daily summary sent for {label}")
        
        logger.info(f"Daily summaries sent")
        
    except Exception as e:
        logger.error(f"Failed to send daily summary: {str(e)}")


async def send_user_daily_summary(account: str, label: str, user_bets: list, now_arizona):
    """Send daily summary for a specific user"""
    try:
        # Filter out garbage entries (bets with $0 wager or no ticket)
        user_bets = [b for b in user_bets if b.get('wager_amount', 0) > 0 and b.get('bet_slip_id')]
        
        if not user_bets:
            message = f"""
📊 *{label} - DAILY SUMMARY*
📅 {now_arizona.strftime('%B %d, %Y')}

No bets placed today.

_Have a good night! 🌙_
            """
        else:
            total_wagered = sum(b.get('wager_amount', 0) for b in user_bets)
            
            # Calculate results
            won_bets = [b for b in user_bets if b.get('result') == 'won']
            lost_bets = [b for b in user_bets if b.get('result') == 'lost']
            push_bets = [b for b in user_bets if b.get('result') == 'push']
            pending_bets = [b for b in user_bets if not b.get('result') or b.get('result') == 'pending']
            
            total_won = sum(b.get('win_amount', 0) for b in won_bets)
            total_lost = sum(b.get('wager_amount', 0) for b in lost_bets)
            net_profit = total_won - total_lost
            
            # Build bet list with results
            bet_lines = []
            for i, bet in enumerate(user_bets[:15], 1):  # Limit to 15 bets per user
                game = bet.get('game', 'Unknown')[:25]
                bet_type = bet.get('bet_type', '')[:12]
                odds = format_american_odds(bet.get('odds', -110))
                wager = bet.get('wager_amount', 0)
                result = bet.get('result', '')
                
                # Result emoji
                if result == 'won':
                    result_emoji = "✅"
                elif result == 'lost':
                    result_emoji = "❌"
                elif result == 'push':
                    result_emoji = "↔️"
                else:
                    result_emoji = "⏳"
                
                bet_lines.append(f"{result_emoji} {game} | {bet_type} {odds} | ${wager}")
            
            bets_text = "\n".join(bet_lines)
            if len(user_bets) > 15:
                bets_text += f"\n_... and {len(user_bets) - 15} more bets_"
            
            # Profit/Loss indicator
            if net_profit > 0:
                profit_text = f"📈 *Net Profit:* +${net_profit:,.2f} MXN"
            elif net_profit < 0:
                profit_text = f"📉 *Net Loss:* -${abs(net_profit):,.2f} MXN"
            else:
                profit_text = f"➡️ *Net:* $0.00 MXN"
            
            message = f"""
📊 *{label} - DAILY SUMMARY*
📅 {now_arizona.strftime('%B %d, %Y')}

📈 *Results:*
• Total Bets: {len(user_bets)}
• ✅ Won: {len(won_bets)} (${total_won:,.2f})
• ❌ Lost: {len(lost_bets)} (${total_lost:,.2f})
• ↔️ Push: {len(push_bets)}
• ⏳ Pending: {len(pending_bets)}

💰 *Financials:*
• Total Wagered: ${total_wagered:,.2f} MXN
{profit_text}

🎯 *Today's Bets:*
{bets_text}

_⚠️ Numbers may differ slightly from plays888 due to timing_
_Have a good night! 🌙_
            """
        
        await telegram_bot.send_message(
            chat_id=telegram_chat_id,
            text=message.strip(),
            parse_mode=ParseMode.MARKDOWN
        )
        logger.info(f"Daily summary sent for {label} ({len(user_bets)} bets)")
        
    except Exception as e:
        logger.error(f"Failed to send daily summary for {label}: {str(e)}")


async def send_activity_summary():
    """Send daily activity summary showing all check times - runs at 10:59 PM Arizona"""
    if not telegram_bot or not telegram_chat_id:
        logger.info("Telegram not configured, skipping activity summary")
        return
    
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        now_arizona = datetime.now(arizona_tz)
        today_date = now_arizona.strftime('%Y-%m-%d')
        
        # Get all checks from today for TIPSTER (jac083)
        today_checks = await db.activity_log.find({
            "date": today_date,
            "type": "bet_check",
            "account": "jac083"
        }, {"_id": 0}).sort("timestamp", 1).to_list(1000)
        
        if not today_checks:
            check_times_text = "No checks performed today."
        else:
            # Group checks by hour for cleaner display
            check_times = [c.get('timestamp_arizona', 'Unknown') for c in today_checks]
            
            # Format as a list
            if len(check_times) <= 30:
                check_times_text = " • ".join(check_times)
            else:
                # If many checks, show summary
                check_times_text = f"{len(check_times)} checks performed\n"
                check_times_text += f"First: {check_times[0]} | Last: {check_times[-1]}\n"
                check_times_text += " • ".join(check_times[-10:])  # Show last 10
        
        message = f"""
🔄 *TIPSTER ACTIVITY SUMMARY*
📅 {now_arizona.strftime('%B %d, %Y')}

👤 *Account:* TIPSTER (jac083)
📡 *System Checks:* {len(today_checks)}

⏰ *Check Times (Arizona):*
{check_times_text}

✅ *System Status:* Active
🕐 *Sleep Hours:* 10:00 PM - 7:00 AM

_Betting summaries follow..._
        """
        
        await telegram_bot.send_message(
            chat_id=telegram_chat_id,
            text=message.strip(),
            parse_mode=ParseMode.MARKDOWN
        )
        logger.info(f"Activity summary sent to Telegram ({len(today_checks)} checks)")
        
    except Exception as e:
        logger.error(f"Failed to send activity summary: {str(e)}")


async def daily_cleanup():
    """Clean up old data at 9 AM Arizona time"""
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        now_arizona = datetime.now(arizona_tz)
        
        # Clean up yesterday's activity logs (activity logs only kept for 1 day)
        yesterday = (now_arizona - timedelta(days=1)).strftime('%Y-%m-%d')
        activity_deleted = await db.activity_log.delete_many({"date": {"$lte": yesterday}})
        logger.info(f"Cleaned up {activity_deleted.deleted_count} old activity logs")
        
        # Clean up bet history older than 7 days
        seven_days_ago = (now_arizona - timedelta(days=7))
        old_bets_deleted = await db.bet_history.delete_many({
            "placed_at": {"$lt": seven_days_ago.isoformat()}
        })
        logger.info(f"Cleaned up {old_bets_deleted.deleted_count} old bet records (>7 days)")
        
    except Exception as e:
        logger.error(f"Failed to run daily cleanup: {str(e)}")


async def check_bet_results():
    """Check plays888.co for settled bet results and update database"""
    logger.info("Checking for settled bet results...")
    
    try:
        # Get all connections
        connections = await db.connections.find({"is_connected": True}, {"_id": 0}).to_list(100)
        
        if not connections:
            logger.info("No active connections, skipping results check")
            return
        
        for conn in connections:
            await check_results_for_account(conn)
            
    except Exception as e:
        logger.error(f"Error checking bet results: {str(e)}")


async def check_results_for_account(conn: dict):
    """Check settled bets for a single account"""
    username = conn["username"]
    password = decrypt_password(conn["password_encrypted"])
    
    logger.info(f"Checking results for account: {username}")
    results_service = None
    
    try:
        results_service = Plays888Service()
        await results_service.initialize()
        
        # Login
        login_result = await results_service.login(username, password)
        if not login_result["success"]:
            logger.error(f"Results login failed for {username}: {login_result['message']}")
            await results_service.close()
            return
        
        # Navigate to Bet History / Graded Bets page
        # Common URLs to try: GradedBets.aspx, BetHistory.aspx, History.aspx
        history_urls = [
            'https://www.plays888.co/wager/GradedBets.aspx',
            'https://www.plays888.co/wager/BetHistory.aspx',
            'https://www.plays888.co/wager/History.aspx',
            'https://www.plays888.co/wager/SettledBets.aspx'
        ]
        
        page_loaded = False
        for url in history_urls:
            try:
                await results_service.page.goto(url, timeout=15000)
                await results_service.page.wait_for_timeout(3000)
                
                # Check if page has content (not a 404 or redirect)
                content = await results_service.page.content()
                if 'Ticket' in content or 'ticket' in content:
                    logger.info(f"Found history page: {url}")
                    page_loaded = True
                    break
            except:
                continue
        
        if not page_loaded:
            logger.info(f"Could not find bet history page for {username}")
            await results_service.close()
            return
        
        # Extract settled bets from the page
        # First, let's get debug info - look for rows with Ticket numbers
        debug_info = await results_service.page.evaluate('''() => {
            const rows = document.querySelectorAll('table tr');
            const debug = [];
            // Look for rows containing ticket info
            for (let i = 0; i < rows.length; i++) {
                const text = rows[i].textContent;
                // Only log rows that might have ticket info
                if (text.includes('Ticket') || text.includes('337') || text.includes('Result')) {
                    debug.push({
                        idx: i,
                        text: text.substring(0, 400),
                        cells: rows[i].querySelectorAll('td').length
                    });
                }
            }
            return {rowCount: rows.length, ticketRows: debug};
        }''')
        logger.info(f"History page: {debug_info['rowCount']} total rows, {len(debug_info['ticketRows'])} with ticket/result info")
        for row in debug_info['ticketRows'][:5]:
            logger.info(f"Row {row['idx']} ({row['cells']} cells): {row['text']}")
        
        settled_bets = await results_service.page.evaluate('''() => {
            const bets = [];
            const rows = document.querySelectorAll('table tr');
            
            for (let i = 0; i < rows.length; i++) {
                const row = rows[i];
                const cells = row.querySelectorAll('td');
                const rowText = row.textContent || '';
                
                // Look for ticket numbers - handle "Ticket #: 337..." or "Ticket#: 337..."
                const ticketMatch = rowText.match(/Ticket\\s*#?\\s*:?\\s*(\\d{9})/i);
                
                if (ticketMatch && cells.length >= 3) {
                    const ticket = ticketMatch[1];
                    
                    // Look for result indicators
                    // plays888 shows "WINWIN" or "LOSELOSE" at the end of each row
                    let result = 'pending';
                    const rowTextUpper = rowText.toUpperCase();
                    
                    // Check for specific patterns in plays888 format
                    if (rowTextUpper.includes('WINWIN') || rowTextUpper.endsWith('WIN')) {
                        result = 'won';
                    } 
                    else if (rowTextUpper.includes('LOSELOSE') || rowTextUpper.endsWith('LOSE') || rowTextUpper.endsWith('LOSS')) {
                        result = 'lost';
                    } 
                    else if (rowTextUpper.includes('PUSHPUSH') || rowTextUpper.includes('PUSH')) {
                        result = 'push';
                    } 
                    else if (rowTextUpper.includes('CANCEL') || rowTextUpper.includes('VOID')) {
                        result = 'cancelled';
                    }
                    
                    // Extract win amount - format is "2000.00WINWIN" or "-2200.00LOSELOSE"
                    let winAmount = 0;
                    
                    // Look for amount right before WIN or LOSE
                    const amountBeforeResult = rowText.match(/([\\d,]+\\.\\d+)(?:WINWIN|WIN)/i);
                    if (amountBeforeResult && result === 'won') {
                        winAmount = parseFloat(amountBeforeResult[1].replace(/,/g, ''));
                    }
                    
                    // Also get the wager amount from Risk/Win format "2200.00 / 2000.00"
                    const riskWinMatch = rowText.match(/([\\d,]+\\.\\d+)\\s*\\/\\s*([\\d,]+\\.\\d+)/);
                    let wagerAmount = 0;
                    if (riskWinMatch) {
                        wagerAmount = parseFloat(riskWinMatch[1].replace(/,/g, ''));
                        if (result === 'won' && winAmount === 0) {
                            winAmount = parseFloat(riskWinMatch[2].replace(/,/g, ''));
                        }
                    }
                    
                    if (result !== 'pending') {
                        bets.push({
                            ticket: ticket,
                            result: result,
                            winAmount: winAmount,
                            rowText: rowText.substring(0, 200)  // For debugging
                        });
                    }
                }
            }
            return bets;
        }''')
        
        logger.info(f"Found {len(settled_bets)} settled bets for {username}")
        
        # Update database with results
        results_updated = 0
        for bet in settled_bets:
            ticket_num = bet.get('ticket')
            result = bet.get('result')
            win_amount = bet.get('winAmount', 0)
            
            # Find and update the bet in database
            existing_bet = await db.bet_history.find_one({"bet_slip_id": ticket_num})
            
            if existing_bet and existing_bet.get('result') != result:
                # Update the result
                await db.bet_history.update_one(
                    {"bet_slip_id": ticket_num},
                    {"$set": {
                        "result": result,
                        "win_amount": win_amount,
                        "result_updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                results_updated += 1
                logger.info(f"Updated Ticket#{ticket_num}: {result}")
                
                # Send Telegram notification for result
                await send_result_notification(existing_bet, result, win_amount, username)
        
        await results_service.close()
        logger.info(f"Results check complete for {username}: {results_updated} bets updated")
        
    except Exception as e:
        logger.error(f"Error checking results for {username}: {str(e)}")
        if results_service:
            try:
                await results_service.close()
            except:
                pass


async def send_result_notification(bet: dict, result: str, win_amount: float, account: str = None):
    """Update the compilation when a bet result is determined"""
    if not telegram_bot or not telegram_chat_id:
        return
    
    try:
        ticket = bet.get('bet_slip_id', 'N/A')
        
        # Update the compilation with the result
        await update_bet_result_in_compilation(account, ticket, result, win_amount)
        logger.info(f"Compilation updated for Ticket#{ticket}: {result}")
        
    except Exception as e:
        logger.error(f"Failed to update result in compilation: {str(e)}")


# Playwright automation service
class Plays888Service:
    def __init__(self):
        self.browser = None
        self.context = None
        self.page = None
        self.playwright = None
        
    async def initialize(self):
        try:
            if not self.playwright:
                self.playwright = await async_playwright().start()
            if not self.browser:
                # Launch in headless mode with flags to help with JavaScript execution
                self.browser = await self.playwright.chromium.launch(
                    headless=True,
                    args=[
                        '--disable-blink-features=AutomationControlled',
                        '--disable-dev-shm-usage',
                        '--no-sandbox'
                    ]
                )
            if not self.context:
                # Set a realistic viewport and user agent
                self.context = await self.browser.new_context(
                    viewport={'width': 1920, 'height': 1080},
                    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
                )
            if not self.page:
                self.page = await self.context.new_page()
        except Exception as e:
            logger.error(f"Failed to initialize Playwright: {str(e)}")
            raise
    
    async def close(self):
        if self.page:
            await self.page.close()
        if self.context:
            await self.context.close()
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()
    
    async def login(self, username: str, password: str) -> Dict[str, Any]:
        try:
            await self.initialize()
            
            if not self.page:
                return {"success": False, "message": "Failed to initialize browser"}
            
            logger.info(f"Navigating to plays888.co for user {username}")
            await self.page.goto('https://www.plays888.co', timeout=30000)
            
            # Wait for page to load
            await self.page.wait_for_timeout(2000)
            
            # Look for login form
            try:
                # Try to find username/login input
                await self.page.wait_for_selector('input[type="text"], input[name*="user"], input[name*="login"]', timeout=5000)
                await self.page.fill('input[type="text"], input[name*="user"], input[name*="login"]', username)
                
                # Find password input
                await self.page.fill('input[type="password"]', password)
                
                # Look for submit button
                await self.page.click('button[type="submit"], input[type="submit"], button:has-text("Acceder")')
                
                # Wait for navigation or error
                await self.page.wait_for_timeout(3000)
                
                # Check if login was successful
                current_url = self.page.url
                page_content = await self.page.content()
                
                # Basic success check
                if 'error' not in page_content.lower() and 'invalid' not in page_content.lower():
                    logger.info(f"Login successful for {username}")
                    return {"success": True, "message": "Connected to plays888.co"}
                else:
                    return {"success": False, "message": "Login failed - check credentials"}
                    
            except PlaywrightTimeoutError:
                logger.error("Could not find login form")
                return {"success": False, "message": "Could not find login form on website"}
                
        except Exception as e:
            logger.error(f"Login error: {str(e)}")
            return {"success": False, "message": f"Login error: {str(e)}"}
    
    async def get_balance(self) -> Optional[float]:
        try:
            # Look for balance element - this would need to be adjusted based on actual website structure
            balance_text = await self.page.locator('text=/balance|saldo/i').first.text_content(timeout=5000)
            # Extract number from text
            import re
            match = re.search(r'\d+\.?\d*', balance_text)
            if match:
                return float(match.group())
        except Exception as e:
            logger.error(f"Error getting balance: {str(e)}")
        return None
    
    async def get_opportunities(self) -> List[Dict[str, Any]]:
        """Scrape available betting opportunities from the website"""
        opportunities = []
        try:
            # This is a mock implementation - actual implementation would scrape the site
            # Navigate to sports betting page
            await self.page.goto('https://www.plays888.co/Deportes.html', timeout=30000)
            await self.page.wait_for_timeout(2000)
            
            # Mock opportunities for demo - American odds
            opportunities = [
                {
                    "event_name": "Team A vs Team B",
                    "odds": +150,  # American odds
                    "sport": "Soccer",
                    "bet_type": "Match Winner"
                },
                {
                    "event_name": "Team C vs Team D",
                    "odds": -110,  # American odds
                    "sport": "Basketball",
                    "bet_type": "Match Winner"
                },
                {
                    "event_name": "Lakers vs Warriors",
                    "odds": +200,  # American odds
                    "sport": "Basketball",
                    "bet_type": "Spread"
                },
            ]
            
        except Exception as e:
            logger.error(f"Error getting opportunities: {str(e)}")
        
        return opportunities
    
    async def place_specific_bet(self, game: str, bet_type: str, line: str, odds: int, wager: float, league: str = "NATIONAL HOCKEY LEAGUE - OT INCLUDED") -> Dict[str, Any]:
        """
        Place a specific bet on plays888.co
        NOTE: User must be connecting from Phoenix, Arizona IP for accurate location
        """
        try:
            if not self.page:
                return {"success": False, "message": "Browser not initialized"}
            
            logger.info(f"Placing bet: {game} - {bet_type} {line} @ {odds} for ${wager}")
            
            # Step 1: Navigate to plays888.co (should already be logged in)
            await self.page.goto('https://www.plays888.co/wager/Welcome.aspx', timeout=30000)
            await self.page.wait_for_load_state('networkidle')
            await self.page.wait_for_timeout(2000)
            
            await self.page.screenshot(path="/tmp/step1_welcome.png")
            logger.info("Step 1: Welcome page loaded")
            
            # Step 2: Click "Straight" in the left sidebar
            try:
                # Look for the Straight link in the sidebar
                await self.page.click('a:has-text("Straight")', timeout=10000)
                await self.page.wait_for_timeout(2000)
                logger.info("Step 2: Clicked 'Straight' in sidebar")
            except Exception as e:
                logger.error(f"Could not find Straight link: {str(e)}")
                return {"success": False, "message": f"Could not find 'Straight' section: {str(e)}"}
            
            await self.page.screenshot(path="/tmp/step2_league_selection.png")
            
            # Step 3: Check the league checkbox and click Continue
            try:
                # Find the checkbox for the specified league
                # The checkbox is next to the league text
                await self.page.click(f'text=/{league}/i', timeout=10000)
                await self.page.wait_for_timeout(1000)
                logger.info(f"Step 3: Checked '{league}' checkbox")
                
                # Click Continue button at bottom - use force to bypass overlays
                await self.page.click('input[value="Continue"]', force=True, timeout=5000)
                logger.info("Clicked Continue, waiting for games to load...")
                
                # CRITICAL: Wait for the games table to actually populate
                # The URL stays the same but content changes via AJAX
                # Wait for the page to finish loading
                await self.page.wait_for_load_state('networkidle', timeout=15000)
                
                # Wait longer and check for game content to appear
                logger.info("Waiting for game data to populate via AJAX...")
                games_loaded = False
                for i in range(15):  # Try up to 15 times (30 seconds)
                    await self.page.wait_for_timeout(2000)
                    
                    # Check multiple indicators that games have loaded
                    button_count = await self.page.locator('input[type="submit"][value*="+"], input[type="submit"][value*="-"]').count()
                    
                    # Also check for team names or game content
                    page_text = await self.page.content()
                    has_game_content = 'vs' in page_text.lower() or 'vrs' in page_text.lower()
                    
                    logger.info(f"Attempt {i+1}: Found {button_count} betting buttons, has_game_content={has_game_content}")
                    
                    if button_count > 10 and has_game_content:  # Games have loaded
                        games_loaded = True
                        logger.info(f"Games loaded successfully! Found {button_count} betting options")
                        break
                
                if not games_loaded:
                    # Take screenshot for debugging
                    await self.page.screenshot(path="/tmp/games_not_loaded.png")
                    return {
                        "success": False,
                        "message": "Games did not load after waiting 30 seconds. Check /tmp/games_not_loaded.png"
                    }
                
                logger.info("Step 3: Games loaded successfully")
            except Exception as e:
                logger.error(f"Could not select league: {str(e)}")
                return {"success": False, "message": f"Could not select league: {str(e)}"}
            
            await self.page.screenshot(path="/tmp/step3_games_list.png")
            
            # Check current URL
            current_url = self.page.url
            logger.info(f"Current URL after Continue: {current_url}")
            
            # Step 4: Find the game and click the odds button
            try:
                # Format odds - need to handle both positive and negative
                # e.g., +110, -125, o150+110
                odds_text = f"+{odds}" if odds > 0 else str(odds)
                
                # Save HTML for debugging
                page_content = await self.page.content()
                with open("/tmp/step3_games_html.html", "w", encoding="utf-8") as f:
                    f.write(page_content)
                logger.info("Saved games page HTML for debugging")
                
                # Find all betting buttons (input type="submit")
                all_inputs = await self.page.query_selector_all('input[type="submit"]')
                logger.info(f"Found {len(all_inputs)} total input buttons on page")
                
                # Look for the specific odds button
                # The odds might be part of a larger string like "o150+110" or just "+110"
                clicked = False
                matched_button = None
                
                for input_elem in all_inputs:
                    try:
                        value = await input_elem.get_attribute('value')
                        # Check if this button contains our odds
                        if value and odds_text in value:
                            matched_button = input_elem
                            logger.info(f"Found matching button with value: {value}")
                            await input_elem.click(force=True)
                            clicked = True
                            logger.info(f"Step 4: Clicked odds button containing '{odds_text}' (full value: {value})")
                            break
                    except Exception as e:
                        logger.error(f"Error checking button: {str(e)}")
                        continue
                
                if not clicked:
                    # Log all button values for debugging
                    logger.error("Could not find matching odds button. Available buttons:")
                    for i, input_elem in enumerate(all_inputs[:20]):  # Log first 20
                        try:
                            value = await input_elem.get_attribute('value')
                            logger.error(f"  Button {i}: {value}")
                        except:
                            pass
                    
                    return {
                        "success": False, 
                        "message": f"Could not find odds button containing '{odds_text}' for game. Found {len(all_inputs)} buttons total. Check /tmp/step3_games_html.html and logs for details."
                    }
                
                # Wait for button selection to register
                await self.page.wait_for_timeout(2000)
                
                # Click Continue button to go to bet slip
                await self.page.click('input[value="Continue"]', force=True, timeout=5000)
                await self.page.wait_for_load_state('networkidle')
                await self.page.wait_for_timeout(2000)
                logger.info("Step 4: Clicked Continue after selecting odds")
                
            except Exception as e:
                logger.error(f"Could not find/click odds: {str(e)}")
                return {"success": False, "message": f"Could not find/click odds: {str(e)}"}
            
            await self.page.screenshot(path="/tmp/step4_betslip.png")
            
            # Step 5: On bet slip page - Select "To Win Amount" and enter amount
            try:
                # Wait for bet slip page to load
                await self.page.wait_for_selector('input[value="To Win Amount"]', timeout=10000)
                logger.info("Bet slip page loaded")
                
                await self.page.screenshot(path="/tmp/step4_betslip.png")
                
                # Select "To Win Amount" radio button
                await self.page.click('input[value="To Win Amount"]', force=True, timeout=5000)
                await self.page.wait_for_timeout(500)
                logger.info("Step 5: Selected 'To Win Amount' radio button")
                
                # Find the text input field for amount and clear it first
                # Look for visible text input (not password, not hidden)
                input_field = await self.page.query_selector('input[type="text"]:not([style*="display: none"]):not([style*="display:none"])')
                if input_field:
                    await input_field.click()
                    await input_field.fill('')  # Clear first
                    await input_field.fill(str(int(wager)))
                    logger.info(f"Step 5: Entered wager amount: ${wager}")
                else:
                    logger.error("Could not find amount input field")
                    return {"success": False, "message": "Could not find amount input field"}
                
                await self.page.wait_for_timeout(1000)
                await self.page.screenshot(path="/tmp/step5_amount_entered.png")
                
                # Click Continue to go to confirmation
                await self.page.click('input[value="Continue"]', force=True, timeout=5000)
                await self.page.wait_for_load_state('networkidle')
                await self.page.wait_for_timeout(2000)
                logger.info("Step 5: Clicked Continue, going to confirmation")
                
            except Exception as e:
                logger.error(f"Could not enter wager amount: {str(e)}")
                return {"success": False, "message": f"Could not enter wager amount: {str(e)}"}
            
            await self.page.screenshot(path="/tmp/step5_confirmation.png")
            
            # Step 6: On confirmation page - Click Confirm button to place bet
            try:
                # Wait for confirmation page to load
                await self.page.wait_for_selector('input[value="Confirm"]', timeout=10000)
                logger.info("Confirmation page loaded")
                
                await self.page.screenshot(path="/tmp/step5_confirmation_page.png")
                
                # Click Confirm to place the bet
                await self.page.click('input[value="Confirm"]', force=True, timeout=5000)
                await self.page.wait_for_load_state('networkidle')
                await self.page.wait_for_timeout(3000)
                logger.info("Step 6: Clicked Confirm button - bet should be placed!")
                
            except Exception as e:
                logger.error(f"Could not click Confirm: {str(e)}")
                await self.page.screenshot(path="/tmp/confirm_error.png")
                return {"success": False, "message": f"Could not click Confirm: {str(e)}"}
            
            await self.page.screenshot(path="/tmp/step6_success_page.png")
            
            # Step 7: Verify bet was placed by checking for Ticket# on success page
            try:
                # Check if we're on the ConfirmWager.aspx success page
                current_url = self.page.url
                logger.info(f"Final URL: {current_url}")
                
                # Look for "Ticket#" text on the page
                page_content = await self.page.content()
                with open("/tmp/success_page.html", "w", encoding="utf-8") as f:
                    f.write(page_content)
                
                # Extract ticket number
                import re
                ticket_match = re.search(r'Ticket#?[:\s]*(\d+)', page_content, re.IGNORECASE)
                
                if ticket_match or "ConfirmWager" in current_url:
                    ticket_number = ticket_match.group(1) if ticket_match else "Check screenshot"
                    
                    logger.info(f"🎉 BET PLACED SUCCESSFULLY! Ticket#: {ticket_number}")
                    
                    return {
                        "success": True,
                        "message": f"✅ Bet placed successfully: {game} - {bet_type} {line} @ {odds} for ${wager} MXN",
                        "ticket_number": ticket_number,
                        "bet_details": {
                            "game": game,
                            "bet_type": bet_type,
                            "line": line,
                            "odds": odds,
                            "wager": wager,
                            "league": league
                        },
                        "verification": "Check /tmp/step6_success_page.png for visual confirmation"
                    }
                else:
                    logger.error("Could not verify bet placement")
                    return {
                        "success": False,
                        "message": "Reached final page but could not verify bet placement (no Ticket# found)",
                        "current_url": current_url,
                        "screenshots": "Check /tmp/step*.png and /tmp/success_page.html for debugging"
                    }
                    
            except Exception as e:
                logger.error(f"Error verifying bet: {str(e)}")
                return {"success": False, "message": f"Error verifying bet: {str(e)}"}
                
        except Exception as e:
            logger.error(f"Error placing bet: {str(e)}")
            await self.page.screenshot(path="/tmp/error.png")
            return {"success": False, "message": f"Error: {str(e)}", "screenshot": "/tmp/error.png"}

    async def scrape_totals(self, league: str = "NBA") -> List[Dict[str, Any]]:
        """
        Scrape over/under totals from plays888.co for a specific league
        Returns list of games with team names and total lines
        """
        import re
        
        try:
            if not self.page:
                logger.error("Browser not initialized")
                return []
            
            logger.info(f"Scraping {league} totals from plays888.co")
            
            # Navigate directly to the CreateSports page (Straight bets)
            await self.page.goto('https://www.plays888.co/wager/CreateSports.aspx?WT=0', timeout=30000)
            await self.page.wait_for_load_state('networkidle')
            await self.page.wait_for_timeout(2000)
            
            # Determine checkbox ID and continue button based on league
            # For NCAAB: Need to check BOTH "NCAA BASKETBALL - MEN" and "NCAA BASKETBALL - MEN EXTRA GAMES"
            if league.upper() == "NBA":
                checkbox_id = "lg_3"  # NBA checkbox
                card_heading = "heading4"  # Baloncesto section
                checkbox_ids = ["lg_3"]
            elif league.upper() == "NHL":
                checkbox_id = "lg_1166"  # NHL - OT INCLUDED checkbox
                card_heading = "heading12"  # Hockey section (adjust if needed)
                checkbox_ids = ["lg_1166"]
            elif league.upper() == "NFL":
                checkbox_id = "lg_5207"  # NFL - GAME LINES checkbox
                card_heading = "heading1"  # Futbol Americano section
                checkbox_ids = ["lg_5207"]
            elif league.upper() in ["NCAAB", "CBB"]:
                # NCAAB has two sections on plays888:
                # 1. NCAA BASKETBALL - MEN
                # 2. NCAA BASKETBALL - MEN EXTRA GAMES
                # We need to scrape from BOTH
                checkbox_id = "lg_1"  # Primary NCAA BASKETBALL - MEN
                card_heading = "heading3"  # Baloncesto Universitario section
                checkbox_ids = ["lg_1"]  # Will also try "lg_1268" for EXTRA GAMES
            else:
                logger.error(f"Unsupported league: {league}")
                return []
            
            # Check the league checkbox(es) using JavaScript
            checkbox_result = await self.page.evaluate(f'''
                () => {{
                    const checkboxIds = {checkbox_ids};
                    let checked = 0;
                    for (const id of checkboxIds) {{
                        const checkbox = document.getElementById(id);
                        if (checkbox) {{
                            checkbox.checked = true;
                            checkbox.dispatchEvent(new Event('change', {{ bubbles: true }}));
                            checked++;
                        }}
                    }}
                    return checked > 0;
                }}
            ''')
            
            if not checkbox_result:
                logger.error(f"Could not find checkbox for {league}")
                return []
            
            await self.page.wait_for_timeout(500)
            
            # Click Continue button using JavaScript (ASP.NET postback)
            continue_result = await self.page.evaluate(f'''
                () => {{
                    // Find checkbox and its card
                    const checkbox = document.getElementById("{checkbox_id}");
                    if (!checkbox) return false;
                    
                    const card = checkbox.closest('.card');
                    if (card) {{
                        const continueBtn = card.querySelector('input[value="Continue"]');
                        if (continueBtn) {{
                            continueBtn.click();
                            return true;
                        }}
                    }}
                    
                    // Fallback: trigger ASP.NET postback directly
                    if (typeof __doPostBack !== 'undefined') {{
                        const buttons = document.querySelectorAll('input[value="Continue"]');
                        for (const btn of buttons) {{
                            const rect = btn.getBoundingClientRect();
                            if (rect.top >= 0 && rect.bottom <= window.innerHeight) {{
                                btn.click();
                                return true;
                            }}
                        }}
                    }}
                    return false;
                }}
            ''')
            
            await self.page.wait_for_timeout(5000)
            await self.page.wait_for_load_state('networkidle', timeout=15000)
            
            # Take screenshot for debugging
            await self.page.screenshot(path=f"/tmp/plays888_{league.lower()}_totals.png")
            
            # Get page text for parsing
            page_text = await self.page.inner_text('body')
            
            # Check if error message appeared
            if "select at least one League" in page_text:
                logger.error(f"Failed to select {league} - league selection error")
                return []
            
            # Parse games from the page text
            # NBA format: "541 BROOKLYN NETS" (3 digit game number)
            # NHL format: "53 PITTSBURGH PENGUINS" (2 digit game number)
            games = []
            lines = page_text.split('\n')
            
            i = 0
            current_away_team = None
            current_time = None
            
            while i < len(lines):
                line = lines[i].strip()
                
                # Look for game number + team name pattern (2-3 digit game number)
                team_match = re.match(r'^\d{2,3}\s+(.+)$', line)
                if team_match:
                    team_name = team_match.group(1).strip()
                    
                    # Look ahead for the total line (o/u pattern)
                    # NBA uses format like o219-110 (3 digits), NHL uses o6-120 or o5½-110 (1-2 digits)
                    for j in range(i + 1, min(i + 5, len(lines))):
                        next_line = lines[j].strip()
                        # Updated regex to match 1-3 digit totals
                        total_match = re.match(r'^[ou](\d{1,3}[½]?)[-+]\d+$', next_line, re.IGNORECASE)
                        if total_match:
                            total_str = total_match.group(1).replace('½', '.5')
                            total = float(total_str)
                            
                            if current_away_team:
                                # This is the home team, complete the game
                                games.append({
                                    "away": current_away_team,
                                    "home": team_name,
                                    "total": total,
                                    "time": current_time
                                })
                                current_away_team = None
                                current_time = None
                            else:
                                # This is the away team
                                current_away_team = team_name
                            break
                
                # Look for time pattern (e.g., "4:10 PM")
                time_match = re.match(r'^(\d{1,2}:\d{2}\s*[AP]M)$', line, re.IGNORECASE)
                if time_match and current_away_team:
                    current_time = time_match.group(1)
                
                i += 1
            
            logger.info(f"Found {len(games)} games with totals for {league}")
            return games
            
        except Exception as e:
            logger.error(f"Error scraping totals: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    async def scrape_open_bets(self) -> List[Dict[str, Any]]:
        """
        Scrape open/pending bets from plays888.co
        Returns list of open bets with game info and bet type (over/under)
        """
        import re
        
        # List of official NHL teams to filter out non-NHL hockey leagues
        NHL_TEAMS = [
            'ANAHEIM DUCKS', 'ARIZONA COYOTES', 'BOSTON BRUINS', 'BUFFALO SABRES',
            'CALGARY FLAMES', 'CAROLINA HURRICANES', 'CHICAGO BLACKHAWKS', 'COLORADO AVALANCHE',
            'COLUMBUS BLUE JACKETS', 'DALLAS STARS', 'DETROIT RED WINGS', 'EDMONTON OILERS',
            'FLORIDA PANTHERS', 'LOS ANGELES KINGS', 'MINNESOTA WILD', 'MONTREAL CANADIENS',
            'NASHVILLE PREDATORS', 'NEW JERSEY DEVILS', 'NEW YORK ISLANDERS', 'NEW YORK RANGERS',
            'OTTAWA SENATORS', 'PHILADELPHIA FLYERS', 'PITTSBURGH PENGUINS', 'SAN JOSE SHARKS',
            'SEATTLE KRAKEN', 'ST. LOUIS BLUES', 'TAMPA BAY LIGHTNING', 'TORONTO MAPLE LEAFS',
            'UTAH MAMMOTH', 'VANCOUVER CANUCKS', 'VEGAS GOLDEN KNIGHTS', 'WASHINGTON CAPITALS',
            'WINNIPEG JETS'
        ]
        
        # NBA teams for filtering
        NBA_TEAMS = [
            'ATLANTA HAWKS', 'BOSTON CELTICS', 'BROOKLYN NETS', 'CHARLOTTE HORNETS',
            'CHICAGO BULLS', 'CLEVELAND CAVALIERS', 'DALLAS MAVERICKS', 'DENVER NUGGETS',
            'DETROIT PISTONS', 'GOLDEN STATE WARRIORS', 'HOUSTON ROCKETS', 'INDIANA PACERS',
            'LOS ANGELES CLIPPERS', 'LOS ANGELES LAKERS', 'MEMPHIS GRIZZLIES', 'MIAMI HEAT',
            'MILWAUKEE BUCKS', 'MINNESOTA TIMBERWOLVES', 'NEW ORLEANS PELICANS', 'NEW YORK KNICKS',
            'OKLAHOMA CITY THUNDER', 'ORLANDO MAGIC', 'PHILADELPHIA 76ERS', 'PHOENIX SUNS',
            'PORTLAND TRAIL BLAZERS', 'SACRAMENTO KINGS', 'SAN ANTONIO SPURS', 'TORONTO RAPTORS',
            'UTAH JAZZ', 'WASHINGTON WIZARDS'
        ]
        
        # NFL teams for filtering
        NFL_TEAMS = [
            'ARIZONA CARDINALS', 'ATLANTA FALCONS', 'BALTIMORE RAVENS', 'BUFFALO BILLS',
            'CAROLINA PANTHERS', 'CHICAGO BEARS', 'CINCINNATI BENGALS', 'CLEVELAND BROWNS',
            'DALLAS COWBOYS', 'DENVER BRONCOS', 'DETROIT LIONS', 'GREEN BAY PACKERS',
            'HOUSTON TEXANS', 'INDIANAPOLIS COLTS', 'JACKSONVILLE JAGUARS', 'KANSAS CITY CHIEFS',
            'LAS VEGAS RAIDERS', 'LOS ANGELES CHARGERS', 'LOS ANGELES RAMS', 'MIAMI DOLPHINS',
            'MINNESOTA VIKINGS', 'NEW ENGLAND PATRIOTS', 'NEW ORLEANS SAINTS', 'NEW YORK GIANTS',
            'NEW YORK JETS', 'PHILADELPHIA EAGLES', 'PITTSBURGH STEELERS', 'SAN FRANCISCO 49ERS',
            'SEATTLE SEAHAWKS', 'TAMPA BAY BUCCANEERS', 'TENNESSEE TITANS', 'WASHINGTON COMMANDERS'
        ]
        
        def is_nhl_team(team_name):
            team_upper = team_name.upper()
            return any(nhl_team in team_upper or team_upper in nhl_team for nhl_team in NHL_TEAMS)
        
        def is_nba_team(team_name):
            team_upper = team_name.upper()
            return any(nba_team in team_upper or team_upper in nba_team for nba_team in NBA_TEAMS)
        
        def is_nfl_team(team_name):
            team_upper = team_name.upper()
            return any(nfl_team in team_upper or team_upper in nfl_team for nfl_team in NFL_TEAMS)
        
        try:
            if not self.page:
                logger.error("Browser not initialized")
                return []
            
            logger.info("Scraping open bets from plays888.co")
            
            # Navigate to Open Bets page
            await self.page.goto('https://www.plays888.co/wager/OpenBets.aspx', timeout=30000)
            await self.page.wait_for_load_state('networkidle')
            await self.page.wait_for_timeout(2000)
            
            # Get page content
            page_text = await self.page.inner_text('body')
            
            # Parse open bets - collect raw bets first
            raw_bets = []
            
            # Pattern to match totals: TOTAL o6-110 or u6-110
            # Format: (TEAM1 vrs TEAM2) or (TEAM1 REG.TIME vrs TEAM2 REG.TIME)
            lines = page_text.split('\n')
            
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                
                # Look for TOTAL lines with o/u (Over/Under bets)
                total_match = re.search(r'TOTAL\s+([ou])(\d+\.?\d*)(½)?[-+]\d+', line, re.IGNORECASE)
                # Look for SPREAD bets like "[837] OAKLAND +2-110" or "[811] DETROIT +11-110"
                spread_match = re.search(r'\[\d+\]\s*([A-Z\s\.]+?)\s*([+-]\d+\.?\d*)(½)?[-+]\d+', line, re.IGNORECASE)
                # Look for live betting format: "Over 161.5 -106" or "Under 161.5 -106"
                live_total_match = re.search(r'(Over|Under)\s+(\d+\.?\d*)\s*[-+]\d+', line, re.IGNORECASE)
                
                if total_match:
                    bet_type = 'OVER' if total_match.group(1).lower() == 'o' else 'UNDER'
                    total_line = float(total_match.group(2))
                    # Add 0.5 if there's a ½ symbol
                    if total_match.group(3) == '½':
                        total_line += 0.5
                    
                    # Look for team names in nearby lines
                    teams_text = ""
                    for j in range(max(0, i-2), min(len(lines), i+3)):
                        if 'vrs' in lines[j].lower() or ' vs ' in lines[j].lower():
                            teams_text = lines[j]
                            break
                    
                    # Extract team names from "(TEAM1 vrs TEAM2)" or "(TEAM1 REG.TIME vrs TEAM2 REG.TIME)"
                    # Also support live betting format: "TEAM1 vs TEAM2 / Game / Total"
                    # FIXED: Handle team names with parentheses like "Miami (Ohio)" or "Miami (OH)"
                    # Use a pattern that properly handles nested parentheses
                    paren_match = re.search(r'\((.+?)\s+(?:REG\.TIME\s+)?vrs\s+(.+)\)', teams_text, re.IGNORECASE)
                    teams_match = paren_match
                    
                    # If no match, try live betting format: "Team1 vs Team2 / Game / Total"
                    if not teams_match:
                        live_match = re.search(r'([A-Za-z\s\.]+?)\s+vs\s+([A-Za-z\s\.]+?)\s*/\s*Game', teams_text, re.IGNORECASE)
                        if live_match:
                            teams_match = live_match
                    if teams_match:
                        away_team = teams_match.group(1).strip().replace(' REG.TIME', '')
                        home_team = teams_match.group(2).strip().replace(' REG.TIME', '')
                        
                        # Check for international/European basketball (not NCAAB)
                        intl_patterns = ['CSM ', 'BC ', 'KK ', 'TARGU', 'PLOIESTI', 'BUCHAREST', 'CLUJ', 
                                        'SIBIU', 'ORADEA', 'STEAUA', 'DINAMO', 'RAPID', 'FCSB']
                        is_international = any(pattern in away_team.upper() or pattern in home_team.upper() 
                                              for pattern in intl_patterns)
                        
                        # Determine sport based on team names and context
                        sport = None
                        if is_nhl_team(away_team) and is_nhl_team(home_team):
                            sport = 'NHL'
                        elif is_nba_team(away_team) and is_nba_team(home_team):
                            sport = 'NBA'
                        elif is_nfl_team(away_team) and is_nfl_team(home_team):
                            sport = 'NFL'
                        elif is_international:
                            sport = 'INTL_BASKETBALL'  # Romanian/European basketball
                        else:
                            # Assume NCAAB/CBB for college basketball
                            sport = 'NCAAB'
                        
                        # Look for risk amount
                        risk_match = re.search(r'(\d{1,},?\d+\.?\d*)\s*/\s*(\d{1,},?\d+\.?\d*)', lines[i+1] if i+1 < len(lines) else '')
                        risk_amount = 0
                        win_amount = 0
                        if risk_match:
                            risk_amount = float(risk_match.group(1).replace(',', ''))
                            win_amount = float(risk_match.group(2).replace(',', ''))
                        
                        raw_bets.append({
                            "sport": sport,
                            "away_team": away_team,
                            "home_team": home_team,
                            "bet_type": bet_type,
                            "total_line": total_line,
                            "risk": risk_amount,
                            "to_win": win_amount
                        })
                
                elif spread_match:
                    # Handle spread bets (e.g., "CALIFORNIA +9-130", "MARQUETTE +2-130")
                    # These bets don't have "(TEAM vrs TEAM)" format - only show the team being bet on
                    team_name = spread_match.group(1).strip()
                    spread_value = spread_match.group(2)
                    if spread_match.group(3) == '½':
                        spread_value = str(float(spread_value) + 0.5 if float(spread_value) > 0 else float(spread_value) - 0.5)
                    
                    bet_type = f"{team_name} {spread_value}"  # e.g., "CALIFORNIA +9"
                    
                    # Look for team names in nearby lines - might not find "(vrs)" for spread bets
                    teams_text = ""
                    for j in range(max(0, i-2), min(len(lines), i+3)):
                        if 'vrs' in lines[j].lower():
                            teams_text = lines[j]
                            break
                    
                    # Extract team names if found
                    away_team = team_name  # Default: bet team is the team
                    home_team = "OPPONENT"  # Placeholder if we can't find matchup
                    
                    teams_match = re.search(r'\(([^)]+)\s+(?:REG\.TIME\s+)?vrs\s+([^)]+?)(?:\s+REG\.TIME)?\)', teams_text, re.IGNORECASE)
                    if teams_match:
                        away_team = teams_match.group(1).strip().replace(' REG.TIME', '')
                        home_team = teams_match.group(2).strip().replace(' REG.TIME', '')
                    
                    # Check for international/European basketball (not NCAAB)
                    # Romanian, European teams often have patterns like "CSM", "BC", "KK", etc.
                    intl_patterns = ['CSM ', 'BC ', 'KK ', 'TARGU', 'PLOIESTI', 'BUCHAREST', 'CLUJ', 
                                    'SIBIU', 'ORADEA', 'STEAUA', 'DINAMO', 'RAPID', 'FCSB']
                    is_international = any(pattern in team_name.upper() or pattern in away_team.upper() or pattern in home_team.upper() 
                                          for pattern in intl_patterns)
                    
                    # Determine sport based on context (look for CBB, NHL, NBA, NFL labels)
                    sport = 'INTL_BASKETBALL' if is_international else None
                    sport_from_context = False
                    for j in range(max(0, i-5), min(len(lines), i+1)):
                        context_line = lines[j].upper()
                        if 'NHL' in context_line:
                            sport = 'NHL'
                            sport_from_context = True
                            break
                        elif 'NBA' in context_line:
                            sport = 'NBA'
                            sport_from_context = True
                            break
                        elif 'NFL' in context_line:
                            sport = 'NFL'
                            sport_from_context = True
                            break
                        elif 'CBB' in context_line or 'COLLEGE' in context_line:
                            sport = 'NCAAB'
                            sport_from_context = True
                            break
                    
                    # Only check team names if no sport found from context
                    # This prevents NBA team name matching for college teams (e.g., Memphis)
                    if not sport_from_context:
                        if is_nhl_team(team_name):
                            sport = 'NHL'
                        elif is_nba_team(team_name):
                            sport = 'NBA'
                        elif is_nfl_team(team_name):
                            sport = 'NFL'
                        else:
                            sport = 'NCAAB'  # Default to college basketball
                    
                    # Look for risk amount in the next line
                    risk_match = re.search(r'(\d{1,},?\d+\.?\d*)\s*/\s*(\d{1,},?\d+\.?\d*)', lines[i+1] if i+1 < len(lines) else '')
                    risk_amount = 0
                    win_amount = 0
                    if risk_match:
                        risk_amount = float(risk_match.group(1).replace(',', ''))
                        win_amount = float(risk_match.group(2).replace(',', ''))
                    
                    raw_bets.append({
                        "sport": sport,
                        "away_team": away_team,
                        "home_team": home_team,
                        "bet_type": bet_type,  # e.g., "CALIFORNIA +9"
                        "spread_line": spread_value,
                        "risk": risk_amount,
                        "to_win": win_amount,
                        "is_spread": True  # Flag for spread bets
                    })
                
                elif live_total_match and not total_match:
                    # Handle live betting format: "Over 161.5 -106" or "Team vs Team / Game / Total / Over 161.5"
                    bet_type = 'OVER' if live_total_match.group(1).lower() == 'over' else 'UNDER'
                    total_line = float(live_total_match.group(2))
                    
                    # Look for team names in nearby lines (check wider range for live bets)
                    teams_text = ""
                    for j in range(max(0, i-5), min(len(lines), i+3)):
                        if ' vs ' in lines[j].lower() and 'game' in lines[j].lower():
                            teams_text = lines[j]
                            break
                    
                    # Extract team names from "Team1 vs Team2 / Game / Total"
                    # Handle team names with numbers like "76ers"
                    live_teams_match = re.search(r'(?:\d+\s*-\s*)?([A-Za-z0-9\s\.\']+?)\s+vs\s+([A-Za-z0-9\s\.\']+?)\s*/\s*Game', teams_text, re.IGNORECASE)
                    
                    if live_teams_match:
                        away_team = live_teams_match.group(1).strip()
                        home_team = live_teams_match.group(2).strip()
                        
                        # Check for international/European basketball
                        intl_patterns = ['CSM ', 'BC ', 'KK ', 'TARGU', 'PLOIESTI', 'BUCHAREST', 'CLUJ', 
                                        'SIBIU', 'ORADEA', 'STEAUA', 'DINAMO', 'RAPID', 'FCSB']
                        is_international = any(pattern in away_team.upper() or pattern in home_team.upper() 
                                              for pattern in intl_patterns)
                        
                        # Determine sport - check nearby lines for sport markers
                        # IMPORTANT: Check NBA/NHL/NFL FIRST before RBL (which is a generic marker)
                        sport = None
                        for j in range(max(0, i-5), min(len(lines), i+8)):
                            line_check = lines[j].upper()
                            # Check for explicit sport markers (these are more specific than RBL)
                            if 'BASKETBALL / NBA' in line_check or '/ NBA' in line_check:
                                sport = 'NBA'
                                break
                            elif 'HOCKEY / NHL' in line_check or '/ NHL' in line_check:
                                sport = 'NHL'
                                break
                            elif 'FOOTBALL / NFL' in line_check or '/ NFL' in line_check:
                                sport = 'NFL'
                                break
                        
                        # If no explicit sport found, check for RBL (college basketball marker)
                        if not sport:
                            for j in range(max(0, i-3), min(len(lines), i+5)):
                                if 'RBL' in lines[j] or 'College Basketball' in lines[j]:
                                    sport = 'NCAAB'
                                    break
                        
                        # Fall back to team name matching if still no sport
                        if not sport:
                            if is_nhl_team(away_team) and is_nhl_team(home_team):
                                sport = 'NHL'
                            elif is_nba_team(away_team) and is_nba_team(home_team):
                                sport = 'NBA'
                            elif is_international:
                                sport = 'INTL_BASKETBALL'
                            else:
                                sport = 'NCAAB'  # Default to college basketball for live bets
                        
                        # Look for risk amount
                        risk_match = re.search(r'(\d{1,},?\d+\.?\d*)\s*/\s*(\d{1,},?\d+\.?\d*)', lines[i+1] if i+1 < len(lines) else '')
                        risk_amount = 0
                        win_amount = 0
                        if risk_match:
                            risk_amount = float(risk_match.group(1).replace(',', ''))
                            win_amount = float(risk_match.group(2).replace(',', ''))
                        
                        logger.info(f"[OpenBets] Found live bet: {away_team} vs {home_team} - {bet_type} {total_line} (sport={sport})")
                        
                        raw_bets.append({
                            "sport": sport,
                            "away_team": away_team,
                            "home_team": home_team,
                            "bet_type": bet_type,
                            "total_line": total_line,
                            "risk": risk_amount,
                            "to_win": win_amount,
                            "is_live": True  # Flag for live bets
                        })
                
                i += 1
            
            # Log all raw bets before consolidation
            logger.info(f"[OpenBets] Found {len(raw_bets)} raw bets before consolidation")
            for bet in raw_bets:
                if bet.get('is_spread'):
                    logger.info(f"[OpenBets] Raw spread bet: {bet.get('bet_type')} - sport={bet.get('sport')}")
            
            # Consolidate duplicate bets (same game + bet type) and count them
            open_bets = []
            bet_counts = {}
            
            for bet in raw_bets:
                # Create a unique key for each game + bet type combination
                key = f"{bet['sport']}:{bet['away_team']}:{bet['home_team']}:{bet['bet_type']}"
                
                if key in bet_counts:
                    bet_counts[key]['count'] += 1
                    bet_counts[key]['total_risk'] += bet['risk']
                    bet_counts[key]['total_win'] += bet['to_win']
                else:
                    bet_counts[key] = {
                        'bet': bet,
                        'count': 1,
                        'total_risk': bet['risk'],
                        'total_win': bet['to_win']
                    }
            
            # Build final open_bets list with count
            for key, data in bet_counts.items():
                bet = data['bet'].copy()
                bet['bet_count'] = data['count']
                bet['total_risk'] = data['total_risk']
                bet['total_win'] = data['total_win']
                open_bets.append(bet)
                # Log ALL bets for debugging
                logger.info(f"[OpenBets] Bet: {bet.get('away_team')} vs {bet.get('home_team')} - {bet.get('bet_type')} (sport={bet.get('sport')})")
            
            logger.info(f"Found {len(open_bets)} unique open bets (from {len(raw_bets)} total)")
            return open_bets
            
        except Exception as e:
            logger.error(f"Error scraping open bets: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    async def scrape_settled_bets_with_lines(self, league: str = "NBA") -> List[Dict[str, Any]]:
        """
        Scrape settled/graded bets from plays888.co History page
        Returns list of settled bets with their ORIGINAL bet lines (bet-time lines)
        This is crucial for tracking what line the user bet at vs the closing line
        """
        import re
        
        # Team lists for filtering
        NBA_TEAMS = ['ATLANTA HAWKS', 'BOSTON CELTICS', 'BROOKLYN NETS', 'CHARLOTTE HORNETS',
            'CHICAGO BULLS', 'CLEVELAND CAVALIERS', 'DALLAS MAVERICKS', 'DENVER NUGGETS',
            'DETROIT PISTONS', 'GOLDEN STATE WARRIORS', 'HOUSTON ROCKETS', 'INDIANA PACERS',
            'LOS ANGELES CLIPPERS', 'LOS ANGELES LAKERS', 'MEMPHIS GRIZZLIES', 'MIAMI HEAT',
            'MILWAUKEE BUCKS', 'MINNESOTA TIMBERWOLVES', 'NEW ORLEANS PELICANS', 'NEW YORK KNICKS',
            'OKLAHOMA CITY THUNDER', 'ORLANDO MAGIC', 'PHILADELPHIA 76ERS', 'PHOENIX SUNS',
            'PORTLAND TRAIL BLAZERS', 'SACRAMENTO KINGS', 'SAN ANTONIO SPURS', 'TORONTO RAPTORS',
            'UTAH JAZZ', 'WASHINGTON WIZARDS']
        
        NHL_TEAMS = ['ANAHEIM DUCKS', 'ARIZONA COYOTES', 'BOSTON BRUINS', 'BUFFALO SABRES',
            'CALGARY FLAMES', 'CAROLINA HURRICANES', 'CHICAGO BLACKHAWKS', 'COLORADO AVALANCHE',
            'COLUMBUS BLUE JACKETS', 'DALLAS STARS', 'DETROIT RED WINGS', 'EDMONTON OILERS',
            'FLORIDA PANTHERS', 'LOS ANGELES KINGS', 'MINNESOTA WILD', 'MONTREAL CANADIENS',
            'NASHVILLE PREDATORS', 'NEW JERSEY DEVILS', 'NEW YORK ISLANDERS', 'NEW YORK RANGERS',
            'OTTAWA SENATORS', 'PHILADELPHIA FLYERS', 'PITTSBURGH PENGUINS', 'SAN JOSE SHARKS',
            'SEATTLE KRAKEN', 'ST. LOUIS BLUES', 'TAMPA BAY LIGHTNING', 'TORONTO MAPLE LEAFS',
            'UTAH MAMMOTH', 'VANCOUVER CANUCKS', 'VEGAS GOLDEN KNIGHTS', 'WASHINGTON CAPITALS',
            'WINNIPEG JETS']
        
        NFL_TEAMS = ['ARIZONA CARDINALS', 'ATLANTA FALCONS', 'BALTIMORE RAVENS', 'BUFFALO BILLS',
            'CAROLINA PANTHERS', 'CHICAGO BEARS', 'CINCINNATI BENGALS', 'CLEVELAND BROWNS',
            'DALLAS COWBOYS', 'DENVER BRONCOS', 'DETROIT LIONS', 'GREEN BAY PACKERS',
            'HOUSTON TEXANS', 'INDIANAPOLIS COLTS', 'JACKSONVILLE JAGUARS', 'KANSAS CITY CHIEFS',
            'LAS VEGAS RAIDERS', 'LOS ANGELES CHARGERS', 'LOS ANGELES RAMS', 'MIAMI DOLPHINS',
            'MINNESOTA VIKINGS', 'NEW ENGLAND PATRIOTS', 'NEW ORLEANS SAINTS', 'NEW YORK GIANTS',
            'NEW YORK JETS', 'PHILADELPHIA EAGLES', 'PITTSBURGH STEELERS', 'SAN FRANCISCO 49ERS',
            'SEATTLE SEAHAWKS', 'TAMPA BAY BUCCANEERS', 'TENNESSEE TITANS', 'WASHINGTON COMMANDERS']
        
        def detect_league(team_name):
            team_upper = team_name.upper()
            for t in NBA_TEAMS:
                if t in team_upper or team_upper in t:
                    return 'NBA'
            for t in NHL_TEAMS:
                if t in team_upper or team_upper in t:
                    return 'NHL'
            for t in NFL_TEAMS:
                if t in team_upper or team_upper in t:
                    return 'NFL'
            return None
        
        try:
            if not self.page:
                logger.error("Browser not initialized")
                return []
            
            logger.info(f"Scraping settled bets with lines for {league}")
            
            # Navigate to History page (contains settled bets with original lines)
            await self.page.goto('https://www.plays888.co/wager/History.aspx', timeout=30000)
            await self.page.wait_for_load_state('networkidle')
            await self.page.wait_for_timeout(3000)
            
            # Get page text
            page_text = await self.page.inner_text('body')
            
            # Parse settled bets - look for TOTAL lines with results
            settled_bets = []
            lines = page_text.split('\n')
            
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                
                # Look for TOTAL with o/u (bet line) - e.g., "TOTAL o233-110" or "TOTAL u6-110"
                total_match = re.search(r'TOTAL\s+([ou])(\d+\.?\d*)[½]?[-+]\d+', line, re.IGNORECASE)
                if total_match:
                    bet_type = 'OVER' if total_match.group(1).lower() == 'o' else 'UNDER'
                    bet_line = float(total_match.group(2).replace('½', '.5'))
                    
                    # Look for team names in nearby lines
                    teams_text = ""
                    for j in range(max(0, i-3), min(len(lines), i+3)):
                        if 'vrs' in lines[j].lower():
                            teams_text = lines[j]
                            break
                    
                    # Extract team names
                    teams_match = re.search(r'\(([^)]+)\s+(?:REG\.TIME\s+)?vrs\s+([^)]+?)(?:\s+REG\.TIME)?\)', teams_text, re.IGNORECASE)
                    if teams_match:
                        away_team = teams_match.group(1).strip().replace(' REG.TIME', '')
                        home_team = teams_match.group(2).strip().replace(' REG.TIME', '')
                        
                        # Detect sport
                        detected_league = detect_league(away_team) or detect_league(home_team)
                        
                        # Only add if matches requested league
                        if detected_league == league.upper():
                            # Look for result (WIN/LOSE/PUSH) in nearby lines
                            result = 'pending'
                            for j in range(i, min(len(lines), i+5)):
                                result_line = lines[j].upper()
                                if 'WINWIN' in result_line or result_line.endswith('WIN'):
                                    result = 'won'
                                    break
                                elif 'LOSELOSE' in result_line or result_line.endswith('LOSE') or 'LOSS' in result_line:
                                    result = 'lost'
                                    break
                                elif 'PUSH' in result_line:
                                    result = 'push'
                                    break
                            
                            # Look for date (MM/DD/YYYY format)
                            bet_date = None
                            for j in range(max(0, i-5), min(len(lines), i+5)):
                                date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', lines[j])
                                if date_match:
                                    bet_date = date_match.group(1)
                                    break
                            
                            settled_bets.append({
                                "sport": detected_league,
                                "away_team": away_team,
                                "home_team": home_team,
                                "bet_type": bet_type,
                                "bet_line": bet_line,  # This is the original line when bet was placed
                                "result": result,
                                "bet_date": bet_date
                            })
                
                i += 1
            
            logger.info(f"Found {len(settled_bets)} settled {league} bets with original lines")
            return settled_bets
            
        except Exception as e:
            logger.error(f"Error scraping settled bets: {str(e)}")
            import traceback
            traceback.print_exc()
            return []


async def scrape_scoresandodds(league: str, date_str: str) -> List[Dict[str, Any]]:
    """
    Scrape game data from scoresandodds.com for a specific league and date
    Uses Playwright for dynamic content scraping
    
    Args:
        league: 'NBA', 'NHL', or 'NFL'
        date_str: Date in 'YYYY-MM-DD' format
    
    Returns:
        List of games with teams, total lines, and final scores (if available)
    """
    import re
    from playwright.async_api import async_playwright
    
    # Build URL
    league_paths = {
        'NBA': 'nba',
        'NHL': 'nhl',
        'NFL': 'nfl',
        'NCAAB': 'ncaab'
    }
    league_path = league_paths.get(league.upper(), 'nba')
    url = f"https://www.scoresandodds.com/{league_path}?date={date_str}"
    
    logger.info(f"Scraping scoresandodds.com: {url}")
    
    # Team name normalization for matching
    team_name_map = {
        # NBA
        'CAVALIERS': 'Cleveland', 'CAVS': 'Cleveland', 'CLEVELAND': 'Cleveland',
        'KNICKS': 'New York', 'NEW YORK': 'New York',
        'SPURS': 'San Antonio', 'SAN ANTONIO': 'San Antonio',
        'THUNDER': 'Okla City', 'OKLAHOMA CITY': 'Okla City', 'OKC': 'Okla City',
        'MAVERICKS': 'Dallas', 'DALLAS': 'Dallas', 'MAVS': 'Dallas',
        'WARRIORS': 'Golden State', 'GOLDEN STATE': 'Golden State', 'GSW': 'Golden State',
        'ROCKETS': 'Houston', 'HOUSTON': 'Houston',
        'LAKERS': 'LA Lakers', 'LOS ANGELES LAKERS': 'LA Lakers', 'LAL': 'LA Lakers',
        'TIMBERWOLVES': 'Minnesota', 'MINNESOTA': 'Minnesota', 'WOLVES': 'Minnesota',
        'NUGGETS': 'Denver', 'DENVER': 'Denver',
        'CELTICS': 'Boston', 'BOSTON': 'Boston',
        '76ERS': 'Philadelphia', 'SIXERS': 'Philadelphia', 'PHILADELPHIA': 'Philadelphia',
        'PACERS': 'Indiana', 'INDIANA': 'Indiana',
        'NETS': 'Brooklyn', 'BROOKLYN': 'Brooklyn',
        'MAGIC': 'Orlando', 'ORLANDO': 'Orlando',
        'HORNETS': 'Charlotte', 'CHARLOTTE': 'Charlotte',
        'BUCKS': 'Milwaukee', 'MILWAUKEE': 'Milwaukee',
        'RAPTORS': 'Toronto', 'TORONTO': 'Toronto',
        'HEAT': 'Miami', 'MIAMI': 'Miami',
        'HAWKS': 'Atlanta', 'ATLANTA': 'Atlanta',
        'BULLS': 'Chicago', 'CHICAGO': 'Chicago',
        'GRIZZLIES': 'Memphis', 'MEMPHIS': 'Memphis',
        'PISTONS': 'Detroit', 'DETROIT': 'Detroit',
        'CLIPPERS': 'LA Clippers', 'LOS ANGELES CLIPPERS': 'LA Clippers', 'LAC': 'LA Clippers',
        'SUNS': 'Phoenix', 'PHOENIX': 'Phoenix',
        'KINGS': 'Sacramento', 'SACRAMENTO': 'Sacramento',
        'BLAZERS': 'Portland', 'TRAIL BLAZERS': 'Portland', 'PORTLAND': 'Portland',
        'JAZZ': 'Utah', 'UTAH': 'Utah',
        'PELICANS': 'New Orleans', 'NEW ORLEANS': 'New Orleans',
        'WIZARDS': 'Washington', 'WASHINGTON': 'Washington',
    }
    
    def normalize_team(name):
        if not name:
            return name
        name_upper = name.upper().strip()
        for key, val in team_name_map.items():
            if key in name_upper:
                return val
        return name.strip()
    
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            await page.goto(url, timeout=30000)
            await page.wait_for_load_state('networkidle')
            await page.wait_for_timeout(2000)
            
            games = []
            
            # Get all game rows from the page
            # scoresandodds.com has event-card elements for each game
            game_elements = await page.query_selector_all('[class*="event-card"]')
            
            # Extract team names from data-abbr attributes
            team_abbrs = await page.evaluate('''() => {
                const abbrs = [];
                document.querySelectorAll('[data-abbr]').forEach(el => {
                    const abbr = el.getAttribute('data-abbr');
                    if (abbr && abbr.length > 2 && !['Next', 'Time', 'Rot#', 'Live', 'All', 'Picks', 'Details', 'ML'].includes(abbr)) {
                        // Skip numeric values (rotation numbers)
                        if (isNaN(abbr)) {
                            abbrs.push(abbr);
                        }
                    }
                });
                return abbrs;
            }''')
            
            # Filter to get only team names (not divisions/conferences)
            team_names = [t for t in team_abbrs if not any(x in t.lower() for x in ['conference', 'eastern', 'western', 'pacific', 'atlantic', 'central', 'southeast', 'southwest', 'northwest'])]
            
            # Get totals from the page - try multiple selectors
            totals = await page.evaluate('''() => {
                const totals = [];
                
                // Try different selectors for totals
                const selectors = [
                    '[data-field="current-total"]',
                    '[class*="total"]',
                    '.event-card-book-value'
                ];
                
                // First try the specific selectors
                for (const selector of selectors) {
                    document.querySelectorAll(selector).forEach(el => {
                        const text = el.innerText.trim();
                        const match = text.match(/[ou]?(\\d+\\.?\\d*)/);
                        if (match && parseFloat(match[1]) > 100) {
                            totals.push(parseFloat(match[1]));
                        }
                    });
                    if (totals.length > 0) break;
                }
                
                // If still no totals, search the whole page for o/u patterns
                if (totals.length === 0) {
                    const bodyText = document.body.innerText;
                    const matches = bodyText.match(/[ou](\\d{3}\\.?\\d*)/g);
                    if (matches) {
                        matches.forEach(m => {
                            const num = parseFloat(m.substring(1));
                            if (num > 100 && num < 300) {
                                totals.push(num);
                            }
                        });
                    }
                }
                
                return totals;
            }''')
            
            # Get final scores if available (for completed games)
            scores = await page.evaluate('''() => {
                const scores = [];
                // Try multiple selectors for scores
                const scoreSelectors = [
                    '[data-field="score"]',
                    '.score',
                    '.game-score',
                    '.final-score',
                    '.team-score',
                    '.score-value',
                    '.box-score td.score'
                ];
                
                for (const selector of scoreSelectors) {
                    document.querySelectorAll(selector).forEach(el => {
                        const text = el.innerText.trim();
                        if (text && !isNaN(text) && parseInt(text) > 0) {
                            scores.push(parseInt(text));
                        }
                    });
                    if (scores.length > 0) break;
                }
                
                // Also try getting scores from the main game container
                if (scores.length === 0) {
                    document.querySelectorAll('.game-row, .game-container, .matchup, tr').forEach(row => {
                        const scoreEls = row.querySelectorAll('.score, [class*="score"], td');
                        scoreEls.forEach(el => {
                            const text = el.innerText.trim();
                            const num = parseInt(text);
                            // NBA/NFL scores typically 70-150, NHL 0-10
                            if (!isNaN(num) && num >= 0 && num <= 200) {
                                // Check if it's in a score context (not line numbers)
                                const parent = el.parentElement;
                                if (parent && !parent.innerText.includes('o/u') && !parent.innerText.includes('+') && !parent.innerText.includes('-')) {
                                    scores.push(num);
                                }
                            }
                        });
                    });
                }
                
                return scores;
            }''')
            
            # Get game times (filter out non-time values)
            times = await page.evaluate('''() => {
                const times = [];
                document.querySelectorAll('[data-field="time"]').forEach(el => {
                    const text = el.innerText.trim();
                    // Only include if it looks like a time (contains AM or PM or : or numbers)
                    if (text && (text.includes('AM') || text.includes('PM') || text.includes(':')) && !text.includes('Sort')) {
                        times.push(text);
                    }
                });
                return times;
            }''')
            
            await browser.close()
            
            logger.info(f"Found {len(team_names)} teams, {len(totals)} totals, {len(scores)} scores, {len(times)} times")
            
            # Pair teams into games (every 2 teams = 1 game)
            # totals list has pairs: over line, under line (same value)
            unique_totals = []
            for i in range(0, len(totals), 2):
                if i < len(totals):
                    unique_totals.append(totals[i])
            
            for i in range(0, len(team_names) - 1, 2):
                game_idx = i // 2
                away_team = normalize_team(team_names[i])
                home_team = normalize_team(team_names[i + 1])
                
                game = {
                    "away_team": away_team,
                    "home_team": home_team,
                    "total": unique_totals[game_idx] if game_idx < len(unique_totals) else None,
                    "time": times[game_idx] if game_idx < len(times) else "",
                    "date": date_str
                }
                
                # Add scores if available (for completed games)
                if len(scores) >= (i + 2):
                    away_score = scores[i]
                    home_score = scores[i + 1]
                    game["away_score"] = away_score
                    game["home_score"] = home_score
                    game["final_score"] = away_score + home_score
                
                games.append(game)
            
            logger.info(f"Scraped {len(games)} games from scoresandodds.com for {league} on {date_str}")
            return games
            
    except Exception as e:
        logger.error(f"Error scraping scoresandodds.com: {e}")
        import traceback
        traceback.print_exc()
        return []


def convert_time_to_arizona(time_str: str) -> str:
    """
    Convert UTC time from CBS Sports to Arizona time (UTC-7).
    CBS shows times in the viewer's local timezone, which is UTC for the server.
    """
    import re
    
    if not time_str or time_str == 'FINAL' or ':' not in time_str:
        return time_str
    
    try:
        # Parse the time (e.g., "SAT 12:00AM" or "12:00AM")
        time_match = re.search(r'(\d+):(\d+)\s*(AM|PM|am|pm)', time_str)
        if not time_match:
            return time_str
            
        hour = int(time_match.group(1))
        minute = int(time_match.group(2))
        ampm = time_match.group(3).upper()
        
        # Convert to 24-hour format
        if ampm == 'PM' and hour != 12:
            hour += 12
        elif ampm == 'AM' and hour == 12:
            hour = 0
        
        # The server time is UTC, subtract 7 hours for Arizona
        utc_hour = hour
        az_hour = (utc_hour - 7) % 24
        
        # Determine day change
        day_prefix = ''
        time_upper = time_str.upper()
        if 'SAT' in time_upper:
            day_prefix = 'FRI ' if utc_hour < 7 else 'SAT '
        elif 'SUN' in time_upper:
            day_prefix = 'SAT ' if utc_hour < 7 else 'SUN '
        elif 'FRI' in time_upper:
            day_prefix = 'THU ' if utc_hour < 7 else 'FRI '
        elif 'THU' in time_upper:
            day_prefix = 'WED ' if utc_hour < 7 else 'THU '
        elif 'WED' in time_upper:
            day_prefix = 'TUE ' if utc_hour < 7 else 'WED '
        elif 'TUE' in time_upper:
            day_prefix = 'MON ' if utc_hour < 7 else 'TUE '
        elif 'MON' in time_upper:
            day_prefix = 'SUN ' if utc_hour < 7 else 'MON '
        
        # Convert back to 12-hour format
        az_ampm = 'AM' if az_hour < 12 else 'PM'
        az_hour_12 = az_hour % 12
        if az_hour_12 == 0:
            az_hour_12 = 12
        
        return f"{day_prefix}{az_hour_12}:{minute:02d}{az_ampm}"
    except Exception:
        return time_str


async def scrape_cbssports_ncaab(target_date: str) -> List[Dict[str, Any]]:
    """
    Scrape NCAAB games and lines from CBS Sports.
    This is more reliable than scoresandodds for NCAAB.
    
    Args:
        target_date: Date in 'YYYY-MM-DD' format
    
    Returns:
        List of games with teams, totals, spreads, times
    """
    from playwright.async_api import async_playwright
    
    # Convert date format from YYYY-MM-DD to YYYYMMDD for CBS URL
    # Use 'all' for college basketball (not FBS which is football)
    date_for_url = target_date.replace("-", "")
    url = f"https://www.cbssports.com/college-basketball/scoreboard/all/{date_for_url}/?layout=compact"
    
    logger.info(f"Scraping CBS Sports NCAAB: {url}")
    
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            await page.goto(url, timeout=60000)
            await page.wait_for_load_state("domcontentloaded", timeout=30000)
            await page.wait_for_timeout(3000)
            
            # Extract games using proper DOM structure
            # CBS Sports: First .team element is AWAY, second .team element is HOME
            games = await page.evaluate("""() => {
                const games = [];
                const cards = document.querySelectorAll('.single-score-card');
                
                cards.forEach((card) => {
                    try {
                        // Get team elements - first is AWAY, second is HOME
                        const teamElements = card.querySelectorAll('.team.team--collegebasketball');
                        if (teamElements.length < 2) return;
                        
                        // Extract team names from the team-name-link or innerText
                        const awayTeamEl = teamElements[0].querySelector('.team-name-link');
                        const homeTeamEl = teamElements[1].querySelector('.team-name-link');
                        
                        let awayTeam = awayTeamEl ? awayTeamEl.innerText.trim() : teamElements[0].innerText.split('\\n')[0].trim();
                        let homeTeam = homeTeamEl ? homeTeamEl.innerText.trim() : teamElements[1].innerText.split('\\n')[0].trim();
                        
                        // Remove rank numbers from team names (e.g., "1 Duke" -> "Duke")
                        awayTeam = awayTeam.replace(/^\\d+\\s+/, '').trim();
                        homeTeam = homeTeam.replace(/^\\d+\\s+/, '').trim();
                        
                        if (!awayTeam || !homeTeam) return;
                        
                        const rawText = card.innerText;
                        
                        // Get time
                        let time = '';
                        const timeMatch = rawText.match(/(\\d+:\\d+\\s*(AM|PM)?|[A-Z]{3}\\s+\\d+:\\d+\\s*(AM|PM)?)/i);
                        if (timeMatch) time = timeMatch[0].trim();
                        
                        // Get total (over/under line) from away team's odds cell
                        let total = null;
                        const awayOddsEl = card.querySelector('.in-progress-odds-away');
                        if (awayOddsEl) {
                            const awayOddsText = awayOddsEl.innerText.trim();
                            const totalMatch = awayOddsText.match(/o(\\d+\\.?\\d*)/);
                            if (totalMatch) total = parseFloat(totalMatch[1]);
                        }
                        // Fallback
                        if (!total) {
                            const totalMatch = rawText.match(/o(\\d+\\.?\\d*)/);
                            if (totalMatch) total = parseFloat(totalMatch[1]);
                        }
                        
                        // Get spread from home team's odds cell (favorite's spread)
                        let spread = null;
                        let spreadTeam = null;
                        const homeOddsEl = card.querySelector('.in-progress-odds-home');
                        if (homeOddsEl) {
                            const homeOddsText = homeOddsEl.innerText.trim();
                            const spreadMatch = homeOddsText.match(/([+-]\\d+\\.?\\d*)/);
                            if (spreadMatch) {
                                spread = parseFloat(spreadMatch[1]);
                                // Negative spread = this team is favorite
                                if (spread < 0) {
                                    spreadTeam = homeTeam;
                                } else {
                                    spreadTeam = awayTeam;
                                }
                            }
                        }
                        // Fallback
                        if (!spread) {
                            const spreadMatch = rawText.match(/([+-]\\d+\\.?\\d*)/);
                            if (spreadMatch) spread = parseFloat(spreadMatch[1]);
                        }
                        
                        // Get scores if game is finished (look for FINAL text and total scores)
                        const scores = [];
                        const rawTextLines = rawText.split('\\n');
                        const isFinal = rawText.toLowerCase().includes('final');
                        
                        if (isFinal) {
                            for (let i = 0; i < rawTextLines.length; i++) {
                                const line = rawTextLines[i].trim();
                                if (line === awayTeam || line === homeTeam) {
                                    for (let j = i + 1; j < Math.min(i + 4, rawTextLines.length); j++) {
                                        const scoreLine = rawTextLines[j];
                                        if (scoreLine.includes('\\t') || /^[\\d\\s]+$/.test(scoreLine.replace(/\\t/g, ' '))) {
                                            const parts = scoreLine.trim().split(/\\s+/);
                                            const lastNum = parts[parts.length - 1];
                                            const total = parseInt(lastNum);
                                            if (!isNaN(total) && total >= 60 && total <= 200) {
                                                scores.push(total);
                                                break;
                                            }
                                        }
                                    }
                                }
                            }
                            
                            if (scores.length < 2) {
                                const allNums = rawText.match(/\\b(\\d{2,3})\\b/g) || [];
                                const finalScores = allNums.filter(n => {
                                    const num = parseInt(n);
                                    return num >= 40 && num <= 150;
                                });
                                if (finalScores.length >= 2) {
                                    scores.length = 0;
                                    scores.push(parseInt(finalScores[finalScores.length - 2]));
                                    scores.push(parseInt(finalScores[finalScores.length - 1]));
                                }
                            }
                            
                            if (scores.length < 2) {
                                const scoreEls = card.querySelectorAll('.total, .score, [class*="score"]');
                                scoreEls.forEach(el => {
                                    const num = parseInt(el.innerText.trim());
                                    if (!isNaN(num) && num >= 40 && num <= 150) {
                                        scores.push(num);
                                    }
                                });
                            }
                        }
                        
                        const game = {
                            away_team: awayTeam,
                            home_team: homeTeam,
                            time: isFinal ? 'FINAL' : time,
                            total: total,
                            opening_line: total,
                            spread: spread,
                            spread_team: spreadTeam,  // Team the spread belongs to (favorite)
                            opening_spread: spread,
                            opening_spread_team: spreadTeam
                        };
                        
                        // Add scores if available
                        if (scores.length >= 2) {
                            game.away_score = scores[0];
                            game.home_score = scores[1];
                            game.final_score = scores[0] + scores[1];
                        }
                        
                        games.push(game);
                    } catch(e) {
                        console.error('Error parsing card:', e);
                    }
                });
                
                return games;
            }""")
            
            await browser.close()
            
            # Convert times to Arizona timezone
            for game in games:
                game['time'] = convert_time_to_arizona(game.get('time', ''))
            
            logger.info(f"Scraped {len(games)} NCAAB games from CBS Sports for {target_date}")
            return games
            
    except Exception as e:
        logger.error(f"Error scraping CBS Sports NCAAB: {e}")
        import traceback
        traceback.print_exc()
        return []


async def scrape_cbssports_nba(target_date: str) -> List[Dict[str, Any]]:
    """
    Scrape NBA games and lines from CBS Sports.
    Fallback when scoresandodds.com is blocked.
    
    Args:
        target_date: Date in 'YYYY-MM-DD' format
    
    Returns:
        List of games with teams, totals, spreads, times
    """
    from playwright.async_api import async_playwright
    
    # Convert date format from YYYY-MM-DD to YYYYMMDD for CBS URL
    # Use ?layout=compact to ensure betting lines are shown
    date_for_url = target_date.replace("-", "")
    url = f"https://www.cbssports.com/nba/scoreboard/{date_for_url}/?layout=compact"
    
    logger.info(f"Scraping CBS Sports NBA: {url}")
    
    # NBA team name mapping
    nba_team_map = {
        'WIZARDS': 'Washington',
        'BULLS': 'Chicago',
        'KNICKS': 'New York',
        'THUNDER': 'Okla City',
        'BUCKS': 'Milwaukee',
        'NUGGETS': 'Denver',
        'WARRIORS': 'Golden State',
        'MAVERICKS': 'Dallas',
        'SUNS': 'Phoenix',
        'SPURS': 'San Antonio',
        'CAVALIERS': 'Cleveland',
        'HEAT': 'Miami',
        'CELTICS': 'Boston',
        'LAKERS': 'LA Lakers',
        'CLIPPERS': 'LA Clippers',
        'KINGS': 'Sacramento',
        'TRAIL BLAZERS': 'Portland',
        'BLAZERS': 'Portland',
        'JAZZ': 'Utah',
        'TIMBERWOLVES': 'Minnesota',
        'ROCKETS': 'Houston',
        'PELICANS': 'New Orleans',
        'GRIZZLIES': 'Memphis',
        'HAWKS': 'Atlanta',
        'MAGIC': 'Orlando',
        'PACERS': 'Indiana',
        '76ERS': 'Philadelphia',
        'NETS': 'Brooklyn',
        'RAPTORS': 'Toronto',
        'HORNETS': 'Charlotte',
        'PISTONS': 'Detroit',
    }
    
    def normalize_nba_team(name):
        if not name:
            return name
        name_upper = name.upper().strip()
        # Check for exact match first (more specific)
        if name_upper in nba_team_map:
            return nba_team_map[name_upper]
        # Then check for partial matches, but handle HORNETS/NETS conflict
        # by checking longer keys first
        for key in sorted(nba_team_map.keys(), key=len, reverse=True):
            if key in name_upper:
                return nba_team_map[key]
        return name.strip()
    
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            await page.goto(url, timeout=60000)
            await page.wait_for_load_state("domcontentloaded", timeout=30000)
            await page.wait_for_timeout(3000)
            
            # Extract games using proper DOM structure
            # CBS Sports: First .team element is AWAY, second .team element is HOME
            games = await page.evaluate("""() => {
                const games = [];
                const cards = document.querySelectorAll('.single-score-card');
                
                cards.forEach((card) => {
                    try {
                        // Get team elements - first is AWAY, second is HOME
                        const teamElements = card.querySelectorAll('.team.team--nba');
                        if (teamElements.length < 2) return;
                        
                        // Extract team names from the team-name-link or innerText
                        const awayTeamEl = teamElements[0].querySelector('.team-name-link');
                        const homeTeamEl = teamElements[1].querySelector('.team-name-link');
                        
                        const awayTeam = awayTeamEl ? awayTeamEl.innerText.trim() : teamElements[0].innerText.split('\\n')[0].trim();
                        const homeTeam = homeTeamEl ? homeTeamEl.innerText.trim() : teamElements[1].innerText.split('\\n')[0].trim();
                        
                        if (!awayTeam || !homeTeam) return;
                        
                        const rawText = card.innerText;
                        
                        // Get time - look for time pattern at start
                        let time = '';
                        const timeMatch = rawText.match(/(\\d+:\\d+\\s*(am|pm)|[A-Z]{3}\\s+\\d+:\\d+\\s*(am|pm)?)/i);
                        if (timeMatch) time = timeMatch[0].trim();
                        
                        // Get total (over/under line) from away team's odds cell
                        let total = null;
                        const awayOddsEl = card.querySelector('.in-progress-odds-away');
                        if (awayOddsEl) {
                            const awayOddsText = awayOddsEl.innerText.trim();
                            const totalMatch = awayOddsText.match(/o(\\d+\\.?\\d*)/);
                            if (totalMatch) total = parseFloat(totalMatch[1]);
                        }
                        // Fallback to rawText parsing
                        if (!total) {
                            const totalMatch = rawText.match(/o(\\d+\\.?\\d*)/);
                            if (totalMatch) total = parseFloat(totalMatch[1]);
                        }
                        
                        // Get spread from home team's odds cell (favorite's spread is shown)
                        // Format: -10.5 means home team is favored by 10.5
                        let spread = null;
                        let spreadTeam = null;  // Which team the spread belongs to
                        const homeOddsEl = card.querySelector('.in-progress-odds-home');
                        if (homeOddsEl) {
                            const homeOddsText = homeOddsEl.innerText.trim();
                            const spreadMatch = homeOddsText.match(/([+-]\\d+\\.?\\d*)/);
                            if (spreadMatch) {
                                spread = parseFloat(spreadMatch[1]);
                                // Negative spread = this team is favorite
                                // Home odds cell shows home team's line
                                if (spread < 0) {
                                    spreadTeam = homeTeam;  // Home team is favorite
                                } else {
                                    spreadTeam = awayTeam;  // Away team is favorite (spread shown as +)
                                }
                            }
                        }
                        // Fallback
                        if (!spread) {
                            const spreadMatch = rawText.match(/([+-]\\d+\\.?\\d*)/);
                            if (spreadMatch) spread = parseFloat(spreadMatch[1]);
                        }
                        
                        // Get scores if game is finished (look for FINAL text and total scores)
                        const scores = [];
                        const rawTextLines = rawText.split('\\n');
                        const isFinal = rawText.toLowerCase().includes('final');
                        
                        if (isFinal) {
                            // Look for team rows with total scores
                            for (let i = 0; i < rawTextLines.length; i++) {
                                const line = rawTextLines[i].trim();
                                if (line === awayTeam || line === homeTeam) {
                                    for (let j = i + 1; j < Math.min(i + 4, rawTextLines.length); j++) {
                                        const scoreLine = rawTextLines[j];
                                        if (scoreLine.includes('\\t') || /^[\\d\\s]+$/.test(scoreLine.replace(/\\t/g, ' '))) {
                                            const parts = scoreLine.trim().split(/\\s+/);
                                            const lastNum = parts[parts.length - 1];
                                            const total = parseInt(lastNum);
                                            if (!isNaN(total) && total >= 60 && total <= 200) {
                                                scores.push(total);
                                                break;
                                            }
                                        }
                                    }
                                }
                            }
                            
                            if (scores.length < 2) {
                                const allNums = rawText.match(/\\b(\\d{2,3})\\b/g) || [];
                                const finalScores = allNums.filter(n => {
                                    const num = parseInt(n);
                                    return num >= 70 && num <= 200;
                                });
                                if (finalScores.length >= 2) {
                                    scores.length = 0;
                                    scores.push(parseInt(finalScores[finalScores.length - 2]));
                                    scores.push(parseInt(finalScores[finalScores.length - 1]));
                                }
                            }
                        }
                        
                        const game = {
                            away_team: awayTeam,
                            home_team: homeTeam,
                            time: isFinal ? 'FINAL' : time,
                            total: total,
                            opening_line: total,
                            spread: spread,
                            spread_team: spreadTeam,  // Team the spread belongs to (favorite)
                            opening_spread: spread,   // Opening spread line
                            opening_spread_team: spreadTeam
                        };
                        
                        if (scores.length >= 2) {
                            game.away_score = scores[0];
                            game.home_score = scores[1];
                            game.final_score = scores[0] + scores[1];
                        }
                        
                        games.push(game);
                    } catch(e) {
                        console.error('Error parsing card:', e);
                    }
                });
                
                return games;
            }""")
            
            await browser.close()
            
            # Normalize team names and convert times to Arizona timezone
            for game in games:
                game['away_team'] = normalize_nba_team(game.get('away_team', ''))
                game['home_team'] = normalize_nba_team(game.get('home_team', ''))
                game['time'] = convert_time_to_arizona(game.get('time', ''))
            
            logger.info(f"Scraped {len(games)} NBA games from CBS Sports for {target_date}")
            return games
            
    except Exception as e:
        logger.error(f"Error scraping CBS Sports NBA: {e}")
        import traceback
        traceback.print_exc()
        return []


async def scrape_cbssports_nhl(target_date: str) -> List[Dict[str, Any]]:
    """
    Scrape NHL games and lines from CBS Sports.
    Fallback when scoresandodds.com is blocked.
    
    Args:
        target_date: Date in 'YYYY-MM-DD' format
    
    Returns:
        List of games with teams, totals, spreads, times
    """
    from playwright.async_api import async_playwright
    
    # Convert date format
    date_for_url = target_date.replace("-", "")
    url = f"https://www.cbssports.com/nhl/scoreboard/{date_for_url}/"
    
    logger.info(f"Scraping CBS Sports NHL: {url}")
    
    # NHL team name mapping
    nhl_team_map = {
        'CAPITALS': 'Washington',
        'BRUINS': 'Boston',
        'RANGERS': 'NY Rangers',
        'ISLANDERS': 'NY Islanders',
        'FLYERS': 'Philadelphia',
        'PENGUINS': 'Pittsburgh',
        'DEVILS': 'New Jersey',
        'BLUE JACKETS': 'Columbus',
        'HURRICANES': 'Carolina',
        'PANTHERS': 'Florida',
        'LIGHTNING': 'Tampa Bay',
        'MAPLE LEAFS': 'Toronto',
        'CANADIENS': 'Montreal',
        'SENATORS': 'Ottawa',
        'SABRES': 'Buffalo',
        'RED WINGS': 'Detroit',
        'BLACKHAWKS': 'Chicago',
        'BLUES': 'St. Louis',
        'PREDATORS': 'Nashville',
        'STARS': 'Dallas',
        'WILD': 'Minnesota',
        'JETS': 'Winnipeg',
        'AVALANCHE': 'Colorado',
        'UTAH HC': 'Utah',
        'UTAH': 'Utah',
        'MAMMOTH': 'Utah',
        'GOLDEN KNIGHTS': 'Vegas',
        'KRAKEN': 'Seattle',
        'SHARKS': 'San Jose',
        'KINGS': 'LA Kings',
        'DUCKS': 'Anaheim',
        'OILERS': 'Edmonton',
        'FLAMES': 'Calgary',
        'CANUCKS': 'Vancouver',
    }
    
    def normalize_nhl_team(name):
        if not name:
            return name
        name_upper = name.upper().strip()
        for key, val in nhl_team_map.items():
            if key in name_upper:
                return val
        return name.strip()
    
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            await page.goto(url, timeout=60000)
            await page.wait_for_load_state("domcontentloaded", timeout=30000)
            await page.wait_for_timeout(3000)
            
            # Extract games using proper DOM structure
            # CBS Sports: First .team element is AWAY, second .team element is HOME
            games = await page.evaluate("""() => {
                const games = [];
                const cards = document.querySelectorAll('.single-score-card');
                
                cards.forEach((card) => {
                    try {
                        // Get team elements - first is AWAY, second is HOME
                        const teamElements = card.querySelectorAll('.team.team--nhl');
                        if (teamElements.length < 2) return;
                        
                        // Extract team names from the team-name-link or innerText
                        const awayTeamEl = teamElements[0].querySelector('.team-name-link');
                        const homeTeamEl = teamElements[1].querySelector('.team-name-link');
                        
                        const awayTeam = awayTeamEl ? awayTeamEl.innerText.trim() : teamElements[0].innerText.split('\\n')[0].trim();
                        const homeTeam = homeTeamEl ? homeTeamEl.innerText.trim() : teamElements[1].innerText.split('\\n')[0].trim();
                        
                        if (!awayTeam || !homeTeam) return;
                        
                        const rawText = card.innerText;
                        
                        // Get time
                        let time = '';
                        const timeMatch = rawText.match(/(\\d+:\\d+\\s*(am|pm)|[A-Z]{3}\\s+\\d+:\\d+\\s*(am|pm)?)/i);
                        if (timeMatch) time = timeMatch[0].trim();
                        
                        // Get total from away team's odds cell
                        let total = null;
                        const awayOddsEl = card.querySelector('.in-progress-odds-away');
                        if (awayOddsEl) {
                            const awayOddsText = awayOddsEl.innerText.trim();
                            const totalMatch = awayOddsText.match(/o(\\d+\\.?\\d*)/);
                            if (totalMatch) total = parseFloat(totalMatch[1]);
                        }
                        // Fallback
                        if (!total) {
                            const totalMatch = rawText.match(/o(\\d+\\.?\\d*)/);
                            if (totalMatch) total = parseFloat(totalMatch[1]);
                        }
                        
                        // Get moneyline from home team's odds cell
                        // NHL uses moneyline (e.g., -141) instead of spread
                        let moneyline = null;
                        let moneylineTeam = null;
                        const homeOddsEl = card.querySelector('.in-progress-odds-home');
                        if (homeOddsEl) {
                            const homeOddsText = homeOddsEl.innerText.trim();
                            // Moneyline format: -141, +130, etc (3-digit number with sign)
                            const mlMatch = homeOddsText.match(/([+-]\\d{3,})/);
                            if (mlMatch) {
                                moneyline = parseInt(mlMatch[1]);
                                // Negative moneyline = favorite
                                if (moneyline < 0) {
                                    moneylineTeam = homeTeam;  // Home team is favorite
                                } else {
                                    moneylineTeam = awayTeam;  // Away team is favorite
                                }
                            }
                        }
                        // Fallback - look for 3+ digit number with sign
                        if (!moneyline) {
                            const mlMatch = rawText.match(/([+-]\\d{3,})/);
                            if (mlMatch) moneyline = parseInt(mlMatch[1]);
                        }
                        
                        // Get scores if game is finished
                        const scores = [];
                        const rawTextLines = rawText.split('\\n');
                        const isFinal = rawText.toLowerCase().includes('final');
                        
                        if (isFinal) {
                            for (let i = 0; i < rawTextLines.length; i++) {
                                const line = rawTextLines[i].trim();
                                if (line === awayTeam || line === homeTeam) {
                                    for (let j = i + 1; j < Math.min(i + 4, rawTextLines.length); j++) {
                                        const scoreLine = rawTextLines[j];
                                        if (scoreLine.includes('\\t') || /^[\\d\\s]+$/.test(scoreLine.replace(/\\t/g, ' '))) {
                                            const parts = scoreLine.trim().split(/\\s+/);
                                            const lastNum = parts[parts.length - 1];
                                            const total = parseInt(lastNum);
                                            if (!isNaN(total) && total >= 0 && total <= 15) {
                                                scores.push(total);
                                                break;
                                            }
                                        }
                                    }
                                }
                            }
                            
                            if (scores.length < 2) {
                                const lines = rawText.split('\\n');
                                for (let i = 0; i < lines.length; i++) {
                                    const line = lines[i].trim();
                                    if (/^\\d+-\\d+-\\d+$/.test(line)) {
                                        if (i + 1 < lines.length) {
                                            const scoreLine = lines[i + 1].trim();
                                            const parts = scoreLine.split(/[\\t\\s]+/);
                                            if (parts.length >= 3) {
                                                const total = parseInt(parts[parts.length - 1]);
                                                if (!isNaN(total) && total >= 0 && total <= 15) {
                                                    scores.push(total);
                                                }
                                            }
                                        }
                                    }
                                }
                            }
                        }
                        
                        const game = {
                            away_team: awayTeam,
                            home_team: homeTeam,
                            time: isFinal ? 'FINAL' : time,
                            total: total,
                            opening_line: total,
                            moneyline: moneyline,
                            moneyline_team: moneylineTeam,  // Team the moneyline belongs to (favorite)
                            opening_moneyline: moneyline,
                            opening_moneyline_team: moneylineTeam
                        };
                        
                        if (scores.length >= 2) {
                            game.away_score = scores[0];
                            game.home_score = scores[1];
                            game.final_score = scores[0] + scores[1];
                        }
                        
                        games.push(game);
                    } catch(e) {
                        console.error('Error parsing card:', e);
                    }
                });
                
                return games;
            }""")
            
            await browser.close()
            
            # Normalize team names and convert times to Arizona timezone
            for game in games:
                game['away_team'] = normalize_nhl_team(game.get('away_team', ''))
                game['home_team'] = normalize_nhl_team(game.get('home_team', ''))
                game['time'] = convert_time_to_arizona(game.get('time', ''))
            
            logger.info(f"Scraped {len(games)} NHL games from CBS Sports for {target_date}")
            return games
            
    except Exception as e:
        logger.error(f"Error scraping CBS Sports NHL: {e}")
        import traceback
        traceback.print_exc()
        return []


async def scrape_covers_consensus(league: str, target_date: str) -> Dict[str, Dict]:
    """
    Scrape betting consensus percentages and spreads from Covers.com
    Returns a dict mapping team abbreviations to their consensus data
    
    Args:
        league: 'NBA', 'NHL', or 'NCAAB'
        target_date: Date in 'YYYY-MM-DD' format
    
    Returns:
        Dict like: {'CLE': {'consensus_pct': 65, 'spread': -5.5}, 'IND': {'consensus_pct': 35, 'spread': 5.5}}
    """
    from playwright.async_api import async_playwright
    import re
    
    # Map league to Covers.com URL format
    league_map = {
        'NBA': 'nba',
        'NHL': 'nhl', 
        'NCAAB': 'ncaab'
    }
    
    league_url = league_map.get(league.upper(), 'nba')
    url = f"https://contests.covers.com/consensus/topconsensus/{league_url}/overall/{target_date}"
    
    logger.info(f"[Covers Consensus] Scraping {url}")
    
    consensus_data = {}
    
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            await page.goto(url, timeout=60000)
            await page.wait_for_load_state("networkidle")
            await page.wait_for_timeout(3000)
            
            # Extract data from the table using specific column structure
            rows_data = await page.evaluate("""() => {
                const results = [];
                const rows = document.querySelectorAll('table tbody tr');
                
                rows.forEach(row => {
                    const cells = row.querySelectorAll('td');
                    if (cells.length >= 5) {
                        results.push({
                            matchup: cells[0].innerText,
                            consensus: cells[2].innerText,
                            sides: cells[3].innerText
                        });
                    }
                });
                
                return results;
            }""")
            
            await browser.close()
            
            # Parse each row
            # Format: Matchup = "NBA\n AWAY\n HOME", Consensus = "AWAY%\nHOME%", Sides = "AWAY_SPREAD\nHOME_SPREAD"
            for row in rows_data:
                matchup_parts = row.get('matchup', '').replace('\n', ' ').split()
                consensus_parts = row.get('consensus', '').replace('\n', ' ').split()
                sides_parts = row.get('sides', '').replace('\n', ' ').split()
                
                # Filter out league name and get teams
                teams = [p for p in matchup_parts if p.upper() not in ['NBA', 'NHL', 'NCAAB', 'NFL', 'MLB'] and len(p) >= 2]
                
                if len(teams) >= 2:
                    away_team = teams[0].upper()
                    home_team = teams[1].upper()
                    
                    # Parse percentages (e.g., "58%" -> 58)
                    away_pct = None
                    home_pct = None
                    for i, part in enumerate(consensus_parts):
                        pct_match = re.match(r'(\d+)%?', part.replace('%', ''))
                        if pct_match:
                            pct_val = int(pct_match.group(1))
                            if away_pct is None:
                                away_pct = pct_val
                            elif home_pct is None:
                                home_pct = pct_val
                    
                    # Parse spreads (e.g., "-4.5", "+4.5")
                    away_spread = None
                    home_spread = None
                    for i, part in enumerate(sides_parts):
                        spread_match = re.match(r'([+-]?\d+\.?\d*)', part)
                        if spread_match:
                            spread_val = float(spread_match.group(1))
                            if away_spread is None:
                                away_spread = spread_val
                            elif home_spread is None:
                                home_spread = spread_val
                    
                    consensus_data[away_team] = {
                        'consensus_pct': away_pct,
                        'spread': away_spread,
                        'opponent': home_team
                    }
                    consensus_data[home_team] = {
                        'consensus_pct': home_pct,
                        'spread': home_spread,
                        'opponent': away_team
                    }
                    logger.debug(f"[Covers] {away_team} {away_pct}% @ {away_spread} vs {home_team} {home_pct}% @ {home_spread}")
            
            logger.info(f"[Covers Consensus] Scraped {len(consensus_data)} team consensus entries for {league} on {target_date}")
            return consensus_data
            
    except Exception as e:
        logger.error(f"[Covers Consensus] Error scraping: {e}")
        import traceback
        traceback.print_exc()
        return {}


async def scrape_cbssports_ncaab_with_team_urls(target_date: str) -> Tuple[List[Dict], Dict[str, str]]:
    """
    Scrape NCAAB games from CBS Sports including team URLs for Last 3 PPG lookup.
    
    Returns:
        Tuple of (games list, team_urls dict)
    """
    from playwright.async_api import async_playwright
    
    date_for_url = target_date.replace("-", "")
    url = f"https://www.cbssports.com/college-basketball/scoreboard/FBS/{date_for_url}/?layout=compact"
    
    logger.info(f"Scraping CBS Sports NCAAB with team URLs: {url}")
    
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            await page.goto(url, timeout=60000)
            await page.wait_for_load_state("domcontentloaded", timeout=30000)
            await page.wait_for_timeout(3000)
            
            # Extract games with team URLs
            data = await page.evaluate("""() => {
                const games = [];
                const teamUrls = {};
                const cards = document.querySelectorAll('.single-score-card');
                
                cards.forEach((card) => {
                    try {
                        // Get team links
                        const teamLinks = card.querySelectorAll('a[href*="/college-basketball/teams/"]');
                        const teams = [];
                        
                        teamLinks.forEach(a => {
                            const href = a.getAttribute('href');
                            const teamSpan = a.querySelector('.team-location-name, .team');
                            let name = teamSpan ? teamSpan.innerText.trim() : a.innerText.trim();
                            name = name.replace(/^\\d+\\s*/, '').split('\\n')[0].trim();
                            
                            if (name && name.length > 1 && name.length < 30 && href && href.includes('/teams/')) {
                                teams.push({ name: name, url: href });
                                teamUrls[name] = href;
                            }
                        });
                        
                        if (teams.length >= 2) {
                            const rawText = card.innerText;
                            
                            // Get total line
                            let total = null;
                            const totalMatch = rawText.match(/o(\\d+\\.?\\d*)/);
                            if (totalMatch) total = parseFloat(totalMatch[1]);
                            
                            // Get time
                            let time = '';
                            const lines = rawText.split('\\n');
                            for (const line of lines) {
                                if (/^\\d+:\\d+\\s*(AM|PM)?/i.test(line.trim()) || /^(WED|THU|FRI|SAT|SUN|MON|TUE)/.test(line.trim())) {
                                    time = line.trim();
                                    break;
                                }
                            }
                            
                            // Get spread
                            let spread = null;
                            const spreadMatches = rawText.match(/([+-]\\d+\\.?\\d*)/g);
                            if (spreadMatches && spreadMatches.length > 0) {
                                spread = parseFloat(spreadMatches[spreadMatches.length - 1]);
                            }
                            
                            games.push({
                                away_team: teams[0].name,
                                away_url: teams[0].url,
                                home_team: teams[teams.length - 1].name,
                                home_url: teams[teams.length - 1].url,
                                total: total,
                                opening_line: total,
                                time: time,
                                spread: spread
                            });
                        }
                    } catch(e) {}
                });
                
                return { games: games, teamUrls: teamUrls };
            }""")
            
            await browser.close()
            
            logger.info(f"Scraped {len(data['games'])} games with {len(data['teamUrls'])} team URLs")
            return data['games'], data['teamUrls']
            
    except Exception as e:
        logger.error(f"Error scraping CBS Sports NCAAB with team URLs: {e}")
        import traceback
        traceback.print_exc()
        return [], {}


async def scrape_ncaab_team_last3_ppg(team_urls: Dict[str, str], max_concurrent: int = 5) -> Dict[str, Dict]:
    """
    Scrape Last 3 game scores for NCAAB teams from CBS Sports team pages.
    
    Args:
        team_urls: Dict mapping team names to their CBS Sports URLs
        max_concurrent: Maximum concurrent browser pages
        
    Returns:
        Dict mapping team names to their Last 3 stats
    """
    from playwright.async_api import async_playwright
    import re
    
    logger.info(f"Scraping Last 3 PPG for {len(team_urls)} NCAAB teams...")
    
    team_stats = {}
    semaphore = asyncio.Semaphore(max_concurrent)
    
    async def scrape_single_team(browser, team_name: str, team_url: str):
        """Scrape a single team's Last 3 scores"""
        async with semaphore:
            try:
                full_url = f"https://www.cbssports.com{team_url}"
                page = await browser.new_page()
                await page.goto(full_url, timeout=20000)
                await page.wait_for_load_state("domcontentloaded")
                await page.wait_for_timeout(800)
                
                # Get schedule section text
                schedule_text = await page.evaluate("""() => {
                    const body = document.body.innerText;
                    const scheduleStart = body.indexOf('Schedule');
                    const scheduleEnd = body.indexOf('Full Schedule');
                    if (scheduleStart > -1 && scheduleEnd > -1) {
                        return body.substring(scheduleStart, scheduleEnd);
                    }
                    return body.substring(0, 2000);
                }""")
                
                await page.close()
                
                # Parse completed games
                # Format: "W 83-69" (team won, scored 83) or "L 80-58" (team lost, scored 58)
                completed_scores = []
                for line in schedule_text.split('\n'):
                    match = re.search(r'\b([WL])\s+(\d+)-(\d+)\b', line)
                    if match:
                        result = match.group(1)
                        score1 = int(match.group(2))
                        score2 = int(match.group(3))
                        # W = team won, their score is LEFT (score1)
                        # L = team lost, their score is RIGHT (score2)
                        team_score = score1 if result == 'W' else score2
                        completed_scores.append(team_score)
                
                # Get last 3 scores (reverse since schedule is chronological)
                if completed_scores:
                    last3 = completed_scores[-3:] if len(completed_scores) >= 3 else completed_scores
                    last3.reverse()  # Most recent first
                    
                    return team_name, {
                        'last3_scores': last3,
                        'last3_total': sum(last3),
                        'last3_avg': round(sum(last3) / len(last3), 1),
                        'games_played': len(completed_scores)
                    }
            except Exception as e:
                logger.debug(f"Error scraping {team_name}: {e}")
            return team_name, None
    
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            
            # Process teams in batches
            teams_list = list(team_urls.items())
            batch_size = 10
            
            for i in range(0, len(teams_list), batch_size):
                batch = teams_list[i:i+batch_size]
                tasks = [scrape_single_team(browser, name, url) for name, url in batch]
                results = await asyncio.gather(*tasks, return_exceptions=True)
                
                for result in results:
                    if isinstance(result, tuple) and result[1] is not None:
                        team_name, stats = result
                        team_stats[team_name] = stats
                
                logger.info(f"  Scraped {min(i+batch_size, len(teams_list))}/{len(teams_list)} teams...")
            
            await browser.close()
            
    except Exception as e:
        logger.error(f"Error in batch scraping: {e}")
    
    logger.info(f"Got Last 3 PPG for {len(team_stats)}/{len(team_urls)} teams")
    return team_stats


plays888_service = Plays888Service()

# Bet monitoring scheduler
scheduler = AsyncIOScheduler()
monitoring_enabled = False

# Random interval settings (in minutes)
MIN_INTERVAL = 5
MAX_INTERVAL = 5

# Track last check time for watchdog
last_check_time = None

async def save_next_check_time(next_time):
    """Save next scheduled check time to database for persistence across restarts"""
    try:
        await db.monitor_state.update_one(
            {"type": "next_check"},
            {"$set": {"next_check_time": next_time, "updated_at": datetime.now(timezone.utc)}},
            upsert=True
        )
    except Exception as e:
        logger.error(f"Failed to save next check time: {e}")

async def get_next_check_time():
    """Get next scheduled check time from database"""
    try:
        state = await db.monitor_state.find_one({"type": "next_check"}, {"_id": 0})
        if state:
            return state.get("next_check_time")
    except Exception as e:
        logger.error(f"Failed to get next check time: {e}")
    return None

def schedule_next_check():
    """DEPRECATED - Now using monitoring_loop() instead"""
    # Do nothing - monitoring is handled by the background loop
    pass

async def monitor_and_reschedule():
    """Run monitoring - called by manual check API"""
    await run_monitoring_cycle()


async def send_check_notification(check_time, new_bets_found):
    """Send a notification to Telegram that a check was performed - auto-deletes after 30 min"""
    try:
        telegram_config = await db.telegram_config.find_one({}, {"_id": 0})
        if not telegram_config or not telegram_config.get("bot_token"):
            return
        
        bot = Bot(token=telegram_config["bot_token"])
        chat_id = telegram_config["chat_id"]
        
        # Generate random next interval for display
        next_interval = random.randint(MIN_INTERVAL, MAX_INTERVAL)
        
        # Build message - handle None safely
        if new_bets_found is None:
            new_bets_found = {}
        enano_new = new_bets_found.get("jac075", 0) or 0
        tipster_new = new_bets_found.get("jac083", 0) or 0
        
        if enano_new > 0 or tipster_new > 0:
            status = f"🆕 New bets: ENANO={enano_new}, TIPSTER={tipster_new}"
        else:
            status = "✅ No new bets"
        
        message = f"""🔄 *CHECK COMPLETE*
⏰ {check_time.strftime('%I:%M %p')} Arizona
{status}
⏭️ Next check in ~{next_interval} min"""
        
        sent_msg = await bot.send_message(
            chat_id=chat_id,
            text=message,
            parse_mode=ParseMode.MARKDOWN
        )
        logger.info(f"Check notification sent to Telegram")
        
        # Schedule auto-deletion after 30 minutes
        asyncio.create_task(delete_message_later(bot, chat_id, sent_msg.message_id, 15))
        
    except Exception as e:
        logger.error(f"Failed to send check notification: {str(e)}")


async def watchdog_check():
    """Watchdog to ensure monitoring is running - runs every 5 minutes"""
    global last_check_time, monitoring_enabled
    
    if not monitoring_enabled:
        return
    
    # Check if we're in sleep hours
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    now_arizona = datetime.now(arizona_tz)
    current_hour = now_arizona.hour
    current_minute = now_arizona.minute
    current_time_minutes = current_hour * 60 + current_minute
    
    sleep_start = 22 * 60 + 0   # 10:00 PM
    sleep_end = 7 * 60 + 0      # 7:00 AM
    
    if current_time_minutes >= sleep_start or current_time_minutes < sleep_end:
        return  # Don't check during sleep hours
    
    # Check if last check was more than 20 minutes ago
    if last_check_time:
        minutes_since_last = (datetime.now(timezone.utc) - last_check_time).total_seconds() / 60
        
        if minutes_since_last > 20:
            logger.warning(f"Watchdog: No check for {minutes_since_last:.0f} minutes! Triggering immediate check...")
            
            # Re-schedule monitoring
            schedule_next_check()
            
            # Trigger immediate check
            asyncio.create_task(monitor_and_reschedule())
    else:
        # First watchdog run, initialize last_check_time
        last_check_time = datetime.now(timezone.utc)



async def update_opportunity_with_bet(sport: str, game: str, bet_type: str, bet_line: float, ticket_num: str, account: str):
    """
    #3.75 BET LINE CAPTURE: Update opportunities collection when a bet is placed.
    This ensures the bet_line is stored with the game for display in 'Today's Plays'.
    """
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    today = datetime.now(arizona_tz).strftime('%Y-%m-%d')
    
    try:
        # Determine league from sport code
        league_map = {
            'NBA': 'nba_opportunities',
            'NHL': 'nhl_opportunities',
            'NFL': 'nfl_opportunities',
            'CBB': 'ncaab_opportunities',
            'NCAAB': 'ncaab_opportunities',
            'RBL': 'ncaab_opportunities',  # College Basketball Extra on plays888
        }
        
        # Try to match sport to collection
        collection_name = None
        for key, value in league_map.items():
            if key.upper() in sport.upper():
                collection_name = value
                break
        
        if not collection_name:
            logger.warning(f"Could not determine league for sport: {sport}")
            return
        
        collection = db[collection_name]
        
        # Get today's document
        doc = await collection.find_one({"date": today})
        if not doc or not doc.get('games'):
            logger.warning(f"No {collection_name} data for {today}")
            return
        
        games = doc.get('games', [])
        plays = doc.get('plays', [])
        updated = False
        
        # Normalize game name for matching (e.g., "Milwaukee @ Charlotte" or "MILWAUKEE")
        game_lower = game.lower()
        
        # Find the matching game
        for g in games:
            away = g.get('away_team', g.get('away', '')).lower()
            home = g.get('home_team', g.get('home', '')).lower()
            
            # Check if game matches (partial match is ok)
            if away in game_lower or home in game_lower or game_lower in f"{away} @ {home}":
                # Update game with bet info
                g['has_bet'] = True
                g['bet_slip_id'] = ticket_num
                g['bet_account'] = account
                
                # #3.75: Store the bet line (line at which bet was placed)
                if bet_line:
                    g['bet_line'] = bet_line
                    logger.info(f"Stored bet_line={bet_line} for {away} @ {home}")
                
                # Determine bet type (OVER/UNDER)
                if 'over' in bet_type.lower() or bet_type.lower().startswith('o'):
                    g['bet_type'] = 'OVER'
                elif 'under' in bet_type.lower() or bet_type.lower().startswith('u'):
                    g['bet_type'] = 'UNDER'
                else:
                    g['bet_type'] = bet_type
                
                updated = True
                logger.info(f"Updated opportunity {away} @ {home} with bet info: has_bet=True, bet_line={bet_line}")
                
                # Also update/add to plays list
                play_exists = False
                for p in plays:
                    if away in p.get('game', '').lower() or home in p.get('game', '').lower():
                        p['has_bet'] = True
                        p['bet_line'] = bet_line
                        p['bet_type'] = g['bet_type']
                        play_exists = True
                        break
                
                if not play_exists:
                    # Add new play entry
                    plays.append({
                        "game": f"{g.get('away_team', g.get('away'))} @ {g.get('home_team', g.get('home'))}",
                        "total": g.get('total', g.get('live_line', g.get('opening_line'))),
                        "bet_line": bet_line,
                        "combined_ppg": g.get('combined_ppg', g.get('gpg_avg')),
                        "edge": g.get('edge'),
                        "recommendation": g.get('recommendation', g['bet_type']),
                        "has_bet": True,
                        "bet_type": g['bet_type'],
                        "bet_count": 1
                    })
                
                break
        
        if updated:
            # Save back to database
            await collection.update_one(
                {"date": today},
                {"$set": {"games": games, "plays": plays}}
            )
            logger.info(f"Saved bet info to {collection_name} for {today}")
        else:
            logger.warning(f"Could not find matching game for: {game}")
            
    except Exception as e:
        logger.error(f"Error updating opportunity with bet: {str(e)}")



async def monitor_open_bets():
    """Background job to monitor plays888.co for new bets"""
    global monitoring_enabled
    
    new_bets_count = {"jac075": 0, "jac083": 0}
    
    if not monitoring_enabled:
        return new_bets_count
    
    # Check if we're in sleep hours (10:45 PM - 6:00 AM Arizona time)
    # Arizona is UTC-7 (no daylight saving)
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    now_arizona = datetime.now(arizona_tz)
    current_hour = now_arizona.hour
    current_minute = now_arizona.minute
    current_time_minutes = current_hour * 60 + current_minute
    
    # Sleep window: 10:00 PM to 7:00 AM (active 7am-10pm)
    sleep_start = 22 * 60 + 0   # 10:00 PM = 1260 minutes
    sleep_end = 7 * 60 + 0       # 7:00 AM = 420 minutes
    
    if current_time_minutes >= sleep_start or current_time_minutes < sleep_end:
        logger.info(f"Sleep hours ({now_arizona.strftime('%I:%M %p')} Arizona) - skipping bet check")
        return new_bets_count
    
    logger.info(f"Checking plays888.co for new bets... ({now_arizona.strftime('%I:%M %p')} Arizona)")
    
    try:
        # Get ALL active connections (multiple accounts)
        connections = await db.connections.find({"is_connected": True}, {"_id": 0}).to_list(100)
        
        if not connections:
            logger.info("No active connections, skipping bet monitoring")
            return new_bets_count
        
        # Monitor each account
        for conn in connections:
            new_count = await monitor_single_account(conn)
            username = conn.get("username", "")
            new_bets_count[username] = new_count
            
    except Exception as e:
        logger.error(f"Error in bet monitoring: {str(e)}")
    
    return new_bets_count

async def monitor_single_account(conn: dict):
    """Monitor a single account for new bets - returns count of new bets found"""
    username = conn["username"]
    password = decrypt_password(conn["password_encrypted"])
    new_bets_count = 0
    
    logger.info(f"Monitoring account: {username}")
    monitor_service = None
    
    try:
        # Create a new service instance for monitoring
        monitor_service = Plays888Service()
        await monitor_service.initialize()
        
        # Login
        login_result = await monitor_service.login(username, password)
        if not login_result["success"]:
            logger.error(f"Monitor login failed for {username}: {login_result['message']}")
            await monitor_service.close()
            return new_bets_count  # Return 0, not None
        
        # Navigate to Open Bets page
        await monitor_service.page.goto('https://www.plays888.co/wager/OpenBets.aspx', timeout=30000)
        await monitor_service.page.wait_for_timeout(5000)  # Wait longer for table to load
        
        # Extract open bets by parsing the table rows using Playwright
        import re
        
        # Try to extract bet data from table rows
        bets_data = await monitor_service.page.evaluate('''() => {
            const bets = [];
            // Find all table rows in the bets table
            const rows = document.querySelectorAll('table tr');
            
            for (let i = 0; i < rows.length; i++) {
                const row = rows[i];
                const cells = row.querySelectorAll('td');
                
                // Table structure based on plays888.co:
                // 0: GameDate (contains Ticket#)
                // 1: User/Phone
                // 2: Date Placed
                // 3: Sport (CBB, NBA, NHL, SOC, etc)
                // 4: Description (bet details, game name)
                // 5: Risk/Win amounts
                
                if (cells.length >= 5) {
                    // Extract ticket number from first column
                    const ticketCell = cells[0].textContent || '';
                    const ticketMatch = ticketCell.match(/Ticket#?[:\\s-]*(\\d+)/i);
                    
                    // Extract game time from first column (format: "Ticket#:340605842 Jan 19 03:15 PM")
                    let gameTime = '';
                    let gameDate = '';
                    const timeMatch = ticketCell.match(/(\\d{1,2}:\\d{2}\\s*(?:AM|PM))/i);
                    if (timeMatch) {
                        gameTime = timeMatch[1].trim();
                    }
                    // Extract date too (e.g., "Jan 19")
                    const dateMatch = ticketCell.match(/([A-Za-z]{3})\\s+(\\d{1,2})/);
                    if (dateMatch) {
                        gameDate = dateMatch[1] + ' ' + dateMatch[2];
                    }
                    
                    if (ticketMatch) {
                        const ticket = ticketMatch[1];
                        
                        // Column indices - adjusted based on actual table structure
                        const sport = cells[3] ? cells[3].textContent.trim() : '';
                        const description = cells[4] ? cells[4].textContent.trim() : '';
                        const riskWin = cells[5] ? cells[5].textContent.trim() : '';
                        
                        // Extract country/league from description
                        // Format examples: "(Israel)", "(Uruguay)", "(Norway)", "(Croatia)"
                        // Or from patterns like "Basketball / Israel" or "Hockey / Norway"
                        let country = '';
                        const countryMatch = description.match(/\\(([A-Za-z]+)\\)\\s*$/);
                        if (countryMatch) {
                            country = countryMatch[1];
                        } else {
                            // Try to find country after slash (e.g., "Basketball / Israel")
                            const slashCountry = description.match(/\\/\\s*([A-Za-z]+)\\s*(?:\\/|$)/);
                            if (slashCountry && slashCountry[1].length <= 15) {
                                // Only take short names that could be countries
                                const possibleCountry = slashCountry[1].trim();
                                if (!['Game', 'Total', 'Spread', 'Money', 'Over', 'Under', 'NBA', 'NHL', 'NFL', 'CBB', 'NCAAB'].includes(possibleCountry)) {
                                    country = possibleCountry;
                                }
                            }
                        }
                        // Check sport column for international leagues
                        if (!country && sport) {
                            const sportUpper = sport.toUpperCase();
                            // Map sport codes to countries/regions
                            const sportCountryMap = {
                                'RBL': '', // Will need to extract from description
                                'SOC': '', // Soccer - varies
                                'HKI': '', // Hockey international
                                'TEN': '', // Tennis
                            };
                        }
                        
                        // Parse risk/win amounts (format: "1100.00 / 1000.00" or "500.00 / 9500.00")
                        let wager = 0;
                        let toWin = 0;
                        const riskMatch = riskWin.match(/([\\d,]+\\.?\\d*)\\s*\\/\\s*([\\d,]+\\.?\\d*)/);
                        if (riskMatch) {
                            wager = parseFloat(riskMatch[1].replace(/,/g, ''));
                            toWin = parseFloat(riskMatch[2].replace(/,/g, ''));
                        }
                        
                        // Calculate American odds from Risk/Win amounts
                        // This works for ALL bet types (Straight, Parlay, Teaser, etc.)
                        let odds = 0;
                        if (wager > 0 && toWin > 0) {
                            if (toWin >= wager) {
                                // Positive odds: (Win / Risk) * 100
                                odds = Math.round((toWin / wager) * 100);
                            } else {
                                // Negative odds: -(Risk / Win) * 100
                                odds = Math.round(-(wager / toWin) * 100);
                            }
                        }
                        
                        // Better parsing of the description field
                        // Description format examples:
                        // "STRAIGHT BET[2526] MILWAUKEE BUCKS 2H +1½-110"
                        // "STRAIGHT BET[297497888] Dallas Mavericks..."
                        // "PARLAY[123] Team A vs Team B..."
                        // "G278190881 - STRAIGHT BET 302234312 - Cleveland Cavaliers vs Minnesota Timberwolves / Game / Total / Under 242.5 -113"
                        
                        let game = '';
                        let betType = '';
                        let totalLine = null;
                        
                        // NEW FORMAT: Handle slash-separated format
                        // "Team A vs Team B / Game / Total / Under 242.5"
                        if (description.includes('/')) {
                            const slashParts = description.split('/').map(p => p.trim());
                            for (const part of slashParts) {
                                // Look for team matchup
                                if ((part.includes(' vs ') || part.includes(' vrs ')) && !game) {
                                    // Extract just the team names
                                    const teamMatch = part.match(/([A-Za-z][A-Za-z\\s]+)\\s+(?:vs|vrs)\\s+([A-Za-z][A-Za-z\\s]+)/i);
                                    if (teamMatch) {
                                        game = teamMatch[1].trim() + ' vs ' + teamMatch[2].trim();
                                    } else {
                                        game = part;
                                    }
                                }
                                // Look for Over/Under with line
                                const overUnderMatch = part.match(/(Over|Under)\\s+([\\d.½]+)/i);
                                if (overUnderMatch) {
                                    const ouType = overUnderMatch[1].toUpperCase().charAt(0);  // O or U
                                    const line = overUnderMatch[2].replace('½', '.5');
                                    betType = 'TOTAL ' + ouType + line;
                                    totalLine = parseFloat(line);
                                }
                            }
                        }
                        
                        // First, extract game name from parentheses (TEAM A vs TEAM B)
                        const vsMatch = description.match(/\\(([^)]*(?:vs|vrs)[^)]*)\\)/i);
                        if (vsMatch) {
                            game = vsMatch[1].trim();
                        }
                        
                        // Check bet type
                        if (description.toUpperCase().includes('PARLAY')) {
                            betType = 'PARLAY';
                        } else if (description.toUpperCase().includes('TEASER')) {
                            betType = 'TEASER';
                        } else {
                            // Look for TOTAL over/under
                            const totalMatch = description.match(/TOTAL\\s+([ou][\\d.½]+)/i);
                            if (totalMatch) {
                                betType = 'TOTAL ' + totalMatch[1].toUpperCase();
                            }
                            
                            // Look for spread like "+1½" or "-5.5" with team name
                            // Format after bracket: "TEAM NAME +/-SPREAD"
                            const afterBracket = description.match(/\\]\\s*(.+)/);
                            if (afterBracket) {
                                const betDetails = afterBracket[1].trim();
                                
                                // Extract team name and spread/line
                                // Pattern: "TEAM NAME +/-NUMBER" or "TEAM NAME 2H +/-NUMBER"
                                const teamSpreadMatch = betDetails.match(/^([A-Za-z][A-Za-z0-9\\s\\.]+?)\\s*(2H\\s*)?([+-][\\d½\\.]+)/);
                                if (teamSpreadMatch) {
                                    const teamName = teamSpreadMatch[1].trim();
                                    const halfIndicator = teamSpreadMatch[2] ? '2H ' : '';
                                    const spread = teamSpreadMatch[3];
                                    
                                    if (!betType) {
                                        betType = teamName + ' ' + halfIndicator + spread;
                                    }
                                    if (!game) {
                                        game = teamName;
                                    }
                                }
                                
                                // If still no game, use the bet details directly (cleaned up)
                                if (!game && betDetails) {
                                    // Remove odds from end like "-110" and clean up
                                    game = betDetails.replace(/[+-]\\d{3}$/, '').trim();
                                    // Limit length
                                    if (game.length > 50) {
                                        game = game.substring(0, 47) + '...';
                                    }
                                }
                            }
                        }
                        
                        // Fallback: if still no info, use cleaned description
                        if (!game) {
                            // Remove "STRAIGHT BET", IDs, etc and get the actual content
                            let cleanDesc = description
                                .replace(/STRAIGHT\\s*BET/gi, '')
                                .replace(/\\[[^\\]]+\\]/g, '')  // Remove [ID]
                                .replace(/[+-]\\d{3}$/g, '')   // Remove trailing odds
                                .trim();
                            if (cleanDesc.length > 50) {
                                cleanDesc = cleanDesc.substring(0, 47) + '...';
                            }
                            game = cleanDesc || 'Unknown';
                        }
                        
                        if (!betType) {
                            betType = 'Straight';
                        }
                        
                        bets.push({
                            ticket: ticket,
                            description: description,
                            sport: sport,
                            game: game,
                            betType: betType,
                            totalLine: totalLine,
                            gameTime: gameTime,
                            gameDate: gameDate,
                            country: country,
                            odds: odds,
                            wager: wager,
                            toWin: toWin
                        });
                    }
                }
            }
            return bets;
        }''')
        
        logger.info(f"Extracted {len(bets_data)} open bets from plays888.co table")
        
        # Check each bet against our database
        for bet_info in bets_data:
            ticket_num = bet_info.get('ticket', '')
            
            if not ticket_num:
                continue
                
            # #3.5 DUPLICATE PREVENTION: Check if this ticket already exists in our database
            existing_bet = await db.bet_history.find_one({"bet_slip_id": ticket_num})
            
            if existing_bet:
                # Skip - ticket already recorded
                continue
            
            game = bet_info.get('game', '') or 'Unknown Game'
            bet_type = bet_info.get('betType', '') or 'Unknown'
            
            # #3.5 ADDITIONAL DUPLICATE CHECK: Check for same game + bet_type + account placed today
            today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
            duplicate_bet = await db.bet_history.find_one({
                "game": {"$regex": game.replace("@", ".*"), "$options": "i"},
                "bet_type": {"$regex": bet_type, "$options": "i"},
                "account": username,
                "placed_at": {"$gte": today_start.isoformat()}
            })
            
            if duplicate_bet:
                logger.warning(f"#3.5 DUPLICATE PREVENTION: Skipping duplicate bet on same game: {game} - {bet_type}")
                continue
                
            # New bet detected! 
            logger.info(f"New bet detected: Ticket#{ticket_num}")
            logger.info(f"Bet details: {bet_info}")
            
            odds = bet_info.get('odds', -110)
            wager = bet_info.get('wager', 0)
            to_win = bet_info.get('toWin', 0)
            sport = bet_info.get('sport', '')
            description = bet_info.get('description', '')
            total_line_from_parser = bet_info.get('totalLine')  # From slash format parsing
            game_time = adjust_time_for_arizona(bet_info.get('gameTime', ''))  # Game time adjusted to Arizona
            game_date = bet_info.get('gameDate', '')  # Game date from first column
            country = bet_info.get('country', '')  # Country/league from description
            
            # Fix sport detection from description if RBL is just a generic marker
            # Description may contain "Basketball / NBA" or "Hockey / NHL" etc.
            if sport == 'RBL' and description:
                desc_upper = description.upper()
                if 'BASKETBALL / NBA' in desc_upper or '/ NBA' in desc_upper:
                    sport = 'NBA'
                    logger.info(f"Corrected sport from RBL to NBA based on description")
                elif 'BASKETBALL / NCAAB' in desc_upper or 'COLLEGE' in desc_upper or '/ CBB' in desc_upper:
                    sport = 'NCAAB'
                elif 'HOCKEY / NHL' in desc_upper or '/ NHL' in desc_upper:
                    sport = 'NHL'
                elif 'FOOTBALL / NFL' in desc_upper or '/ NFL' in desc_upper:
                    sport = 'NFL'
            
            # #3.75 BET LINE CAPTURE: Extract the line from bet_type (e.g., "TOTAL o228" -> 228)
            bet_line = total_line_from_parser  # Use parser's extracted line first
            if bet_line is None and bet_type:
                # Pattern: "TOTAL o228" or "TOTAL u6.5" or "o228" or "u6"
                line_match = re.search(r'[ou](\d+\.?\d*)', bet_type, re.IGNORECASE)
                if line_match:
                    bet_line = float(line_match.group(1))
                    logger.info(f"Captured bet line: {bet_line} from bet_type: {bet_type}")
            
            # If game is still unknown, use part of description
            if game == 'Unknown Game' and description:
                # Try to extract game from description
                game = description[:50] + '...' if len(description) > 50 else description
            
            # Store in database with account info for filtering
            bet_doc = {
                "id": str(uuid.uuid4()),
                "opportunity_id": "mobile_detected",
                "rule_id": "mobile_detected",
                "wager_amount": wager,
                "odds": odds,
                "status": "placed",
                "placed_at": datetime.now(timezone.utc).isoformat(),
                "result": None,
                "game": game,
                "bet_type": bet_type,
                "line": bet_type,
                "total_line": bet_line,  # #3.75: Store the numeric line value
                "bet_slip_id": ticket_num,
                "account": username,
                "notes": f"Account: {username}. Auto-detected from plays888.co. Sport: {sport}"
            }
            await db.bet_history.insert_one(bet_doc)
            
            # #3.75 BET LINE CAPTURE: Update opportunities with bet info
            await update_opportunity_with_bet(sport, game, bet_type, bet_line, ticket_num, username)
            
            # Send Telegram notification with actual details
            await send_telegram_notification({
                "game": game,
                "bet_type": bet_type,
                "line": bet_type,
                "odds": odds,
                "wager": wager,
                "potential_win": to_win,
                "ticket_number": ticket_num,
                "game_time": game_time,
                "game_date": game_date,
                "country": country,
                "status": "Placed",
                "league": f"{sport} - Detected from mobile/web"
            }, account=username)
            new_bets_count += 1
        
        await monitor_service.close()
        logger.info(f"Bet monitoring check complete for {username}")
        return new_bets_count
        
    except Exception as e:
        logger.error(f"Error monitoring account {username}: {str(e)}")
        if monitor_service:
            try:
                await monitor_service.close()
            except:
                pass
        return new_bets_count


# API Routes
@api_router.get("/")
async def root():
    return {"message": "Betting Automation API", "status": "running"}


@api_router.post("/connection/setup")
async def setup_connection(connection: ConnectionCreate):
    """Setup connection credentials for plays888.co"""
    try:
        # Encrypt password
        encrypted_password = encrypt_password(connection.password)
        
        # Test connection
        login_result = await plays888_service.login(connection.username, connection.password)
        
        # Store in database
        conn_doc = {
            "id": str(uuid.uuid4()),
            "username": connection.username,
            "password_encrypted": encrypted_password,
            "website": "plays888.co",
            "is_connected": login_result["success"],
            "last_connection": datetime.now(timezone.utc).isoformat() if login_result["success"] else None,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Delete existing connections for this user
        await db.connections.delete_many({"username": connection.username})
        
        # Insert new connection
        await db.connections.insert_one(conn_doc)
        
        # Close browser after test
        await plays888_service.close()
        
        return {"success": login_result["success"], "message": login_result["message"]}
        
    except Exception as e:
        logger.error(f"Setup connection error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/connection/status")
async def get_connection_status():
    """Get current connection status"""
    try:
        conn = await db.connections.find_one({}, {"_id": 0}, sort=[("created_at", -1)])
        
        if not conn:
            return AccountStatus(is_connected=False)
        
        return AccountStatus(
            is_connected=conn.get("is_connected", False),
            username=conn.get("username"),
            last_sync=datetime.fromisoformat(conn["last_connection"]) if conn.get("last_connection") else None
        )
    except Exception as e:
        logger.error(f"Get connection status error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/rules")
async def create_rule(rule: BettingRuleCreate):
    """Create a new betting rule"""
    try:
        rule_doc = {
            "id": str(uuid.uuid4()),
            "name": rule.name,
            "enabled": rule.enabled,
            "min_odds": rule.min_odds,
            "max_odds": rule.max_odds,
            "wager_amount": rule.wager_amount,
            "auto_place": rule.auto_place,
            "sport": rule.sport,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.betting_rules.insert_one(rule_doc)
        return {"success": True, "rule_id": rule_doc["id"]}
    except Exception as e:
        logger.error(f"Create rule error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/rules")
async def get_rules():
    """Get all betting rules"""
    try:
        rules = await db.betting_rules.find({}, {"_id": 0}).to_list(100)
        
        for rule in rules:
            if isinstance(rule.get('created_at'), str):
                rule['created_at'] = datetime.fromisoformat(rule['created_at'])
        
        return rules
    except Exception as e:
        logger.error(f"Get rules error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/rules/{rule_id}")
async def delete_rule(rule_id: str):
    """Delete a betting rule"""
    try:
        result = await db.betting_rules.delete_one({"id": rule_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Rule not found")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete rule error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============== TELEGRAM CLEANUP ==============

@api_router.post("/telegram/cleanup-now")
async def cleanup_telegram_messages():
    """Manually trigger cleanup of scheduled message deletions"""
    try:
        await process_pending_deletions()
        
        # Count remaining scheduled deletions
        remaining = await db.scheduled_deletions.count_documents({})
        
        return {
            "success": True,
            "message": "Cleanup completed",
            "remaining_scheduled": remaining
        }
    except Exception as e:
        logger.error(f"Cleanup error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/telegram/scheduled-deletions")
async def get_scheduled_deletions():
    """View all scheduled message deletions"""
    try:
        deletions = await db.scheduled_deletions.find({}, {"_id": 0}).to_list(100)
        now = datetime.now(timezone.utc)
        
        for d in deletions:
            if d.get('delete_at'):
                d['delete_at'] = d['delete_at'].isoformat()
                d['time_remaining'] = str(d.get('delete_at', now) - now) if isinstance(d.get('delete_at'), datetime) else 'N/A'
            if d.get('created_at'):
                d['created_at'] = d['created_at'].isoformat()
        
        return {
            "count": len(deletions),
            "deletions": deletions
        }
    except Exception as e:
        logger.error(f"Get scheduled deletions error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/telegram/bulk-cleanup")
async def bulk_cleanup_telegram(start_msg_id: int, end_msg_id: int):
    """Delete a range of messages from Telegram chat to clean up clutter"""
    try:
        telegram_config = await db.telegram_config.find_one({}, {"_id": 0})
        if not telegram_config or not telegram_config.get("bot_token") or not telegram_config.get("chat_id"):
            raise HTTPException(status_code=400, detail="Telegram not configured")
        
        bot = Bot(token=telegram_config["bot_token"])
        chat_id = telegram_config["chat_id"]
        
        deleted_count = 0
        failed_count = 0
        
        for msg_id in range(start_msg_id, end_msg_id + 1):
            try:
                await bot.delete_message(chat_id=chat_id, message_id=msg_id)
                deleted_count += 1
            except Exception as e:
                failed_count += 1
                # Message may already be deleted or doesn't exist
                pass
        
        # Also clear scheduled deletions for these messages
        await db.scheduled_deletions.delete_many({
            "message_id": {"$gte": start_msg_id, "$lte": end_msg_id}
        })
        
        return {
            "success": True,
            "deleted": deleted_count,
            "failed": failed_count,
            "range": f"{start_msg_id} to {end_msg_id}"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Bulk cleanup error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/rules/opportunities")
async def get_rules_opportunities():
    """Get current betting opportunities that match rules (from plays888)"""
    try:
        # Get connection
        conn = await db.connections.find_one({}, {"_id": 0}, sort=[("created_at", -1)])
        
        if not conn or not conn.get("is_connected"):
            return {"opportunities": [], "message": "Not connected to plays888.co"}
        
        # Get active rules
        rules = await db.betting_rules.find({"enabled": True}, {"_id": 0}).to_list(100)
        
        if not rules:
            return {"opportunities": [], "message": "No active betting rules"}
        
        # Login and get opportunities
        username = conn["username"]
        password = decrypt_password(conn["password_encrypted"])
        
        await plays888_service.login(username, password)
        raw_opportunities = await plays888_service.get_opportunities()
        await plays888_service.close()
        
        # Match opportunities with rules
        matched_opportunities = []
        for opp in raw_opportunities:
            for rule in rules:
                # Check if opportunity matches rule
                matches = True
                if rule.get("min_odds") and opp["odds"] < rule["min_odds"]:
                    matches = False
                if rule.get("max_odds") and opp["odds"] > rule["max_odds"]:
                    matches = False
                if rule.get("sport") and opp["sport"].lower() != rule["sport"].lower():
                    matches = False
                
                if matches:
                    potential_win = calculate_american_odds_payout(rule["wager_amount"], opp["odds"])
                    opp_doc = {
                        "id": str(uuid.uuid4()),
                        "event_name": opp["event_name"],
                        "odds": opp["odds"],
                        "sport": opp["sport"],
                        "bet_type": opp["bet_type"],
                        "available": True,
                        "matched_rule_id": rule["id"],
                        "matched_rule_name": rule["name"],
                        "wager_amount": rule["wager_amount"],
                        "potential_win": potential_win,
                        "auto_place": rule["auto_place"],
                        "discovered_at": datetime.now(timezone.utc).isoformat()
                    }
                    matched_opportunities.append(opp_doc)
                    break
        
        return {"opportunities": matched_opportunities}
    except Exception as e:
        logger.error(f"Get opportunities error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/open-bets/{account}")
async def get_open_bets(account: str = "ENANO"):
    """Get open/pending bets from plays888.co for a specific account"""
    scrape_service = Plays888Service()
    
    try:
        # Get connection credentials based on account
        # ENANO = jac075, TIPSTER = jac083
        if account.upper() == "ENANO":
            username = "jac075"
        elif account.upper() == "TIPSTER":
            username = "jac083"
        else:
            raise HTTPException(status_code=400, detail=f"Unknown account: {account}")
        
        # Get password from connections
        conn = await db.connections.find_one({"username": username}, {"_id": 0})
        
        if not conn:
            raise HTTPException(status_code=400, detail=f"No connection found for {account}")
        
        password = decrypt_password(conn["password_encrypted"])
        
        # Login and scrape open bets
        await scrape_service.login(username, password)
        open_bets = await scrape_service.scrape_open_bets()
        
        return {
            "success": True,
            "account": account.upper(),
            "open_bets": open_bets,
            "count": len(open_bets)
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get open bets error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        await scrape_service.close()


@api_router.post("/scrape/totals/{league}")
async def scrape_plays888_totals(league: str):
    """Scrape over/under totals from plays888.co for NBA or NHL"""
    # Create a new service instance for each request
    scrape_service = Plays888Service()
    
    try:
        # Get connection credentials
        conn = await db.connections.find_one({}, {"_id": 0}, sort=[("created_at", -1)])
        
        if not conn:
            raise HTTPException(status_code=400, detail="No plays888 connection configured")
        
        username = conn["username"]
        password = decrypt_password(conn["password_encrypted"])
        
        # Login and scrape
        await scrape_service.login(username, password)
        games = await scrape_service.scrape_totals(league.upper())
        
        return {
            "success": True,
            "league": league.upper(),
            "games": games,
            "count": len(games),
            "screenshot": f"/tmp/plays888_{league.lower()}_totals.png"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Scrape totals error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        await scrape_service.close()


@api_router.post("/bets/place")
async def place_bet(bet_request: PlaceBetRequest):
    """Place a bet manually"""
    try:
        # Get opportunity from recent scan
        # In production, this would place actual bet via Playwright
        
        bet_doc = {
            "id": str(uuid.uuid4()),
            "opportunity_id": bet_request.opportunity_id,
            "rule_id": "manual",
            "wager_amount": bet_request.wager_amount,
            "odds": 0,
            "status": "placed",
            "placed_at": datetime.now(timezone.utc).isoformat(),
            "result": None
        }
        
        await db.bet_history.insert_one(bet_doc)
        
        return {"success": True, "message": "Bet placed successfully", "bet_id": bet_doc["id"]}
    except Exception as e:
        logger.error(f"Place bet error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


class SpecificBetRequest(BaseModel):
    game: str
    bet_type: str
    line: str
    odds: int
    wager: float
    league: str = "NATIONAL HOCKEY LEAGUE - OT INCLUDED"  # Default to NHL

class ManualBetRecord(BaseModel):
    game: str
    bet_type: str
    line: str
    odds: int
    wager: float
    bet_slip_id: Optional[str] = None
    notes: Optional[str] = None


@api_router.post("/bets/place-specific")
async def place_specific_bet(bet_request: SpecificBetRequest):
    """Place a specific bet on plays888.co"""
    # Create a new service instance for this bet
    bet_service = Plays888Service()
    
    try:
        # Get connection credentials
        conn = await db.connections.find_one({}, {"_id": 0}, sort=[("created_at", -1)])
        
        if not conn or not conn.get("is_connected"):
            raise HTTPException(status_code=400, detail="Not connected to plays888.co")
        
        # Login and place bet
        username = conn["username"]
        password = decrypt_password(conn["password_encrypted"])
        
        await bet_service.initialize()
        login_result = await bet_service.login(username, password)
        
        if not login_result["success"]:
            await bet_service.close()
            raise HTTPException(status_code=400, detail=f"Login failed: {login_result['message']}")
        
        # Place the bet
        result = await bet_service.place_specific_bet(
            game=bet_request.game,
            bet_type=bet_request.bet_type,
            line=bet_request.line,
            odds=bet_request.odds,
            wager=bet_request.wager,
            league=bet_request.league
        )
        
        await bet_service.close()
        
        # Store in history
        if result["success"]:
            bet_doc = {
                "id": str(uuid.uuid4()),
                "opportunity_id": "specific",
                "rule_id": "manual",
                "wager_amount": bet_request.wager,
                "odds": bet_request.odds,
                "status": "placed",
                "placed_at": datetime.now(timezone.utc).isoformat(),
                "result": None,
                "game": bet_request.game,
                "bet_type": bet_request.bet_type,
                "line": bet_request.line
            }
            await db.bet_history.insert_one(bet_doc)
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Place specific bet error: {str(e)}")
        await bet_service.close()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/bets/history")
async def get_bet_history():
    """Get betting history"""
    try:
        bets = await db.bet_history.find({}, {"_id": 0}).sort("placed_at", -1).to_list(100)
        
        for bet in bets:
            if isinstance(bet.get('placed_at'), str):
                bet['placed_at'] = datetime.fromisoformat(bet['placed_at'])
        
        return bets
    except Exception as e:
        logger.error(f"Get bet history error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/bets/record-manual")
async def record_manual_bet(bet: ManualBetRecord):
    """Record a manually placed bet"""
    try:
        potential_win = calculate_american_odds_payout(bet.wager, bet.odds)
        
        bet_doc = {
            "id": str(uuid.uuid4()),
            "opportunity_id": "manual",
            "rule_id": "manual",
            "wager_amount": bet.wager,
            "odds": bet.odds,
            "status": "placed",
            "placed_at": datetime.now(timezone.utc).isoformat(),
            "result": None,
            "game": bet.game,
            "bet_type": bet.bet_type,
            "line": bet.line,
            "bet_slip_id": bet.bet_slip_id,
            "notes": bet.notes
        }
        
        await db.bet_history.insert_one(bet_doc)
        
        # Send Telegram notification
        await send_telegram_notification({
            "game": bet.game,
            "bet_type": bet.bet_type,
            "line": bet.line,
            "odds": bet.odds,
            "wager": bet.wager,
            "potential_win": potential_win,
            "ticket_number": bet.bet_slip_id or bet_doc["id"],
            "status": "Placed",
            "league": bet.notes if "via extension" not in (bet.notes or "") else ""
        })
        
        return {
            "success": True,
            "message": f"Bet recorded: {bet.game} - {bet.bet_type} {bet.line} @ {bet.odds} for ${bet.wager}",
            "bet_id": bet_doc["id"]
        }
    except Exception as e:
        logger.error(f"Record manual bet error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


class TelegramConfig(BaseModel):
    bot_token: str
    chat_id: int

@api_router.post("/telegram/config")
async def configure_telegram(config: TelegramConfig):
    """Configure Telegram bot for notifications"""
    global telegram_bot, telegram_chat_id
    
    try:
        # Test the bot token
        test_bot = Bot(token=config.bot_token)
        bot_info = await test_bot.get_me()
        
        # Store configuration in memory
        telegram_bot = test_bot
        telegram_chat_id = config.chat_id
        
        # Persist to MongoDB (upsert - update or insert)
        await db.telegram_config.delete_many({})  # Remove old config
        await db.telegram_config.insert_one({
            "bot_token": config.bot_token,
            "chat_id": config.chat_id,
            "bot_username": bot_info.username,
            "bot_name": bot_info.first_name,
            "configured_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Telegram config saved to database for @{bot_info.username}")
        
        # Send test message
        await telegram_bot.send_message(
            chat_id=telegram_chat_id,
            text="✅ *Telegram Notifications Enabled*\\n\\nYou will receive notifications when bets are placed\\.",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        
        return {
            "success": True,
            "message": f"Telegram configured successfully. Bot: @{bot_info.username}",
            "bot_name": bot_info.first_name
        }
    except Exception as e:
        logger.error(f"Telegram configuration error: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Failed to configure Telegram: {str(e)}")

@api_router.get("/telegram/status")
async def telegram_status():
    """Check Telegram configuration status"""
    global telegram_bot, telegram_chat_id
    
    # If not in memory, try loading from database
    if not telegram_bot or not telegram_chat_id:
        try:
            config = await db.telegram_config.find_one({}, {"_id": 0})
            if config:
                telegram_bot = Bot(token=config["bot_token"])
                telegram_chat_id = int(config["chat_id"])
                logger.info("Telegram reloaded from database")
        except Exception as e:
            logger.error(f"Error loading Telegram from database: {e}")
    
    if telegram_bot and telegram_chat_id:
        try:
            bot_info = await telegram_bot.get_me()
            return {
                "configured": True,
                "bot_username": bot_info.username,
                "bot_name": bot_info.first_name,
                "chat_id": telegram_chat_id
            }
        except:
            return {"configured": False, "error": "Bot token invalid"}
    return {"configured": False}

@api_router.post("/telegram/test")
async def test_telegram():
    """Send a test notification using the new compilation system"""
    if not telegram_bot or not telegram_chat_id:
        raise HTTPException(status_code=400, detail="Telegram not configured")
    
    try:
        # Test the new compilation-based notification
        await send_telegram_notification({
            "game": "FALCONS vs CARDINALS",
            "description": "Atlanta Falcons @ Arizona Cardinals",
            "bet_type": "TOTAL UNDER 48",
            "line": "u48",
            "odds": -110,
            "wager": 2200,
            "potential_win": 2000,
            "ticket_number": f"TEST{datetime.now().strftime('%H%M%S')}",
            "status": "Placed"
        }, account="jac075")  # Test with ENANO label
        return {"success": True, "message": "Test bet added to compilation"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/telegram/test-compilation")
async def test_compilation():
    """Test the full compilation workflow with multiple bets"""
    if not telegram_bot or not telegram_chat_id:
        raise HTTPException(status_code=400, detail="Telegram not configured")
    
    try:
        import time
        timestamp = int(time.time())
        
        # Add first bet
        await send_telegram_notification({
            "game": "FALCONS vs CARDINALS",
            "description": "Atlanta Falcons @ Arizona Cardinals",
            "bet_type": "TOTAL UNDER 48",
            "odds": -110,
            "wager": 2200,
            "potential_win": 2000,
            "ticket_number": f"DEMO{timestamp}A"
        }, account="jac075")
        
        await asyncio.sleep(2)
        
        # Add second bet
        await send_telegram_notification({
            "game": "JAGUARS vs BRONCOS",
            "description": "Jacksonville Jaguars @ Denver Broncos",
            "bet_type": "TOTAL OVER 47",
            "odds": -110,
            "wager": 2200,
            "potential_win": 2000,
            "ticket_number": f"DEMO{timestamp}B"
        }, account="jac075")
        
        await asyncio.sleep(2)
        
        # Mark first bet as won
        await update_bet_result_in_compilation("jac075", f"DEMO{timestamp}A", "won", 2000)
        
        return {"success": True, "message": "Test compilation workflow completed - check Telegram!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/telegram/daily-summary")
async def trigger_daily_summary():
    """Manually trigger the daily betting summary"""
    if not telegram_bot or not telegram_chat_id:
        raise HTTPException(status_code=400, detail="Telegram not configured")
    
    try:
        await send_daily_summary()
        return {"success": True, "message": "Daily summary sent"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/telegram/send-compilations")
async def send_all_compilations():
    """Send compilation messages: ENANO comparison view + TIPSTER detailed view"""
    if not telegram_bot or not telegram_chat_id:
        raise HTTPException(status_code=400, detail="Telegram not configured")
    
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        today = datetime.now(arizona_tz).strftime('%Y-%m-%d')
        
        # Find all compilations for today
        compilations = await db.daily_compilations.find({"date": today}).to_list(10)
        
        if not compilations:
            return {"success": True, "message": "No compilations to send"}
        
        # Delete ALL old messages first
        for comp in compilations:
            for field in ['message_id_short', 'message_id_detailed', 'message_id']:
                old_id = comp.get(field)
                if old_id:
                    try:
                        await telegram_bot.delete_message(chat_id=telegram_chat_id, message_id=old_id)
                    except Exception:
                        pass
        
        message_ids = {}
        sent_count = 0
        
        # 1. ENANO comparison view (detailed comparison against TIPSTER)
        enano_comparison = await build_enano_comparison_message()
        if enano_comparison:
            sent = await telegram_bot.send_message(
                chat_id=telegram_chat_id,
                text=enano_comparison,
                parse_mode=ParseMode.MARKDOWN
            )
            message_ids["jac075_detailed"] = sent.message_id
            sent_count += 1
        
        # 2. TIPSTER detailed view
        tipster_detail = await build_compilation_message("jac083", detailed=True)
        if tipster_detail:
            sent = await telegram_bot.send_message(
                chat_id=telegram_chat_id,
                text=tipster_detail,
                parse_mode=ParseMode.MARKDOWN
            )
            message_ids["jac083_detailed"] = sent.message_id
            sent_count += 1
        
        # Update database with new message IDs (no more short messages)
        await db.daily_compilations.update_one(
            {"account": "jac075", "date": today},
            {"$set": {
                "message_id_detailed": message_ids.get("jac075_detailed"),
                "message_id_short": None,
                "message_id": None
            }}
        )
        await db.daily_compilations.update_one(
            {"account": "jac083", "date": today},
            {"$set": {
                "message_id_detailed": message_ids.get("jac083_detailed"),
                "message_id_short": None,
                "message_id": None
            }}
        )
        
        return {"success": True, "message": f"Sent {sent_count} compilation messages"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/telegram/activity-summary")
async def trigger_activity_summary():
    """Manually trigger the activity summary"""
    if not telegram_bot or not telegram_chat_id:
        raise HTTPException(status_code=400, detail="Telegram not configured")
    
    try:
        await send_activity_summary()
        return {"success": True, "message": "Activity summary sent"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/bets/check-results")
async def trigger_results_check():
    """Manually trigger a check for settled bet results"""
    try:
        await check_bet_results()
        return {"success": True, "message": "Results check completed"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/stats")
async def get_stats():
    """Get statistics for dashboard"""
    try:
        total_bets = await db.bet_history.count_documents({})
        active_rules = await db.betting_rules.count_documents({"enabled": True})
        
        return {
            "total_bets": total_bets,
            "active_rules": active_rules,
            "win_rate": 0,
            "total_profit": 0
        }
    except Exception as e:
        logger.error(f"Get stats error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/accounts")
async def get_accounts():
    """Get list of connected accounts"""
    try:
        connections = await db.connections.find({"is_connected": True}, {"_id": 0}).to_list(100)
        accounts = []
        for conn in connections:
            username = conn.get("username", "")
            label = ACCOUNT_LABELS.get(username, username)
            accounts.append({
                "username": username,
                "label": label,
                "is_connected": conn.get("is_connected", False)
            })
        return {"accounts": accounts}
    except Exception as e:
        logger.error(f"Get accounts error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Cache for account summaries (5-minute TTL)
account_summary_cache = {}
CACHE_TTL_SECONDS = 300  # 5 minutes

@api_router.get("/accounts/{username}/summary")
async def get_account_summary(username: str, force_refresh: bool = False):
    """Get daily/weekly profit summary for a specific account (cached for 5 min)"""
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        now_arizona = datetime.now(arizona_tz)
        current_time = datetime.now(timezone.utc)
        
        # Check cache (unless force refresh requested)
        if not force_refresh and username in account_summary_cache:
            cached = account_summary_cache[username]
            cache_age = (current_time - cached['cached_at']).total_seconds()
            if cache_age < CACHE_TTL_SECONDS:
                logger.info(f"Returning cached summary for {username} (age: {cache_age:.0f}s)")
                return cached['data']
        
        # Find the connection
        conn = await db.connections.find_one({"username": username, "is_connected": True}, {"_id": 0})
        if not conn:
            raise HTTPException(status_code=404, detail=f"Account {username} not found or not connected")
        
        password = decrypt_password(conn["password_encrypted"])
        label = ACCOUNT_LABELS.get(username, username)
        
        # Get the daily totals from plays888
        totals = await get_plays888_daily_totals(username, password)
        
        day_names = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']
        today_day = day_names[now_arizona.weekday()]
        
        # Get bet count from the scraped data (from plays888.co)
        total_bets = totals.get('total_bets', 0) if totals else 0
        
        if not totals or not totals.get('daily_profits'):
            result = {
                "username": username,
                "label": label,
                "success": False,
                "error": "Could not retrieve data from plays888.co",
                "daily_profits": [],
                "week_total": 0,
                "today_profit": 0,
                "today_day": today_day,
                "date": now_arizona.strftime('%B %d, %Y'),
                "total_bets": total_bets
            }
        else:
            # Find today's profit
            today_profit = 0
            for day_data in totals['daily_profits']:
                if day_data['day'].lower() == today_day:
                    today_profit = day_data['profit']
                    break
            
            result = {
                "username": username,
                "label": label,
                "success": True,
                "daily_profits": totals['daily_profits'],
                "week_total": totals.get('week_total', 0),
                "today_profit": today_profit,
                "today_day": today_day,
                "date": now_arizona.strftime('%B %d, %Y'),
                "total_bets": total_bets
            }
        
        # Store in cache
        account_summary_cache[username] = {
            'data': result,
            'cached_at': current_time
        }
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get account summary error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/accounts/all/summaries")
async def get_all_account_summaries():
    """Get summaries for all connected accounts (uses cache)"""
    try:
        connections = await db.connections.find({"is_connected": True}, {"_id": 0}).to_list(100)
        summaries = []
        
        for conn in connections:
            username = conn.get("username", "")
            try:
                # Use the cached endpoint logic
                summary = await get_account_summary(username, force_refresh=False)
                summaries.append(summary)
            except Exception as e:
                logger.error(f"Error getting summary for {username}: {e}")
                summaries.append({
                    "username": username,
                    "label": ACCOUNT_LABELS.get(username, username),
                    "success": False,
                    "error": str(e)
                })
        
        return {"summaries": summaries}
    except Exception as e:
        logger.error(f"Get all summaries error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/monitoring/start")
async def start_monitoring():
    """Start the bet monitoring system"""
    global monitoring_enabled
    
    # Check if connection is configured
    conn = await db.connections.find_one({}, {"_id": 0})
    if not conn or not conn.get("is_connected"):
        raise HTTPException(status_code=400, detail="Please configure plays888.co connection first")
    
    monitoring_enabled = True
    
    # Save monitoring config to DB for auto-start on restart
    await db.monitor_config.delete_many({})
    await db.monitor_config.insert_one({
        "auto_start": True,
        "interval_minutes": 15,  # Production interval
        "enabled_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Schedule the job with random intervals (7-15 minutes)
    # Uses a wrapper that reschedules with random delay after each run
    schedule_next_check()
    
    if not scheduler.running:
        scheduler.start()
    
    logger.info("Bet monitoring started - checking every 5 minutes (paused 10:00 PM - 7:00 AM Arizona)")
    
    return {
        "success": True,
        "message": "Bet monitoring started. Will check plays888.co every 5 minutes (paused during sleep hours 10:00 PM - 7:00 AM Arizona).",
        "interval": "7-15 minutes (random)"
    }

@api_router.post("/monitoring/stop")
async def stop_monitoring():
    """Stop the bet monitoring system"""
    global monitoring_enabled
    
    monitoring_enabled = False
    
    # Update DB to disable auto-start
    await db.monitor_config.update_one(
        {},
        {"$set": {"auto_start": False}},
        upsert=True
    )
    
    if scheduler.running:
        scheduler.remove_job('bet_monitor')
    
    logger.info("Bet monitoring stopped")
    
    return {
        "success": True,
        "message": "Bet monitoring stopped"
    }

@api_router.get("/monitoring/status")
async def monitoring_status():
    """Get monitoring system status"""
    next_check = None
    try:
        if monitoring_enabled and scheduler.running:
            job = scheduler.get_job('bet_monitor')
            if job and job.next_run_time:
                next_check = job.next_run_time.isoformat()
    except Exception as e:
        logger.error(f"Error getting next check time: {e}")
    
    return {
        "enabled": monitoring_enabled,
        "interval": "7-15 minutes (random)",
        "sleep_hours": "10:00 PM - 7:00 AM Arizona",
        "running": scheduler.running,
        "next_check": next_check
    }

@api_router.post("/monitoring/check-now")
async def check_now():
    """Manually trigger a bet check immediately"""
    if not monitoring_enabled:
        raise HTTPException(status_code=400, detail="Monitoring is not enabled. Please start monitoring first.")
    
    # Run the full check cycle (which includes the notification) in background
    asyncio.create_task(monitor_and_reschedule())
    
    return {
        "success": True,
        "message": "Manual bet check triggered. Results will be logged and notifications sent if new bets found."
    }


# ============== NBA OPPORTUNITIES SCRAPING ==============

async def scrape_nba_odds():
    """Scrape NBA odds from TeamRankings"""
    import re
    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.get("https://www.teamrankings.com/nba/odds/")
        html = response.text
        
        games = []
        # Parse the HTML to extract game data
        # Looking for tables with game info
        
        # Find today's date section
        today_pattern = r'<h2[^>]*>([^<]*December[^<]*2025[^<]*)</h2>(.*?)(?=<h2|$)'
        today_match = re.search(today_pattern, html, re.DOTALL | re.IGNORECASE)
        
        if not today_match:
            # Try to find any games
            logger.warning("Could not find today's games section")
            return games
        
        games_html = today_match.group(2)
        
        # Parse each game row
        game_pattern = r'(\d{1,2}:\d{2}\s*[AP]M\s*EST).*?<a[^>]*>([^<]+)</a>.*?<a[^>]*>([^<]+)</a>.*?(\d+\.?\d*)\s*\|'
        game_matches = re.findall(game_pattern, games_html, re.DOTALL | re.IGNORECASE)
        
        # Simpler approach - just get the key data
        time_pattern = r'<tr[^>]*>.*?(\d{1,2}:\d{2}\s*[AP]M\s*EST)'
        team_pattern = r'teamrankings\.com/nba/team/([^"]+)"[^>]*>([^<]+)</a>'
        total_pattern = r'>(\d{3}\.?\d?)</td>'
        
        return games

async def scrape_nba_ppg_rankings(target_date: str = None):
    """
    Scrape NBA Points Per Game rankings and values from teamrankings.com
    
    Args:
        target_date: The date string (YYYY-MM-DD) for which to get PPG data.
                    This should be the date of the games being analyzed.
                    The teamrankings data for a given date reflects stats BEFORE that day's games.
    
    Returns: {
        'season_ranks': {team: rank},
        'season_values': {team: ppg_value},
        'last3_ranks': {team: rank},
        'last3_values': {team: ppg_value}
    }
    """
    import re
    
    # Team name mapping
    team_map = {
        'denver-nuggets': 'Denver', 'oklahoma-city-thunder': 'Okla City',
        'houston-rockets': 'Houston', 'new-york-knicks': 'New York',
        'miami-heat': 'Miami', 'utah-jazz': 'Utah', 'san-antonio-spurs': 'San Antonio',
        'chicago-bulls': 'Chicago', 'detroit-pistons': 'Detroit', 'atlanta-hawks': 'Atlanta',
        'cleveland-cavaliers': 'Cleveland', 'minnesota-timberwolves': 'Minnesota',
        'orlando-magic': 'Orlando', 'los-angeles-lakers': 'LA Lakers',
        'portland-trail-blazers': 'Portland', 'philadelphia-76ers': 'Philadelphia',
        'boston-celtics': 'Boston', 'new-orleans-pelicans': 'New Orleans',
        'memphis-grizzlies': 'Memphis', 'charlotte-hornets': 'Charlotte',
        'phoenix-suns': 'Phoenix', 'golden-state-warriors': 'Golden State',
        'toronto-raptors': 'Toronto', 'milwaukee-bucks': 'Milwaukee',
        'dallas-mavericks': 'Dallas', 'washington-wizards': 'Washington',
        'sacramento-kings': 'Sacramento', 'los-angeles-clippers': 'LA Clippers',
        'indiana-pacers': 'Indiana', 'brooklyn-nets': 'Brooklyn'
    }
    
    result = {
        'season_ranks': {},
        'season_values': {},
        'last3_ranks': {},
        'last3_values': {}
    }
    
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Connection': 'keep-alive',
                'Cache-Control': 'max-age=0'
            }
            
            # Determine the date to use for scraping
            # If target_date provided, use it; otherwise use today's Arizona date
            from zoneinfo import ZoneInfo
            arizona_tz = ZoneInfo('America/Phoenix')
            
            if target_date:
                scrape_date = target_date
            else:
                # Default to today's date (Arizona time)
                scrape_date = datetime.now(arizona_tz).strftime('%Y-%m-%d')
            
            # The teamrankings data for a given date shows stats AS OF that date
            # (i.e., BEFORE that day's games have been played)
            ppg_url = f"https://www.teamrankings.com/nba/stat/points-per-game?date={scrape_date}"
            logger.info(f"Scraping NBA PPG from teamrankings.com with date={scrape_date}")
            
            season_response = await client.get(ppg_url, headers=headers)
            season_html = season_response.text
            
            # Parse season rankings - pattern: rank, team slug, team name, 2024-25 PPG, Last 3 PPG
            # Table has: Rank | Team | 2024-25 | Last 3 | Last 1 | Home | Away | ...
            row_pattern = r'<tr[^>]*>\s*<td[^>]*>(\d+)</td>\s*<td[^>]*>.*?team/([^"]+)"[^>]*>[^<]*</a>\s*</td>\s*<td[^>]*>([\d.]+)</td>\s*<td[^>]*>([\d.]+)</td>'
            
            matches = re.findall(row_pattern, season_html, re.DOTALL)
            
            logger.info(f"Found {len(matches)} teams in teamrankings.com PPG table (date={scrape_date})")
            
            for rank, slug, season_ppg, last3_ppg in matches:
                team_name = team_map.get(slug)
                if team_name:
                    result['season_ranks'][team_name] = int(rank)
                    result['season_values'][team_name] = float(season_ppg)
                    result['last3_ranks'][team_name] = int(rank)  # Will update based on Last 3 values
                    result['last3_values'][team_name] = float(last3_ppg)
            
            # Create ranking based on Last 3 values (sorted descending)
            if result['last3_values']:
                last3_sorted = sorted(result['last3_values'].items(), key=lambda x: x[1], reverse=True)
                for i, (team, _) in enumerate(last3_sorted, 1):
                    result['last3_ranks'][team] = i
            
            logger.info(f"Scraped PPG data: {len(result['season_values'])} teams")
            return result
            
    except Exception as e:
        logger.error(f"Error scraping NBA PPG rankings: {e}")
        import traceback
        traceback.print_exc()
        return result

async def scrape_ncaab_ppg_rankings(target_date: str = None):
    """
    Scrape NCAAB (NCAA Basketball) Points Per Game rankings and values from teamrankings.com
    
    Args:
        target_date: The date string (YYYY-MM-DD) for which to get PPG data.
                    This should be the date of the games being analyzed.
                    The teamrankings data for a given date reflects stats BEFORE that day's games.
    
    Returns: {
        'season_ranks': {team: rank},
        'season_values': {team: ppg_value},
        'last3_ranks': {team: rank},
        'last3_values': {team: ppg_value}
    }
    """
    import re
    
    result = {
        'season_ranks': {},
        'season_values': {},
        'last3_ranks': {},
        'last3_values': {}
    }
    
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Connection': 'keep-alive',
                'Cache-Control': 'max-age=0'
            }
            
            # Determine the date to use for scraping
            from zoneinfo import ZoneInfo
            arizona_tz = ZoneInfo('America/Phoenix')
            
            if target_date:
                scrape_date = target_date
            else:
                # Default to today's date (Arizona time)
                scrape_date = datetime.now(arizona_tz).strftime('%Y-%m-%d')
            
            # Scrape NCAAB PPG from teamrankings with the target date
            # The data for a given date shows stats BEFORE that day's games
            ppg_url = f"https://www.teamrankings.com/ncaa-basketball/stat/points-per-game?date={scrape_date}"
            logger.info(f"Scraping NCAAB PPG from teamrankings.com with date={scrape_date}")
            
            response = await client.get(ppg_url, headers=headers)
            html = response.text
            
            # Parse rankings - pattern: rank, team slug, team name, 2024-25 PPG, Last 3 PPG
            # Table columns: Rank | Team | 2024-25 | Last 3 | Last 1 | Home | Away | ...
            row_pattern = r'<tr[^>]*>\s*<td[^>]*>(\d+)</td>\s*<td[^>]*>.*?team/([^"]+)"[^>]*>([^<]*)</a>\s*</td>\s*<td[^>]*>([\d.]+)</td>\s*<td[^>]*>([\d.]+)</td>'
            
            matches = re.findall(row_pattern, html, re.DOTALL)
            
            logger.info(f"Found {len(matches)} NCAAB teams in teamrankings.com PPG table (date={scrape_date})")
            
            for rank, slug, display_name, season_ppg, last3_ppg in matches:
                # Use display name as-is (clean it up)
                team_name = display_name.strip()
                if team_name:
                    result['season_ranks'][team_name] = int(rank)
                    result['season_values'][team_name] = float(season_ppg)
                    result['last3_values'][team_name] = float(last3_ppg)
            
            # Create ranking based on Last 3 values
            last3_sorted = sorted(result['last3_values'].items(), key=lambda x: x[1], reverse=True)
            for i, (team, _) in enumerate(last3_sorted, 1):
                result['last3_ranks'][team] = i
            
            logger.info(f"Scraped NCAAB PPG data: {len(result['season_values'])} teams")
            return result
            
    except Exception as e:
        logger.error(f"Error scraping NCAAB PPG rankings: {e}")
        import traceback
        traceback.print_exc()
        return result

async def get_ncaab_opportunities():
    """Get NCAAB betting opportunities with PPG analysis"""
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    today = datetime.now(arizona_tz).strftime('%Y-%m-%d')
    
    # Check if we have cached data for today
    cached = await db.ncaab_opportunities.find_one({"date": today})
    if cached:
        return cached
    
    # Scrape PPG rankings with today's date (pre-game values)
    ppg_data = await scrape_ncaab_ppg_rankings(today)
    
    # Scrape today's games from scoresandodds
    games = await scrape_scoresandodds("NCAAB", today)
    
    # Process games and add PPG analysis
    processed_games = []
    
    # NCAAB has ~365 teams, so ranking thresholds are different
    # Top 25% = rank 1-91 (Green)
    # 25-50% = rank 92-182 (Yellow) 
    # 50-75% = rank 183-273 (Red)
    # Bottom 25% = rank 274-365 (Blue)
    def get_ncaab_dot_color(rank):
        if rank is None:
            return '⚪'  # Unknown - not in top 365
        if rank <= 92:
            return '🟢'  # Top tier (1-92)
        elif rank <= 184:
            return '🔵'  # Second tier (93-184)
        elif rank <= 276:
            return '🟡'  # Third tier (185-276)
        elif rank <= 365:
            return '🔴'  # Fourth tier (277-365)
        else:
            return '⚪'  # Unknown - not in top 365
    
    for game in games:
        away_team = game.get('away', '')
        home_team = game.get('home', '')
        
        # Get PPG data
        away_ppg_rank = ppg_data['season_ranks'].get(away_team)
        away_ppg_value = ppg_data['season_values'].get(away_team)
        away_last3_rank = ppg_data['last3_ranks'].get(away_team)
        away_last3_value = ppg_data['last3_values'].get(away_team)
        
        home_ppg_rank = ppg_data['season_ranks'].get(home_team)
        home_ppg_value = ppg_data['season_values'].get(home_team)
        home_last3_rank = ppg_data['last3_ranks'].get(home_team)
        home_last3_value = ppg_data['last3_values'].get(home_team)
        
        # Calculate combined PPG using the correct formula:
        # (Team1 Season PPG + Team2 Season PPG + Team1 L3 PPG + Team2 L3 PPG) / 2
        combined_ppg = None
        if away_ppg_value and home_ppg_value and away_last3_value and home_last3_value:
            combined_ppg = round((away_ppg_value + home_ppg_value + away_last3_value + home_last3_value) / 2, 1)
        elif away_ppg_value and home_ppg_value:
            # Fallback to season only if L3 not available
            combined_ppg = round((away_ppg_value + home_ppg_value), 1)
        
        # Calculate edge (same as NBA)
        edge = None
        line = game.get('total')
        if combined_ppg and line:
            try:
                edge = round(combined_ppg - float(line), 1)
            except:
                pass
        
        # Generate recommendation (NCAAB uses edge threshold of 9)
        recommendation = ''
        if edge is not None:
            if edge >= 10:
                recommendation = 'OVER'
            elif edge <= -9:
                recommendation = 'UNDER'
        
        # Generate dot colors
        away_dots = get_ncaab_dot_color(away_ppg_rank) + get_ncaab_dot_color(away_last3_rank)
        home_dots = get_ncaab_dot_color(home_ppg_rank) + get_ncaab_dot_color(home_last3_rank)
        
        processed_game = {
            **game,
            'away_team': away_team,
            'home_team': home_team,
            'away_ppg_rank': away_ppg_rank,
            'away_ppg_value': away_ppg_value,
            'away_last3_rank': away_last3_rank,
            'away_last3_value': away_last3_value,
            'home_ppg_rank': home_ppg_rank,
            'home_ppg_value': home_ppg_value,
            'home_last3_rank': home_last3_rank,
            'home_last3_value': home_last3_value,
            'combined_ppg': combined_ppg,
            'edge': edge,
            'recommendation': recommendation,
            'away_dots': away_dots,
            'home_dots': home_dots,
            'opening_line': line
        }
        processed_games.append(processed_game)
    
    # Save to database
    doc = {
        "date": today,
        "games": processed_games,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.ncaab_opportunities.replace_one(
        {"date": today},
        doc,
        upsert=True
    )
    
    return doc

async def get_nba_opportunities():
    """Get NBA betting opportunities with PPG analysis"""
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    today = datetime.now(arizona_tz).strftime('%Y-%m-%d')
    
    # Check if we have cached data for today
    cached = await db.nba_opportunities.find_one({"date": today})
    if cached:
        return cached.get('games', [])
    
    # If no cache, return empty (data will be refreshed by scheduled job)
    return []

async def refresh_nba_opportunities_scheduled():
    """Scheduled job to refresh NBA opportunities data daily at 10:30 PM Arizona"""
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    today = datetime.now(arizona_tz).strftime('%Y-%m-%d')
    
    logger.info(f"[Scheduled] Refreshing NBA opportunities for {today}")
    
    try:
        # PPG Rankings (Season) - would be scraped from teamrankings.com in production
        ppg_season = {
            'Denver': 1, 'Okla City': 2, 'Houston': 3, 'New York': 4, 'Miami': 5,
            'Utah': 6, 'San Antonio': 7, 'Chicago': 8, 'Detroit': 9, 'Atlanta': 10,
            'Cleveland': 11, 'Minnesota': 12, 'Orlando': 13, 'LA Lakers': 14, 'Portland': 15,
            'Philadelphia': 16, 'Boston': 17, 'New Orleans': 18, 'Memphis': 19, 'Charlotte': 20,
            'Phoenix': 21, 'Golden State': 22, 'Toronto': 23, 'Milwaukee': 24, 'Dallas': 25,
            'Washington': 26, 'Sacramento': 27, 'LA Clippers': 28, 'Indiana': 29, 'Brooklyn': 30
        }
        
        # PPG Rankings (Last 3 games)
        ppg_last3 = {
            'Chicago': 1, 'Utah': 2, 'New Orleans': 3, 'Atlanta': 4, 'San Antonio': 5,
            'Portland': 6, 'Houston': 7, 'Orlando': 8, 'Dallas': 9, 'Memphis': 10,
            'Denver': 11, 'Philadelphia': 12, 'New York': 13, 'Sacramento': 14, 'Golden State': 15,
            'LA Lakers': 16, 'Cleveland': 17, 'Miami': 18, 'Boston': 19, 'Okla City': 20,
            'Detroit': 21, 'Charlotte': 22, 'Washington': 23, 'Phoenix': 24, 'Brooklyn': 25,
            'Toronto': 26, 'Indiana': 27, 'Minnesota': 28, 'LA Clippers': 29, 'Milwaukee': 30
        }
        
        # PPG Values (Season) - Actual points per game averages
        ppg_season_values = {
            'Denver': 125.8, 'Okla City': 121.5, 'New York': 120.6, 'Utah': 120.5,
            'Houston': 120.3, 'Miami': 120.2, 'San Antonio': 119.9, 'Cleveland': 119.5,
            'Detroit': 118.8, 'Chicago': 118.8, 'Atlanta': 118.8, 'Minnesota': 118.8,
            'LA Lakers': 117.2, 'Orlando': 117.1, 'Portland': 116.6, 'Boston': 116.5,
            'Philadelphia': 115.7, 'Charlotte': 115.6, 'Memphis': 115.3, 'New Orleans': 115.2,
            'Phoenix': 115.1, 'Golden State': 114.9, 'Toronto': 114.4, 'Washington': 113.6,
            'Dallas': 113.5, 'Milwaukee': 112.7, 'Sacramento': 111.7, 'LA Clippers': 111.5,
            'Brooklyn': 109.3, 'Indiana': 109.2
        }
        
        # PPG Values (Last 3 games)
        ppg_last3_values = {
            'Chicago': 136.3, 'Utah': 128.7, 'New Orleans': 127.3, 'Atlanta': 124.7, 'San Antonio': 123.3,
            'Portland': 122.7, 'Houston': 122.0, 'Orlando': 121.7, 'Dallas': 121.0, 'Memphis': 120.7,
            'Denver': 120.3, 'Philadelphia': 119.7, 'New York': 119.0, 'Sacramento': 118.3, 'Golden State': 117.7,
            'LA Lakers': 117.3, 'Cleveland': 117.0, 'Miami': 116.7, 'Boston': 115.7, 'Okla City': 114.7,
            'Detroit': 114.0, 'Charlotte': 113.7, 'Washington': 112.0, 'Phoenix': 111.0, 'Brooklyn': 110.7,
            'Toronto': 110.0, 'Indiana': 108.7, 'Minnesota': 107.3, 'LA Clippers': 106.0, 'Milwaukee': 104.0
        }
        
        # Today's games (would be scraped from teamrankings.com/nba/odds/)
        games_raw = [
            {"time": "7:00 PM", "away": "Charlotte", "home": "Cleveland", "total": 239.5},
            {"time": "7:30 PM", "away": "Indiana", "home": "Boston", "total": 226.5},
            {"time": "8:00 PM", "away": "Dallas", "home": "New Orleans", "total": 240.5},
            {"time": "9:00 PM", "away": "Utah", "home": "Denver", "total": 250.5},
            {"time": "9:30 PM", "away": "Memphis", "home": "Okla City", "total": 232.5},
            {"time": "10:00 PM", "away": "Detroit", "home": "Portland", "total": 234.5},
            {"time": "10:00 PM", "away": "Orlando", "home": "Golden State", "total": 227.5},
        ]
        
        # Calculate averages and recommendations
        games = []
        plays = []
        
        for i, g in enumerate(games_raw, 1):
            away_season = ppg_season.get(g['away'], 15)
            away_last3 = ppg_last3.get(g['away'], 15)
            away_avg = (away_season + away_last3) / 2
            
            home_season = ppg_season.get(g['home'], 15)
            home_last3 = ppg_last3.get(g['home'], 15)
            home_avg = (home_season + home_last3) / 2
            
            game_avg = (away_avg + home_avg) / 2
            
            # Get actual PPG values
            away_season_ppg = ppg_season_values.get(g['away'], 115.0)
            away_last3_ppg = ppg_last3_values.get(g['away'], 115.0)
            home_season_ppg = ppg_season_values.get(g['home'], 115.0)
            home_last3_ppg = ppg_last3_values.get(g['home'], 115.0)
            
            # Calculate combined PPG
            season_total = away_season_ppg + home_season_ppg
            last3_total = away_last3_ppg + home_last3_ppg
            combined_ppg = (season_total + last3_total) / 2
            
            # Determine recommendation based on PPG vs Line
            if combined_ppg >= g['total'] + 0.5:
                recommendation = "OVER"
                color = "green"
            elif combined_ppg <= g['total'] - 0.5:
                recommendation = "UNDER"
                color = "red"
            else:
                recommendation = None
                color = "neutral"
            
            game_data = {
                "game_num": i,
                "time": g['time'],
                "away_team": g['away'],
                "away_ppg_rank": away_season,
                "away_last3_rank": away_last3,
                "away_avg": round(away_avg, 1),
                "away_season_ppg": round(away_season_ppg, 1),
                "away_last3_ppg": round(away_last3_ppg, 1),
                "home_team": g['home'],
                "home_ppg_rank": home_season,
                "home_last3_rank": home_last3,
                "home_avg": round(home_avg, 1),
                "home_season_ppg": round(home_season_ppg, 1),
                "home_last3_ppg": round(home_last3_ppg, 1),
                "total": g['total'],
                "game_avg": round(game_avg, 1),
                "combined_ppg": round(combined_ppg, 1),
                "recommendation": recommendation,
                "color": color
            }
            games.append(game_data)
            
            if recommendation:
                plays.append({
                    "game": f"{g['away']} @ {g['home']}",
                    "total": g['total'],
                    "game_avg": round(game_avg, 1),
                    "recommendation": recommendation,
                    "color": color
                })
        
        # Save to database with ppg_populated flag
        await db.nba_opportunities.update_one(
            {"date": today},
            {"$set": {
                "date": today,
                "last_updated": datetime.now(arizona_tz).strftime('%I:%M %p'),
                "games": games,
                "plays": plays,
                "ppg_populated": True
            }},
            upsert=True
        )
        
        logger.info(f"[Scheduled] NBA opportunities refreshed successfully: {len(games)} games, {len(plays)} plays")
        
    except Exception as e:
        logger.error(f"[Scheduled] Error refreshing NBA opportunities: {e}")

# ============================================
# PROCESS NOTEBOOK - SCHEDULED JOBS
# ============================================

async def scrape_tomorrows_opening_lines():
    """
    #1 - 8:00 PM Arizona: Scrape tomorrow's games and opening lines from CBS Sports
    These are the opening lines for each game.
    Creates full game documents in the opportunities collection.
    """
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    tomorrow = (datetime.now(arizona_tz) + timedelta(days=1)).strftime('%Y-%m-%d')
    
    logger.info(f"[8PM Job] Scraping tomorrow's opening lines for {tomorrow}")
    
    try:
        # NFL eliminated - only process NBA, NHL, NCAAB
        leagues = ['NBA', 'NHL', 'NCAAB']
        
        for league in leagues:
            try:
                # Use CBS Sports for all leagues (more reliable for opening lines)
                if league == 'NCAAB':
                    games = await scrape_cbssports_ncaab(tomorrow)
                elif league == 'NBA':
                    games = await scrape_cbssports_nba(tomorrow)
                elif league == 'NHL':
                    games = await scrape_cbssports_nhl(tomorrow)
                else:
                    games = await scrape_scoresandodds(league.upper(), tomorrow)
                
                if games:
                    # Create full game documents for the opportunities collection
                    collection_name = f"{league.lower()}_opportunities"
                    collection = db[collection_name]
                    
                    processed_games = []
                    for game in games:
                        away = game.get('away_team') or game.get('away')
                        home = game.get('home_team') or game.get('home')
                        total = game.get('total')
                        time = game.get('time', 'TBD')
                        spread = game.get('spread')
                        spread_team = game.get('spread_team')
                        moneyline = game.get('moneyline')
                        moneyline_team = game.get('moneyline_team')
                        
                        if away and home:
                            game_doc = {
                                'away_team': away,
                                'home_team': home,
                                'total': total,
                                'opening_line': total,  # Store as opening line
                                'time': time,
                                'date': tomorrow
                            }
                            
                            # Add spread for NBA/NCAAB
                            if league in ['NBA', 'NCAAB']:
                                game_doc['spread'] = spread
                                game_doc['spread_team'] = spread_team
                                game_doc['opening_spread'] = spread
                                game_doc['opening_spread_team'] = spread_team
                            
                            # Add moneyline for NHL
                            if league == 'NHL':
                                game_doc['moneyline'] = moneyline
                                game_doc['moneyline_team'] = moneyline_team
                                game_doc['opening_moneyline'] = moneyline
                                game_doc['opening_moneyline_team'] = moneyline_team
                            
                            processed_games.append(game_doc)
                            
                            # Also store in opening_lines collection for tracking
                            if total:
                                await store_opening_line(league, tomorrow, away, home, total)
                    
                    if processed_games:
                        # All leagues now use CBS Sports
                        data_source = "cbssports.com opening lines"
                        
                        # Create or update the opportunities document
                        await collection.update_one(
                            {"date": tomorrow},
                            {
                                "$set": {
                                    "date": tomorrow,
                                    "games": processed_games,
                                    "last_updated": datetime.now(arizona_tz).strftime('%I:%M %p'),
                                    "data_source": data_source
                                }
                            },
                            upsert=True
                        )
                        logger.info(f"[8PM Job] Created {len(processed_games)} {league} games for {tomorrow}")
                    else:
                        logger.warning(f"[8PM Job] No valid {league} games to store for {tomorrow}")
                else:
                    logger.warning(f"[8PM Job] No {league} games found for {tomorrow}")
                    
            except Exception as e:
                logger.error(f"[8PM Job] Error scraping {league}: {e}")
                import traceback
                traceback.print_exc()
                
    except Exception as e:
        logger.error(f"[8PM Job] Error in scrape_tomorrows_opening_lines: {e}")


async def populate_ppg_and_dots_for_tomorrow():
    """
    #2 - After Scraping: Fill PPG Data & 4-Dot System for Tomorrow's Games
    
    Formula: GPG Avg = (Team1 Season PPG + Team2 Season PPG + Team1 L3 PPG + Team2 L3 PPG) / 2
    Edge = GPG Avg - Line
    
    4 Dots System:
    - 🟢 Green: Rank 1-8 (Top tier)
    - 🔵 Blue: Rank 9-16 (Upper middle)
    - 🟡 Yellow: Rank 17-24 (Lower middle)
    - 🔴 Red: Rank 25-32 (Bottom tier)
    """
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    tomorrow = (datetime.now(arizona_tz) + timedelta(days=1)).strftime('%Y-%m-%d')
    
    logger.info(f"[8PM Job #2] Populating PPG and dots for tomorrow ({tomorrow})")
    
    def get_dot_color(rank: int) -> str:
        """Get dot color based on rank (1-32)"""
        if rank <= 8:
            return "🟢"  # Green: Top tier
        elif rank <= 16:
            return "🔵"  # Blue: Upper middle
        elif rank <= 24:
            return "🟡"  # Yellow: Lower middle
        else:
            return "🔴"  # Red: Bottom tier
    
    # ==================== NBA PPG DATA ====================
    nba_ppg_season = {
        'Boston': 1, 'Okla City': 2, 'Indiana': 3, 'Milwaukee': 4, 'Golden State': 5,
        'Denver': 6, 'Minnesota': 7, 'Sacramento': 8, 'New York': 9, 'Dallas': 10,
        'Memphis': 11, 'Cleveland': 12, 'Houston': 13, 'LA Lakers': 14, 'Detroit': 15,
        'Phoenix': 16, 'San Antonio': 17, 'Atlanta': 18, 'New Orleans': 19, 'Brooklyn': 20,
        'LA Clippers': 21, 'Toronto': 22, 'Philadelphia': 23, 'Portland': 24, 'Washington': 25,
        'Chicago': 26, 'Utah': 27, 'Orlando': 28, 'Charlotte': 29, 'Miami': 30
    }
    nba_ppg_last3 = {
        'Indiana': 1, 'Boston': 2, 'Memphis': 3, 'Golden State': 4, 'Detroit': 5,
        'Washington': 6, 'New York': 7, 'Milwaukee': 8, 'Houston': 9, 'Phoenix': 10,
        'Denver': 11, 'LA Lakers': 12, 'Sacramento': 13, 'Okla City': 14, 'Minnesota': 15,
        'Dallas': 16, 'San Antonio': 17, 'Cleveland': 18, 'Brooklyn': 19, 'Toronto': 20,
        'Atlanta': 21, 'New Orleans': 22, 'LA Clippers': 23, 'Portland': 24, 'Philadelphia': 25,
        'Chicago': 26, 'Utah': 27, 'Charlotte': 28, 'Orlando': 29, 'Miami': 30
    }
    nba_ppg_season_values = {
        'Denver': 125.8, 'Okla City': 121.5, 'New York': 120.6, 'Utah': 120.5,
        'Houston': 120.3, 'Miami': 120.2, 'San Antonio': 119.9, 'Cleveland': 119.5,
        'Detroit': 118.8, 'Chicago': 118.8, 'Atlanta': 118.8, 'Minnesota': 118.8,
        'LA Lakers': 117.2, 'Orlando': 117.1, 'Portland': 116.6, 'Boston': 116.5,
        'Philadelphia': 115.7, 'Charlotte': 115.6, 'Memphis': 115.3, 'New Orleans': 115.2,
        'Phoenix': 115.1, 'Golden State': 114.9, 'Toronto': 114.4, 'Washington': 113.6,
        'Dallas': 113.5, 'Milwaukee': 112.7, 'Sacramento': 111.7, 'LA Clippers': 111.5,
        'Brooklyn': 109.3, 'Indiana': 109.2
    }
    nba_ppg_last3_values = {
        'Denver': 132.7, 'Okla City': 113.7, 'New York': 119.3, 'Utah': 128.7,
        'Houston': 114.7, 'Miami': 119.7, 'San Antonio': 120.3, 'Cleveland': 121.7,
        'Detroit': 121.3, 'Chicago': 112.7, 'Atlanta': 119.7, 'Minnesota': 120.0,
        'LA Lakers': 109.7, 'Orlando': 114.0, 'Portland': 107.7, 'Boston': 117.0,
        'Philadelphia': 104.0, 'Charlotte': 126.0, 'Memphis': 124.7, 'New Orleans': 113.3,
        'Phoenix': 123.3, 'Golden State': 124.3, 'Toronto': 123.3, 'Washington': 121.0,
        'Dallas': 118.0, 'Milwaukee': 109.0, 'Sacramento': 113.7, 'LA Clippers': 119.7,
        'Brooklyn': 101.7, 'Indiana': 99.3
    }
    
    # ==================== NHL GPG DATA ====================
    # Updated Dec 29, 2025 from ESPN (Season GPG) and StatMuse (Last 3 Games GPG)
    # Season GPG Rankings (1-32)
    nhl_gpg_season = {
        'Colorado': 1, 'Dallas': 2, 'Edmonton': 3, 'Anaheim': 4, 'Carolina': 5,
        'Montreal': 6, 'Tampa Bay': 7, 'Ottawa': 8, 'Toronto': 9, 'Washington': 10,
        'Vegas': 11, 'Florida': 12, 'Pittsburgh': 13, 'Boston': 14, 'Detroit': 15,
        'Buffalo': 16, 'Minnesota': 17, 'San Jose': 18, 'Utah': 19, 'Columbus': 20,
        'Winnipeg': 21, 'Philadelphia': 22, 'Nashville': 23, 'Vancouver': 24, 'NY Islanders': 25,
        'Chicago': 26, 'New Jersey': 27, 'Calgary': 28, 'Los Angeles': 29, 'Seattle': 30,
        'NY Rangers': 31, 'St. Louis': 32
    }
    # Last 3 Games GPG Rankings (1-32)
    nhl_gpg_last3 = {
        'Toronto': 1, 'Vegas': 2, 'Montreal': 3, 'Ottawa': 4, 'Pittsburgh': 5,
        'Tampa Bay': 6, 'Colorado': 7, 'Dallas': 8, 'Edmonton': 9, 'Carolina': 10,
        'Buffalo': 11, 'San Jose': 12, 'Columbus': 13, 'Calgary': 14, 'Seattle': 15,
        'St. Louis': 16, 'Washington': 17, 'Florida': 18, 'Detroit': 19, 'Philadelphia': 20,
        'Vancouver': 21, 'Los Angeles': 22, 'Winnipeg': 23, 'NY Rangers': 24, 'Minnesota': 25,
        'Nashville': 26, 'Chicago': 27, 'Anaheim': 28, 'NY Islanders': 29, 'Boston': 30,
        'Utah': 31, 'New Jersey': 32
    }
    # Season GPG Values (from ESPN)
    nhl_gpg_season_values = {
        'Colorado': 3.97, 'Dallas': 3.49, 'Edmonton': 3.38, 'Anaheim': 3.32, 'Carolina': 3.30,
        'Montreal': 3.29, 'Tampa Bay': 3.29, 'Ottawa': 3.27, 'Toronto': 3.26, 'Washington': 3.18,
        'Vegas': 3.17, 'Florida': 3.16, 'Pittsburgh': 3.14, 'Boston': 3.10, 'Detroit': 3.08,
        'Buffalo': 3.05, 'Minnesota': 3.03, 'San Jose': 3.00, 'Utah': 2.97, 'Columbus': 2.92,
        'Winnipeg': 2.92, 'Philadelphia': 2.89, 'Nashville': 2.78, 'Vancouver': 2.78, 'NY Islanders': 2.77,
        'Chicago': 2.76, 'New Jersey': 2.71, 'Calgary': 2.61, 'Los Angeles': 2.59, 'Seattle': 2.58,
        'NY Rangers': 2.55, 'St. Louis': 2.51
    }
    # Last 3 Games GPG Values (from StatMuse)
    nhl_gpg_last3_values = {
        'Toronto': 5.00, 'Vegas': 5.00, 'Montreal': 4.33, 'Ottawa': 4.33, 'Pittsburgh': 4.33,
        'Tampa Bay': 4.00, 'Colorado': 3.67, 'Dallas': 3.67, 'Edmonton': 3.67, 'Carolina': 3.67,
        'Buffalo': 3.33, 'San Jose': 3.33, 'Columbus': 3.33, 'Calgary': 3.33, 'Seattle': 3.33,
        'St. Louis': 3.33, 'Washington': 3.00, 'Florida': 3.00, 'Detroit': 3.00, 'Philadelphia': 3.00,
        'Vancouver': 3.00, 'Los Angeles': 3.00, 'Winnipeg': 2.67, 'NY Rangers': 2.67, 'Minnesota': 2.33,
        'Nashville': 2.33, 'Chicago': 2.33, 'Anaheim': 2.00, 'NY Islanders': 2.00, 'Boston': 1.67,
        'Utah': 1.67, 'New Jersey': 1.67
    }
    
    # ==================== NFL PPG DATA ====================
    # From teamrankings.com/nfl/stat/points-per-game (Dec 27, 2025)
    # Season Rankings (2025 PPG)
    nfl_ppg_season = {
        'LA Rams': 1, 'Seattle': 2, 'Buffalo': 3, 'Detroit': 4, 'Dallas': 5,
        'Indianapolis': 6, 'Jacksonville': 7, 'New England': 8, 'San Francisco': 9, 'Chicago': 10,
        'Pittsburgh': 11, 'Green Bay': 12, 'Cincinnati': 13, 'Baltimore': 14, 'Denver': 15,
        'LA Chargers': 16, 'Philadelphia': 17, 'Tampa Bay': 18, 'Houston': 19, 'Kansas City': 20,
        'Arizona': 21, 'Miami': 22, 'NY Giants': 23, 'Washington': 24, 'Minnesota': 25,
        'Atlanta': 26, 'Carolina': 27, 'NY Jets': 28, 'New Orleans': 29, 'Tennessee': 30,
        'Cleveland': 31, 'Las Vegas': 32
    }
    # Last 3 Rankings (sorted by Last 3 PPG descending)
    nfl_ppg_last3 = {
        'LA Rams': 1, 'Jacksonville': 2, 'San Francisco': 3, 'Buffalo': 4, 'Seattle': 5,
        'New England': 6, 'Pittsburgh': 7, 'Houston': 8, 'Tennessee': 9, 'Cincinnati': 10,
        'Philadelphia': 11, 'Denver': 12, 'Minnesota': 13, 'Chicago': 14, 'Dallas': 15,
        'LA Chargers': 16, 'New Orleans': 17, 'Carolina': 18, 'Baltimore': 19, 'Green Bay': 20,
        'Miami': 21, 'Washington': 22, 'Detroit': 23, 'Tampa Bay': 24, 'Atlanta': 25,
        'Indianapolis': 26, 'Arizona': 27, 'Cleveland': 28, 'NY Giants': 29, 'Tennessee': 30,
        'Las Vegas': 31, 'NY Jets': 32, 'Kansas City': 33
    }
    # Season PPG Values (2025)
    nfl_ppg_season_values = {
        'LA Rams': 30.5, 'Seattle': 29.5, 'Buffalo': 28.9, 'Detroit': 28.9, 'Dallas': 28.4,
        'Indianapolis': 27.9, 'Jacksonville': 27.3, 'New England': 27.3, 'San Francisco': 26.1, 'Chicago': 25.8,
        'Pittsburgh': 24.3, 'Green Bay': 24.3, 'Cincinnati': 23.9, 'Baltimore': 23.9, 'Denver': 23.9,
        'LA Chargers': 23.3, 'Philadelphia': 23.3, 'Tampa Bay': 23.1, 'Houston': 23.1, 'Kansas City': 21.9,
        'Arizona': 21.4, 'Miami': 21.1, 'NY Giants': 20.9, 'Washington': 20.8, 'Minnesota': 20.5,
        'Atlanta': 20.5, 'Carolina': 19.1, 'NY Jets': 18.8, 'New Orleans': 17.0, 'Tennessee': 16.7,
        'Cleveland': 16.4, 'Las Vegas': 14.5
    }
    # Last 3 PPG Values
    nfl_ppg_last3_values = {
        'LA Rams': 41.0, 'Jacksonville': 39.3, 'San Francisco': 37.0, 'Buffalo': 32.3, 'Seattle': 31.0,
        'New England': 30.7, 'Pittsburgh': 28.0, 'Houston': 27.7, 'Tennessee': 27.0, 'Cincinnati': 26.3,
        'Philadelphia': 26.3, 'Denver': 24.7, 'Minnesota': 24.3, 'Chicago': 24.7, 'Dallas': 24.3,
        'LA Chargers': 24.0, 'New Orleans': 24.3, 'Carolina': 23.7, 'Baltimore': 23.3, 'Green Bay': 23.3,
        'Miami': 23.3, 'Washington': 23.3, 'Detroit': 22.7, 'Tampa Bay': 22.7, 'Atlanta': 21.3,
        'Indianapolis': 20.7, 'Arizona': 18.7, 'Cleveland': 17.3, 'NY Giants': 16.3, 'Tennessee': 27.0,
        'Las Vegas': 12.7, 'NY Jets': 12.0, 'Kansas City': 11.7
    }
    
    # Edge thresholds by league (NFL eliminated)
    edge_thresholds = {'NBA': 8, 'NHL': 0.6}
    
    for league in ['NBA', 'NHL']:
        try:
            collection_name = f"{league.lower()}_opportunities"
            collection = db[collection_name]
            
            # Fetch existing tomorrow data
            cached = await collection.find_one({"date": tomorrow}, {"_id": 0})
            
            if not cached or not cached.get('games'):
                logger.warning(f"[8PM Job #2] No {league} games found for {tomorrow}, skipping PPG population")
                continue
            
            games = cached['games']
            logger.info(f"[8PM Job #2] Processing {len(games)} {league} games for PPG and dots")
            
            # Select appropriate PPG data based on league
            if league == 'NBA':
                ppg_season = nba_ppg_season
                ppg_last3 = nba_ppg_last3
                ppg_season_values = nba_ppg_season_values
                ppg_last3_values = nba_ppg_last3_values
            elif league == 'NHL':
                ppg_season = nhl_gpg_season
                ppg_last3 = nhl_gpg_last3
                ppg_season_values = nhl_gpg_season_values
                ppg_last3_values = nhl_gpg_last3_values
            else:  # NFL
                ppg_season = nfl_ppg_season
                ppg_last3 = nfl_ppg_last3
                ppg_season_values = nfl_ppg_season_values
                ppg_last3_values = nfl_ppg_last3_values
            
            edge_threshold = edge_thresholds[league]
            
            # Process each game
            for game in games:
                away = game.get('away_team') or game.get('away', '')
                home = game.get('home_team') or game.get('home', '')
                total = game.get('total')
                
                # Get rankings
                away_season_rank = ppg_season.get(away, 16)
                away_last3_rank = ppg_last3.get(away, 16)
                home_season_rank = ppg_season.get(home, 16)
                home_last3_rank = ppg_last3.get(home, 16)
                
                # Get PPG values
                away_season_ppg = ppg_season_values.get(away, 100.0 if league == 'NBA' else (3.0 if league == 'NHL' else 20.0))
                away_last3_ppg = ppg_last3_values.get(away, 100.0 if league == 'NBA' else (3.0 if league == 'NHL' else 20.0))
                home_season_ppg = ppg_season_values.get(home, 100.0 if league == 'NBA' else (3.0 if league == 'NHL' else 20.0))
                home_last3_ppg = ppg_last3_values.get(home, 100.0 if league == 'NBA' else (3.0 if league == 'NHL' else 20.0))
                
                # Calculate combined PPG using the formula:
                # (Team1 Season PPG + Team2 Season PPG + Team1 L3 PPG + Team2 L3 PPG) / 2
                combined_ppg = round((away_season_ppg + home_season_ppg + away_last3_ppg + home_last3_ppg) / 2, 1)
                
                # Calculate edge
                edge = round(combined_ppg - total, 1) if total else None
                
                # Determine recommendation
                recommendation = None
                color = "neutral"
                if edge is not None:
                    if edge >= edge_threshold:
                        recommendation = "OVER"
                        color = "green"
                    elif edge <= -edge_threshold:
                        recommendation = "UNDER"
                        color = "red"
                
                # Generate 4 dots (away_season, away_l3, home_season, home_l3)
                dots = f"{get_dot_color(away_season_rank)}{get_dot_color(away_last3_rank)}{get_dot_color(home_season_rank)}{get_dot_color(home_last3_rank)}"
                
                # Update game data
                game['away_ppg_rank'] = away_season_rank
                game['away_last3_rank'] = away_last3_rank
                game['away_season_ppg'] = away_season_ppg
                game['away_last3_ppg'] = away_last3_ppg
                game['home_ppg_rank'] = home_season_rank
                game['home_last3_rank'] = home_last3_rank
                game['home_season_ppg'] = home_season_ppg
                game['home_last3_ppg'] = home_last3_ppg
                game['combined_ppg'] = combined_ppg
                game['edge'] = edge
                game['recommendation'] = recommendation
                game['color'] = color
                game['dots'] = dots
                game['away_dots'] = f"{get_dot_color(away_season_rank)}{get_dot_color(away_last3_rank)}"
                game['home_dots'] = f"{get_dot_color(home_season_rank)}{get_dot_color(home_last3_rank)}"
                
                logger.debug(f"[8PM Job #2] {away} @ {home}: PPG={combined_ppg}, Edge={edge}, Dots={dots}")
            
            # Save updated data
            await collection.update_one(
                {"date": tomorrow},
                {"$set": {"games": games, "ppg_populated": True, "ppg_updated_at": datetime.now(timezone.utc).isoformat()}},
                upsert=True
            )
            
            logger.info(f"[8PM Job #2] ✅ {league}: Updated {len(games)} games with PPG and dots")
            
        except Exception as e:
            logger.error(f"[8PM Job #2] Error processing {league}: {e}")
            import traceback
            traceback.print_exc()
    
    # ==================== NCAAB PPG PROCESSING ====================
    # NCAAB: Scrape Last 3 PPG directly from CBS Sports team pages
    try:
        logger.info(f"[8PM Job #2] Processing NCAAB PPG using CBS Sports Last 3 scores...")
        
        collection = db.ncaab_opportunities
        cached = await collection.find_one({"date": tomorrow}, {"_id": 0})
        
        if cached and cached.get('games'):
            games = cached['games']
            logger.info(f"[8PM Job #2] Found {len(games)} NCAAB games for {tomorrow}")
            
            # Method 1: Get season PPG from TeamRankings (for rankings/dots)
            ppg_data = await scrape_ncaab_ppg_rankings(tomorrow)
            logger.info(f"[8PM Job #2] Scraped season PPG for {len(ppg_data.get('season_values', {}))} teams from TeamRankings")
            
            # Method 2: Scrape games with team URLs to get Last 3 from CBS Sports
            _, team_urls = await scrape_cbssports_ncaab_with_team_urls(tomorrow)
            logger.info(f"[8PM Job #2] Got {len(team_urls)} team URLs from CBS Sports")
            
            # Scrape Last 3 PPG for each team from CBS Sports
            team_last3_stats = await scrape_ncaab_team_last3_ppg(team_urls, max_concurrent=5)
            logger.info(f"[8PM Job #2] Scraped Last 3 PPG for {len(team_last3_stats)} teams from CBS Sports")
            
            # NCAAB has ~365 teams, dot thresholds are percentile-based
            def get_ncaab_dot_color(rank):
                if rank is None:
                    return '⚪'  # Unknown
                if rank <= 92:  # Top 25%
                    return '🟢'
                elif rank <= 184:  # 25-50%
                    return '🔵'
                elif rank <= 276:  # 50-75%
                    return '🟡'
                else:  # Bottom 25%
                    return '🔴'
            
            # NCAAB edge threshold is 9
            ncaab_edge_threshold = 10
            
            # Helper function to find team in PPG data with fuzzy matching
            def find_team_data(team_name, data_dict):
                if not team_name:
                    return None
                    
                def normalize(name):
                    name = name.strip()
                    name = name.replace(' St.', '').replace(' St', '')
                    name = name.replace(' State', '').replace('-Fort Kent', '')
                    name = name.replace('N. ', 'North ').replace('S. ', 'South ')
                    name = name.replace('E. ', 'East ').replace('W. ', 'West ')
                    name = name.replace('C. ', 'Central ')
                    return name.lower().strip()
                
                search_name = normalize(team_name)
                
                # Direct match
                if team_name in data_dict:
                    return data_dict[team_name]
                
                # Try lowercase
                for key, val in data_dict.items():
                    if key.lower() == team_name.lower():
                        return val
                
                # Try normalized match
                for key, val in data_dict.items():
                    if normalize(key) == search_name:
                        return val
                
                # Try partial match
                for key, val in data_dict.items():
                    key_norm = normalize(key)
                    if search_name in key_norm or key_norm in search_name:
                        return val
                
                return None
            
            for game in games:
                away = game.get('away_team') or game.get('away', '')
                home = game.get('home_team') or game.get('home', '')
                total = game.get('total')
                
                # Get season PPG and ranks from TeamRankings
                away_season_ppg = find_team_data(away, ppg_data.get('season_values', {}))
                away_season_rank = find_team_data(away, ppg_data.get('season_ranks', {}))
                home_season_ppg = find_team_data(home, ppg_data.get('season_values', {}))
                home_season_rank = find_team_data(home, ppg_data.get('season_ranks', {}))
                
                # Get Last 3 PPG from CBS Sports scrape
                away_last3_stats = find_team_data(away, team_last3_stats)
                home_last3_stats = find_team_data(home, team_last3_stats)
                
                away_last3_ppg = away_last3_stats['last3_avg'] if away_last3_stats else None
                home_last3_ppg = home_last3_stats['last3_avg'] if home_last3_stats else None
                
                # Calculate Last 3 ranks (based on season rank for now - could improve later)
                away_last3_rank = away_season_rank  # Use season rank as proxy
                home_last3_rank = home_season_rank
                
                # Calculate combined PPG
                # Formula: (Team1 Season PPG + Team2 Season PPG + Team1 L3 PPG + Team2 L3 PPG) / 2
                combined_ppg = None
                if away_season_ppg and home_season_ppg and away_last3_ppg and home_last3_ppg:
                    combined_ppg = round((away_season_ppg + home_season_ppg + away_last3_ppg + home_last3_ppg) / 2, 1)
                elif away_season_ppg and home_season_ppg:
                    # Fallback: just use season PPG x 2
                    combined_ppg = round(away_season_ppg + home_season_ppg, 1)
                
                # Calculate edge
                edge = None
                if combined_ppg and total:
                    try:
                        edge = round(combined_ppg - float(total), 1)
                    except:
                        pass
                
                # Determine recommendation
                recommendation = None
                color = "neutral"
                if edge is not None:
                    if edge >= ncaab_edge_threshold:
                        recommendation = "OVER"
                        color = "green"
                    elif edge <= -ncaab_edge_threshold:
                        recommendation = "UNDER"
                        color = "red"
                
                # Generate dots
                away_dots = f"{get_ncaab_dot_color(away_season_rank)}{get_ncaab_dot_color(away_last3_rank)}"
                home_dots = f"{get_ncaab_dot_color(home_season_rank)}{get_ncaab_dot_color(home_last3_rank)}"
                dots = f"{away_dots}{home_dots}"
                
                # Update game data
                game['away_ppg_rank'] = away_season_rank
                game['away_last3_rank'] = away_last3_rank
                game['away_season_ppg'] = away_season_ppg
                game['away_last3_ppg'] = away_last3_ppg
                game['away_last3_scores'] = away_last3_stats['last3_scores'] if away_last3_stats else None
                game['home_ppg_rank'] = home_season_rank
                game['home_last3_rank'] = home_last3_rank
                game['home_season_ppg'] = home_season_ppg
                game['home_last3_ppg'] = home_last3_ppg
                game['home_last3_scores'] = home_last3_stats['last3_scores'] if home_last3_stats else None
                game['combined_ppg'] = combined_ppg
                game['edge'] = edge
                game['recommendation'] = recommendation
                game['color'] = color
                game['dots'] = dots
                game['away_dots'] = away_dots
                game['home_dots'] = home_dots
            
            # Save updated data
            await collection.update_one(
                {"date": tomorrow},
                {"$set": {"games": games, "ppg_populated": True, "ppg_source": "cbs_sports_last3", "ppg_updated_at": datetime.now(timezone.utc).isoformat()}},
                upsert=True
            )
            
            # Count games with valid PPG data
            games_with_ppg = sum(1 for g in games if g.get('combined_ppg') is not None)
            games_with_last3 = sum(1 for g in games if g.get('away_last3_ppg') and g.get('home_last3_ppg'))
            games_with_rec = sum(1 for g in games if g.get('recommendation'))
            logger.info(f"[8PM Job #2] ✅ NCAAB: Updated {len(games)} games ({games_with_ppg} with PPG, {games_with_last3} with Last 3, {games_with_rec} with recommendations)")
        else:
            logger.warning(f"[8PM Job #2] No NCAAB games found for {tomorrow}")
            
    except Exception as e:
        logger.error(f"[8PM Job #2] Error processing NCAAB: {e}")
        import traceback
        traceback.print_exc()
            
    except Exception as e:
        logger.error(f"[8PM Job #2] Error processing NCAAB: {e}")
        import traceback
        traceback.print_exc()
    
    logger.info(f"[8PM Job #2] Completed PPG and dots population for tomorrow ({tomorrow})")


async def execute_8pm_process():
    """
    Execute the full 8pm process: #1 (scrape opening lines) + #2 (populate PPG and dots)
    This is called by the scheduled job at 8pm Arizona time or manually via API.
    """
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    now_arizona = datetime.now(arizona_tz)
    tomorrow = (now_arizona + timedelta(days=1)).strftime('%Y-%m-%d')
    
    logger.info(f"=" * 60)
    logger.info(f"[8PM PROCESS] Starting at {now_arizona.strftime('%I:%M %p')} Arizona time")
    logger.info(f"[8PM PROCESS] Target date: {tomorrow}")
    logger.info(f"=" * 60)
    
    # Step #1: Scrape tomorrow's opening lines from ScoresAndOdds
    logger.info(f"[8PM PROCESS] Step #1: Scraping tomorrow's opening lines...")
    await scrape_tomorrows_opening_lines()
    
    # Step #2: Populate PPG and dots
    logger.info(f"[8PM PROCESS] Step #2: Populating PPG and 4-dot analysis...")
    await populate_ppg_and_dots_for_tomorrow()
    
    logger.info(f"[8PM PROCESS] ✅ Completed successfully")
    return {"status": "success", "date": tomorrow, "timestamp": now_arizona.isoformat()}


async def morning_data_refresh():
    """
    5:00 AM Arizona: Morning data refresh
    #3 - Switch to Plays888 live lines (handled by flag)
    #4 - Get yesterday's scores from ScoresAndOdds + Mark edge HITs/MISSes
    #5 - Get bet results from Plays888 History
    #6 - Update betting and edge records
    """
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    yesterday = (datetime.now(arizona_tz) - timedelta(days=1)).strftime('%Y-%m-%d')
    today = datetime.now(arizona_tz).strftime('%Y-%m-%d')
    
    logger.info(f"[5AM Job] Starting morning data refresh for {today}")
    logger.info(f"[5AM Job] Getting yesterday's ({yesterday}) results")
    
    try:
        # #4 - Get yesterday's scores from ScoresAndOdds and mark edge HITs/MISSes
        # NFL eliminated - only process NBA, NHL, NCAAB
        leagues = ['NBA', 'NHL', 'NCAAB']
        edge_results = {}  # Track edge performance by league
        
        for league in leagues:
            try:
                # Scrape yesterday's results from ScoresAndOdds (or CBS for NCAAB)
                logger.info(f"[#4 Process] Scraping {league} results for {yesterday}")
                if league == 'NCAAB':
                    results = await scrape_cbssports_ncaab(yesterday)
                else:
                    results = await scrape_scoresandodds(league.upper(), yesterday)
                
                if results:
                    logger.info(f"[#4 Process] Got {len(results)} {league} games")
                    
                    # Update the opportunities collection with final scores
                    collection_name = f"{league.lower()}_opportunities"
                    collection = db[collection_name]
                    
                    dates_to_check = [yesterday]
                    
                    edge_hits = 0
                    edge_misses = 0
                    games_updated = 0
                    
                    for check_date in dates_to_check:
                        cached = await collection.find_one({"date": check_date}, {"_id": 0})
                        if cached and cached.get('games'):
                            games = cached['games']
                            
                            for game in games:
                                # Match with scraped results and update final_score
                                for result in results:
                                    r_away = (result.get('away_team') or result.get('away', '')).lower()
                                    r_home = (result.get('home_team') or result.get('home', '')).lower()
                                    g_away = (game.get('away_team') or game.get('away', '')).lower()
                                    g_home = (game.get('home_team') or game.get('home', '')).lower()
                                    
                                    # Use multiple matching strategies
                                    # 1. Exact or substring match
                                    match1 = (r_away in g_away or g_away in r_away) and (r_home in g_home or g_home in r_home)
                                    
                                    # 2. Last word match (team nickname)
                                    r_away_last = r_away.split()[-1] if r_away else ''
                                    r_home_last = r_home.split()[-1] if r_home else ''
                                    g_away_last = g_away.split()[-1] if g_away else ''
                                    g_home_last = g_home.split()[-1] if g_home else ''
                                    match2 = (r_away_last == g_away_last) and (r_home_last == g_home_last)
                                    
                                    # 3. First word match (city name)
                                    r_away_first = r_away.split()[0] if r_away else ''
                                    r_home_first = r_home.split()[0] if r_home else ''
                                    g_away_first = g_away.split()[0] if g_away else ''
                                    g_home_first = g_home.split()[0] if g_home else ''
                                    match3 = (r_away_first == g_away_first) and (r_home_first == g_home_first)
                                    
                                    if match1 or match2 or match3:
                                        final_score = result.get('final_score')
                                        
                                        if final_score:
                                            game['final_score'] = final_score
                                            game['away_score'] = result.get('away_score')
                                            game['home_score'] = result.get('home_score')
                                            games_updated += 1
                                            
                                            # Determine if the total went OVER or UNDER
                                            line = game.get('total')
                                            if line:
                                                actual_result = 'OVER' if final_score > line else 'UNDER'
                                                game['actual_result'] = actual_result
                                                
                                                # Check if our edge recommendation was correct
                                                recommendation = game.get('recommendation')
                                                if recommendation:
                                                    edge_hit = (recommendation == actual_result)
                                                    game['result_hit'] = edge_hit
                                                    
                                                    if edge_hit:
                                                        edge_hits += 1
                                                    else:
                                                        edge_misses += 1
                                                    
                                                    logger.info(f"[#4] {g_away} @ {g_home}: Final={final_score}, Line={line}, Result={actual_result}, Rec={recommendation}, HIT={edge_hit}")
                                        break
                            
                            # Save updated games back to database
                            await collection.update_one(
                                {"date": check_date},
                                {"$set": {
                                    "games": games,
                                    "scores_updated": True,
                                    "scores_updated_at": datetime.now(timezone.utc).isoformat()
                                }}
                            )
                    
                    edge_results[league] = {"hits": edge_hits, "misses": edge_misses}
                    logger.info(f"[#4 Process] {league}: Updated {games_updated} games, Edge: {edge_hits}-{edge_misses}")
                else:
                    logger.warning(f"[#4 Process] No {league} results found for {yesterday}")
                        
            except Exception as e:
                logger.error(f"[#4 Process] Error getting {league} results: {e}")
                import traceback
                traceback.print_exc()
        
        # #5 - Get bet results from Plays888 History
        try:
            conn = await db.connections.find_one({}, {"_id": 0}, sort=[("created_at", -1)])
            if conn and conn.get("is_connected"):
                username = conn["username"]
                password = decrypt_password(conn["password_encrypted"])
                
                scraper = Plays888Service()
                await scraper.login(username, password)
                
                # Get bet history
                bets = await scraper.scrape_open_bets()  # This gets history too
                await scraper.close()
                
                if bets:
                    # Update settled bets
                    settled = [b for b in bets if b.get('result') in ['won', 'lost', 'push']]
                    logger.info(f"[5AM Job] Found {len(settled)} settled bets")
                    
                    # #6 - Update records based on settled bets
                    for league in leagues:
                        league_bets = [b for b in settled if league.upper() in (b.get('sport') or '').upper()]
                        if league_bets:
                            wins = sum(1 for b in league_bets if b.get('result') == 'won')
                            losses = sum(1 for b in league_bets if b.get('result') == 'lost')
                            
                            if wins > 0 or losses > 0:
                                logger.info(f"[5AM Job] Found {league} bet results: {wins} wins, {losses} losses")
                                
        except Exception as e:
            logger.error(f"[5AM Job] Error getting bet results: {e}")
        
        # #6 - Update Records from 12/22/25 to yesterday
        try:
            logger.info(f"[#6 Process] Updating betting and edge records from 12/22/25")
            records_result = await update_records_from_start_date("2025-12-22")
            logger.info(f"[#6 Process] Records updated: {records_result}")
        except Exception as e:
            logger.error(f"[5AM Job] Error updating records: {e}")
        
        # Summary
        logger.info(f"[5AM Job] Morning data refresh completed")
        logger.info(f"[5AM Job] Edge Results Summary: {edge_results}")
        
        return {"status": "success", "edge_results": edge_results}
        
    except Exception as e:
        logger.error(f"[5AM Job] Error in morning_data_refresh: {e}")
        return {"status": "error", "error": str(e)}


# ============== PROCESS #6: UPDATE RECORDS FROM 12/22/25 ==============

def parse_final_score(final_score):
    """
    Parse final_score which could be:
    - A number (e.g., 225)
    - A string number (e.g., "225")
    - Individual scores (e.g., "110-115" or "70-55")
    Returns the total score as a float, or None if invalid.
    """
    if final_score is None:
        return None
    
    # If it's already a number, return it
    if isinstance(final_score, (int, float)):
        return float(final_score)
    
    # If it's a string, try to parse it
    if isinstance(final_score, str):
        # Check if it contains a hyphen (individual scores)
        if '-' in final_score:
            try:
                parts = final_score.split('-')
                return float(parts[0]) + float(parts[1])
            except:
                return None
        else:
            try:
                return float(final_score)
            except:
                return None
    
    return None


async def calculate_records_from_start_date(start_date: str = "2025-12-22"):
    """
    Calculate betting records and edge records from start_date to yesterday.
    
    Betting Record: W-L record of actual user bets (from plays888.co History)
    Edge Record: W-L record of system recommendations (edge-based predictions)
    
    IMPORTANT: Edge is calculated as PPG_Avg - Line (using bet_line or live line)
    - Positive edge (>= threshold) = OVER recommendation
    - Negative edge (<= -threshold) = UNDER recommendation
    
    Args:
        start_date: Start date in 'YYYY-MM-DD' format (default: 12/22/25)
    
    Returns:
        Dictionary with calculated records for each league
    """
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    
    # Get today's date - we'll include games that have final scores
    today = datetime.now(arizona_tz).strftime('%Y-%m-%d')
    
    logger.info(f"[#6] Calculating records from {start_date} to {today}")
    
    results = {
        "NBA": {"betting": {"wins": 0, "losses": 0}, "edge": {"hits": 0, "misses": 0, "over_hits": 0, "over_misses": 0, "under_hits": 0, "under_misses": 0}, "public": {"hits": 0, "misses": 0, "games": []}, "dates_processed": []},
        "NHL": {"betting": {"wins": 0, "losses": 0}, "edge": {"hits": 0, "misses": 0, "over_hits": 0, "over_misses": 0, "under_hits": 0, "under_misses": 0}, "public": {"hits": 0, "misses": 0, "games": []}, "dates_processed": []},
        "NCAAB": {"betting": {"wins": 0, "losses": 0}, "edge": {"hits": 0, "misses": 0, "over_hits": 0, "over_misses": 0, "under_hits": 0, "under_misses": 0}, "public": {"hits": 0, "misses": 0, "games": []}, "dates_processed": []},
    }
    
    # Public Record threshold: only consider games where consensus >= 57%
    PUBLIC_CONSENSUS_THRESHOLD = 57
    
    # Generate list of dates from start_date to today (inclusive)
    start = datetime.strptime(start_date, '%Y-%m-%d')
    end = datetime.strptime(today, '%Y-%m-%d')
    
    dates = []
    current = start
    while current <= end:
        dates.append(current.strftime('%Y-%m-%d'))
        current += timedelta(days=1)
    
    logger.info(f"[#6] Processing {len(dates)} dates: {dates}")
    
    # Process NBA records (threshold: 8)
    NBA_THRESHOLD = 8
    for date in dates:
        doc = await db.nba_opportunities.find_one({"date": date})
        if doc and doc.get('games'):
            date_over_hits = 0
            date_over_misses = 0
            date_under_hits = 0
            date_under_misses = 0
            date_bet_wins = 0
            date_bet_losses = 0
            
            # First, check for actual_bet_record (accurate count from History page)
            actual_record = doc.get('actual_bet_record')
            if actual_record:
                date_bet_wins = actual_record.get('wins', 0)
                date_bet_losses = actual_record.get('losses', 0)
            
            for game in doc['games']:
                # Check if game has final score (completed)
                final_score_raw = game.get('final_score')
                if final_score_raw is None:
                    continue
                
                final_score = parse_final_score(final_score_raw)
                if final_score is None:
                    continue
                
                # Calculate TRUE edge using combined_ppg and the correct line
                combined_ppg = game.get('combined_ppg')
                if combined_ppg is None:
                    continue
                
                # Use bet_line if bet exists, otherwise use live line (total)
                has_bet = game.get('has_bet', False)
                bet_line = game.get('bet_line')
                total = game.get('total')
                
                if has_bet and bet_line:
                    line = float(bet_line)
                elif total:
                    line = float(total)
                else:
                    continue
                
                # Calculate true edge: PPG - Line
                true_edge = round(float(combined_ppg) - line, 2)
                
                # Only count games with edge >= threshold or <= -threshold
                if true_edge >= NBA_THRESHOLD:
                    # OVER recommendation
                    if final_score > line:
                        date_over_hits += 1
                    elif final_score < line:
                        date_over_misses += 1
                    # Push (equal) doesn't count
                elif true_edge <= -NBA_THRESHOLD:
                    # UNDER recommendation
                    if final_score < line:
                        date_under_hits += 1
                    elif final_score > line:
                        date_under_misses += 1
                
                # Betting Record: If no actual_bet_record, count from matched games
                if not actual_record and game.get('has_bet'):
                    if game.get('user_bet_hit') is True:
                        date_bet_wins += 1
                    elif game.get('user_bet_hit') is False:
                        date_bet_losses += 1
            
            # Calculate Public Record for this date (consensus >= 57% threshold)
            date_public_hits = 0
            date_public_misses = 0
            for game in doc['games']:
                # Get consensus percentages
                away_pct = game.get('away_consensus_pct') or 0
                home_pct = game.get('home_consensus_pct') or 0
                
                # Skip if no consensus data
                if away_pct == 0 and home_pct == 0:
                    continue
                
                # Determine public pick and percentage
                is_away_public_pick = away_pct >= home_pct
                public_pct = away_pct if is_away_public_pick else home_pct
                
                # Only include if consensus >= 57%
                if public_pct < PUBLIC_CONSENSUS_THRESHOLD:
                    continue
                
                # Get scores - skip if game not completed
                away_score = game.get('away_score')
                home_score = game.get('home_score')
                if away_score is None or home_score is None:
                    continue
                
                # Get spread - PRIORITY: Covers.com first, CBS Sports as fallback
                # away_spread = Covers.com spread for away team
                # spread = CBS Sports live spread (home team's perspective)
                public_spread = None
                spread_source = None
                
                if is_away_public_pick:
                    # Away team is the public pick
                    if game.get('away_spread') is not None:
                        # Use Covers.com spread directly for away team
                        public_spread = float(game.get('away_spread'))
                        spread_source = "Covers.com"
                    elif game.get('spread') is not None:
                        # Fallback: CBS Sports (invert home spread for away)
                        public_spread = -float(game.get('spread'))
                        spread_source = "CBS Sports"
                else:
                    # Home team is the public pick
                    # Covers.com stores away_spread, so home spread = -away_spread
                    if game.get('away_spread') is not None:
                        public_spread = -float(game.get('away_spread'))
                        spread_source = "Covers.com"
                    elif game.get('spread') is not None:
                        # Fallback: CBS Sports spread directly for home
                        public_spread = float(game.get('spread'))
                        spread_source = "CBS Sports"
                
                if public_spread is None:
                    continue
                
                # Calculate if public pick covered the spread
                try:
                    away_score_f = float(away_score)
                    home_score_f = float(home_score)
                    spread_f = float(public_spread)
                    
                    if is_away_public_pick:
                        # Away team needs to cover: away_score + spread > home_score
                        covered = away_score_f + spread_f > home_score_f
                        push = away_score_f + spread_f == home_score_f
                    else:
                        # Home team needs to cover: home_score + spread > away_score
                        covered = home_score_f + spread_f > away_score_f
                        push = home_score_f + spread_f == away_score_f
                    
                    if not push:
                        if covered:
                            date_public_hits += 1
                            results["NBA"]["public"]["games"].append({
                                "date": date,
                                "game": f"{game.get('away_team')} @ {game.get('home_team')}",
                                "public_pick": game.get('away_team') if is_away_public_pick else game.get('home_team'),
                                "consensus_pct": public_pct,
                                "spread": public_spread,
                                "spread_source": spread_source,
                                "result": "HIT"
                            })
                        else:
                            date_public_misses += 1
                            results["NBA"]["public"]["games"].append({
                                "date": date,
                                "game": f"{game.get('away_team')} @ {game.get('home_team')}",
                                "public_pick": game.get('away_team') if is_away_public_pick else game.get('home_team'),
                                "consensus_pct": public_pct,
                                "spread": public_spread,
                                "spread_source": spread_source,
                                "result": "MISS"
                            })
                except (ValueError, TypeError):
                    continue
            
            results["NBA"]["public"]["hits"] += date_public_hits
            results["NBA"]["public"]["misses"] += date_public_misses
            
            results["NBA"]["edge"]["over_hits"] += date_over_hits
            results["NBA"]["edge"]["over_misses"] += date_over_misses
            results["NBA"]["edge"]["under_hits"] += date_under_hits
            results["NBA"]["edge"]["under_misses"] += date_under_misses
            results["NBA"]["edge"]["hits"] += date_over_hits + date_under_hits
            results["NBA"]["edge"]["misses"] += date_over_misses + date_under_misses
            results["NBA"]["betting"]["wins"] += date_bet_wins
            results["NBA"]["betting"]["losses"] += date_bet_losses
            results["NBA"]["dates_processed"].append({
                "date": date,
                "edge": f"O:{date_over_hits}-{date_over_misses} U:{date_under_hits}-{date_under_misses}",
                "betting": f"{date_bet_wins}W-{date_bet_losses}L",
                "public": f"{date_public_hits}-{date_public_misses}"
            })
            
            logger.info(f"[#6] NBA {date}: Edge O:{date_over_hits}-{date_over_misses} U:{date_under_hits}-{date_under_misses}, Betting {date_bet_wins}W-{date_bet_losses}L, Public {date_public_hits}-{date_public_misses}")
    
    # Process NHL records (threshold: 0.6)
    NHL_THRESHOLD = 0.6
    for date in dates:
        doc = await db.nhl_opportunities.find_one({"date": date})
        if doc and doc.get('games'):
            date_over_hits = 0
            date_over_misses = 0
            date_under_hits = 0
            date_under_misses = 0
            date_bet_wins = 0
            date_bet_losses = 0
            
            # First, check for actual_bet_record
            actual_record = doc.get('actual_bet_record')
            if actual_record:
                date_bet_wins = actual_record.get('wins', 0)
                date_bet_losses = actual_record.get('losses', 0)
            
            for game in doc['games']:
                final_score_raw = game.get('final_score')
                if final_score_raw is None:
                    continue
                
                final_score = parse_final_score(final_score_raw)
                if final_score is None:
                    continue
                
                # Calculate TRUE edge using combined_gpg and the correct line
                combined_gpg = game.get('combined_gpg') or game.get('combined_ppg')
                if combined_gpg is None:
                    continue
                
                # Use bet_line if bet exists, otherwise use live line (total)
                has_bet = game.get('has_bet', False)
                bet_line = game.get('bet_line')
                total = game.get('total')
                
                if has_bet and bet_line:
                    line = float(bet_line)
                elif total:
                    line = float(total)
                else:
                    continue
                
                # Calculate true edge: GPG - Line
                true_edge = round(float(combined_gpg) - line, 2)
                
                # Only count games with edge >= threshold or <= -threshold
                if true_edge >= NHL_THRESHOLD:
                    # OVER recommendation
                    if final_score > line:
                        date_over_hits += 1
                    elif final_score < line:
                        date_over_misses += 1
                elif true_edge <= -NHL_THRESHOLD:
                    # UNDER recommendation
                    if final_score < line:
                        date_under_hits += 1
                    elif final_score > line:
                        date_under_misses += 1
                
                # Betting Record (if no actual_bet_record)
                if not actual_record and game.get('has_bet'):
                    if game.get('user_bet_hit') is True:
                        date_bet_wins += 1
                    elif game.get('user_bet_hit') is False:
                        date_bet_losses += 1
            
            # Calculate Public Record for NHL (consensus >= 57% threshold)
            date_public_hits = 0
            date_public_misses = 0
            for game in doc['games']:
                away_pct = game.get('away_consensus_pct') or 0
                home_pct = game.get('home_consensus_pct') or 0
                
                if away_pct == 0 and home_pct == 0:
                    continue
                
                is_away_public_pick = away_pct >= home_pct
                public_pct = away_pct if is_away_public_pick else home_pct
                
                if public_pct < PUBLIC_CONSENSUS_THRESHOLD:
                    continue
                
                away_score = game.get('away_score')
                home_score = game.get('home_score')
                if away_score is None or home_score is None:
                    continue
                
                # Get spread - PRIORITY: Covers.com first, CBS Sports as fallback
                public_spread = None
                spread_source = None
                
                if is_away_public_pick:
                    if game.get('away_spread') is not None:
                        public_spread = float(game.get('away_spread'))
                        spread_source = "Covers.com"
                    elif game.get('spread') is not None:
                        public_spread = -float(game.get('spread'))
                        spread_source = "CBS Sports"
                else:
                    if game.get('away_spread') is not None:
                        public_spread = -float(game.get('away_spread'))
                        spread_source = "Covers.com"
                    elif game.get('spread') is not None:
                        public_spread = float(game.get('spread'))
                        spread_source = "CBS Sports"
                
                if public_spread is None:
                    continue
                
                try:
                    away_score_f = float(away_score)
                    home_score_f = float(home_score)
                    spread_f = float(public_spread)
                    
                    if is_away_public_pick:
                        covered = away_score_f + spread_f > home_score_f
                        push = away_score_f + spread_f == home_score_f
                    else:
                        covered = home_score_f + spread_f > away_score_f
                        push = home_score_f + spread_f == away_score_f
                    
                    if not push:
                        if covered:
                            date_public_hits += 1
                            results["NHL"]["public"]["games"].append({
                                "date": date,
                                "game": f"{game.get('away_team')} @ {game.get('home_team')}",
                                "public_pick": game.get('away_team') if is_away_public_pick else game.get('home_team'),
                                "consensus_pct": public_pct,
                                "spread": public_spread,
                                "spread_source": spread_source,
                                "result": "HIT"
                            })
                        else:
                            date_public_misses += 1
                            results["NHL"]["public"]["games"].append({
                                "date": date,
                                "game": f"{game.get('away_team')} @ {game.get('home_team')}",
                                "public_pick": game.get('away_team') if is_away_public_pick else game.get('home_team'),
                                "consensus_pct": public_pct,
                                "spread": public_spread,
                                "spread_source": spread_source,
                                "result": "MISS"
                            })
                except (ValueError, TypeError):
                    continue
            
            results["NHL"]["public"]["hits"] += date_public_hits
            results["NHL"]["public"]["misses"] += date_public_misses
            
            results["NHL"]["edge"]["over_hits"] += date_over_hits
            results["NHL"]["edge"]["over_misses"] += date_over_misses
            results["NHL"]["edge"]["under_hits"] += date_under_hits
            results["NHL"]["edge"]["under_misses"] += date_under_misses
            results["NHL"]["edge"]["hits"] += date_over_hits + date_under_hits
            results["NHL"]["edge"]["misses"] += date_over_misses + date_under_misses
            results["NHL"]["betting"]["wins"] += date_bet_wins
            results["NHL"]["betting"]["losses"] += date_bet_losses
            results["NHL"]["dates_processed"].append({
                "date": date,
                "edge": f"O:{date_over_hits}-{date_over_misses} U:{date_under_hits}-{date_under_misses}",
                "betting": f"{date_bet_wins}W-{date_bet_losses}L",
                "public": f"{date_public_hits}-{date_public_misses}"
            })
            
            logger.info(f"[#6] NHL {date}: Edge O:{date_over_hits}-{date_over_misses} U:{date_under_hits}-{date_under_misses}, Betting {date_bet_wins}W-{date_bet_losses}L, Public {date_public_hits}-{date_public_misses}")
    
    # Process NCAAB records (threshold: 10)
    NCAAB_THRESHOLD = 10
    for date in dates:
        doc = await db.ncaab_opportunities.find_one({"date": date})
        if doc and doc.get('games'):
            date_over_hits = 0
            date_over_misses = 0
            date_under_hits = 0
            date_under_misses = 0
            date_bet_wins = 0
            date_bet_losses = 0
            
            # First, check for actual_bet_record (important for NCAAB due to duplicate bets)
            actual_record = doc.get('actual_bet_record')
            if actual_record:
                date_bet_wins = actual_record.get('wins', 0)
                date_bet_losses = actual_record.get('losses', 0)
            
            for game in doc['games']:
                final_score_raw = game.get('final_score')
                if final_score_raw is None:
                    continue
                
                final_score = parse_final_score(final_score_raw)
                if final_score is None:
                    continue
                
                # Calculate TRUE edge using combined_ppg and the correct line
                combined_ppg = game.get('combined_ppg')
                if combined_ppg is None:
                    continue
                
                # Use bet_line if bet exists, otherwise use live line (total)
                has_bet = game.get('has_bet', False)
                bet_line = game.get('bet_line')
                total = game.get('total')
                
                if has_bet and bet_line:
                    line = float(bet_line)
                elif total:
                    line = float(total)
                else:
                    continue
                
                # Calculate true edge: PPG - Line
                true_edge = round(float(combined_ppg) - line, 2)
                
                # Only count games with edge >= threshold or <= -threshold
                if true_edge >= NCAAB_THRESHOLD:
                    # OVER recommendation
                    if final_score > line:
                        date_over_hits += 1
                    elif final_score < line:
                        date_over_misses += 1
                elif true_edge <= -NCAAB_THRESHOLD:
                    # UNDER recommendation
                    if final_score < line:
                        date_under_hits += 1
                    elif final_score > line:
                        date_under_misses += 1
                
                # Betting Record (if no actual_bet_record)
                if not actual_record and game.get('has_bet'):
                    # Check for multi-bet games with bet_results array
                    if game.get('bet_results') and len(game.get('bet_results')) > 0:
                        for br in game.get('bet_results'):
                            if br.get('hit') is True:
                                date_bet_wins += 1
                            elif br.get('hit') is False:
                                date_bet_losses += 1
                    # Also check bet_wins/bet_losses fields
                    elif game.get('bet_wins') is not None or game.get('bet_losses') is not None:
                        date_bet_wins += game.get('bet_wins', 0)
                        date_bet_losses += game.get('bet_losses', 0)
                    # Fallback to single bet result
                    elif game.get('user_bet_hit') is True:
                        date_bet_wins += 1
                    elif game.get('user_bet_hit') is False:
                        date_bet_losses += 1
            
            # Calculate Public Record for NCAAB (consensus >= 57% threshold)
            date_public_hits = 0
            date_public_misses = 0
            for game in doc['games']:
                away_pct = game.get('away_consensus_pct') or 0
                home_pct = game.get('home_consensus_pct') or 0
                
                if away_pct == 0 and home_pct == 0:
                    continue
                
                is_away_public_pick = away_pct >= home_pct
                public_pct = away_pct if is_away_public_pick else home_pct
                
                if public_pct < PUBLIC_CONSENSUS_THRESHOLD:
                    continue
                
                away_score = game.get('away_score')
                home_score = game.get('home_score')
                if away_score is None or home_score is None:
                    continue
                
                # Get spread - PRIORITY: Covers.com first, CBS Sports as fallback
                public_spread = None
                spread_source = None
                
                if is_away_public_pick:
                    if game.get('away_spread') is not None:
                        public_spread = float(game.get('away_spread'))
                        spread_source = "Covers.com"
                    elif game.get('spread') is not None:
                        public_spread = -float(game.get('spread'))
                        spread_source = "CBS Sports"
                else:
                    if game.get('away_spread') is not None:
                        public_spread = -float(game.get('away_spread'))
                        spread_source = "Covers.com"
                    elif game.get('spread') is not None:
                        public_spread = float(game.get('spread'))
                        spread_source = "CBS Sports"
                
                if public_spread is None:
                    continue
                
                try:
                    away_score_f = float(away_score)
                    home_score_f = float(home_score)
                    spread_f = float(public_spread)
                    
                    if is_away_public_pick:
                        covered = away_score_f + spread_f > home_score_f
                        push = away_score_f + spread_f == home_score_f
                    else:
                        covered = home_score_f + spread_f > away_score_f
                        push = home_score_f + spread_f == away_score_f
                    
                    if not push:
                        if covered:
                            date_public_hits += 1
                            results["NCAAB"]["public"]["games"].append({
                                "date": date,
                                "game": f"{game.get('away_team')} @ {game.get('home_team')}",
                                "public_pick": game.get('away_team') if is_away_public_pick else game.get('home_team'),
                                "consensus_pct": public_pct,
                                "spread": public_spread,
                                "spread_source": spread_source,
                                "result": "HIT"
                            })
                        else:
                            date_public_misses += 1
                            results["NCAAB"]["public"]["games"].append({
                                "date": date,
                                "game": f"{game.get('away_team')} @ {game.get('home_team')}",
                                "public_pick": game.get('away_team') if is_away_public_pick else game.get('home_team'),
                                "consensus_pct": public_pct,
                                "spread": public_spread,
                                "spread_source": spread_source,
                                "result": "MISS"
                            })
                except (ValueError, TypeError):
                    continue
            
            results["NCAAB"]["public"]["hits"] += date_public_hits
            results["NCAAB"]["public"]["misses"] += date_public_misses
            
            results["NCAAB"]["edge"]["over_hits"] += date_over_hits
            results["NCAAB"]["edge"]["over_misses"] += date_over_misses
            results["NCAAB"]["edge"]["under_hits"] += date_under_hits
            results["NCAAB"]["edge"]["under_misses"] += date_under_misses
            results["NCAAB"]["edge"]["hits"] += date_over_hits + date_under_hits
            results["NCAAB"]["edge"]["misses"] += date_over_misses + date_under_misses
            results["NCAAB"]["betting"]["wins"] += date_bet_wins
            results["NCAAB"]["betting"]["losses"] += date_bet_losses
            results["NCAAB"]["dates_processed"].append({
                "date": date,
                "edge": f"O:{date_over_hits}-{date_over_misses} U:{date_under_hits}-{date_under_misses}",
                "betting": f"{date_bet_wins}W-{date_bet_losses}L",
                "public": f"{date_public_hits}-{date_public_misses}"
            })
            
            logger.info(f"[#6] NCAAB {date}: Edge O:{date_over_hits}-{date_over_misses} U:{date_under_hits}-{date_under_misses}, Betting {date_bet_wins}W-{date_bet_losses}L, Public {date_public_hits}-{date_public_misses}")
    
    logger.info(f"[#6] Final calculated records: NBA={results['NBA']}, NHL={results['NHL']}, NCAAB={results['NCAAB']}")
    return results


async def update_records_from_start_date(start_date: str = "2025-12-22"):
    """
    Update compound_records (betting) and edge_records collections with calculated values.
    This replaces (not increments) the values to ensure accuracy.
    """
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    
    # Calculate records
    records = await calculate_records_from_start_date(start_date)
    now = datetime.now(arizona_tz).strftime('%Y-%m-%d %I:%M %p')
    
    # Update compound_records (Betting Record) - NFL eliminated
    for league in ["NBA", "NHL", "NCAAB"]:
        betting = records[league]["betting"]
        await db.compound_records.update_one(
            {"league": league},
            {"$set": {
                "league": league,
                "hits": betting["wins"],
                "misses": betting["losses"],
                "last_updated": now,
                "start_date": start_date
            }},
            upsert=True
        )
        logger.info(f"[#6] Updated {league} betting record: {betting['wins']}-{betting['losses']}")
    
    # Update edge_records (Edge/Recommendation Record) with Over/Under breakdown - NFL eliminated
    for league in ["NBA", "NHL", "NCAAB"]:
        edge = records[league]["edge"]
        await db.edge_records.update_one(
            {"league": league},
            {"$set": {
                "league": league,
                "hits": edge["hits"],
                "misses": edge["misses"],
                "over_hits": edge["over_hits"],
                "over_misses": edge["over_misses"],
                "under_hits": edge["under_hits"],
                "under_misses": edge["under_misses"],
                "last_updated": now,
                "start_date": start_date
            }},
            upsert=True
        )
        logger.info(f"[#6] Updated {league} edge record: {edge['hits']}-{edge['misses']} (O:{edge['over_hits']}-{edge['over_misses']} U:{edge['under_hits']}-{edge['under_misses']})")
    
    # Update public_records (Public Consensus Record) - uses 57%+ threshold with Covers.com spread
    for league in ["NBA", "NHL", "NCAAB"]:
        public = records[league]["public"]
        await db.public_records.update_one(
            {"league": league},
            {"$set": {
                "league": league,
                "hits": public["hits"],
                "misses": public["misses"],
                "games": public.get("games", []),
                "last_updated": now,
                "start_date": start_date,
                "threshold": "57%"
            }},
            upsert=True
        )
        logger.info(f"[#6] Updated {league} public record: {public['hits']}-{public['misses']}")
    
    return {
        "status": "success",
        "records": records,
        "updated_at": now,
        "start_date": start_date
    }


@api_router.post("/process/backfill-opening-lines")
async def backfill_opening_lines(start_date: str = "2024-12-22", end_date: str = None):
    """
    Backfill opening_line field for historical games that don't have it.
    Uses the 'total' field as the opening line if opening_line is not set.
    
    Args:
        start_date: Start date in YYYY-MM-DD format (default: 2024-12-22)
        end_date: End date in YYYY-MM-DD format (default: today)
    """
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo
    
    arizona_tz = ZoneInfo('America/Phoenix')
    
    if not end_date:
        end_date = datetime.now(arizona_tz).strftime("%Y-%m-%d")
    
    results = {
        "NBA": {"dates_processed": 0, "games_updated": 0},
        "NHL": {"dates_processed": 0, "games_updated": 0},
        "NCAAB": {"dates_processed": 0, "games_updated": 0}
    }
    
    # Parse dates
    current_date = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    
    while current_date <= end:
        date_str = current_date.strftime("%Y-%m-%d")
        
        for league in ["NBA", "NHL", "NCAAB"]:
            collection_name = f"{league.lower()}_opportunities"
            collection = db[collection_name]
            
            # Find document for this date
            doc = await collection.find_one({"date": date_str})
            
            if doc and doc.get('games'):
                games = doc['games']
                updated = False
                games_updated_count = 0
                
                for game in games:
                    # If opening_line is not set but total exists, use total as opening_line
                    if not game.get('opening_line') and game.get('total'):
                        game['opening_line'] = game['total']
                        updated = True
                        games_updated_count += 1
                    # If neither exists, try to use bet_line or live_line
                    elif not game.get('opening_line') and not game.get('total'):
                        if game.get('bet_line'):
                            game['opening_line'] = game['bet_line']
                            game['total'] = game['bet_line']
                            updated = True
                            games_updated_count += 1
                        elif game.get('live_line'):
                            game['opening_line'] = game['live_line']
                            game['total'] = game['live_line']
                            updated = True
                            games_updated_count += 1
                
                if updated:
                    await collection.update_one(
                        {"date": date_str},
                        {"$set": {"games": games}}
                    )
                    results[league]["games_updated"] += games_updated_count
                
                results[league]["dates_processed"] += 1
        
        current_date += timedelta(days=1)
    
    logger.info(f"[Backfill Opening Lines] Results: {results}")
    
    return {
        "success": True,
        "start_date": start_date,
        "end_date": end_date,
        "results": results
    }


@api_router.post("/process/backfill-consensus")
async def backfill_consensus_data(start_date: str = "2025-12-22", end_date: str = None, league: str = "NBA"):
    """
    Backfill consensus percentage data from Covers.com for historical dates.
    This scrapes the public betting consensus and spread data for each date.
    
    Args:
        start_date: Start date in YYYY-MM-DD format (default: 2025-12-22)
        end_date: End date in YYYY-MM-DD format (default: yesterday)
        league: League to backfill (NBA, NHL, NCAAB)
    """
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo
    
    arizona_tz = ZoneInfo('America/Phoenix')
    
    if not end_date:
        # Default to yesterday
        end_date = (datetime.now(arizona_tz) - timedelta(days=1)).strftime("%Y-%m-%d")
    
    league = league.upper()
    collection_name = f"{league.lower()}_opportunities"
    
    results = {
        "league": league,
        "start_date": start_date,
        "end_date": end_date,
        "dates_processed": 0,
        "games_updated": 0,
        "dates_details": []
    }
    
    # Team abbreviation mapping for matching
    team_abbrev_map = {
        # NBA
        'ATLANTA': 'ATL', 'BOSTON': 'BOS', 'BROOKLYN': 'BKN', 'CHARLOTTE': 'CHA',
        'CHICAGO': 'CHI', 'CLEVELAND': 'CLE', 'DALLAS': 'DAL', 'DENVER': 'DEN',
        'DETROIT': 'DET', 'GOLDEN STATE': 'GSW', 'HOUSTON': 'HOU', 'INDIANA': 'IND',
        'LA CLIPPERS': 'LAC', 'LA LAKERS': 'LAL', 'MEMPHIS': 'MEM', 'MIAMI': 'MIA',
        'MILWAUKEE': 'MIL', 'MINNESOTA': 'MIN', 'NEW ORLEANS': 'NOP', 'NEW YORK': 'NYK',
        'OKLA CITY': 'OKC', 'OKLAHOMA CITY': 'OKC', 'ORLANDO': 'ORL', 'PHILADELPHIA': 'PHI',
        'PHOENIX': 'PHX', 'PORTLAND': 'POR', 'SACRAMENTO': 'SAC', 'SAN ANTONIO': 'SAS',
        'TORONTO': 'TOR', 'UTAH': 'UTA', 'WASHINGTON': 'WAS',
        # NHL
        'ANAHEIM': 'ANA', 'ARIZONA': 'ARI', 'BUFFALO': 'BUF', 'CALGARY': 'CGY',
        'CAROLINA': 'CAR', 'COLORADO': 'COL', 'COLUMBUS': 'CBJ', 'EDMONTON': 'EDM',
        'FLORIDA': 'FLA', 'LOS ANGELES': 'LAK', 'MONTREAL': 'MTL', 'NASHVILLE': 'NSH',
        'NEW JERSEY': 'NJD', 'NY ISLANDERS': 'NYI', 'NY RANGERS': 'NYR', 'OTTAWA': 'OTT',
        'PITTSBURGH': 'PIT', 'SAN JOSE': 'SJS', 'SEATTLE': 'SEA', 'ST LOUIS': 'STL',
        'ST. LOUIS': 'STL', 'TAMPA BAY': 'TBL', 'VANCOUVER': 'VAN', 'VEGAS': 'VGK',
        'WINNIPEG': 'WPG',
    }
    
    # Parse dates
    current_date = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    
    while current_date <= end:
        date_str = current_date.strftime("%Y-%m-%d")
        date_display = current_date.strftime("%m/%d")
        
        logger.info(f"[Backfill Consensus] Processing {league} for {date_str}")
        
        # Scrape consensus data from Covers.com
        consensus_data = await scrape_covers_consensus(league, date_str)
        
        if not consensus_data:
            logger.warning(f"[Backfill Consensus] No consensus data for {league} on {date_str}")
            results["dates_details"].append({"date": date_display, "status": "no_data", "games_updated": 0})
            current_date += timedelta(days=1)
            continue
        
        # Find document for this date
        doc = await db[collection_name].find_one({"date": date_str})
        
        if doc and doc.get('games'):
            games = doc['games']
            updated = False
            games_updated_count = 0
            
            for game in games:
                away_team = game.get('away_team', '').upper()
                home_team = game.get('home_team', '').upper()
                
                # Get abbreviations
                away_abbrev = team_abbrev_map.get(away_team, away_team[:3])
                home_abbrev = team_abbrev_map.get(home_team, home_team[:3])
                
                # Try to match with consensus data
                away_consensus = consensus_data.get(away_abbrev, {})
                home_consensus = consensus_data.get(home_abbrev, {})
                
                if away_consensus.get('consensus_pct') or home_consensus.get('consensus_pct'):
                    game['away_consensus_pct'] = away_consensus.get('consensus_pct')
                    game['home_consensus_pct'] = home_consensus.get('consensus_pct')
                    
                    # Store spreads from Covers if available (and we don't have spread already)
                    # home_consensus['spread'] is the HOME team's spread from Covers
                    if home_consensus.get('spread') is not None and game.get('spread') is None:
                        game['spread'] = home_consensus.get('spread')
                    if away_consensus.get('spread') is not None:
                        game['away_spread'] = away_consensus.get('spread')
                    
                    # Determine public pick
                    if away_consensus.get('consensus_pct') and home_consensus.get('consensus_pct'):
                        if away_consensus['consensus_pct'] > home_consensus['consensus_pct']:
                            game['public_pick'] = game['away_team']
                            game['public_pick_pct'] = away_consensus['consensus_pct']
                        else:
                            game['public_pick'] = game['home_team']
                            game['public_pick_pct'] = home_consensus['consensus_pct']
                    
                    updated = True
                    games_updated_count += 1
            
            if updated:
                await db[collection_name].update_one(
                    {"date": date_str},
                    {"$set": {"games": games}}
                )
                results["games_updated"] += games_updated_count
            
            results["dates_processed"] += 1
            results["dates_details"].append({"date": date_display, "status": "success", "games_updated": games_updated_count})
            logger.info(f"[Backfill Consensus] {date_str}: Updated {games_updated_count} games")
        else:
            results["dates_details"].append({"date": date_display, "status": "no_games", "games_updated": 0})
        
        current_date += timedelta(days=1)
        
        # Small delay between requests to be nice to the server
        await asyncio.sleep(2)
    
    logger.info(f"[Backfill Consensus] Results: {results}")
    
    return {
        "success": True,
        "results": results
    }


@api_router.post("/process/backfill-scores")
async def backfill_scores_from_cbs(start_date: str = "2025-12-22", end_date: str = None, league: str = "NBA"):
    """
    Backfill team scores from CBS Sports for historical dates.
    This scrapes individual team scores (away_score, home_score) and updates the database.
    
    Args:
        start_date: Start date in YYYY-MM-DD format (default: 2025-12-22)
        end_date: End date in YYYY-MM-DD format (default: yesterday)
        league: League to backfill (NBA, NHL, NCAAB)
    """
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo
    
    arizona_tz = ZoneInfo('America/Phoenix')
    
    if not end_date:
        end_date = (datetime.now(arizona_tz) - timedelta(days=1)).strftime("%Y-%m-%d")
    
    league = league.upper()
    collection_name = f"{league.lower()}_opportunities"
    
    results = {
        "league": league,
        "start_date": start_date,
        "end_date": end_date,
        "dates_processed": 0,
        "games_updated": 0,
        "dates_details": []
    }
    
    # Parse dates
    current_date = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    
    while current_date <= end:
        date_str = current_date.strftime("%Y-%m-%d")
        date_display = current_date.strftime("%m/%d")
        
        logger.info(f"[Backfill Scores] Processing {league} for {date_str}")
        
        # Scrape scores from CBS Sports
        if league == "NBA":
            cbs_games = await scrape_cbssports_nba(date_str)
        elif league == "NHL":
            cbs_games = await scrape_cbssports_nhl(date_str)
        elif league == "NCAAB":
            cbs_games = await scrape_cbssports_ncaab(date_str)
        else:
            cbs_games = []
        
        if not cbs_games:
            logger.warning(f"[Backfill Scores] No CBS data for {league} on {date_str}")
            results["dates_details"].append({"date": date_display, "status": "no_cbs_data", "games_updated": 0})
            current_date += timedelta(days=1)
            continue
        
        # Find document for this date
        doc = await db[collection_name].find_one({"date": date_str})
        
        if doc and doc.get('games'):
            games = doc['games']
            updated = False
            games_updated_count = 0
            
            for game in games:
                away_team = game.get('away_team', '').lower()
                home_team = game.get('home_team', '').lower()
                
                # Find matching CBS game
                for cbs_game in cbs_games:
                    cbs_away = cbs_game.get('away_team', '').lower()
                    cbs_home = cbs_game.get('home_team', '').lower()
                    
                    # Match games by team names
                    if ((away_team in cbs_away or cbs_away in away_team) and
                        (home_team in cbs_home or cbs_home in home_team)):
                        
                        # Update scores if CBS has them
                        if cbs_game.get('away_score') is not None:
                            game['away_score'] = cbs_game['away_score']
                        if cbs_game.get('home_score') is not None:
                            game['home_score'] = cbs_game['home_score']
                        if cbs_game.get('final_score') is not None:
                            game['final_score'] = cbs_game['final_score']
                        
                        # Also update spread if missing
                        if cbs_game.get('spread') is not None and game.get('spread') is None:
                            game['spread'] = cbs_game['spread']
                        
                        updated = True
                        games_updated_count += 1
                        logger.debug(f"[Backfill Scores] {away_team} @ {home_team}: {game.get('away_score')}-{game.get('home_score')}")
                        break
            
            if updated:
                await db[collection_name].update_one(
                    {"date": date_str},
                    {"$set": {"games": games}}
                )
                results["games_updated"] += games_updated_count
            
            results["dates_processed"] += 1
            results["dates_details"].append({"date": date_display, "status": "success", "games_updated": games_updated_count})
            logger.info(f"[Backfill Scores] {date_str}: Updated {games_updated_count} games with scores")
        else:
            results["dates_details"].append({"date": date_display, "status": "no_db_games", "games_updated": 0})
        
        current_date += timedelta(days=1)
        
        # Delay between requests
        await asyncio.sleep(2)
    
    logger.info(f"[Backfill Scores] Results: {results}")
    
    return {
        "success": True,
        "results": results
    }


@api_router.post("/process/scrape-openers")
async def scrape_opening_lines_endpoint(target_date: str = None):
    """
    Process #1 - 8:00 PM Job: Scrape tomorrow's opening lines from ScoresAndOdds/CBS Sports.
    Creates game documents in the opportunities collections with opening lines.
    
    Args:
        target_date: Target date in 'YYYY-MM-DD' format. Defaults to tomorrow (Arizona time).
    """
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    
    # Default to tomorrow if no date specified
    if not target_date:
        target_date = (datetime.now(arizona_tz) + timedelta(days=1)).strftime('%Y-%m-%d')
    
    logger.info(f"[8PM Job] Starting scrape-openers for date: {target_date}")
    
    results = {
        "date": target_date,
        "leagues": {},
        "status": "success",
        "timestamp": datetime.now(arizona_tz).isoformat()
    }
    
    try:
        # NFL eliminated - only process NBA, NHL, NCAAB
        leagues = ['NBA', 'NHL', 'NCAAB']
        
        for league in leagues:
            league_result = {
                "games_scraped": 0,
                "games_stored": 0,
                "data_source": "",
                "status": "success",
                "error": None
            }
            
            try:
                # Use CBS Sports as primary source for all leagues (scoresandodds often blocks)
                # Try CBS Sports first, fallback to scoresandodds if CBS fails
                games = []
                
                if league == 'NBA':
                    games = await scrape_cbssports_nba(target_date)
                    if games:
                        league_result["data_source"] = "cbssports.com"
                    else:
                        # Fallback to scoresandodds
                        games = await scrape_scoresandodds(league.upper(), target_date)
                        if games:
                            league_result["data_source"] = "scoresandodds.com"
                            
                elif league == 'NHL':
                    games = await scrape_cbssports_nhl(target_date)
                    if games:
                        league_result["data_source"] = "cbssports.com"
                    else:
                        # Fallback to scoresandodds
                        games = await scrape_scoresandodds(league.upper(), target_date)
                        if games:
                            league_result["data_source"] = "scoresandodds.com"
                            
                elif league == 'NCAAB':
                    games = await scrape_cbssports_ncaab(target_date)
                    league_result["data_source"] = "cbssports.com"
                
                league_result["games_scraped"] = len(games) if games else 0
                
                if games:
                    # Create full game documents for the opportunities collection
                    collection_name = f"{league.lower()}_opportunities"
                    collection = db[collection_name]
                    
                    processed_games = []
                    for game in games:
                        away = game.get('away_team') or game.get('away')
                        home = game.get('home_team') or game.get('home')
                        total = game.get('total')
                        time = game.get('time', 'TBD')
                        spread = game.get('spread')
                        spread_team = game.get('spread_team')
                        moneyline = game.get('moneyline')
                        moneyline_team = game.get('moneyline_team')
                        
                        if away and home:
                            game_doc = {
                                'away_team': away,
                                'home_team': home,
                                'total': total,
                                'opening_line': total,  # Store as opening line
                                'time': time,
                                'date': target_date,
                                'status': 'scheduled'
                            }
                            
                            # Add spread for NBA/NCAAB
                            if league in ['NBA', 'NCAAB']:
                                game_doc['spread'] = spread
                                game_doc['spread_team'] = spread_team
                                game_doc['opening_spread'] = spread
                                game_doc['opening_spread_team'] = spread_team
                            
                            # Add moneyline for NHL
                            if league == 'NHL':
                                game_doc['moneyline'] = moneyline
                                game_doc['moneyline_team'] = moneyline_team
                                game_doc['opening_moneyline'] = moneyline
                                game_doc['opening_moneyline_team'] = moneyline_team
                            
                            processed_games.append(game_doc)
                            
                            # Also store in opening_lines collection for tracking
                            if total:
                                await store_opening_line(league, target_date, away, home, total)
                    
                    if processed_games:
                        # Create or update the opportunities document
                        await collection.update_one(
                            {"date": target_date},
                            {
                                "$set": {
                                    "date": target_date,
                                    "games": processed_games,
                                    "last_updated": datetime.now(arizona_tz).strftime('%I:%M %p'),
                                    "data_source": f"{league_result['data_source']} opening lines"
                                }
                            },
                            upsert=True
                        )
                        league_result["games_stored"] = len(processed_games)
                        logger.info(f"[8PM Job] Created {len(processed_games)} {league} games for {target_date}")
                    else:
                        league_result["status"] = "warning"
                        league_result["error"] = "No valid games to store"
                        logger.warning(f"[8PM Job] No valid {league} games to store for {target_date}")
                else:
                    league_result["status"] = "warning"
                    league_result["error"] = "No games found"
                    logger.warning(f"[8PM Job] No {league} games found for {target_date}")
                    
            except Exception as e:
                league_result["status"] = "error"
                league_result["error"] = str(e)
                logger.error(f"[8PM Job] Error scraping {league}: {e}")
                import traceback
                traceback.print_exc()
            
            results["leagues"][league] = league_result
        
        # Determine overall status
        error_count = sum(1 for l in results["leagues"].values() if l["status"] == "error")
        if error_count == len(leagues):
            results["status"] = "error"
        elif error_count > 0:
            results["status"] = "partial"
        
        return results
        
    except Exception as e:
        logger.error(f"[8PM Job] Error in scrape_opening_lines_endpoint: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/process/update-records")
async def trigger_update_records(start_date: str = "2025-12-22"):
    """
    Process #6: Manually trigger record update from start_date to yesterday.
    
    Args:
        start_date: Start date in 'YYYY-MM-DD' format (default: 12/22/25)
    """
    try:
        result = await update_records_from_start_date(start_date)
        return result
    except Exception as e:
        logger.error(f"Error updating records: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/records/summary")
async def get_records_summary():
    """
    Get a summary of both betting and edge records for all leagues.
    Returns NBA, NHL, NCAAB, and NFL
    """
    try:
        summary = {}
        
        for league in ["NBA", "NHL", "NCAAB", "NFL"]:
            # Get betting record
            betting = await db.compound_records.find_one({"league": league}, {"_id": 0})
            # Get edge record  
            edge = await db.edge_records.find_one({"league": league}, {"_id": 0})
            
            summary[league] = {
                "betting_record": f"{betting.get('hits', 0)}-{betting.get('misses', 0)}" if betting else "0-0",
                "edge_record": f"{edge.get('hits', 0)}-{edge.get('misses', 0)}" if edge else "0-0",
                "edge_over": f"{edge.get('over_hits', 0)}-{edge.get('over_misses', 0)}" if edge else "0-0",
                "edge_under": f"{edge.get('under_hits', 0)}-{edge.get('under_misses', 0)}" if edge else "0-0",
                "betting_last_updated": betting.get('last_updated') if betting else None,
                "edge_last_updated": edge.get('last_updated') if edge else None,
                "start_date": betting.get('start_date', '2025-12-22') if betting else '2025-12-22'
            }
        
        return summary
    except Exception as e:
        logger.error(f"Error getting records summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============== EXCEL EXPORT ==============

@api_router.get("/export/excel")
async def export_to_excel(
    league: str = "NBA",
    start_date: str = "2025-12-22",
    end_date: str = None
):
    """
    Export opportunities data to Excel with colored dots and all analysis data.
    Matches the user's custom format with 4-dot analysis.
    
    Args:
        league: NBA, NHL, or NFL
        start_date: Start date in 'YYYY-MM-DD' format
        end_date: End date in 'YYYY-MM-DD' format (defaults to yesterday)
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from zoneinfo import ZoneInfo
    
    try:
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # Default end_date to yesterday
        if not end_date:
            end_date = (datetime.now(arizona_tz) - timedelta(days=1)).strftime('%Y-%m-%d')
        
        league = league.upper()
        collection_name = f"{league.lower()}_opportunities"
        collection = db[collection_name]
        
        # Generate list of dates
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        
        dates = []
        current = start
        while current <= end:
            dates.append(current.strftime('%Y-%m-%d'))
            current += timedelta(days=1)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{league} Analysis"
        
        # Define colors for dots - using RGB fills (matching user's file exactly)
        dot_colors = {
            'green': PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),
            'yellow': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
            'red': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            'blue': PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid'),
        }
        
        # Font colors for readability
        white_font = Font(color='FFFFFF', bold=True)
        white_font_normal = Font(color='FFFFFF')
        
        # Result colors (matching user's file)
        hit_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
        miss_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light pink
        
        # Header style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Divider row fill (dark gray - same as header)
        divider_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
        
        # 4-Dot Result colors (matching user's file exactly)
        over_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green for OVER
        under_fill = PatternFill(start_color='EAB200', end_color='EAB200', fill_type='solid')  # Orange/Gold for UNDER
        no_bet_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')  # Blue for NO BET
        
        # Headers - columns A through AI (35 columns)
        headers = [
            'Date', '#', 'Time',                                    # A, B, C
            'Away PPG', 'Away L3', 'Away Dots', 'Away Team',        # D, E, F, G
            'Home PPG', 'Home L3', 'Home Dots', 'Home Team',        # H, I, J, K
            'Line', 'Final', 'Diff',                                 # L, M, N
            'PPG Avg', 'Edge', 'Rec',                                # O, P, Q
            'Result', 'Edge Hit',                                    # R, S
            'Bet', 'Type', 'Bet Result', 'Record',                   # T, U, V, W
            '',                                                      # X - spacer
            '', '', 'Away Dots', 'Away Team', '', '',               # Y, Z, AA, AB, AC, AD
            '',                                                      # AE - spacer
            '4-Dot Result',                                          # AF
            '',                                                      # AG - spacer
            '4-Dot Hit', '4-Dot Record'                              # AH, AI
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Set column widths
        col_widths = {
            'A': 12, 'B': 4, 'C': 10,
            'D': 10, 'E': 10, 'F': 8, 'G': 14,
            'H': 10, 'I': 10, 'J': 8, 'K': 14,
            'L': 8, 'M': 8, 'N': 8,
            'O': 10, 'P': 8, 'Q': 10,
            'R': 10, 'S': 10,
            'T': 6, 'U': 10, 'V': 12, 'W': 10,
            'X': 2.5,
            'Y': 3, 'Z': 3, 'AA': 10, 'AB': 14, 'AC': 3, 'AD': 3,
            'AE': 2.5,
            'AF': 12,
            'AG': 2.5,
            'AH': 10, 'AI': 12
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        row_num = 2
        prev_date = None
        four_dot_wins = 0
        four_dot_losses = 0
        betting_wins = 0
        betting_losses = 0
        
        # Helper to get color name from dot emoji
        def get_color_from_emoji(emoji):
            if emoji == '🟢':
                return 'green'
            elif emoji == '🟡':
                return 'yellow'
            elif emoji == '🔴':
                return 'red'
            elif emoji == '🔵':
                return 'blue'
            return None
        
        # Helper to parse dots string
        def parse_dots_to_colors(dots_str):
            colors = []
            if not dots_str:
                return [None, None]
            for char in str(dots_str):
                color = get_color_from_emoji(char)
                if color:
                    colors.append(color)
            while len(colors) < 2:
                colors.append(None)
            return colors
        
        # 4-DOT LOGIC (verified against user's Excel file):
        # 1. If GREEN + BLUE >= 3: OVER
        # 2. If RED + YELLOW >= 3: UNDER
        # 3. In exact 2-2 ties:
        #    a. If GREEN >= 2 AND at least one GREEN on AWAY team: OVER
        #    b. If YELLOW >= 2 AND BLUE >= 2 (no GREEN, no RED): UNDER
        #    c. Otherwise: NO BET
        def calculate_4dot_result(away_colors, home_colors):
            """Calculate 4-dot result using verified logic from user's Excel"""
            all_colors = away_colors + home_colors
            
            green_count = all_colors.count('green')
            blue_count = all_colors.count('blue')
            red_count = all_colors.count('red')
            yellow_count = all_colors.count('yellow')
            
            over_score = green_count + blue_count
            under_score = red_count + yellow_count
            
            # Rule 1: >= 3 in either direction wins
            if over_score >= 3:
                return 'OVER'
            if under_score >= 3:
                return 'UNDER'
            
            # Rule 2: Handle 2-2 ties
            away_green = away_colors.count('green')
            
            # 2a: GREEN >= 2 AND at least one on away team = OVER
            if green_count >= 2 and away_green >= 1:
                return 'OVER'
            
            # 2b: YELLOW >= 2 AND BLUE >= 2 with no GREEN and no RED = UNDER
            if yellow_count >= 2 and blue_count >= 2 and green_count == 0 and red_count == 0:
                return 'UNDER'
            
            # 2c: Everything else is NO BET
            return 'NO BET'
        
        # Process each date
        for date in dates:
            doc = await collection.find_one({"date": date})
            if not doc or not doc.get('games'):
                continue
            
            # Add divider row between dates
            if prev_date is not None and prev_date != date:
                for col in range(1, 36):  # A to AI (35 columns)
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = divider_fill
                row_num += 1
            
            prev_date = date
            games = doc['games']
            
            for idx, game in enumerate(games, 1):
                # Get ranking data
                away_ppg_rank = game.get('away_ppg_rank', '')
                away_last3_rank = game.get('away_last3_rank', '')
                home_ppg_rank = game.get('home_ppg_rank', '')
                home_last3_rank = game.get('home_last3_rank', '')
                
                # Get dots and parse colors
                away_dots_str = game.get('away_dots', '')
                home_dots_str = game.get('home_dots', '')
                away_colors = parse_dots_to_colors(away_dots_str)
                home_colors = parse_dots_to_colors(home_dots_str)
                
                # Calculate diff (final - line)
                final_score = game.get('final_score', '')
                line = game.get('total', game.get('opening_line', ''))
                diff = ''
                if final_score and line:
                    try:
                        diff_val = round(float(final_score) - float(line), 1)
                        diff = f"+{diff_val}" if diff_val > 0 else str(diff_val)
                    except:
                        pass
                
                # Determine edge hit
                edge_hit = ''
                if game.get('edge_hit') == True or game.get('result_hit') == True:
                    edge_hit = 'HIT'
                elif game.get('edge_hit') == False or game.get('result_hit') == False:
                    edge_hit = 'MISS'
                
                # Determine bet result and update betting record
                bet_result = ''
                bet_type_display = game.get('bet_type', '')
                has_bet = game.get('user_bet') or game.get('has_bet')
                if has_bet:
                    if game.get('multiple_bets') and game.get('bet_results'):
                        bet_type_display = 'O+U'
                        wins = game['bet_results'].count('won')
                        losses = game['bet_results'].count('lost')
                        bet_result = f"{wins}W-{losses}L"
                        # Update betting record for multiple bets
                        betting_wins += wins
                        betting_losses += losses
                    elif game.get('user_bet_hit') == True:
                        bet_result = 'won'
                        betting_wins += 1
                    elif game.get('user_bet_hit') == False:
                        bet_result = 'lost'
                        betting_losses += 1
                    else:
                        bet_result = game.get('bet_result', '')
                        # Also check bet_result string for won/lost
                        if bet_result == 'won':
                            betting_wins += 1
                        elif bet_result == 'lost':
                            betting_losses += 1
                
                # Betting record: Show cumulative record whenever there's a bet
                betting_record = ''
                if has_bet:
                    betting_record = f"{betting_wins}-{betting_losses}"
                
                # Calculate 4-dot result
                four_dot_result = calculate_4dot_result(away_colors, home_colors)
                
                # Calculate 4-dot hit
                four_dot_hit = ''
                actual_result = game.get('actual_result', '')
                if four_dot_result in ['OVER', 'UNDER'] and actual_result:
                    if four_dot_result == actual_result:
                        four_dot_hit = 'HIT'
                        four_dot_wins += 1
                    else:
                        four_dot_hit = 'MISS'
                        four_dot_losses += 1
                
                # 4-dot record: Show current cumulative record whenever 4-Dot is NOT "NO BET"
                # The record reflects all completed games up to this point
                four_dot_record = ''
                if four_dot_result != 'NO BET':
                    four_dot_record = f"{four_dot_wins}-{four_dot_losses}"
                
                # Build row data (35 columns: A-AI)
                row_data = [
                    date,                                                    # A - Date
                    idx,                                                     # B - #
                    game.get('time', ''),                                    # C - Time
                    away_ppg_rank,                                           # D - Away PPG
                    away_last3_rank,                                         # E - Away L3
                    away_dots_str,                                           # F - Away Dots
                    game.get('away_team', game.get('away', '')),             # G - Away Team
                    home_ppg_rank,                                           # H - Home PPG
                    home_last3_rank,                                         # I - Home L3
                    home_dots_str,                                           # J - Home Dots
                    game.get('home_team', game.get('home', '')),             # K - Home Team
                    line,                                                    # L - Line
                    final_score,                                             # M - Final
                    diff,                                                    # N - Diff
                    game.get('combined_ppg', game.get('ppg_avg', '')),       # O - PPG Avg
                    game.get('edge', ''),                                    # P - Edge
                    game.get('recommendation', ''),                          # Q - Rec
                    actual_result,                                           # R - Result
                    edge_hit,                                                # S - Edge Hit
                    '💰' if has_bet else '',                                  # T - Bet
                    bet_type_display,                                        # U - Type
                    bet_result,                                              # V - Bet Result
                    betting_record,                                          # W - Record (cumulative betting record)
                    '',                                                      # X - spacer
                    '',                                                      # Y - Away dot 1 color
                    '',                                                      # Z - Away dot 2 color
                    away_dots_str,                                           # AA - Away Dots
                    game.get('away_team', game.get('away', '')),             # AB - Away Team
                    '',                                                      # AC - Home dot 1 color
                    '',                                                      # AD - Home dot 2 color
                    '',                                                      # AE - spacer
                    four_dot_result,                                         # AF - 4-Dot Result
                    '',                                                      # AG - spacer
                    four_dot_hit,                                            # AH - 4-Dot Hit
                    four_dot_record                                          # AI - 4-Dot Record
                ]
                
                # Write row data
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                    # Apply colors to ranking columns (D, E, H, I)
                    if col == 4 and away_colors[0] in dot_colors:  # D - Away PPG
                        cell.fill = dot_colors[away_colors[0]]
                        if away_colors[0] in ['blue', 'red']:
                            cell.font = white_font
                    elif col == 5 and away_colors[1] and away_colors[1] in dot_colors:  # E - Away L3
                        cell.fill = dot_colors[away_colors[1]]
                        if away_colors[1] in ['blue', 'red']:
                            cell.font = white_font
                    elif col == 8 and home_colors[0] in dot_colors:  # H - Home PPG
                        cell.fill = dot_colors[home_colors[0]]
                        if home_colors[0] in ['blue', 'red']:
                            cell.font = white_font
                    elif col == 9 and home_colors[1] and home_colors[1] in dot_colors:  # I - Home L3
                        cell.fill = dot_colors[home_colors[1]]
                        if home_colors[1] in ['blue', 'red']:
                            cell.font = white_font
                    
                    # Edge Hit colors (S - col 19)
                    elif col == 19:
                        if value == 'HIT':
                            cell.fill = hit_fill
                        elif value == 'MISS':
                            cell.fill = miss_fill
                    
                    # Bet Result colors (V - col 22)
                    elif col == 22:
                        if value == 'won':
                            cell.fill = hit_fill
                        elif value == 'lost':
                            cell.fill = miss_fill
                    
                    # Color cells for dot visualization (Y, Z, AC, AD)
                    elif col == 25 and away_colors[0] in dot_colors:  # Y - Away dot 1
                        cell.fill = dot_colors[away_colors[0]]
                    elif col == 26 and away_colors[1] and away_colors[1] in dot_colors:  # Z - Away dot 2
                        cell.fill = dot_colors[away_colors[1]]
                    elif col == 29 and home_colors[0] in dot_colors:  # AC - Home dot 1
                        cell.fill = dot_colors[home_colors[0]]
                    elif col == 30 and home_colors[1] and home_colors[1] in dot_colors:  # AD - Home dot 2
                        cell.fill = dot_colors[home_colors[1]]
                    
                    # 4-Dot Result colors (AF - col 32)
                    elif col == 32:
                        if value == 'OVER':
                            cell.fill = over_fill
                        elif value == 'UNDER':
                            cell.fill = under_fill
                        elif value == 'NO BET':
                            cell.fill = no_bet_fill
                            cell.font = white_font_normal
                    
                    # 4-Dot Hit colors (AH - col 34)
                    elif col == 34:
                        if value == 'HIT':
                            cell.fill = hit_fill
                        elif value == 'MISS':
                            cell.fill = miss_fill
                
                row_num += 1
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Generate filename
        filename = f"{league}_Analysis_{start_date}_to_{end_date}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Access-Control-Expose-Headers": "Content-Disposition",
                "Cache-Control": "no-cache"
            }
        )
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============== COMPOUND RECORD TRACKING ==============

@api_router.get("/opportunities/record/{league}")
async def get_compound_record(league: str):
    """Get the compound record for a league (NBA or NHL)"""
    try:
        league = league.upper()
        record = await db.compound_records.find_one({"league": league}, {"_id": 0})
        
        if record:
            return {
                "league": league,
                "hits": record.get('hits', 0),
                "misses": record.get('misses', 0),
                "last_updated": record.get('last_updated')
            }
        
        return {
            "league": league,
            "hits": 0,
            "misses": 0,
            "last_updated": None
        }
    except Exception as e:
        logger.error(f"Error getting compound record: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/opportunities/record/{league}/update")
async def update_compound_record(league: str, hits: int = 0, misses: int = 0, reset: bool = False):
    """Update the compound record for a league. Use reset=True to reset to 0-0"""
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        league = league.upper()
        
        if reset:
            await db.compound_records.update_one(
                {"league": league},
                {"$set": {
                    "league": league,
                    "hits": 0,
                    "misses": 0,
                    "last_updated": datetime.now(arizona_tz).strftime('%Y-%m-%d %I:%M %p')
                }},
                upsert=True
            )
            return {"success": True, "message": f"{league} record reset to 0-0"}
        
        # Add to existing record
        await db.compound_records.update_one(
            {"league": league},
            {
                "$inc": {"hits": hits, "misses": misses},
                "$set": {"last_updated": datetime.now(arizona_tz).strftime('%Y-%m-%d %I:%M %p')}
            },
            upsert=True
        )
        
        # Get updated record
        record = await db.compound_records.find_one({"league": league}, {"_id": 0})
        return {
            "success": True,
            "league": league,
            "hits": record.get('hits', 0),
            "misses": record.get('misses', 0)
        }
    except Exception as e:
        logger.error(f"Error updating compound record: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/opportunities/record/set")
async def set_compound_record(league: str, hits: int = 0, misses: int = 0):
    """Set the compound record for a league to specific values"""
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        league = league.upper()
        
        await db.compound_records.update_one(
            {"league": league},
            {"$set": {
                "league": league,
                "hits": hits,
                "misses": misses,
                "last_updated": datetime.now(arizona_tz).strftime('%Y-%m-%d %I:%M %p')
            }},
            upsert=True
        )
        
        return {
            "success": True,
            "league": league,
            "hits": hits,
            "misses": misses
        }
    except Exception as e:
        logger.error(f"Error setting compound record: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/opportunities")
async def get_opportunities(day: str = "today"):
    """Get NBA betting opportunities. day parameter: 'yesterday', 'today', 'tomorrow', or a specific date 'YYYY-MM-DD'"""
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # Check if day is a specific date format (YYYY-MM-DD)
        if len(day) == 10 and day[4] == '-' and day[7] == '-':
            target_date = day
        elif day == "tomorrow":
            target_date = (datetime.now(arizona_tz) + timedelta(days=1)).strftime('%Y-%m-%d')
        elif day == "yesterday":
            target_date = (datetime.now(arizona_tz) - timedelta(days=1)).strftime('%Y-%m-%d')
        else:
            target_date = datetime.now(arizona_tz).strftime('%Y-%m-%d')
        
        # Get cached NBA opportunities
        cached = await db.nba_opportunities.find_one({"date": target_date}, {"_id": 0})
        
        # Get compound record
        record = await db.compound_records.find_one({"league": "NBA"}, {"_id": 0})
        compound_record = {
            "hits": record.get('hits', 0) if record else 0,
            "misses": record.get('misses', 0) if record else 0
        }
        
        if cached and cached.get('games'):
            return {
                "success": True,
                "date": target_date,
                "last_updated": cached.get('last_updated'),
                "games": cached.get('games', []),
                "plays": cached.get('plays', []),
                "compound_record": compound_record,
                "data_source": cached.get('data_source', 'hardcoded'),
                "actual_bet_record": cached.get('actual_bet_record')
            }
        
        return {
            "success": True,
            "date": target_date,
            "message": "No opportunities data yet. Data refreshes daily before 10:45 PM Arizona.",
            "games": [],
            "plays": [],
            "compound_record": compound_record,
            "data_source": None
        }
    except Exception as e:
        logger.error(f"Error getting opportunities: {e}")
        raise HTTPException(status_code=500, detail=str(e))


async def scrape_todays_history_bets(page, league: str, target_date: str) -> List[Dict[str, Any]]:
    """
    Scrape today's history bets (already settled) from plays888.co History page.
    Returns bets that match today's date.
    
    This is important because:
    - Games that already finished today have settled bets in History
    - We need to track both open AND settled bets to show all bets placed for today
    """
    import re
    from datetime import datetime
    
    try:
        logger.info(f"[History] Scraping history bets for {league} on {target_date}")
        
        # Navigate to History page
        await page.goto('https://www.plays888.co/wager/History.aspx', timeout=30000)
        await page.wait_for_load_state('networkidle')
        await page.wait_for_timeout(2000)
        
        # Get page text
        page_text = await page.inner_text('body')
        lines = page_text.split('\n')
        
        # Parse today's date for matching (e.g., "Dec 31" or "Jan 03")
        date_obj = datetime.strptime(target_date, '%Y-%m-%d')
        # Create both patterns - with and without leading zero
        today_pattern_no_zero = date_obj.strftime('%b ') + str(date_obj.day)  # "Jan 3"
        today_pattern_with_zero = date_obj.strftime('%b %d')  # "Jan 03"
        
        raw_bets = []
        i = 0
        
        while i < len(lines):
            line = lines[i].strip()
            
            # Check if this section is for today's date (match both patterns)
            is_today = False
            for j in range(max(0, i-10), min(len(lines), i+3)):
                if today_pattern_no_zero in lines[j] or today_pattern_with_zero in lines[j]:
                    is_today = True
                    break
            
            if not is_today:
                i += 1
                continue
            
            # Check sport context
            sport = None
            for j in range(max(0, i-5), min(len(lines), i+1)):
                ctx = lines[j].upper()
                if 'CBB' in ctx or 'COLLEGE' in ctx:
                    sport = 'NCAAB'
                    break
                elif 'NBA' in ctx:
                    sport = 'NBA'
                    break
                elif 'NHL' in ctx:
                    sport = 'NHL'
                    break
                elif 'NFL' in ctx or 'NCAAF' in ctx:
                    sport = 'NFL'
                    break
            
            # Filter by league
            if sport and sport != league.upper():
                i += 1
                continue
            
            # Look for TOTAL bets: "TOTAL o143-110" or "TOTAL u159-110"
            total_match = re.search(r'TOTAL\s+([ou])(\d+\.?\d*)(½)?[-+]\d+', line, re.IGNORECASE)
            
            # Also handle "o/u TOTAL 146½ -10" format (need to look at additional context for over/under)
            ou_total_match = re.search(r'o/u\s+TOTAL\s+(\d+\.?\d*)(½)?\s*[-+]\d+', line, re.IGNORECASE)
            
            if total_match:
                bet_type = 'OVER' if total_match.group(1).lower() == 'o' else 'UNDER'
                bet_line = float(total_match.group(2))
                if total_match.group(3) == '½':
                    bet_line += 0.5
                
                # Find team matchup
                teams_text = ""
                for j in range(max(0, i-2), min(len(lines), i+3)):
                    if 'vrs' in lines[j].lower():
                        teams_text = lines[j]
                        break
                
                teams_match = re.search(r'\(([^)]+)\s+(?:REG\.TIME\s+)?vrs\s+([^)]+?)(?:\s+REG\.TIME)?\)', teams_text, re.IGNORECASE)
                if teams_match:
                    away_team = teams_match.group(1).strip().replace(' REG.TIME', '')
                    home_team = teams_match.group(2).strip().replace(' REG.TIME', '')
                    
                    raw_bets.append({
                        "sport": sport or 'NCAAB',
                        "away_team": away_team,
                        "home_team": home_team,
                        "bet_type": bet_type,
                        "total_line": bet_line,
                        "is_spread": False
                    })
                    logger.info(f"[History] Found TOTAL bet: {bet_type} {bet_line} ({away_team} vs {home_team})")
            
            elif ou_total_match:
                # For "o/u TOTAL X" format, we need to look for additional context
                # Check nearby lines for "UNDER" or "OVER" indicator, or use edge to determine
                bet_line = float(ou_total_match.group(1))
                if ou_total_match.group(2) == '½':
                    bet_line += 0.5
                
                # Default to UNDER for college basketball (more common)
                # But look for explicit indicators
                bet_type = 'UNDER'  # Default to UNDER
                
                # Look in surrounding lines for OVER/UNDER indication
                context_lines = ' '.join(lines[max(0, i-3):min(len(lines), i+3)]).upper()
                if ' OVER ' in context_lines or 'OVER ' in line.upper():
                    bet_type = 'OVER'
                elif ' UNDER ' in context_lines or 'UNDER ' in line.upper():
                    bet_type = 'UNDER'
                
                # Find team matchup
                teams_text = ""
                for j in range(max(0, i-2), min(len(lines), i+3)):
                    if 'vrs' in lines[j].lower():
                        teams_text = lines[j]
                        break
                
                teams_match = re.search(r'\(([^)]+)\s+(?:REG\.TIME\s+)?vrs\s+([^)]+?)(?:\s+REG\.TIME)?\)', teams_text, re.IGNORECASE)
                if teams_match:
                    away_team = teams_match.group(1).strip().replace(' REG.TIME', '')
                    home_team = teams_match.group(2).strip().replace(' REG.TIME', '')
                    
                    raw_bets.append({
                        "sport": sport or 'NCAAB',
                        "away_team": away_team,
                        "home_team": home_team,
                        "bet_type": bet_type,
                        "total_line": bet_line,
                        "is_spread": False
                    })
                    logger.info(f"[History] Found o/u TOTAL bet: {bet_type} {bet_line} ({away_team} vs {home_team})")
            
            # Look for SPREAD bets: "[837] TEAM +/-XX-110"
            spread_match = re.search(r'\[\d+\]\s*([A-Z\s\.\']+?)\s*([+-]?\d+\.?\d*)(½)?[-+]\d+', line, re.IGNORECASE)
            if spread_match and 'TOTAL' not in line.upper():
                team_name = spread_match.group(1).strip()
                spread_value = spread_match.group(2)
                if spread_match.group(3) == '½':
                    try:
                        spread_val = float(spread_value)
                        spread_value = str(spread_val + 0.5 if spread_val > 0 else spread_val - 0.5)
                    except:
                        pass
                
                bet_type = f"{team_name} {spread_value}"
                
                # Find team matchup if available
                teams_text = ""
                for j in range(max(0, i-2), min(len(lines), i+3)):
                    if 'vrs' in lines[j].lower():
                        teams_text = lines[j]
                        break
                
                away_team = team_name
                home_team = "OPPONENT"
                
                teams_match = re.search(r'\(([^)]+)\s+(?:REG\.TIME\s+)?vrs\s+([^)]+?)(?:\s+REG\.TIME)?\)', teams_text, re.IGNORECASE)
                if teams_match:
                    away_team = teams_match.group(1).strip().replace(' REG.TIME', '')
                    home_team = teams_match.group(2).strip().replace(' REG.TIME', '')
                
                raw_bets.append({
                    "sport": sport or 'NCAAB',
                    "away_team": away_team,
                    "home_team": home_team,
                    "bet_type": bet_type,
                    "spread_line": spread_value,
                    "is_spread": True
                })
                logger.info(f"[History] Found SPREAD bet: {bet_type} ({away_team} vs {home_team})")
            
            i += 1
        
        logger.info(f"[History] Found {len(raw_bets)} total history bets for today")
        return raw_bets
        
    except Exception as e:
        logger.error(f"[History] Error scraping history bets: {e}")
        import traceback
        traceback.print_exc()
        return []


@api_router.post("/games/update-line")
async def update_game_line(request: dict):
    """
    Update the line (total) for a specific game.
    Used for manually correcting NHL lines.
    
    Request body:
    {
        "league": "nhl",
        "date": "2026-01-07",
        "away_team": "Team A",
        "home_team": "Team B",
        "new_line": 6.5
    }
    """
    try:
        league = request.get('league', '').lower()
        date = request.get('date')
        away_team = request.get('away_team')
        home_team = request.get('home_team')
        new_line = request.get('new_line')
        
        if not all([league, date, away_team, home_team, new_line is not None]):
            raise HTTPException(status_code=400, detail="Missing required fields")
        
        collection_name = f"{league}_opportunities"
        
        # Find the document
        doc = await db[collection_name].find_one({"date": date})
        if not doc:
            raise HTTPException(status_code=404, detail=f"No data found for {date}")
        
        games = doc.get('games', [])
        game_found = False
        new_edge = None
        
        for i, game in enumerate(games):
            db_away = game.get('away_team', '').lower()
            db_home = game.get('home_team', '').lower()
            
            if away_team.lower() in db_away or db_away in away_team.lower():
                if home_team.lower() in db_home or db_home in home_team.lower():
                    # Update ALL line fields to ensure consistency
                    games[i]['total'] = float(new_line)
                    games[i]['opening_line'] = float(new_line)
                    games[i]['live_line'] = float(new_line)  # Also update live_line
                    games[i]['line_manually_edited'] = True
                    
                    # Recalculate Edge: Edge = Combined PPG - Line
                    combined_ppg = game.get('combined_ppg')
                    if combined_ppg is not None:
                        new_edge = round(float(combined_ppg) - float(new_line), 1)
                        games[i]['edge'] = new_edge
                        
                        # Update recommendation based on new edge
                        # NHL threshold is 0.6
                        if league == 'nhl':
                            if new_edge >= 0.6:
                                games[i]['recommendation'] = 'OVER'
                            elif new_edge <= -0.6:
                                games[i]['recommendation'] = 'UNDER'
                            else:
                                games[i]['recommendation'] = None
                        # NBA threshold is 8
                        elif league == 'nba':
                            if new_edge >= 8:
                                games[i]['recommendation'] = 'OVER'
                            elif new_edge <= -8:
                                games[i]['recommendation'] = 'UNDER'
                            else:
                                games[i]['recommendation'] = None
                        # NCAAB threshold is 9
                        elif league == 'ncaab':
                            if new_edge >= 9:
                                games[i]['recommendation'] = 'OVER'
                            elif new_edge <= -9:
                                games[i]['recommendation'] = 'UNDER'
                            else:
                                games[i]['recommendation'] = None
                        
                        logger.info(f"[Line Edit] Recalculated edge: {combined_ppg} - {new_line} = {new_edge}")
                    
                    game_found = True
                    logger.info(f"[Line Edit] Updated {away_team} @ {home_team} on {date} to {new_line}")
                    break
        
        if not game_found:
            raise HTTPException(status_code=404, detail=f"Game not found: {away_team} @ {home_team}")
        
        # Save back to database
        await db[collection_name].update_one(
            {"date": date},
            {"$set": {"games": games}}
        )
        
        return {"success": True, "message": f"Line updated to {new_line}", "new_edge": new_edge}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[Line Edit] Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/opportunities/refresh-lines")
async def refresh_lines_and_bets(league: str = "NBA", day: str = "today"):
    """
    Refresh live betting lines AND open bets from plays888.co.
    Does NOT touch PPG values, opening lines, or any other analysis data.
    
    #3.5 - Bet Duplication Prevention: Checks for existing bets before adding
    #3.75 - Bet Line Capture: Stores the line at which bet was placed
    """
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        now_arizona = datetime.now(arizona_tz)
        
        # Support day parameter: 'today', 'tomorrow', or specific date 'YYYY-MM-DD'
        if day == "tomorrow":
            target_date = (now_arizona + timedelta(days=1)).strftime('%Y-%m-%d')
        elif len(day) == 10 and day[4] == '-' and day[7] == '-':
            target_date = day
        else:
            target_date = now_arizona.strftime('%Y-%m-%d')
        
        logger.info(f"[Refresh Lines & Bets] Refreshing {league} for {target_date}")
        
        # Get current games from database
        collection_name = f"{league.lower()}_opportunities"
        collection = db[collection_name]
        cached = await collection.find_one({"date": target_date}, {"_id": 0})
        
        if not cached or not cached.get('games'):
            raise HTTPException(status_code=404, detail=f"No {league} data found for {target_date}")
        
        games = cached['games']
        original_count = len(games)
        
        # Fetch live lines from CBS Sports (same source as opening lines)
        # plays888 is ONLY used for scraping bets (open bets / history)
        live_lines = {}
        live_spreads = {}  # For NBA/NCAAB
        live_moneylines = {}  # For NHL
        open_bets = []
        
        try:
            # Use CBS Sports for live lines for ALL leagues
            if league.upper() == "NCAAB":
                live_games = await scrape_cbssports_ncaab(target_date)
            elif league.upper() == "NBA":
                live_games = await scrape_cbssports_nba(target_date)
            elif league.upper() == "NHL":
                live_games = await scrape_cbssports_nhl(target_date)
            else:
                live_games = []
            
            for game in live_games:
                # CBS Sports returns 'away_team'/'home_team', handle both formats
                away = game.get('away') or game.get('away_team', '')
                home = game.get('home') or game.get('home_team', '')
                total = game.get('total')
                
                if total and away and home:
                    key = f"{away.upper()}_{home.upper()}"
                    live_lines[key] = float(total) if isinstance(total, str) else total
                    
                    # Capture spread for NBA/NCAAB
                    if league.upper() in ['NBA', 'NCAAB']:
                        spread = game.get('spread')
                        spread_team = game.get('spread_team')
                        if spread:
                            live_spreads[key] = {'spread': spread, 'spread_team': spread_team}
                    
                    # Capture moneyline for NHL
                    if league.upper() == 'NHL':
                        moneyline = game.get('moneyline')
                        moneyline_team = game.get('moneyline_team')
                        if moneyline:
                            live_moneylines[key] = {'moneyline': moneyline, 'moneyline_team': moneyline_team}
                    
                    logger.info(f"[Refresh Lines] CBS Sports line for {away} @ {home}: {total}")
            
            logger.info(f"[Refresh Lines] Fetched {len(live_lines)} live lines from CBS Sports")
        except Exception as e:
            logger.warning(f"[Refresh Lines] Error fetching lines from CBS Sports: {e}")
        
        # Fetch bets ONLY from ENANO account (jac075) - ignore TIPSTER
        service2 = Plays888Service()
        try:
            await service2.initialize()
            login_result2 = await service2.login("jac075", "acuna2025!")
            
            if login_result2.get('success'):
                # Fetch open bets from ENANO only
                enano_bets = await service2.scrape_open_bets()
                for bet in enano_bets:
                    bet['_account'] = 'ENANO'
                logger.info(f"[Refresh Bets] Found {len(enano_bets)} open bets from ENANO")
                open_bets.extend(enano_bets)
                
                # Also scrape today's history bets (already settled)
                history_bets = await scrape_todays_history_bets(service2.page, league, target_date)
                for bet in history_bets:
                    bet['_account'] = 'ENANO'
                    bet['_source'] = 'history'
                logger.info(f"[Refresh Bets] Found {len(history_bets)} history bets from today")
                open_bets.extend(history_bets)
        except Exception as e:
            logger.warning(f"[Refresh Bets] Error fetching from ENANO: {e}")
        finally:
            await service2.close()
        
        # Deduplicate bets within ENANO account (keep intentional duplicates like Wofford x2)
        # Key: game teams + bet type (normalized)
        deduplicated_bets = []
        seen_bet_keys = {}  # key -> count of this bet
        
        for bet in open_bets:
            away = bet.get('away_team', '').upper()
            home = bet.get('home_team', '').upper()
            bet_type = bet.get('bet_type', '').upper()
            ticket = bet.get('ticket', '')
            source = bet.get('_source', 'open')  # 'open' or 'history'
            
            # Create a normalized key for the bet (game + bet type)
            # Normalize bet type: "TOTAL O143" -> "OVER", "DUKE -26" -> spread
            normalized_type = bet_type
            if 'TOTAL' in bet_type and ('O' in bet_type or 'OVER' in bet_type):
                normalized_type = 'TOTAL_OVER'
            elif 'TOTAL' in bet_type and ('U' in bet_type or 'UNDER' in bet_type):
                normalized_type = 'TOTAL_UNDER'
            elif 'OVER' in bet_type:
                normalized_type = 'TOTAL_OVER'
            elif 'UNDER' in bet_type:
                normalized_type = 'TOTAL_UNDER'
            else:
                # It's a spread bet - normalize by extracting team name
                parts = bet_type.split()
                team = parts[0] if parts else ''
                normalized_type = f"SPREAD_{team}"
            
            bet_key = f"{away}_{home}_{normalized_type}"
            
            # Keep all bets - intentional duplicates (like Wofford x2) should be counted
            deduplicated_bets.append(bet)
            if bet_key not in seen_bet_keys:
                seen_bet_keys[bet_key] = 1
            else:
                seen_bet_keys[bet_key] += 1
                logger.info(f"[Dedup] Keeping duplicate bet #{seen_bet_keys[bet_key]}: {bet_key}")
        
        logger.info(f"[Refresh Bets] Total bets from ENANO: {len(deduplicated_bets)}")
        open_bets = deduplicated_bets
        
        # Update lines and add bets
        lines_updated = 0
        bets_added = 0
        bets_skipped = 0  # #3.5 - Track duplicates
        
        # Create a set of existing bet tickets for deduplication
        existing_bet_tickets = set()
        for game in games:
            if game.get('bet_slip_id'):
                existing_bet_tickets.add(game.get('bet_slip_id'))
        
        for game in games:
            away = game.get('away_team', '')
            home = game.get('home_team', '')
            key = f"{away.upper()}_{home.upper()}"
            
            # Preserve cancelled flag - don't reset bet info for cancelled games
            is_cancelled = game.get('bet_cancelled', False)
            
            # CRITICAL: Check if game is already completed (has final score and result)
            # If so, preserve ALL bet data - do NOT reset
            is_game_completed = game.get('final_score') is not None or game.get('user_bet_hit') is not None
            
            # Preserve existing bet_line before reset - it should NEVER be overwritten by live line
            preserved_bet_line = game.get('bet_line')
            preserved_bet_type = game.get('bet_type')
            preserved_bet_types = game.get('bet_types', [])
            preserved_bet_lines = game.get('bet_lines', [])
            preserved_bet_count = game.get('bet_count', 0)
            preserved_has_bet = game.get('has_bet', False)
            preserved_user_bet = game.get('user_bet', False)
            preserved_user_bet_hit = game.get('user_bet_hit')
            preserved_result = game.get('result')
            preserved_bet_result = game.get('bet_result')
            
            if not is_cancelled and not is_game_completed:
                # Reset bet tracking for this refresh (only if not cancelled AND game not completed)
                game['bet_types'] = []
                game['bet_lines'] = []
                game['bet_count'] = 0
                game['has_bet'] = False
                game['user_bet'] = False
                game['bet_type'] = None
                # Preserve bet_line - it's the original line when bet was placed
                # Only reset if no preserved line (new game with no historical bet)
                game['bet_line'] = preserved_bet_line
            elif is_game_completed:
                # Game is completed - preserve ALL existing bet data
                logger.info(f"[Refresh Lines] Preserving completed game data for {away} @ {home} (final_score={game.get('final_score')}, result={preserved_result})")
                game['bet_line'] = preserved_bet_line
                game['bet_type'] = preserved_bet_type
                game['bet_types'] = preserved_bet_types
                game['bet_lines'] = preserved_bet_lines
                game['bet_count'] = preserved_bet_count
                game['has_bet'] = preserved_has_bet
                game['user_bet'] = preserved_user_bet
                game['user_bet_hit'] = preserved_user_bet_hit
                game['result'] = preserved_result
                game['bet_result'] = preserved_bet_result
            
            # Update live line (but preserve opening_line/total and PPG)
            # opening_line/total = the original line from 8pm scrape (never changes)
            # live_line = current line from plays888 (updated on refresh)
            # For NCAAB, try fuzzy key matching since team names may differ
            matched_line = None
            if key in live_lines:
                matched_line = live_lines[key]
            elif league.upper() == 'NCAAB':
                # Try fuzzy matching for NCAAB
                away_clean = away.upper().replace(".", "").replace("'", "").replace("-", " ").strip()
                home_clean = home.upper().replace(".", "").replace("'", "").replace("-", " ").strip()
                for lk, lv in live_lines.items():
                    lk_clean = lk.replace(".", "").replace("'", "").replace("-", " ").strip()
                    lk_parts = lk_clean.split("_")
                    if len(lk_parts) == 2:
                        lk_away, lk_home = lk_parts
                        # Check if team names are close enough
                        if (away_clean in lk_away or lk_away in away_clean) and (home_clean in lk_home or lk_home in home_clean):
                            matched_line = lv
                            logger.info(f"[Refresh Lines] Fuzzy matched {away} @ {home} -> {lk}")
                            break
            
            if matched_line:
                new_line = matched_line
                old_live_line = game.get('live_line') or game.get('total')
                
                if old_live_line != new_line:
                    # Store in live_line field, NOT total (total = opening line)
                    game['live_line'] = new_line
                    
                    # If this is the first time and opening_line is not set, set it
                    if not game.get('opening_line') and game.get('total'):
                        game['opening_line'] = game.get('total')
                    
                    # CRITICAL: Edge calculation logic
                    # - If bet is placed: use bet_line (the line when bet was taken) - NEVER CHANGES
                    # - If no bet: use live_line (current line)
                    combined_ppg = game.get('combined_ppg')
                    if combined_ppg:
                        # Check for existing bet_line (preserved from before or from database)
                        bet_line_for_edge = game.get('bet_line')
                        if bet_line_for_edge:
                            # Edge is locked to bet_line - DOES NOT CHANGE WITH LIVE LINE
                            game['edge'] = round(combined_ppg - bet_line_for_edge, 1)
                            logger.info(f"[Refresh Lines] Edge locked to bet_line={bet_line_for_edge} for {away} @ {home}")
                        else:
                            # No bet - use live line for edge calculation
                            game['edge'] = round(combined_ppg - new_line, 1)
                        
                        # Update recommendation based on edge
                        edge = game['edge']
                        if league.upper() == 'NCAAB':
                            if edge >= 10:
                                game['recommendation'] = 'OVER'
                            elif edge <= -9:
                                game['recommendation'] = 'UNDER'
                            else:
                                game['recommendation'] = ''
                        elif league.upper() == 'NBA':
                            if edge >= 8:
                                game['recommendation'] = 'OVER'
                            elif edge <= -5:
                                game['recommendation'] = 'UNDER'
                            else:
                                game['recommendation'] = ''
                        else:  # NHL
                            if edge >= 0.6:
                                game['recommendation'] = 'OVER'
                            elif edge <= -0.5:
                                game['recommendation'] = 'UNDER'
                            else:
                                game['recommendation'] = ''
                    
                    lines_updated += 1
                    logger.info(f"[Refresh Lines] Updated live line {away} @ {home}: {old_live_line} -> {new_line} (opening: {game.get('opening_line') or game.get('total')}, bet_line: {game.get('bet_line')})")
            
            # Update live spread (NBA/NCAAB) or moneyline (NHL)
            if key in live_spreads:
                spread_data = live_spreads[key]
                old_spread = game.get('spread')
                new_spread = spread_data.get('spread')
                new_spread_team = spread_data.get('spread_team')
                
                if new_spread and new_spread != old_spread:
                    game['spread'] = new_spread
                    game['spread_team'] = new_spread_team
                    # Don't overwrite opening_spread if it exists
                    if not game.get('opening_spread'):
                        game['opening_spread'] = new_spread
                        game['opening_spread_team'] = new_spread_team
                    logger.info(f"[Refresh Lines] Updated spread for {away} @ {home}: {old_spread} -> {new_spread} ({new_spread_team})")
            
            if key in live_moneylines:
                ml_data = live_moneylines[key]
                old_ml = game.get('moneyline')
                new_ml = ml_data.get('moneyline')
                new_ml_team = ml_data.get('moneyline_team')
                
                if new_ml and new_ml != old_ml:
                    game['moneyline'] = new_ml
                    game['moneyline_team'] = new_ml_team
                    # Don't overwrite opening_moneyline if it exists
                    if not game.get('opening_moneyline'):
                        game['opening_moneyline'] = new_ml
                        game['opening_moneyline_team'] = new_ml_team
                    logger.info(f"[Refresh Lines] Updated moneyline for {away} @ {home}: {old_ml} -> {new_ml} ({new_ml_team})")
            
            # #3.75 - Match open bets to games and store bet_line
            for bet in open_bets:
                bet_game = bet.get('game', '').upper()
                bet_sport = bet.get('sport', '').upper()
                bet_ticket = bet.get('ticket_id')
                bet_description = bet.get('description', '').upper()
                bet_risk = bet.get('risk', 0) or bet.get('total_risk', 0) or 0
                
                # Get team names from bet data
                bet_away = bet.get('away_team', '').upper()
                bet_home = bet.get('home_team', '').upper()
                bet_game = f"{bet_away} {bet_home}"  # Combined for matching
                
                # Check if this bet matches the current league
                is_league_match = False
                logger.debug(f"[Refresh Bets] Checking bet: sport={bet_sport}, away={bet_away}, home={bet_home}")
                if league.upper() == 'NBA' and 'NBA' in bet_sport:
                    is_league_match = True
                elif league.upper() == 'NHL':
                    # NHL bets can be marked as "NHL" or "SOC" (for Regulation Time Only)
                    if 'NHL' in bet_sport:
                        is_league_match = True
                    elif bet_sport == 'SOC' and ('NHL' in bet_description or 'REGULATION' in bet_description):
                        is_league_match = True
                        logger.info(f"[Refresh Bets] Matched SOC bet as NHL: {bet_away} vs {bet_home}")
                elif league.upper() == 'NCAAB' and ('CBB' in bet_sport or 'NCAA' in bet_sport or 'COLLEGE' in bet_sport or 'NCAAB' in bet_sport or 'RBL' in bet_sport):
                    # RBL = College Basketball Extra (special lines on plays888)
                    is_league_match = True
                
                if not is_league_match:
                    continue
                
                # Filter by bet amount:
                # - NBA/NHL: Only count bets with $2,000+ risk (ignore $1,000 bets)
                # - NCAAB: Count all bets including $1,000
                # NOTE: If risk is 0 (parsing issue), still process the bet
                min_bet_amount = 1800 if league.upper() in ['NBA', 'NHL'] else 0  # Use 1800 to account for rounding
                if bet_risk > 0 and bet_risk < min_bet_amount:
                    logger.info(f"[Refresh Bets] Skipping low-amount bet for {league}: ${bet_risk} < ${min_bet_amount} ({bet_away} vs {bet_home})")
                    continue
                
                # Helper function to normalize team names for matching
                def normalize_team_name(name):
                    """Normalize team name for flexible matching"""
                    if not name:
                        return ''
                    name = name.upper()
                    
                    # Handle specific college team mappings FIRST
                    # Use word-boundary-safe replacements to avoid partial matches
                    # (e.g., "EASTERN" shouldn't match "N ILLINOIS" pattern)
                    
                    # Directional replacements - match full team names or use regex for safety
                    directional_mappings = [
                        # Illinois schools (check full patterns first)
                        (r'\bE\.?\s*ILLINOIS\b', 'EASTERN ILLINOIS'),
                        (r'\bW\.?\s*ILLINOIS\b', 'WESTERN ILLINOIS'),
                        (r'\bN\.?\s*ILLINOIS\b', 'NORTHERN ILLINOIS'),
                        (r'\bS\.?\s*ILLINOIS\b', 'SOUTHERN ILLINOIS'),
                        # Kentucky schools
                        (r'\bE\.?\s*KENTUCKY\b', 'EASTERN KENTUCKY'),
                        (r'\bW\.?\s*KENTUCKY\b', 'WESTERN KENTUCKY'),
                        (r'\bN\.?\s*KENTUCKY\b', 'NORTHERN KENTUCKY'),
                        # Michigan schools
                        (r'\bE\.?\s*MICHIGAN\b', 'EASTERN MICHIGAN'),
                        (r'\bW\.?\s*MICHIGAN\b', 'WESTERN MICHIGAN'),
                        # Iowa
                        (r'\bN\.?\s*IOWA\b', 'NORTHERN IOWA'),
                        # Other states
                        (r'\bN\.?\s*ARIZONA\b', 'NORTHERN ARIZONA'),
                        (r'\bN\.?\s*COLORADO\b', 'NORTHERN COLORADO'),
                        (r'\bN\.?\s*DAKOTA\b', 'NORTH DAKOTA'),
                        (r'\bS\.?\s*DAKOTA\b', 'SOUTH DAKOTA'),
                        (r'\bW\.?\s*VIRGINIA\b', 'WEST VIRGINIA'),
                        (r'\bN\.?\s*CAROLINA\b', 'NORTH CAROLINA'),
                        (r'\bS\.?\s*CAROLINA\b', 'SOUTH CAROLINA'),
                        (r'\bE\.?\s*CAROLINA\b', 'EAST CAROLINA'),
                        (r'\bW\.?\s*CAROLINA\b', 'WESTERN CAROLINA'),
                        # Missouri schools
                        (r'\bSE\.?\s*MISSOURI\s*ST\.?(?:\s|$)', 'SOUTHEAST MISSOURI STATE '),
                        (r'\bSE\.?\s*MISSOURI\b', 'SOUTHEAST MISSOURI'),
                        (r'\bMISSOURI\s*ST\.?\b', 'MISSOURI STATE'),
                    ]
                    
                    import re
                    for pattern, replacement in directional_mappings:
                        name = re.sub(pattern, replacement, name)
                    
                    # Other specific mappings (simple string replacements)
                    simple_mappings = {
                        # Connecticut / UConn - normalize both ways (check CONNECTICUT first before CONN)
                        'CONNECTICUT': 'UCONN',
                        # Virginia Commonwealth / VCU - normalize both ways
                        'VA COMMONWEALTH': 'VCU',
                        'VIRGINIA COMMONWEALTH': 'VCU',
                        'VA. COMMONWEALTH': 'VCU',
                        # Saint / St abbreviations - normalize to ST
                        'SAINT LOUIS': 'ST LOUIS',
                        'SAINT MARYS': 'ST MARYS',
                        'SAINT JOHNS': 'ST JOHNS',
                        'SAINT JOSEPHS': 'ST JOSEPHS',
                        'SAINT PETERS': 'ST PETERS',
                        'SAINT BONAVENTURE': 'ST BONAVENTURE',
                        # Tennessee schools
                        'UT MARTIN': 'TENNESSEE MARTIN',
                        'TENNESSEE-MARTIN': 'TENNESSEE MARTIN',
                        'MIDDLE TENN': 'MIDDLE TENNESSEE',
                        'E. TENNESSEE': 'EAST TENNESSEE',
                        'TENN. TECH': 'TENNESSEE TECH',
                        # Florida schools
                        'FGCU': 'FLORIDA GULF COAST',
                        'FLA GULF COAST': 'FLORIDA GULF COAST',
                        'FAU': 'FLORIDA ATLANTIC',
                        'FLORIDA ATL': 'FLORIDA ATLANTIC',
                        'FIU': 'FLORIDA INTERNATIONAL',
                        'UCF': 'CENTRAL FLORIDA',
                        # Arkansas schools
                        'CENT. ARKANSAS': 'CENTRAL ARKANSAS',
                        'CENT ARKANSAS': 'CENTRAL ARKANSAS',
                        'ARK. STATE': 'ARKANSAS STATE',
                        'ARK STATE': 'ARKANSAS STATE',
                        # Other specific mappings
                        'LOYOLA CHI.': 'LOYOLA CHICAGO',
                        'LOYOLA CHI': 'LOYOLA CHICAGO',
                        'BOSTON U.': 'BOSTON UNIVERSITY',
                        'BOSTON U': 'BOSTON UNIVERSITY',
                        'INDIANA ST.': 'INDIANA STATE',
                        'INDIANA ST': 'INDIANA STATE',
                        'MONTANA ST.': 'MONTANA STATE',
                        'MONTANA ST': 'MONTANA STATE',
                        'CAL POLY SLO': 'CAL POLY',
                        'ILLINOIS ST.': 'ILLINOIS STATE',
                        'ILLINOIS ST': 'ILLINOIS STATE',
                        # SMU / Southern Methodist
                        'SOUTHERN METHODIST': 'SMU',
                        'S. METHODIST': 'SMU',
                        # BYU
                        'BRIGHAM YOUNG': 'BYU',
                        # TCU
                        'TEXAS CHRISTIAN': 'TCU',
                        # USC / Southern California
                        'SOUTHERN CALIFORNIA': 'USC',
                        'SOUTHERN CAL': 'USC',
                        # Ole Miss
                        'MISSISSIPPI': 'OLE MISS',
                        # North Alabama
                        'NORTH ALABAMA': 'N ALABAMA',
                        'N. ALABAMA': 'N ALABAMA',
                        # Fairleigh Dickinson / FDU
                        'FAIRLEIGH DICKINSON': 'FDU',
                        'FAIRLEIGH': 'FDU',
                        # Chicago State
                        'CHICAGO STATE': 'CHICAGO ST',
                        'CHICAGO ST.': 'CHICAGO ST',
                        # Eastern Kentucky
                        'EASTERN KENTUCKY': 'E KENTUCKY',
                        'E. KENTUCKY': 'E KENTUCKY',
                        # UMass Lowell
                        'UMASS LOWELL': 'UMASS LOWELL',
                        'MASSACHUSETTS LOWELL': 'UMASS LOWELL',
                        # UC San Diego
                        'UC SAN DIEGO': 'UC SAN DIEGO',
                        'UCSD': 'UC SAN DIEGO',
                        # CS Fullerton
                        'CS FULLERTON': 'CS FULLERTON',
                        'CAL STATE FULLERTON': 'CS FULLERTON',
                        'CAL ST FULLERTON': 'CS FULLERTON',
                        'CAL ST. FULLERTON': 'CS FULLERTON',
                        # SE Missouri State
                        'SE MISSOURI STATE': 'SE MISSOURI ST',
                        'SE MISSOURI ST.': 'SE MISSOURI ST',
                        'SOUTHEAST MISSOURI STATE': 'SE MISSOURI ST',
                        'SOUTHEAST MISSOURI ST': 'SE MISSOURI ST',
                        'SEMO': 'SE MISSOURI ST',
                        # Southern Indiana
                        'SOUTHERN INDIANA': 'SO INDIANA',
                        'S INDIANA': 'SO INDIANA',
                        'S. INDIANA': 'SO INDIANA',
                        'SO. INDIANA': 'SO INDIANA',
                    }
                    for abbrev, full in simple_mappings.items():
                        if abbrev in name:
                            name = name.replace(abbrev, full)
                    
                    # Miami Ohio normalization - do this AFTER other mappings to avoid double replacement
                    # Normalize all variations to just "MIAMI OHIO"
                    import re as re_inner
                    name = re_inner.sub(r'MIAMI\s*\(OHIO\)', 'MIAMI OHIO', name)
                    name = re_inner.sub(r'MIAMI\s*\(OH\)', 'MIAMI OHIO', name)
                    name = re_inner.sub(r'MIAMI-OHIO', 'MIAMI OHIO', name)
                    # Only replace "MIAMI OH" if it's not already "MIAMI OHIO"
                    if 'MIAMI OHIO' not in name:
                        name = re_inner.sub(r'MIAMI\s+OH\b', 'MIAMI OHIO', name)
                    
                    # NHL team name mappings (city abbreviations to full names)
                    nhl_mappings = {
                        'LA KINGS': 'LOS ANGELES KINGS',
                        'NY RANGERS': 'NEW YORK RANGERS',
                        'NY ISLANDERS': 'NEW YORK ISLANDERS',
                        'TB LIGHTNING': 'TAMPA BAY LIGHTNING',
                        'NJ DEVILS': 'NEW JERSEY DEVILS',
                        'SJ SHARKS': 'SAN JOSE SHARKS',
                    }
                    for abbrev, full in nhl_mappings.items():
                        if abbrev in name:
                            name = name.replace(abbrev, full)
                    
                    # NBA team name mappings
                    nba_mappings = {
                        'OKLAHOMA CITY THUNDER': 'OKLA CITY',
                        'OKLAHOMA CITY': 'OKLA CITY',
                        'THUNDER': 'OKLA CITY',
                        'OKC': 'OKLA CITY',
                        'GOLDEN STATE WARRIORS': 'GOLDEN STATE',
                        'WARRIORS': 'GOLDEN STATE',
                        'LA LAKERS': 'LA LAKERS',
                        'LOS ANGELES LAKERS': 'LA LAKERS',
                        'LAKERS': 'LA LAKERS',
                        'LA CLIPPERS': 'LA CLIPPERS',
                        'LOS ANGELES CLIPPERS': 'LA CLIPPERS',
                        'CLIPPERS': 'LA CLIPPERS',
                        'NEW YORK KNICKS': 'NEW YORK',
                        'KNICKS': 'NEW YORK',
                        'BROOKLYN NETS': 'BROOKLYN',
                        'NETS': 'BROOKLYN',
                        'ATLANTA HAWKS': 'ATLANTA',
                        'HAWKS': 'ATLANTA',
                        'BOSTON CELTICS': 'BOSTON',
                        'CELTICS': 'BOSTON',
                        'CHICAGO BULLS': 'CHICAGO',
                        'BULLS': 'CHICAGO',
                        'CLEVELAND CAVALIERS': 'CLEVELAND',
                        'CAVALIERS': 'CLEVELAND',
                        'CAVS': 'CLEVELAND',
                        'DALLAS MAVERICKS': 'DALLAS',
                        'MAVERICKS': 'DALLAS',
                        'MAVS': 'DALLAS',
                        'DENVER NUGGETS': 'DENVER',
                        'NUGGETS': 'DENVER',
                        'DETROIT PISTONS': 'DETROIT',
                        'PISTONS': 'DETROIT',
                        'HOUSTON ROCKETS': 'HOUSTON',
                        'ROCKETS': 'HOUSTON',
                        'INDIANA PACERS': 'INDIANA',
                        'PACERS': 'INDIANA',
                        'MEMPHIS GRIZZLIES': 'MEMPHIS',
                        'GRIZZLIES': 'MEMPHIS',
                        'MIAMI HEAT': 'MIAMI',
                        'HEAT': 'MIAMI',
                        'MILWAUKEE BUCKS': 'MILWAUKEE',
                        'BUCKS': 'MILWAUKEE',
                        'MINNESOTA TIMBERWOLVES': 'MINNESOTA',
                        'TIMBERWOLVES': 'MINNESOTA',
                        'NEW ORLEANS PELICANS': 'NEW ORLEANS',
                        'PELICANS': 'NEW ORLEANS',
                        'ORLANDO MAGIC': 'ORLANDO',
                        'MAGIC': 'ORLANDO',
                        'PHILADELPHIA 76ERS': 'PHILADELPHIA',
                        '76ERS': 'PHILADELPHIA',
                        'SIXERS': 'PHILADELPHIA',
                        'PHOENIX SUNS': 'PHOENIX',
                        'SUNS': 'PHOENIX',
                        'PORTLAND TRAIL BLAZERS': 'PORTLAND',
                        'TRAIL BLAZERS': 'PORTLAND',
                        'BLAZERS': 'PORTLAND',
                        'SACRAMENTO KINGS': 'SACRAMENTO',
                        'KINGS': 'SACRAMENTO',
                        'SAN ANTONIO SPURS': 'SAN ANTONIO',
                        'SPURS': 'SAN ANTONIO',
                        'TORONTO RAPTORS': 'TORONTO',
                        'RAPTORS': 'TORONTO',
                        'UTAH JAZZ': 'UTAH',
                        'JAZZ': 'UTAH',
                        'WASHINGTON WIZARDS': 'WASHINGTON',
                        'WIZARDS': 'WASHINGTON',
                        'CHARLOTTE HORNETS': 'CHARLOTTE',
                        'HORNETS': 'CHARLOTTE',
                    }
                    for full_name, short in nba_mappings.items():
                        if full_name in name:
                            name = name.replace(full_name, short)
                    
                    # Common abbreviations (applied after specific mappings)
                    # Note: ST. -> STATE for college teams (not SAINT)
                    # Only apply if STATE isn't already in the name
                    if 'STATE' not in name:
                        name = name.replace("ST.", "STATE").replace(" ST ", " STATE ")
                    name = name.replace("VA.", "VIRGINIA").replace("VA ", "VIRGINIA ")
                    name = name.replace("GA.", "GEORGIA").replace("GA ", "GEORGIA ")
                    name = name.replace("CONN.", "CONNECTICUT").replace("CONN ", "CONNECTICUT ")
                    name = name.replace("'S", "S").replace("'", "")  # Remove apostrophes
                    name = name.replace(".", "")  # Remove periods for consistent matching
                    return name
                
                # Check if game matches - compare team names
                game_matches = False
                away_norm = normalize_team_name(away)
                home_norm = normalize_team_name(home)
                bet_away_norm = normalize_team_name(bet_away)
                bet_home_norm = normalize_team_name(bet_home)
                
                # Debug logging for OKC
                if 'OKLA' in bet_away_norm or 'THUNDER' in bet_away or 'GOLDEN' in bet_home:
                    logger.info(f"[DEBUG OKC] Game: {away} @ {home} -> {away_norm} @ {home_norm}")
                    logger.info(f"[DEBUG OKC] Bet: {bet_away} vs {bet_home} -> {bet_away_norm} vs {bet_home_norm}")
                
                # Debug logging for Michigan/Miami (common matching issue)
                if 'MICHIGAN' in bet_away_norm or 'MICHIGAN' in bet_home_norm or 'MIAMI' in bet_away_norm or 'MIAMI' in bet_home_norm:
                    logger.info(f"[DEBUG MICHIGAN/MIAMI] Game: {away} @ {home} -> {away_norm} @ {home_norm}")
                    logger.info(f"[DEBUG MICHIGAN/MIAMI] Bet: {bet_away} vs {bet_home} -> {bet_away_norm} vs {bet_home_norm}")
                
                # Try various matching approaches
                # IMPORTANT: Require BOTH teams to match for accurate bet tracking
                if bet_away_norm and bet_home_norm:
                    # Check away team match
                    away_match = (away_norm == bet_away_norm or 
                                 away_norm in bet_away_norm or bet_away_norm in away_norm or
                                 away_norm == bet_home_norm or
                                 away_norm in bet_home_norm or bet_home_norm in away_norm)
                    
                    # Check home team match  
                    home_match = (home_norm == bet_home_norm or
                                 home_norm in bet_home_norm or bet_home_norm in home_norm or
                                 home_norm == bet_away_norm or
                                 home_norm in bet_away_norm or bet_away_norm in home_norm)
                    
                    # Require both teams to match, but prevent substring false positives
                    # by checking that the matched word is significant (>4 chars)
                    if away_match and home_match:
                        # Verify it's not a false positive by checking word overlap
                        away_words = set(away_norm.split())
                        home_words = set(home_norm.split())
                        bet_away_words = set(bet_away_norm.split())
                        bet_home_words = set(bet_home_norm.split())
                        
                        away_word_match = len(away_words & bet_away_words) >= 1 or len(away_words & bet_home_words) >= 1
                        home_word_match = len(home_words & bet_home_words) >= 1 or len(home_words & bet_away_words) >= 1
                        
                        if away_word_match and home_word_match:
                            game_matches = True
                            logger.info(f"[Refresh Bets] MATCHED: {away} @ {home} -> bet {bet_away} vs {bet_home}")
                        else:
                            # Log near matches that failed word overlap
                            if 'MICHIGAN' in away_norm or 'MIAMI' in home_norm:
                                logger.warning(f"[Refresh Bets] NEAR MISS (word overlap): {away} @ {home} vs bet {bet_away} vs {bet_home}")
                                logger.warning(f"  away_words={away_words}, home_words={home_words}")
                                logger.warning(f"  bet_away_words={bet_away_words}, bet_home_words={bet_home_words}")
                elif bet_game:
                    # Fallback to combined game string
                    bet_game_norm = normalize_team_name(bet_game)
                    if away_norm in bet_game_norm or home_norm in bet_game_norm:
                        game_matches = True
                
                if not game_matches:
                    # Log unmatched bets for debugging
                    if 'MICHIGAN' in bet_away_norm or 'MIAMI' in bet_home_norm:
                        logger.warning(f"[Refresh Bets] UNMATCHED bet: {bet_away} vs {bet_home} (norm: {bet_away_norm} vs {bet_home_norm})")
                    continue
                
                # Skip cancelled games - don't add bets back
                if game.get('bet_cancelled', False):
                    logger.info(f"[Refresh Bets] Skipping cancelled game: {away} @ {home}")
                    continue
                    
                # #3.5 - Bet Duplication Prevention
                if bet_ticket and bet_ticket in existing_bet_tickets:
                    bets_skipped += 1
                    continue
                
                # Add bet to game
                game['has_bet'] = True
                game['user_bet'] = True
                if bet_ticket:
                    existing_bet_tickets.add(bet_ticket)  # Add to set to prevent future duplicates
                
                # #3.75 - Capture the bet line and type
                bet_line = bet.get('total_line') or bet.get('line') or bet.get('total')
                bet_type_raw = bet.get('bet_type', '')
                
                # Determine if this is a spread bet or total bet
                is_spread_bet = False
                is_total_bet = False
                bet_type_display = bet_type_raw
                
                if 'total' in bet_type_raw.lower() or 'over' in bet_type_raw.lower() or 'under' in bet_type_raw.lower():
                    is_total_bet = True
                    if 'over' in bet_type_raw.lower() or bet_type_raw.lower().startswith('o'):
                        bet_type_display = 'OVER'
                    elif 'under' in bet_type_raw.lower() or bet_type_raw.lower().startswith('u'):
                        bet_type_display = 'UNDER'
                else:
                    # It's a spread bet (like "DUKE -26" or "LEHIGH -5")
                    is_spread_bet = True
                    bet_type_display = bet_type_raw.upper()
                
                # Initialize bet tracking arrays if not present
                if 'bet_types' not in game:
                    game['bet_types'] = []
                if 'bet_lines' not in game:
                    game['bet_lines'] = []
                if 'bet_count' not in game:
                    game['bet_count'] = 0
                
                # Add this bet to the arrays
                game['bet_types'].append(bet_type_display)
                if bet_line:
                    try:
                        game['bet_lines'].append(float(bet_line) if isinstance(bet_line, str) else bet_line)
                    except:
                        game['bet_lines'].append(None)
                else:
                    game['bet_lines'].append(None)
                
                game['bet_count'] += bet.get('bet_count', 1)  # Use bet's count (for x2 bets)
                
                # For backward compatibility, set the primary bet_type and bet_line
                # If there's both spread and total, prioritize showing both in bet_type
                if len(game['bet_types']) == 1:
                    game['bet_type'] = bet_type_display
                    if bet_line:
                        try:
                            game['bet_line'] = float(bet_line) if isinstance(bet_line, str) else bet_line
                        except:
                            pass
                else:
                    # Multiple bets on same game - combine them
                    game['bet_type'] = ' + '.join(game['bet_types'])
                    # Keep the first bet_line for display
                
                bets_added += 1
                logger.info(f"[Refresh Bets] Added bet #{game['bet_count']} to {away} @ {home}: {bet_type_display}")
        
        # CRITICAL: After all bets are matched, recalculate edge for games with bet_line
        # This ensures edge is locked to bet_line, not live_line
        for game in games:
            if game.get('has_bet') and game.get('bet_line'):
                combined_ppg = game.get('combined_ppg') or game.get('combined_gpg')
                bet_line = game.get('bet_line')
                if combined_ppg and bet_line:
                    # Lock edge to bet_line - THIS IS THE EDGE THAT MATTERS
                    game['edge'] = round(combined_ppg - bet_line, 1)
                    logger.info(f"[Refresh Bets] Edge recalculated with bet_line={bet_line}: edge={game['edge']} for {game.get('away_team')} @ {game.get('home_team')}")
        
        # Update plays array with games that have bets
        plays = []
        for game in games:
            if game.get('has_bet'):
                away = game.get('away_team', '')
                home = game.get('home_team', '')
                plays.append({
                    "game": f"{away} @ {home}",
                    "total": game.get('total'),
                    "bet_line": game.get('bet_line'),
                    "bet_lines": game.get('bet_lines', []),
                    "combined_ppg": game.get('combined_ppg') or game.get('combined_gpg'),
                    "combined_gpg": game.get('combined_gpg') or game.get('combined_ppg'),
                    "edge": game.get('edge'),
                    "recommendation": game.get('recommendation', ''),
                    "has_bet": True,
                    "bet_type": game.get('bet_type', 'TOTAL'),
                    "bet_types": game.get('bet_types', []),
                    "bet_count": game.get('bet_count', 1)
                })
        
        # Scrape public consensus percentages from Covers.com
        consensus_updated = 0
        try:
            logger.info(f"[Refresh Lines] Scraping consensus data from Covers.com for {league} on {target_date}")
            consensus_data = await scrape_covers_consensus(league.upper(), target_date)
            
            if consensus_data:
                logger.info(f"[Refresh Lines] Got consensus data for {len(consensus_data)} teams")
                
                for game in games:
                    away = game.get('away_team', '').upper()
                    home = game.get('home_team', '').upper()
                    
                    # Try to match teams in consensus data
                    away_consensus = None
                    home_consensus = None
                    
                    # Direct match first
                    if away in consensus_data:
                        away_consensus = consensus_data[away].get('consensus_pct')
                    if home in consensus_data:
                        home_consensus = consensus_data[home].get('consensus_pct')
                    
                    # Fuzzy match if direct match fails
                    if away_consensus is None or home_consensus is None:
                        for team_key, team_data in consensus_data.items():
                            team_key_upper = team_key.upper()
                            # Check if team names are similar
                            if away_consensus is None and (away in team_key_upper or team_key_upper in away):
                                away_consensus = team_data.get('consensus_pct')
                                logger.debug(f"[Consensus] Fuzzy matched {away} -> {team_key}")
                            if home_consensus is None and (home in team_key_upper or team_key_upper in home):
                                home_consensus = team_data.get('consensus_pct')
                                logger.debug(f"[Consensus] Fuzzy matched {home} -> {team_key}")
                    
                    # Update game with consensus data
                    if away_consensus is not None or home_consensus is not None:
                        if away_consensus is not None:
                            game['away_consensus_pct'] = away_consensus
                        if home_consensus is not None:
                            game['home_consensus_pct'] = home_consensus
                        consensus_updated += 1
                        logger.info(f"[Consensus] Updated {away} ({away_consensus}%) @ {home} ({home_consensus}%)")
                
                logger.info(f"[Refresh Lines] Updated consensus for {consensus_updated} games")
            else:
                logger.warning(f"[Refresh Lines] No consensus data returned from Covers.com")
        except Exception as e:
            logger.warning(f"[Refresh Lines] Error fetching consensus data: {e}")
        
        # Save updated games and plays
        # CRITICAL: Calculate bet results for completed games with final scores
        # This ensures results are updated even for live bets that complete during the day
        for game in games:
            if game.get('has_bet') and game.get('final_score') and game.get('bet_line'):
                final_score = game['final_score']
                bet_line = game['bet_line']
                bet_type = game.get('bet_type', '').upper()
                
                # Only calculate if user_bet_hit is not already set (don't overwrite existing results)
                if game.get('user_bet_hit') is None and game.get('bet_result') != 'push':
                    # Check for PUSH first
                    if final_score == bet_line:
                        game['user_bet_hit'] = None
                        game['result'] = 'PUSH'
                        game['bet_result'] = 'push'
                        logger.info(f"[Refresh] PUSH: {game.get('away_team')} @ {game.get('home_team')} - {bet_type} {bet_line} vs {final_score}")
                    elif 'OVER' in bet_type:
                        game['user_bet_hit'] = final_score > bet_line
                        game['result'] = 'OVER' if final_score > bet_line else 'UNDER'
                        logger.info(f"[Refresh] Result: {game.get('away_team')} @ {game.get('home_team')} - OVER {bet_line} vs {final_score} = {'HIT' if game['user_bet_hit'] else 'MISS'}")
                    elif 'UNDER' in bet_type:
                        game['user_bet_hit'] = final_score < bet_line
                        game['result'] = 'UNDER' if final_score < bet_line else 'OVER'
                        logger.info(f"[Refresh] Result: {game.get('away_team')} @ {game.get('home_team')} - UNDER {bet_line} vs {final_score} = {'HIT' if game['user_bet_hit'] else 'MISS'}")
        
        await collection.update_one(
            {"date": target_date},
            {"$set": {
                "games": games,
                "plays": plays,
                "last_updated": now_arizona.strftime('%I:%M %p')
            }}
        )
        
        logger.info(f"[Refresh Lines & Bets] Updated {lines_updated} lines, added {bets_added} bets, skipped {bets_skipped} duplicates, {consensus_updated} consensus, {len(plays)} plays")
        
        return {
            "success": True,
            "league": league,
            "date": target_date,
            "lines_updated": lines_updated,
            "bets_added": bets_added,
            "bets_skipped_duplicates": bets_skipped,
            "consensus_updated": consensus_updated,
            "total_games": original_count,
            "last_updated": now_arizona.strftime('%I:%M %p')
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[Refresh Lines & Bets] Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/opportunities/refresh")
async def refresh_opportunities(day: str = "today", use_live_lines: bool = False):
    """Manually refresh NBA opportunities data. 
    day parameter: 'yesterday', 'today' or 'tomorrow'
    use_live_lines: if True, fetch O/U lines from plays888.co instead of hardcoded values
    
    #3 PROCESS: After 5am Arizona time, TODAY's data automatically uses Plays888 for live lines
    """
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        now_arizona = datetime.now(arizona_tz)
        current_hour = now_arizona.hour
        
        # #3 PROCESS: Automatically use Plays888 for TODAY after 5am Arizona time
        if day == "today" and current_hour >= 5:
            use_live_lines = True
            logger.info(f"[#3 Process] After 5am Arizona ({now_arizona.strftime('%I:%M %p')}), using Plays888 for live lines")
        
        if day == "tomorrow":
            target_date = (now_arizona + timedelta(days=1)).strftime('%Y-%m-%d')
        elif day == "yesterday":
            target_date = (now_arizona - timedelta(days=1)).strftime('%Y-%m-%d')
        else:
            target_date = now_arizona.strftime('%Y-%m-%d')
        
        # #3 PROCESS: For TODAY, preserve PPG values from 8pm job - do NOT re-scrape
        # PPG values should be locked after the 8pm job
        ppg_season = {}
        ppg_last3 = {}
        ppg_season_values = {}
        ppg_last3_values = {}
        
        # Try to get cached PPG data from the database first (from 8pm job)
        cached_ppg = await db.nba_opportunities.find_one({"date": target_date}, {"_id": 0})
        use_cached_ppg = False
        
        if cached_ppg and cached_ppg.get('games') and cached_ppg.get('ppg_populated'):
            # Extract PPG values from cached games
            logger.info(f"[#3 Process] Using LOCKED PPG values from 8pm job (date={target_date})")
            use_cached_ppg = True
            for game in cached_ppg['games']:
                away = game.get('away_team', '')
                home = game.get('home_team', '')
                if away:
                    ppg_season[away] = game.get('away_ppg_rank', 15)
                    ppg_last3[away] = game.get('away_last3_rank', 15)
                    ppg_season_values[away] = game.get('away_season_ppg', 115.0)
                    ppg_last3_values[away] = game.get('away_last3_ppg', 115.0)
                if home:
                    ppg_season[home] = game.get('home_ppg_rank', 15)
                    ppg_last3[home] = game.get('home_last3_rank', 15)
                    ppg_season_values[home] = game.get('home_season_ppg', 115.0)
                    ppg_last3_values[home] = game.get('home_last3_ppg', 115.0)
        
        # If no cached PPG (first time or historical), scrape from TeamRankings
        if not use_cached_ppg:
            try:
                ppg_data = await scrape_nba_ppg_rankings(target_date)
                if ppg_data['season_values']:
                    ppg_season = ppg_data['season_ranks']
                    ppg_last3 = ppg_data['last3_ranks']
                    ppg_season_values = ppg_data['season_values']
                    ppg_last3_values = ppg_data['last3_values']
                    logger.info(f"Scraped fresh PPG data from teamrankings.com (date={target_date}): {len(ppg_season_values)} teams")
                else:
                    raise Exception("No PPG data scraped")
            except Exception as e:
                logger.warning(f"Failed to scrape live PPG, using hardcoded fallback: {e}")
                # Fallback PPG Rankings (Season) - Updated 12/26/2025
                ppg_season = {
                    'Denver': 1, 'Okla City': 2, 'Houston': 3, 'New York': 4, 'Miami': 5,
                    'Utah': 6, 'San Antonio': 7, 'Chicago': 8, 'Detroit': 9, 'Atlanta': 10,
                    'Cleveland': 11, 'Minnesota': 12, 'Orlando': 13, 'LA Lakers': 14, 'Portland': 15,
                    'Philadelphia': 16, 'Boston': 17, 'New Orleans': 18, 'Memphis': 19, 'Charlotte': 20,
                    'Phoenix': 21, 'Golden State': 22, 'Toronto': 23, 'Milwaukee': 24, 'Dallas': 25,
                    'Washington': 26, 'Sacramento': 27, 'LA Clippers': 28, 'Indiana': 29, 'Brooklyn': 30
                }
                
                # Fallback PPG Rankings (Last 3 games)
                ppg_last3 = {
                    'Chicago': 1, 'Utah': 2, 'New Orleans': 3, 'Atlanta': 4, 'San Antonio': 5,
                    'Portland': 6, 'Houston': 7, 'Orlando': 8, 'Dallas': 9, 'Memphis': 10,
                    'Denver': 11, 'Philadelphia': 12, 'New York': 13, 'Sacramento': 14, 'Golden State': 15,
                    'LA Lakers': 16, 'Cleveland': 17, 'Miami': 18, 'Boston': 19, 'Okla City': 20,
                    'Detroit': 21, 'Charlotte': 22, 'Washington': 23, 'Phoenix': 24, 'Brooklyn': 25,
                    'Toronto': 26, 'Indiana': 27, 'Minnesota': 28, 'LA Clippers': 29, 'Milwaukee': 30
                }
                
                ppg_season_values = {
                    'Denver': 125.8, 'Okla City': 121.5, 'New York': 120.6, 'Utah': 120.5,
                    'Houston': 120.3, 'Miami': 120.2, 'San Antonio': 119.9, 'Cleveland': 119.5,
                    'Detroit': 118.8, 'Chicago': 118.8, 'Atlanta': 118.8, 'Minnesota': 118.8,
                    'LA Lakers': 117.2, 'Orlando': 117.1, 'Portland': 116.6, 'Boston': 116.5,
                    'Philadelphia': 115.7, 'Charlotte': 115.6, 'Memphis': 115.3, 'New Orleans': 115.2,
                    'Phoenix': 115.1, 'Golden State': 114.9, 'Toronto': 114.4, 'Washington': 113.6,
                    'Dallas': 113.5, 'Milwaukee': 112.7, 'Sacramento': 111.7, 'LA Clippers': 111.5,
                    'Brooklyn': 109.3, 'Indiana': 109.2
                }
                
                ppg_last3_values = {
                    'Denver': 132.7, 'Okla City': 113.7, 'New York': 119.3, 'Utah': 128.7,
                    'Houston': 114.7, 'Miami': 119.7, 'San Antonio': 120.3, 'Cleveland': 121.7,
                    'Detroit': 121.3, 'Chicago': 112.7, 'Atlanta': 119.7, 'Minnesota': 120.0,
                    'LA Lakers': 109.7, 'Orlando': 114.0, 'Portland': 107.7, 'Boston': 117.0,
                    'Philadelphia': 104.0, 'Charlotte': 126.0, 'Memphis': 124.7, 'New Orleans': 113.3,
                    'Phoenix': 123.3, 'Golden State': 124.3, 'Toronto': 123.3, 'Washington': 121.0,
                    'Dallas': 118.0, 'Milwaukee': 109.0, 'Sacramento': 113.7, 'LA Clippers': 119.7,
                    'Brooklyn': 101.7, 'Indiana': 99.3
                }
        
        games_raw = []
        data_source = "hardcoded"
        open_bets = []
        settled_bets = []
        
        # Always fetch open bets and settled bets for ENANO account
        try:
            enano_conn = await db.connections.find_one({"username": "jac075"}, {"_id": 0})
            if enano_conn:
                scraper = Plays888Service()
                await scraper.login("jac075", decrypt_password(enano_conn["password_encrypted"]))
                open_bets = await scraper.scrape_open_bets()
                # Also get settled bets with original lines for historical data
                if day == "yesterday" or (len(day) == 10 and day[4] == '-'):
                    settled_bets = await scraper.scrape_settled_bets_with_lines("NBA")
                await scraper.close()
                logger.info(f"Fetched {len(open_bets)} open bets, {len(settled_bets)} settled bets for NBA matching")
        except Exception as e:
            logger.error(f"Error fetching bets: {e}")
        
        # DATA SOURCING STRATEGY:
        # - TODAY: Use plays888.co for live lines
        # - YESTERDAY/HISTORICAL: Use scoresandodds.com for final scores
        # - TOMORROW: Use plays888.co or hardcoded data
        
        # For TODAY: Start with full schedule, then update with live lines from plays888.co
        if day == "today":
            games_raw = []
            data_source = "plays888.co"
            
            # #3 PROCESS: After 5am Arizona time, AUTOMATICALLY use Plays888 for live lines
            # This happens regardless of the use_live_lines parameter
            arizona_tz = ZoneInfo('America/Phoenix')
            current_hour = datetime.now(arizona_tz).hour
            auto_use_live = current_hour >= 5  # After 5am Arizona
            
            if use_live_lines or auto_use_live:
                logger.info(f"[#3 Process] After 5am Arizona ({datetime.now(arizona_tz).strftime('%I:%M %p')}), using Plays888 for live lines")
                try:
                    conn = await db.connections.find_one({}, {"_id": 0}, sort=[("created_at", -1)])
                    if conn and conn.get("is_connected"):
                        username = conn["username"]
                        password = decrypt_password(conn["password_encrypted"])
                        
                        scraper = Plays888Service()
                        await scraper.login(username, password)
                        live_games = await scraper.scrape_totals("NBA")
                        await scraper.close()
                        
                        if live_games:
                            # Use Plays888 games as the PRIMARY source
                            for game in live_games:
                                away_short = convert_plays888_team_name(game.get('away', ''))
                                home_short = convert_plays888_team_name(game.get('home', ''))
                                games_raw.append({
                                    "time": game.get('time', ''),
                                    "away": away_short,
                                    "home": home_short,
                                    "total": game.get('total')
                                })
                            
                            data_source = "plays888.co (Live)"
                            logger.info(f"[#3 Process] Using {len(games_raw)} games from Plays888 as primary source")
                except Exception as e:
                    logger.error(f"Error fetching live games from Plays888: {e}")
            
            # #3.85: MERGE live games with cached games to keep started games visible ALL DAY
            if games_raw:
                cached = await db.nba_opportunities.find_one({"date": target_date}, {"_id": 0})
                if cached and cached.get('games'):
                    # Create lookup of current games from Plays888
                    live_matchups = set()
                    for g in games_raw:
                        key = f"{g['away'].upper()}_{g['home'].upper()}"
                        live_matchups.add(key)
                    
                    # Add cached games that are NOT in the live list (started games)
                    added_started = 0
                    for cached_game in cached['games']:
                        away = cached_game.get('away_team', cached_game.get('away', ''))
                        home = cached_game.get('home_team', cached_game.get('home', ''))
                        key = f"{away.upper()}_{home.upper()}"
                        
                        if key not in live_matchups:
                            # This game has started - add it from cache
                            games_raw.append({
                                "time": cached_game.get('time', 'Started'),
                                "away": away,
                                "home": home,
                                "total": cached_game.get('total'),
                                "started": True  # Mark as started
                            })
                            added_started += 1
                    
                    if added_started > 0:
                        logger.info(f"[#3.85] Added {added_started} started NBA games from cache (total: {len(games_raw)} games)")
                        data_source = "plays888.co + cached"
            
            # Fallback: If no games from Plays888, check database cache
            if not games_raw:
                cached = await db.nba_opportunities.find_one({"date": target_date}, {"_id": 0})
                if cached and cached.get('games'):
                    for g in cached['games']:
                        games_raw.append({
                            "time": g.get('time', ''),
                            "away": g.get('away_team', g.get('away', '')),
                            "home": g.get('home_team', g.get('home', '')),
                            "total": g.get('total')
                        })
                    data_source = "cached (from last night's scrape)"
                    logger.info(f"Using {len(games_raw)} cached games from database for {target_date}")
        
        # For TOMORROW: Use scoresandodds.com for schedule/lines
        elif day == "tomorrow":
            try:
                scraped_games = await scrape_scoresandodds("NBA", target_date)
                
                if scraped_games:
                    # STORE OPENING LINES - First time we see these games
                    await store_opening_lines_batch("NBA", target_date, scraped_games)
                    
                    for game in scraped_games:
                        game_entry = {
                            "time": game.get('time', ''),
                            "away": game.get('away_team', ''),
                            "home": game.get('home_team', ''),
                            "total": game.get('total', 220.0),  # Default if no line yet
                        }
                        games_raw.append(game_entry)
                    
                    data_source = "scoresandodds.com"
                    logger.info(f"Fetched {len(games_raw)} tomorrow games from scoresandodds.com for {target_date}")
            except Exception as e:
                logger.error(f"Error scraping tomorrow from scoresandodds.com: {e}")
        
        # For YESTERDAY/HISTORICAL: Return cached data from database with final scores
        elif day == "yesterday" or (len(day) == 10 and day[4] == '-'):
            # Query database for the specific date
            if day == "yesterday":
                target_date = (now_arizona - timedelta(days=1)).strftime('%Y-%m-%d')
            else:
                target_date = day
            
            cached = await db.nba_opportunities.find_one({"date": target_date}, {"_id": 0})
            if cached and cached.get('games'):
                logger.info(f"Returning cached NBA data for {target_date} with {len(cached['games'])} games")
                return {
                    "date": target_date,
                    "games": cached.get('games', []),
                    "plays": cached.get('plays', []),
                    "last_updated": cached.get('last_updated'),
                    "data_source": "cached"
                }
        
        # Fallback to hardcoded data if scraping failed or returned empty data
        if not games_raw:
            if day == "tomorrow":
                # Dec 27 NBA games - Lines from scoresandodds.com
                games_raw = [
                    {"time": "3:30 PM", "away": "Dallas", "home": "Sacramento", "total": 231.5},
                    {"time": "6:00 PM", "away": "Phoenix", "home": "New Orleans", "total": 239.5},
                    {"time": "6:00 PM", "away": "Denver", "home": "Orlando", "total": 235.5},
                    {"time": "6:00 PM", "away": "Milwaukee", "home": "Brooklyn", "total": 222.5},
                    {"time": "6:00 PM", "away": "Indiana", "home": "Miami", "total": 224.5},
                    {"time": "6:30 PM", "away": "New York", "home": "Toronto", "total": 229.5},
                    {"time": "7:00 PM", "away": "Chicago", "home": "Memphis", "total": 232.5},
                    {"time": "8:00 PM", "away": "Utah", "home": "San Antonio", "total": 227.5},
                    {"time": "9:00 PM", "away": "Cleveland", "home": "Houston", "total": 227.5},
                    {"time": "10:00 PM", "away": "LA Clippers", "home": "Golden State", "total": 228.0},
                ]
            elif day == "yesterday":
                # Dec 26, 2025 - NBA games with FINAL SCORES from scoresandodds.com
                games_raw = [
                    {"time": "12:00 PM", "away": "Boston", "home": "Indiana", "total": 222.5, "final_score": 262},
                    {"time": "2:30 PM", "away": "Toronto", "home": "Washington", "total": 225.5, "final_score": 255},
                    {"time": "5:00 PM", "away": "Charlotte", "home": "Orlando", "total": 227.5, "final_score": 225},
                    {"time": "5:00 PM", "away": "Miami", "home": "Atlanta", "total": 250.5, "final_score": 237},
                    {"time": "5:00 PM", "away": "Philadelphia", "home": "Chicago", "total": 241.5, "final_score": 211},
                    {"time": "7:00 PM", "away": "Milwaukee", "home": "Memphis", "total": 228.5, "final_score": 229},
                    {"time": "7:00 PM", "away": "Phoenix", "home": "New Orleans", "total": 241.5, "final_score": 223},
                    {"time": "9:00 PM", "away": "Detroit", "home": "Utah", "total": 245.5, "final_score": 260},
                    {"time": "10:00 PM", "away": "LA Clippers", "home": "Portland", "total": 226.5, "final_score": 222},
                ]
            else:
                # Default fallback - Dec 26 games
                games_raw = [
                    {"time": "12:00 PM", "away": "Boston", "home": "Indiana", "total": 222.0},
                    {"time": "2:30 PM", "away": "Toronto", "home": "Washington", "total": 226.0},
                    {"time": "5:00 PM", "away": "Charlotte", "home": "Orlando", "total": 212.0},
                    {"time": "5:00 PM", "away": "Miami", "home": "Atlanta", "total": 230.0},
                    {"time": "5:00 PM", "away": "Philadelphia", "home": "Chicago", "total": 221.0},
                    {"time": "7:00 PM", "away": "Milwaukee", "home": "Memphis", "total": 223.0},
                    {"time": "7:00 PM", "away": "Phoenix", "home": "New Orleans", "total": 230.0},
                    {"time": "9:00 PM", "away": "Detroit", "home": "Utah", "total": 230.0},
                    {"time": "10:00 PM", "away": "LA Clippers", "home": "Portland", "total": 219.0},
                ]
        
        # Calculate averages and recommendations
        games = []
        plays = []
        
        # Track which games are already in our list (for merging open bets)
        games_in_list = set()
        for g in games_raw:
            games_in_list.add(f"{g['away'].lower()}_{g['home'].lower()}")
        
        # Only add games from open bets for TODAY - NOT for yesterday or historical
        # Open bets are CURRENT bets, not historical ones
        if day == "today":
            for bet in open_bets:
                if bet.get('sport') != 'NBA':
                    continue
                
                bet_away = bet.get('away_team', '').upper()
                bet_home = bet.get('home_team', '').upper()
                
                # Convert bet team names to short names
                away_short = convert_plays888_team_name(bet_away)
                home_short = convert_plays888_team_name(bet_home)
                
                game_key = f"{away_short.lower()}_{home_short.lower()}"
                game_key_reversed = f"{home_short.lower()}_{away_short.lower()}"
                if game_key not in games_in_list and game_key_reversed not in games_in_list:
                    # This bet is for a game not in our list - add it
                    bet_line = bet.get('total_line', 220.0)
                    games_raw.append({
                        "time": "LIVE",  # Mark as live game
                        "away": away_short,
                        "home": home_short,
                        "total": bet_line
                    })
                    games_in_list.add(game_key)
                    logger.info(f"Added game from open bet: {away_short} @ {home_short}")
        
        for i, g in enumerate(games_raw, 1):
            away_season = ppg_season.get(g['away'], 15)
            away_last3 = ppg_last3.get(g['away'], 15)
            away_avg = (away_season + away_last3) / 2
            
            home_season = ppg_season.get(g['home'], 15)
            home_last3 = ppg_last3.get(g['home'], 15)
            home_avg = (home_season + home_last3) / 2
            
            game_avg = (away_avg + home_avg) / 2
            
            # Calculate combined PPG (actual points expected in the game)
            away_season_ppg = ppg_season_values.get(g['away'], 115.0)
            away_last3_ppg = ppg_last3_values.get(g['away'], 115.0)
            home_season_ppg = ppg_season_values.get(g['home'], 115.0)
            home_last3_ppg = ppg_last3_values.get(g['home'], 115.0)
            
            # Combined PPG = average of (season totals + last 3 totals)
            season_total = away_season_ppg + home_season_ppg
            last3_total = away_last3_ppg + home_last3_ppg
            combined_ppg = (season_total + last3_total) / 2
            
            # Check if we have a valid line from plays888.co
            # If total is None or 0, it means "NO LINE" - game not active in plays888
            has_line = g.get('total') and g['total'] > 0
            
            # SIMPLIFIED LOGIC: Determine recommendation based on PPG vs Line comparison
            # If PPG average > Line → OVER (we expect more points than the line)
            # If PPG average < Line → UNDER (we expect fewer points than the line)
            recommendation = None
            color = "neutral"
            
            if has_line:
                edge_value = combined_ppg - g['total']  # Positive = OVER, Negative = UNDER
                
                # Recommend based on which side has the edge
                # Edge must be at least 0.5 points to make a recommendation
                if edge_value >= 0.5:  # PPG is significantly higher than line
                    recommendation = "OVER"
                    color = "green"
                elif edge_value <= -0.5:  # PPG is significantly lower than line
                    recommendation = "UNDER"
                    color = "red"
            
            game_data = {
                "game_num": i,
                "time": g['time'],
                "away_team": g['away'],
                "away_ppg_rank": away_season,
                "away_last3_rank": away_last3,
                "away_avg": round(away_avg, 1),
                "away_season_ppg": round(away_season_ppg, 1),
                "away_last3_ppg": round(away_last3_ppg, 1),
                "home_team": g['home'],
                "home_ppg_rank": home_season,
                "home_last3_rank": home_last3,
                "home_avg": round(home_avg, 1),
                "home_season_ppg": round(home_season_ppg, 1),
                "home_last3_ppg": round(home_last3_ppg, 1),
                "total": g['total'] if has_line else None,  # Show None if no line
                "opening_line": None,  # Will be set below from database
                "has_line": has_line,
                "combined_ppg": round(combined_ppg, 1),
                "game_avg": round(game_avg, 1),
                "recommendation": recommendation,
                "color": color,
                "has_bet": False,
                "bet_type": None,
                "bet_risk": 0,
                "bet_count": 0
            }
            
            # Add dots for NBA
            def get_nba_dot_color(rank: int) -> str:
                if rank <= 8:
                    return "🟢"  # Green: Top tier
                elif rank <= 16:
                    return "🔵"  # Blue: Upper middle
                elif rank <= 24:
                    return "🟡"  # Yellow: Lower middle
                else:
                    return "🔴"  # Red: Bottom tier
            
            game_data["dots"] = f"{get_nba_dot_color(away_season)}{get_nba_dot_color(away_last3)}{get_nba_dot_color(home_season)}{get_nba_dot_color(home_last3)}"
            game_data["away_dots"] = f"{get_nba_dot_color(away_season)}{get_nba_dot_color(away_last3)}"
            game_data["home_dots"] = f"{get_nba_dot_color(home_season)}{get_nba_dot_color(home_last3)}"
            
            # Get stored opening line from database
            stored_opening = await get_opening_line("NBA", target_date, g['away'], g['home'])
            if stored_opening:
                game_data["opening_line"] = stored_opening
            elif has_line:
                # If no stored opening, store current line as opening (first time seeing this game)
                await store_opening_line("NBA", target_date, g['away'], g['home'], g['total'])
                game_data["opening_line"] = g['total']  # First time = opening equals current
            
            # Check if this game has an active bet
            # Also detect "hedged" bets (both OVER and UNDER on same game = cancelled out)
            # IMPORTANT: Only count ENANO bets ($2k+) for betting record, NOT TIPSTER copies ($1k)
            game_bets = []
            enano_bets = []  # Only $2k+ bets count for user record
            
            for bet in open_bets:
                if bet.get('sport') == 'NBA':
                    # Match team names (case-insensitive partial match)
                    bet_away = bet.get('away_team', '').upper()
                    bet_home = bet.get('home_team', '').upper()
                    game_away = g['away'].upper()
                    game_home = g['home'].upper()
                    
                    # Check if teams match (partial match for city names)
                    away_match = any(part in bet_away for part in game_away.split()) or any(part in game_away for part in bet_away.split())
                    home_match = any(part in bet_home for part in game_home.split()) or any(part in game_home for part in bet_home.split())
                    
                    if away_match and home_match:
                        game_bets.append(bet)
                        # Check if it's an ENANO bet ($2k+) - these count for user betting record
                        bet_risk = bet.get('total_risk', bet.get('risk', 0))
                        if bet_risk >= 1500:  # ENANO bets are $2k+, TIPSTER copies are $1k
                            enano_bets.append(bet)
            
            # Check if game is hedged (has both OVER and UNDER bets)
            bet_types = [b.get('bet_type', '').upper() for b in game_bets]
            is_hedged = 'OVER' in bet_types and 'UNDER' in bet_types
            
            # For display purposes, show all bets
            if game_bets and not is_hedged:
                # Game has active bet(s) that are not hedged
                game_data["has_bet"] = True
                game_data["bet_type"] = game_bets[0].get('bet_type')
                game_data["bet_risk"] = sum(b.get('total_risk', b.get('risk', 0)) for b in game_bets)
                game_data["bet_count"] = sum(b.get('bet_count', 1) for b in game_bets)
                # Store the line at which the bet was placed
                game_data["bet_line"] = game_bets[0].get('total_line')
                # Flag if this is an ENANO bet (counts for user record)
                game_data["is_enano_bet"] = len(enano_bets) > 0
            elif is_hedged:
                # Game is hedged (OVER + UNDER = push/cancelled)
                game_data["has_bet"] = False
                game_data["is_hedged"] = True
                game_data["bet_type"] = "HEDGED"
                game_data["bet_risk"] = 0
                game_data["bet_count"] = 0
                game_data["bet_line"] = None
                game_data["is_enano_bet"] = False
            
            # Add result data for yesterday/historical
            if (day == "yesterday" or (len(day) == 10 and day[4] == '-')) and 'final_score' in g:
                game_data["final_score"] = g['final_score']
                # Mark if user actually bet on this game
                # IMPORTANT: Only count ENANO bets ($2k+) for user betting record
                game_data["user_bet"] = g.get('user_bet', False)
                game_data["is_enano_bet"] = g.get('is_enano_bet', g.get('user_bet', False))  # Default: user_bet = ENANO bet for hardcoded data
                # Get user's bet type (OVER or UNDER) if they placed a bet
                user_bet_type = g.get('bet_type', '')
                game_data["bet_type"] = user_bet_type
                
                # IMPORTANT: Store bet-time line (line when bet was placed)
                # This is different from the closing line (g['total'])
                bet_time_line = g.get('bet_line')
                if bet_time_line:
                    game_data["bet_line"] = bet_time_line
                    # Calculate bet-time edge based on original line
                    bet_time_edge = abs(combined_ppg - bet_time_line)
                    game_data["bet_edge"] = round(bet_time_edge, 1)
                
                # Calculate if system recommendation hit (using closing line)
                if recommendation and g['final_score'] is not None and g.get('total'):
                    if recommendation == "OVER":
                        game_data["result_hit"] = g['final_score'] > g['total']
                    else:  # UNDER
                        game_data["result_hit"] = g['final_score'] < g['total']
                else:
                    game_data["result_hit"] = None
                
                # Calculate if USER's bet hit
                # CRITICAL: Use bet_line (line when bet was placed) for user bet evaluation
                if user_bet_type and g['final_score'] is not None:
                    # Use the bet_line if available, otherwise fall back to closing line
                    line_for_evaluation = bet_time_line if bet_time_line else g.get('total')
                    if line_for_evaluation:
                        final_score = g['final_score']
                        # Check for PUSH first
                        if final_score == line_for_evaluation:
                            game_data["user_bet_hit"] = None  # Push
                            game_data["result"] = "PUSH"
                            game_data["bet_result"] = "push"
                        elif user_bet_type.upper() == "OVER":
                            game_data["user_bet_hit"] = final_score > line_for_evaluation
                        elif user_bet_type.upper() == "UNDER":
                            game_data["user_bet_hit"] = final_score < line_for_evaluation
                        else:
                            game_data["user_bet_hit"] = None
                    else:
                        game_data["user_bet_hit"] = None
                else:
                    game_data["user_bet_hit"] = None
            
            # Calculate edge for all games
            # If bet is placed, use bet_line for edge; otherwise use current line
            # Positive edge = PPG > Line = OVER signal
            # Negative edge = PPG < Line = UNDER signal
            if game_data.get('has_bet') and game_data.get('bet_line'):
                edge = combined_ppg - game_data.get('bet_line') if has_line else 0
            else:
                edge = combined_ppg - g['total'] if has_line else 0
            game_data["edge"] = round(edge, 1) if has_line else None
            
            games.append(game_data)
            
            # Only add to plays if this game has an active bet
            if game_data.get("has_bet", False):
                # Calculate bet_edge using the line at which the bet was placed
                bet_line = game_data.get("bet_line")
                if bet_line:
                    bet_edge = combined_ppg - bet_line
                else:
                    bet_edge = edge  # fallback to current edge if no bet_line
                    
                plays.append({
                    "game": f"{g['away']} @ {g['home']}",
                    "total": g['total'],  # Current live line
                    "bet_line": bet_line,  # Line when bet was placed
                    "combined_ppg": round(combined_ppg, 1),
                    "edge": round(bet_edge, 1),  # Edge at bet time
                    "live_edge": round(edge, 1),  # Current live edge
                    "game_avg": round(game_avg, 1),
                    "recommendation": recommendation,
                    "color": color,
                    "has_bet": True,
                    "bet_type": game_data.get("bet_type"),
                    "bet_risk": game_data.get("bet_risk", 0),
                    "bet_count": game_data.get("bet_count", 0)
                })
        
        # Save to database
        await db.nba_opportunities.update_one(
            {"date": target_date},
            {"$set": {
                "date": target_date,
                "last_updated": datetime.now(arizona_tz).strftime('%I:%M %p'),
                "games": games,
                "plays": plays,
                "data_source": data_source
            }},
            upsert=True
        )
        
        return {
            "success": True,
            "message": f"Opportunities refreshed (source: {data_source})",
            "date": target_date,
            "last_updated": datetime.now(arizona_tz).strftime('%I:%M %p'),
            "games": games,
            "plays": plays,
            "data_source": data_source
        }
    except Exception as e:
        logger.error(f"Error refreshing opportunities: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============== NHL OPPORTUNITIES ==============

@api_router.get("/opportunities/nhl")
async def get_nhl_opportunities(day: str = "today"):
    """Get NHL betting opportunities. day parameter: 'yesterday', 'today', 'tomorrow', or a specific date 'YYYY-MM-DD'"""
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # Check if day is a specific date format (YYYY-MM-DD)
        if len(day) == 10 and day[4] == '-' and day[7] == '-':
            target_date = day
        elif day == "tomorrow":
            target_date = (datetime.now(arizona_tz) + timedelta(days=1)).strftime('%Y-%m-%d')
        elif day == "yesterday":
            target_date = (datetime.now(arizona_tz) - timedelta(days=1)).strftime('%Y-%m-%d')
        else:
            target_date = datetime.now(arizona_tz).strftime('%Y-%m-%d')
        
        # Get cached NHL opportunities
        cached = await db.nhl_opportunities.find_one({"date": target_date}, {"_id": 0})
        
        # Get compound record
        record = await db.compound_records.find_one({"league": "NHL"}, {"_id": 0})
        compound_record = {
            "hits": record.get('hits', 0) if record else 0,
            "misses": record.get('misses', 0) if record else 0
        }
        
        if cached and cached.get('games'):
            return {
                "success": True,
                "date": target_date,
                "last_updated": cached.get('last_updated'),
                "games": cached.get('games', []),
                "plays": cached.get('plays', []),
                "compound_record": compound_record,
                "data_source": cached.get('data_source', 'hardcoded'),
                "actual_bet_record": cached.get('actual_bet_record')
            }
        
        return {
            "success": True,
            "date": target_date,
            "message": "No NHL opportunities data yet. Click refresh to load games.",
            "games": [],
            "plays": [],
            "compound_record": compound_record,
            "data_source": None
        }
    except Exception as e:
        logger.error(f"Error getting NHL opportunities: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/opportunities/nhl/refresh")
async def refresh_nhl_opportunities(day: str = "today", use_live_lines: bool = False):
    """Manually refresh NHL opportunities data. 
    day parameter: 'yesterday', 'today' or 'tomorrow'
    use_live_lines: if True, fetch O/U lines from plays888.co instead of hardcoded values
    
    #3 PROCESS: After 5am Arizona time, TODAY's data automatically uses Plays888 for live lines
    """
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        now_arizona = datetime.now(arizona_tz)
        current_hour = now_arizona.hour
        
        # #3 PROCESS: Automatically use Plays888 for TODAY after 5am Arizona time
        if day == "today" and current_hour >= 5:
            use_live_lines = True
            logger.info(f"[#3 Process] NHL: After 5am Arizona ({now_arizona.strftime('%I:%M %p')}), using Plays888 for live lines")
        
        if day == "tomorrow":
            target_date = (now_arizona + timedelta(days=1)).strftime('%Y-%m-%d')
        elif day == "yesterday":
            target_date = (now_arizona - timedelta(days=1)).strftime('%Y-%m-%d')
        else:
            target_date = now_arizona.strftime('%Y-%m-%d')
        
        # NHL GPG Rankings (Season) - from ESPN data (Dec 29, 2025)
        gpg_season = {
            'Colorado': 1, 'Dallas': 2, 'Edmonton': 3, 'Anaheim': 4, 'Carolina': 5,
            'Montreal': 6, 'Tampa Bay': 7, 'Ottawa': 8, 'Toronto': 9, 'Washington': 10,
            'Vegas': 11, 'Florida': 12, 'Pittsburgh': 13, 'Boston': 14, 'Detroit': 15,
            'Buffalo': 16, 'Minnesota': 17, 'San Jose': 18, 'Utah': 19, 'Columbus': 20,
            'Winnipeg': 21, 'Philadelphia': 22, 'Nashville': 23, 'Vancouver': 24, 'NY Islanders': 25,
            'Chicago': 26, 'New Jersey': 27, 'Calgary': 28, 'Los Angeles': 29, 'Seattle': 30,
            'NY Rangers': 31, 'St. Louis': 32
        }
        
        # NHL Goals Last 3 Games Rankings - from StatMuse data (Dec 29, 2025)
        gpg_last3 = {
            'Toronto': 1, 'Vegas': 2, 'Montreal': 3, 'Ottawa': 4, 'Pittsburgh': 5,
            'Tampa Bay': 6, 'Colorado': 7, 'Dallas': 8, 'Edmonton': 9, 'Carolina': 10,
            'Buffalo': 11, 'San Jose': 12, 'Columbus': 13, 'Calgary': 14, 'Seattle': 15,
            'St. Louis': 16, 'Washington': 17, 'Florida': 18, 'Detroit': 19, 'Philadelphia': 20,
            'Vancouver': 21, 'Los Angeles': 22, 'Winnipeg': 23, 'NY Rangers': 24, 'Minnesota': 25,
            'Nashville': 26, 'Chicago': 27, 'Anaheim': 28, 'NY Islanders': 29, 'Boston': 30,
            'Utah': 31, 'New Jersey': 32
        }
        
        # Actual GPG values (Season) - from ESPN (Dec 29, 2025)
        # GPG Season values - from ESPN (Dec 30, 2025) https://www.espn.com/nhl/stats/team
        gpg_season_values = {
            'Colorado': 4.00, 'Dallas': 3.49, 'Edmonton': 3.38, 'Anaheim': 3.33, 'Carolina': 3.29,
            'Montreal': 3.29, 'Tampa Bay': 3.29, 'Toronto': 3.25, 'Ottawa': 3.21, 'Florida': 3.21,
            'Washington': 3.18, 'Pittsburgh': 3.14, 'Vegas': 3.14, 'Buffalo': 3.08, 'Detroit': 3.08,
            'Minnesota': 3.08, 'San Jose': 3.05, 'Boston': 3.05, 'Utah': 2.98, 'Columbus': 2.95,
            'Philadelphia': 2.89, 'Winnipeg': 2.86, 'Nashville': 2.82, 'NY Islanders': 2.77,
            'Chicago': 2.76, 'Vancouver': 2.76, 'New Jersey': 2.71, 'Calgary': 2.59, 'Los Angeles': 2.58,
            'Seattle': 2.57, 'NY Rangers': 2.54, 'St. Louis': 2.50
        }
        
        # GPG Last 3 Games values - from StatMuse (Dec 30, 2025) https://www.statmuse.com/nhl/ask/nhl-most-team-goals-per-game-last-three-games
        gpg_last3_values = {
            'Toronto': 5.00, 'Vegas': 4.67, 'San Jose': 4.33, 'Montreal': 4.33, 'Pittsburgh': 4.33,
            'Florida': 4.00, 'Tampa Bay': 4.00, 'Buffalo': 3.67, 'Colorado': 3.67, 'Minnesota': 3.67,
            'Columbus': 3.67, 'Dallas': 3.67, 'NY Islanders': 3.67, 'Los Angeles': 3.33, 'Washington': 3.33,
            'Edmonton': 3.33, 'Carolina': 3.33, 'NY Rangers': 3.00, 'Philadelphia': 2.94, 'Vancouver': 2.89,
            'New Jersey': 2.20, 'Chicago': 2.00, 'Anaheim': 2.00, 'Nashville': 2.00, 'Boston': 2.00,
            'Utah': 2.00, 'Ottawa': 2.00, 'Winnipeg': 2.00, 'Calgary': 2.00, 'Seattle': 2.00,
            'Detroit': 2.00, 'St. Louis': 2.00
        }
        
        games_raw = []
        data_source = "hardcoded"
        open_bets = []
        
        # #3 PROCESS: After 5am Arizona time, AUTOMATICALLY use Plays888 for live lines
        arizona_tz = ZoneInfo('America/Phoenix')
        current_hour = datetime.now(arizona_tz).hour
        auto_use_live = current_hour >= 5 and day == "today"
        
        if (use_live_lines or auto_use_live) and day == "today":
            logger.info(f"[#3 Process] NHL: After 5am Arizona ({datetime.now(arizona_tz).strftime('%I:%M %p')}), using Plays888 for live lines")
            try:
                # Use ENANO (jac075) account for scraping - this is the primary betting account
                enano_conn = await db.connections.find_one({"username": "jac075"}, {"_id": 0})
                if enano_conn and enano_conn.get("is_connected"):
                    username = enano_conn["username"]
                    password = decrypt_password(enano_conn["password_encrypted"])
                    
                    # Create new scraper instance
                    scraper = Plays888Service()
                    await scraper.login(username, password)
                    live_games = await scraper.scrape_totals("NHL")
                    
                    # Also fetch open bets
                    open_bets = await scraper.scrape_open_bets()
                    await scraper.close()
                    
                    if live_games:
                        # Convert plays888 data to our format
                        for game in live_games:
                            away_short = convert_plays888_team_name(game.get('away', ''))
                            home_short = convert_plays888_team_name(game.get('home', ''))
                            games_raw.append({
                                "time": game.get('time', ''),
                                "away": away_short,
                                "home": home_short,
                                "total": game.get('total', 6.0)
                            })
                        data_source = "plays888.co"
                        logger.info(f"Fetched {len(games_raw)} NHL games from plays888.co")
            except Exception as e:
                logger.error(f"Error fetching live NHL lines: {e}")
        
        # Fallback: Load open bets from database if scraping didn't find any
        if not open_bets:
            db_bets = await db.open_bets.find({"sport": "NHL"}, {"_id": 0}).to_list(100)
            if db_bets:
                open_bets = db_bets
                logger.info(f"Loaded {len(open_bets)} NHL open bets from database")
        
        # #3.85: MERGE live games with cached games to keep started games visible ALL DAY
        if games_raw and day == "today":
            # #3.85: For NHL, DON'T merge with cached games - only show what's on Plays888
            # The old merge was adding games from yesterday's "tomorrow" cache
            # If a game has started and has a bet, it will be added from the open_bets list above
            if games_raw:
                logger.info(f"[NHL] Showing {len(games_raw)} games from Plays888 (no cache merge)")
        
        # Use database cache if live fetch failed or wasn't requested
        if not games_raw:
            # First try to load from database cache (from last night's 8pm scrape)
            cached = await db.nhl_opportunities.find_one({"date": target_date}, {"_id": 0})
            if cached and cached.get('games'):
                for g in cached['games']:
                    games_raw.append({
                        "time": g.get('time', ''),
                        "away": g.get('away_team', g.get('away', '')),
                        "home": g.get('home_team', g.get('home', '')),
                        "total": g.get('total')
                    })
                data_source = "cached (from last night's scrape)"
                logger.info(f"Using {len(games_raw)} cached NHL games from database for {target_date}")
        
        # Add games from open bets that might have started but aren't in the schedule
        # This ensures we show all games with active bets
        # BUT we need to avoid duplicates - games already scraped from Plays888
        if day == "today":
            # Create comprehensive matching keys for existing games
            # Use both city names and team nicknames for matching
            nhl_team_nicknames = {
                'NY ISLANDERS': ['ISLANDERS', 'NY ISLANDERS', 'NEW YORK ISLANDERS'],
                'NY RANGERS': ['RANGERS', 'NY RANGERS', 'NEW YORK RANGERS'],
                'COLUMBUS': ['BLUE JACKETS', 'JACKETS', 'COLUMBUS'],
                'TORONTO': ['MAPLE LEAFS', 'LEAFS', 'TORONTO'],
                'DETROIT': ['RED WINGS', 'WINGS', 'DETROIT'],
                'PITTSBURGH': ['PENGUINS', 'PITTSBURGH'],
                'CHICAGO': ['BLACKHAWKS', 'HAWKS', 'CHICAGO'],
                'PHILADELPHIA': ['FLYERS', 'PHILADELPHIA'],
                'SEATTLE': ['KRAKEN', 'SEATTLE'],
                'MONTREAL': ['CANADIENS', 'HABS', 'MONTREAL'],
                'TAMPA BAY': ['LIGHTNING', 'BOLTS', 'TAMPA BAY', 'TAMPA'],
                'BOSTON': ['BRUINS', 'BOSTON'],
                'BUFFALO': ['SABRES', 'BUFFALO'],
                'CAROLINA': ['HURRICANES', 'CANES', 'CAROLINA'],
                'OTTAWA': ['SENATORS', 'SENS', 'OTTAWA'],
                'NEW JERSEY': ['DEVILS', 'NEW JERSEY'],
                'DALLAS': ['STARS', 'DALLAS'],
                'NASHVILLE': ['PREDATORS', 'PREDS', 'NASHVILLE'],
                'ST. LOUIS': ['BLUES', 'ST. LOUIS', 'ST LOUIS'],
                'ANAHEIM': ['DUCKS', 'ANAHEIM'],
                'LOS ANGELES': ['KINGS', 'LOS ANGELES', 'LA KINGS'],
                'COLORADO': ['AVALANCHE', 'AVS', 'COLORADO'],
                'VEGAS': ['GOLDEN KNIGHTS', 'KNIGHTS', 'VEGAS'],
                'EDMONTON': ['OILERS', 'EDMONTON'],
                'CALGARY': ['FLAMES', 'CALGARY'],
                'SAN JOSE': ['SHARKS', 'SAN JOSE'],
                'VANCOUVER': ['CANUCKS', 'VANCOUVER'],
                'WASHINGTON': ['CAPITALS', 'CAPS', 'WASHINGTON'],
                'FLORIDA': ['PANTHERS', 'FLORIDA'],
                'MINNESOTA': ['WILD', 'MINNESOTA'],
                'WINNIPEG': ['JETS', 'WINNIPEG'],
                'UTAH': ['MAMMOTH', 'UTAH']
            }
            
            def normalize_team_name(team_name):
                """Normalize team name to a standard key for matching"""
                team_upper = team_name.upper().strip()
                # Check if it matches any known team
                for standard_name, aliases in nhl_team_nicknames.items():
                    if team_upper in aliases or standard_name in team_upper:
                        return standard_name
                    # Also check if any alias is contained in the team name
                    for alias in aliases:
                        if alias in team_upper or team_upper in alias:
                            return standard_name
                return team_upper
            
            existing_matchups = set()
            for g in games_raw:
                away_norm = normalize_team_name(g['away'])
                home_norm = normalize_team_name(g['home'])
                existing_matchups.add(f"{away_norm}:{home_norm}")
                logger.debug(f"Existing game: {g['away']} @ {g['home']} -> {away_norm}:{home_norm}")
            
            logger.info(f"NHL: {len(games_raw)} games from Plays888, checking {len(open_bets)} open bets for duplicates")
            
            added_from_bets = set()  # Track what we've added from bets to avoid duplicates
            for bet in open_bets:
                if bet.get('sport') == 'NHL':
                    bet_away = bet.get('away_team', '')
                    bet_home = bet.get('home_team', '')
                    
                    away_norm = normalize_team_name(bet_away)
                    home_norm = normalize_team_name(bet_home)
                    matchup_key = f"{away_norm}:{home_norm}"
                    
                    if matchup_key in existing_matchups:
                        logger.debug(f"Skipping duplicate bet game: {bet_away} @ {bet_home} (already in schedule)")
                        continue
                    
                    if matchup_key in added_from_bets:
                        logger.debug(f"Skipping duplicate bet: {bet_away} @ {bet_home} (already added from bets)")
                        continue
                    
                    # This bet's game isn't in the schedule - add it
                    # Map team names back to short form
                    team_map = {
                        'PENGUINS': 'Pittsburgh', 'MAPLE LEAFS': 'Toronto', 'LEAFS': 'Toronto',
                        'STARS': 'Dallas', 'RED WINGS': 'Detroit', 'WINGS': 'Detroit',
                        'RANGERS': 'NY Rangers', 'CAPITALS': 'Washington',
                        'PANTHERS': 'Florida', 'HURRICANES': 'Carolina',
                        'DEVILS': 'New Jersey', 'ISLANDERS': 'NY Islanders',
                        'SABRES': 'Buffalo', 'SENATORS': 'Ottawa',
                        'CANADIENS': 'Montreal', 'BRUINS': 'Boston',
                        'PREDATORS': 'Nashville', 'WILD': 'Minnesota',
                        'FLYERS': 'Philadelphia', 'BLACKHAWKS': 'Chicago',
                        'MAMMOTH': 'Utah', 'AVALANCHE': 'Colorado',
                        'FLAMES': 'Calgary', 'OILERS': 'Edmonton',
                        'SHARKS': 'San Jose', 'KNIGHTS': 'Vegas', 'GOLDEN KNIGHTS': 'Vegas',
                        'KRAKEN': 'Seattle', 'KINGS': 'Los Angeles', 'DUCKS': 'Anaheim',
                        'BLUE JACKETS': 'Columbus', 'JACKETS': 'Columbus',
                        'LIGHTNING': 'Tampa Bay', 'JETS': 'Winnipeg', 'CANUCKS': 'Vancouver',
                        'BLUES': 'St. Louis'
                    }
                    bet_away_last = bet_away.upper().split()[-1] if bet_away else ''
                    bet_home_last = bet_home.upper().split()[-1] if bet_home else ''
                    away_short = team_map.get(bet_away_last, bet_away_last.title())
                    home_short = team_map.get(bet_home_last, bet_home_last.title())
                    
                    # Use bet line if available
                    bet_line = bet.get('total_line', 6.0)
                    
                    games_raw.append({
                        "time": "In Progress",
                        "away": away_short,
                        "home": home_short,
                        "total": bet_line
                    })
                    added_from_bets.add(matchup_key)
                    logger.info(f"Added missing bet game: {away_short} @ {home_short} (line: {bet_line})")
        
        # Calculate averages and recommendations
        games = []
        plays = []
        
        for i, g in enumerate(games_raw, 1):
            away_season = gpg_season.get(g['away'], 16)
            away_last3 = gpg_last3.get(g['away'], 16)
            away_avg = (away_season + away_last3) / 2
            
            home_season = gpg_season.get(g['home'], 16)
            home_last3 = gpg_last3.get(g['home'], 16)
            home_avg = (home_season + home_last3) / 2
            
            game_avg = (away_avg + home_avg) / 2
            
            # Calculate combined GPG (actual goals expected in the game)
            away_season_gpg = gpg_season_values.get(g['away'], 3.0)
            away_last3_gpg = gpg_last3_values.get(g['away'], 3.0)
            home_season_gpg = gpg_season_values.get(g['home'], 3.0)
            home_last3_gpg = gpg_last3_values.get(g['home'], 3.0)
            
            # Combined GPG = (away_season + home_season + away_l3 + home_l3) / 2
            # This gives expected total goals in the game
            combined_gpg = (away_season_gpg + away_last3_gpg + home_season_gpg + home_last3_gpg) / 2
            
            # Check if we have a valid line from plays888.co
            has_line = g.get('total') and g['total'] > 0
            
            # SIMPLIFIED LOGIC: Determine recommendation based on GPG vs Line comparison
            # If GPG average > Line → OVER (we expect more goals than the line)
            # If GPG average < Line → UNDER (we expect fewer goals than the line)
            recommendation = None
            color = "neutral"
            
            if has_line:
                edge_value = combined_gpg - g['total']  # Positive = OVER, Negative = UNDER
                # Round to 1 decimal place to avoid floating point precision issues
                edge_value = round(edge_value, 1)
                
                # Recommend based on which side has the edge
                # Edge must be at least 0.5 goals to make a recommendation (NHL)
                if edge_value >= 0.5:  # GPG is significantly higher than line
                    recommendation = "OVER"
                    color = "green"
                elif edge_value <= -0.5:  # GPG is significantly lower than line
                    recommendation = "UNDER"
                    color = "red"
            
            game_data = {
                "game_num": i,
                "time": g['time'],
                "away_team": g['away'],
                "away_gpg_rank": away_season,
                "away_last3_rank": away_last3,
                "away_avg": round(away_avg, 1),
                "home_team": g['home'],
                "home_gpg_rank": home_season,
                "home_last3_rank": home_last3,
                "home_avg": round(home_avg, 1),
                "total": g['total'] if has_line else None,  # Current live line from Plays888
                "opening_line": None,  # Will be set from database (8pm scrape)
                "has_line": has_line,
                "combined_gpg": round(combined_gpg, 1),
                "game_avg": round(game_avg, 1),
                "recommendation": recommendation,
                "color": color,
                "has_bet": False,
                "bet_type": None,
                "bet_risk": 0,
                "bet_count": 0
            }
            
            # #3.75: Get opening line from database (from last night's 8pm scrape)
            # This allows comparison: Opening Line vs Current Live Line vs Bet Line
            if day == "today":
                try:
                    cached_today = await db.nhl_opportunities.find_one({"date": target_date}, {"_id": 0})
                    if cached_today and cached_today.get('games'):
                        for cached_game in cached_today['games']:
                            cached_away = (cached_game.get('away_team') or cached_game.get('away', '')).upper()
                            cached_home = (cached_game.get('home_team') or cached_game.get('home', '')).upper()
                            if g['away'].upper() in cached_away or cached_away in g['away'].upper():
                                if g['home'].upper() in cached_home or cached_home in g['home'].upper():
                                    # Found matching game - get opening line
                                    opening = cached_game.get('opening_line') or cached_game.get('total')
                                    if opening:
                                        game_data["opening_line"] = opening
                                    break
                except Exception as e:
                    logger.debug(f"Could not get opening line for {g['away']} @ {g['home']}: {e}")
            
            # Add dots for NHL
            def get_nhl_dot_color(rank: int) -> str:
                if rank <= 8:
                    return "🟢"  # Green: Top tier
                elif rank <= 16:
                    return "🔵"  # Blue: Upper middle
                elif rank <= 24:
                    return "🟡"  # Yellow: Lower middle
                else:
                    return "🔴"  # Red: Bottom tier
            
            game_data["dots"] = f"{get_nhl_dot_color(away_season)}{get_nhl_dot_color(away_last3)}{get_nhl_dot_color(home_season)}{get_nhl_dot_color(home_last3)}"
            game_data["away_dots"] = f"{get_nhl_dot_color(away_season)}{get_nhl_dot_color(away_last3)}"
            game_data["home_dots"] = f"{get_nhl_dot_color(home_season)}{get_nhl_dot_color(home_last3)}"
            
            # Check if this game has an active bet
            # Also detect "hedged" bets (both OVER and UNDER on same game = cancelled out)
            game_bets = []
            for bet in open_bets:
                if bet.get('sport') == 'NHL':
                    # Match team names (case-insensitive partial match)
                    bet_away = bet.get('away_team', '').upper()
                    bet_home = bet.get('home_team', '').upper()
                    game_away = g['away'].upper()
                    game_home = g['home'].upper()
                    
                    # Check if teams match (partial match for city names)
                    away_match = any(part in bet_away for part in game_away.split()) or any(part in game_away for part in bet_away.split())
                    home_match = any(part in bet_home for part in game_home.split()) or any(part in game_home for part in bet_home.split())
                    
                    if away_match and home_match:
                        game_bets.append(bet)
            
            # Check if game is hedged (has both OVER and UNDER bets)
            bet_types = [b.get('bet_type', '').upper() for b in game_bets]
            is_hedged = 'OVER' in bet_types and 'UNDER' in bet_types
            
            if game_bets and not is_hedged:
                # Game has active bet(s) that are not hedged
                game_data["has_bet"] = True
                game_data["bet_type"] = game_bets[0].get('bet_type')
                game_data["bet_risk"] = sum(b.get('total_risk', b.get('risk', 0)) for b in game_bets)
                game_data["bet_count"] = sum(b.get('bet_count', 1) for b in game_bets)
                # Store the line at which the bet was placed
                game_data["bet_line"] = game_bets[0].get('total_line')
            elif is_hedged:
                # Game is hedged (OVER + UNDER = push/cancelled)
                game_data["has_bet"] = False
                game_data["is_hedged"] = True
                game_data["bet_type"] = "HEDGED"
                game_data["bet_risk"] = 0
                game_data["bet_count"] = 0
                game_data["bet_line"] = None
            
            # Add result data for yesterday/historical
            if (day == "yesterday" or (len(day) == 10 and day[4] == '-')) and 'final_score' in g:
                game_data["final_score"] = g['final_score']
                # Mark if user actually bet on this game
                game_data["user_bet"] = g.get('user_bet', False)
                # Get user's bet type (OVER or UNDER) if they placed a bet
                user_bet_type = g.get('bet_type', '')
                game_data["bet_type"] = user_bet_type
                
                # Calculate if system recommendation hit (only if we have final score)
                if recommendation and g['final_score'] is not None:
                    if recommendation == "OVER":
                        game_data["result_hit"] = g['final_score'] > g['total']
                    else:  # UNDER
                        game_data["result_hit"] = g['final_score'] < g['total']
                else:
                    game_data["result_hit"] = None
                
                # Calculate if USER's bet hit (based on their actual bet direction)
                if user_bet_type and g['final_score'] is not None:
                    final_score = g['final_score']
                    bet_line = g.get('bet_line') or g['total']
                    # Check for PUSH first
                    if final_score == bet_line:
                        game_data["user_bet_hit"] = None  # Push
                        game_data["result"] = "PUSH"
                        game_data["bet_result"] = "push"
                    elif user_bet_type.upper() == "OVER":
                        game_data["user_bet_hit"] = final_score > bet_line
                    elif user_bet_type.upper() == "UNDER":
                        game_data["user_bet_hit"] = final_score < bet_line
                    else:
                        game_data["user_bet_hit"] = None
                else:
                    game_data["user_bet_hit"] = None
            
            # Calculate edge for ALL games (for the table)
            # If bet is placed, use bet_line for edge; otherwise use current line
            # Positive edge = GPG > Line = OVER signal
            # Negative edge = GPG < Line = UNDER signal
            if game_data.get('has_bet') and game_data.get('bet_line'):
                edge = combined_gpg - game_data.get('bet_line') if has_line else 0
            else:
                edge = combined_gpg - g['total'] if has_line else 0
            game_data["edge"] = round(edge, 1) if has_line else None
            
            games.append(game_data)
            
            # Only add to plays if this game has an active bet (and not already in plays)
            if game_data.get("has_bet", False) and has_line:
                game_key = f"{g['away']} @ {g['home']}"
                # Check if this game is already in plays to avoid duplicates
                if not any(p.get('game') == game_key for p in plays):
                    # Calculate bet_edge using the line at which the bet was placed
                    bet_line = game_data.get("bet_line")
                    if bet_line:
                        bet_edge = combined_gpg - bet_line
                    else:
                        bet_edge = edge  # fallback to current edge if no bet_line
                        
                    plays.append({
                        "game": game_key,
                        "total": g['total'],  # Current live line
                        "bet_line": bet_line,  # Line when bet was placed
                        "combined_gpg": round(combined_gpg, 1),
                        "edge": round(bet_edge, 1),  # Edge at bet time
                        "live_edge": round(edge, 1),  # Current live edge
                        "game_avg": round(game_avg, 1),
                        "recommendation": recommendation,
                        "color": color,
                        "has_bet": True,
                        "bet_type": game_data.get("bet_type"),
                        "bet_risk": game_data.get("bet_risk", 0),
                        "bet_count": game_data.get("bet_count", 0)
                    })
        
        # Save to database
        await db.nhl_opportunities.update_one(
            {"date": target_date},
            {"$set": {
                "date": target_date,
                "last_updated": datetime.now(arizona_tz).strftime('%I:%M %p'),
                "games": games,
                "plays": plays,
                "data_source": data_source
            }},
            upsert=True
        )
        
        return {
            "success": True,
            "message": f"NHL opportunities refreshed (source: {data_source})",
            "date": target_date,
            "last_updated": datetime.now(arizona_tz).strftime('%I:%M %p'),
            "games": games,
            "plays": plays,
            "data_source": data_source
        }
    except Exception as e:
        logger.error(f"Error refreshing NHL opportunities: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ================= NCAAB OPPORTUNITIES =================

@api_router.get("/opportunities/ncaab")
async def get_ncaab_opportunities_endpoint(day: str = "today"):
    """Get NCAAB (NCAA Basketball) betting opportunities. day parameter: 'yesterday', 'today', 'tomorrow', or a specific date 'YYYY-MM-DD'"""
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # Check if day is a specific date format (YYYY-MM-DD)
        if len(day) == 10 and day[4] == '-' and day[7] == '-':
            target_date = day
        elif day == "tomorrow":
            target_date = (datetime.now(arizona_tz) + timedelta(days=1)).strftime('%Y-%m-%d')
        elif day == "yesterday":
            target_date = (datetime.now(arizona_tz) - timedelta(days=1)).strftime('%Y-%m-%d')
        else:
            target_date = datetime.now(arizona_tz).strftime('%Y-%m-%d')
        
        # Get cached NCAAB opportunities
        cached = await db.ncaab_opportunities.find_one({"date": target_date}, {"_id": 0})
        
        # Get compound record
        record = await db.compound_records.find_one({"league": "NCAAB"}, {"_id": 0})
        compound_record = {
            "hits": record.get('hits', 0) if record else 0,
            "misses": record.get('misses', 0) if record else 0
        }
        
        if cached and cached.get('games'):
            return {
                "success": True,
                "date": target_date,
                "last_updated": cached.get('last_updated'),
                "games": cached.get('games', []),
                "plays": cached.get('plays', []),
                "compound_record": compound_record,
                "data_source": cached.get('data_source', 'scraped'),
                "actual_bet_record": cached.get('actual_bet_record')
            }
        
        return {
            "success": True,
            "date": target_date,
            "message": "No NCAAB opportunities data yet. Click refresh to load games.",
            "games": [],
            "plays": [],
            "compound_record": compound_record,
            "data_source": None
        }
    except Exception as e:
        logger.error(f"Error getting NCAAB opportunities: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/opportunities/ncaab/refresh")
async def refresh_ncaab_opportunities(day: str = "today"):
    """Manually refresh NCAAB opportunities data. 
    day parameter: 'yesterday', 'today' or 'tomorrow'
    """
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        now_arizona = datetime.now(arizona_tz)
        
        if day == "tomorrow":
            target_date = (now_arizona + timedelta(days=1)).strftime('%Y-%m-%d')
        elif day == "yesterday":
            target_date = (now_arizona - timedelta(days=1)).strftime('%Y-%m-%d')
        else:
            target_date = now_arizona.strftime('%Y-%m-%d')
        
        logger.info(f"Refreshing NCAAB opportunities for {target_date}")
        
        # Scrape NCAAB PPG rankings from teamrankings.com with target date (pre-game values)
        ppg_data = await scrape_ncaab_ppg_rankings(target_date)
        
        logger.info(f"Scraped NCAAB PPG data for {len(ppg_data.get('season_ranks', {}))} teams")
        
        # Scrape games from scoresandodds.com
        games = await scrape_scoresandodds("NCAAB", target_date)
        
        logger.info(f"Scraped {len(games)} NCAAB games from scoresandodds.com")
        
        # Process games and add PPG analysis
        processed_games = []
        
        # NCAAB has ~365 teams, so ranking thresholds are different
        # Top 25% = rank 1-91 (Green)
        # 25-50% = rank 92-182 (Yellow) 
        # 50-75% = rank 183-273 (Red)
        # Bottom 25% = rank 274-365 (Blue)
        def get_ncaab_dot_color(rank):
            if rank is None:
                return '⚪'  # Unknown - not in top 365
            if rank <= 92:
                return '🟢'  # Top tier (1-92)
            elif rank <= 184:
                return '🔵'  # Second tier (93-184)
            elif rank <= 276:
                return '🟡'  # Third tier (185-276)
            elif rank <= 365:
                return '🔴'  # Fourth tier (277-365)
            else:
                return '⚪'  # Unknown - not in top 365
        
        for game in games:
            away_team = game.get('away', '')
            home_team = game.get('home', '')
            
            # Try to find team in PPG data (NCAAB teams have varied names)
            def find_team_ppg(team_name, ppg_dict):
                """Find team PPG using fuzzy matching"""
                if not team_name:
                    return None
                # Direct match
                if team_name in ppg_dict:
                    return ppg_dict[team_name]
                # Try lowercase match
                team_lower = team_name.lower()
                for k, v in ppg_dict.items():
                    if k.lower() == team_lower or team_lower in k.lower() or k.lower() in team_lower:
                        return v
                return None
            
            def find_team_rank(team_name, rank_dict):
                """Find team rank using fuzzy matching"""
                if not team_name:
                    return None
                # Direct match
                if team_name in rank_dict:
                    return rank_dict[team_name]
                # Try lowercase match
                team_lower = team_name.lower()
                for k, v in rank_dict.items():
                    if k.lower() == team_lower or team_lower in k.lower() or k.lower() in team_lower:
                        return v
                return None
            
            # Get PPG data with fuzzy matching
            away_ppg_rank = find_team_rank(away_team, ppg_data['season_ranks'])
            away_ppg_value = find_team_ppg(away_team, ppg_data['season_values'])
            away_last3_rank = find_team_rank(away_team, ppg_data['last3_ranks'])
            away_last3_value = find_team_ppg(away_team, ppg_data['last3_values'])
            
            home_ppg_rank = find_team_rank(home_team, ppg_data['season_ranks'])
            home_ppg_value = find_team_ppg(home_team, ppg_data['season_values'])
            home_last3_rank = find_team_rank(home_team, ppg_data['last3_ranks'])
            home_last3_value = find_team_ppg(home_team, ppg_data['last3_values'])
            
            # Calculate combined PPG using the correct formula:
            # (Team1 Season PPG + Team2 Season PPG + Team1 L3 PPG + Team2 L3 PPG) / 2
            combined_ppg = None
            if away_ppg_value and home_ppg_value and away_last3_value and home_last3_value:
                combined_ppg = round((away_ppg_value + home_ppg_value + away_last3_value + home_last3_value) / 2, 1)
            elif away_ppg_value and home_ppg_value:
                # Fallback to season only if L3 not available
                combined_ppg = round((away_ppg_value + home_ppg_value), 1)
            
            # Calculate edge (same as NBA)
            edge = None
            line = game.get('total')
            if combined_ppg and line:
                try:
                    edge = round(combined_ppg - float(line), 1)
                except:
                    pass
            
            # Generate recommendation (NCAAB uses edge threshold of 9)
            recommendation = ''
            if edge is not None:
                if edge >= 10:
                    recommendation = 'OVER'
                elif edge <= -9:
                    recommendation = 'UNDER'
            
            # Generate dot colors
            away_dots = get_ncaab_dot_color(away_ppg_rank) + get_ncaab_dot_color(away_last3_rank)
            home_dots = get_ncaab_dot_color(home_ppg_rank) + get_ncaab_dot_color(home_last3_rank)
            
            processed_game = {
                **game,
                'away_team': away_team,
                'home_team': home_team,
                'away_ppg_rank': away_ppg_rank,
                'away_ppg_value': away_ppg_value,
                'away_last3_rank': away_last3_rank,
                'away_last3_value': away_last3_value,
                'home_ppg_rank': home_ppg_rank,
                'home_ppg_value': home_ppg_value,
                'home_last3_rank': home_last3_rank,
                'home_last3_value': home_last3_value,
                'combined_ppg': combined_ppg,
                'edge': edge,
                'recommendation': recommendation,
                'away_dots': away_dots,
                'home_dots': home_dots,
                'opening_line': line
            }
            processed_games.append(processed_game)
        
        # #3.85 / #3.90: Preserve existing plays and merge games
        existing = await db.ncaab_opportunities.find_one({"date": target_date}, {"_id": 0})
        existing_plays = existing.get('plays', []) if existing else []
        existing_games = existing.get('games', []) if existing else []
        
        # #3 PROCESS: Fetch open bets from Plays888 for NCAAB
        open_bets = []
        if day == "today":
            try:
                enano_conn = await db.connections.find_one({"username": "jac075"}, {"_id": 0})
                if enano_conn and enano_conn.get("is_connected"):
                    scraper = Plays888Service()
                    await scraper.initialize()
                    await scraper.login("jac075", decrypt_password(enano_conn["password_encrypted"]))
                    all_bets = await scraper.scrape_open_bets()
                    await scraper.close()
                    
                    open_bets = [b for b in all_bets if b.get('sport') == 'NCAAB']
                    logger.info(f"[#3 Process] Found {len(open_bets)} NCAAB open bets from Plays888")
            except Exception as e:
                logger.error(f"Error fetching NCAAB open bets: {e}")
        
        # Process open bets into plays
        new_plays = []
        for bet in open_bets:
            away_team = bet.get('away_team', '')
            home_team = bet.get('home_team', '')
            bet_type = bet.get('bet_type', '')
            
            # For spread bets, the bet_type is like "CALIFORNIA +9"
            # For total bets, the bet_type is like "OVER" or "UNDER"
            is_spread = bet.get('is_spread', False)
            
            play = {
                'game': f"{away_team} @ {home_team}",
                'away_team': away_team,
                'home_team': home_team,
                'bet_type': bet_type,
                'bet_line': bet.get('spread_line') if is_spread else bet.get('total_line'),
                'bet_count': 1,
                'is_spread': is_spread,
                'has_bet': True
            }
            new_plays.append(play)
        
        # Merge new plays with existing plays (avoid duplicates)
        merged_plays = existing_plays.copy()
        for new_play in new_plays:
            # Check if this bet already exists
            is_dup = False
            for existing_play in merged_plays:
                if (existing_play.get('bet_type') == new_play.get('bet_type') and 
                    existing_play.get('game') == new_play.get('game')):
                    is_dup = True
                    break
            if not is_dup:
                merged_plays.append(new_play)
        
        logger.info(f"[#3.5] NCAAB: {len(existing_plays)} existing plays + {len(new_plays)} new plays = {len(merged_plays)} total plays (after dedup)")
        
        # Create lookup of new games
        new_matchups = set()
        for g in processed_games:
            key = f"{g.get('away', g.get('away_team', '')).upper()}_{g.get('home', g.get('home_team', '')).upper()}"
            new_matchups.add(key)
        
        # #3.85/#3.90: Add existing games that are NOT in the new list (started games)
        added_started = 0
        for existing_game in existing_games:
            away = existing_game.get('away', existing_game.get('away_team', ''))
            home = existing_game.get('home', existing_game.get('home_team', ''))
            key = f"{away.upper()}_{home.upper()}"
            
            if key not in new_matchups:
                # This game has started or isn't in new scrape - keep it from cache
                existing_game['started'] = True
                processed_games.append(existing_game)
                added_started += 1
        
        if added_started > 0:
            logger.info(f"[#3.85/#3.90] Preserved {added_started} started NCAAB games (total: {len(processed_games)})")
        
        # Preserve bet info on games
        for play in merged_plays:
            play_game = play.get('game', '')
            if ' @ ' not in play_game:
                continue
            play_away = play_game.split(' @ ')[0].lower()
            play_home = play_game.split(' @ ')[1].lower()
            
            for g in processed_games:
                g_away = (g.get('away', g.get('away_team', ''))).lower()
                g_home = (g.get('home', g.get('home_team', ''))).lower()
                if (play_away in g_away or g_away in play_away) and \
                   (play_home in g_home or g_home in play_home):
                    g['has_bet'] = True
                    g['bet_type'] = play.get('bet_type')
                    g['bet_count'] = play.get('bet_count', 1)
                    g['bet_line'] = play.get('bet_line')
                    break
        
        # Save to database
        doc = {
            "date": target_date,
            "games": processed_games,
            "plays": merged_plays,  # #3.85: Preserve existing plays + new open bets
            "last_updated": datetime.now(timezone.utc).isoformat(),
            "data_source": "scraped"
        }
        
        await db.ncaab_opportunities.replace_one(
            {"date": target_date},
            doc,
            upsert=True
        )
        
        logger.info(f"[#3.85/#3.90] Saved NCAAB data with {len(processed_games)} games and {len(merged_plays)} plays")
        
        return {
            "success": True,
            "date": target_date,
            "games_count": len(processed_games),
            "plays_count": len(merged_plays),
            "ppg_teams_found": len(ppg_data.get('season_ranks', {})),
            "games": processed_games,
            "plays": merged_plays,  # Include plays in response!
            "message": f"Refreshed NCAAB data for {target_date}: {len(processed_games)} games, {len(merged_plays)} plays"
        }
        
    except Exception as e:
        logger.error(f"Error refreshing NCAAB opportunities: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/opportunities/ncaab/manual")
async def add_ncaab_manual_data(data: dict):
    """
    Manually add NCAAB games data when scraping is blocked.
    
    Expected payload:
    {
        "date": "2025-12-29",
        "ppg_data": [
            {"rank": 1, "team": "Kansas", "ppg": 89.5, "last3": 92.3},
            ...
        ],
        "games": [
            {"away": "Team A", "home": "Team B", "total": 145.5, "time": "7:00 PM"},
            ...
        ]
    }
    """
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        
        target_date = data.get('date') or datetime.now(arizona_tz).strftime('%Y-%m-%d')
        ppg_list = data.get('ppg_data', [])
        games_list = data.get('games', [])
        
        logger.info(f"Manual NCAAB data entry for {target_date}: {len(ppg_list)} teams, {len(games_list)} games")
        
        # Build PPG lookup dictionaries
        ppg_data = {
            'season_ranks': {},
            'season_values': {},
            'last3_ranks': {},
            'last3_values': {}
        }
        
        for item in ppg_list:
            team = item.get('team', '')
            if team:
                ppg_data['season_ranks'][team] = item.get('rank')
                ppg_data['season_values'][team] = item.get('ppg')
                ppg_data['last3_values'][team] = item.get('last3')
        
        # Create Last3 ranks based on values
        last3_sorted = sorted(ppg_data['last3_values'].items(), key=lambda x: x[1] if x[1] else 0, reverse=True)
        for i, (team, _) in enumerate(last3_sorted, 1):
            ppg_data['last3_ranks'][team] = i
        
        # NCAAB dot colors (same thresholds)
        def get_ncaab_dot_color(rank):
            if rank is None:
                return '⚪'  # Unknown - not in top 365
            if rank <= 92:
                return '🟢'  # Top tier (1-92)
            elif rank <= 184:
                return '🔵'  # Second tier (93-184)
            elif rank <= 276:
                return '🟡'  # Third tier (185-276)
            elif rank <= 365:
                return '🔴'  # Fourth tier (277-365)
            else:
                return '⚪'  # Unknown - not in top 365
        
        # Process games
        processed_games = []
        
        for game in games_list:
            away_team = game.get('away', '')
            home_team = game.get('home', '')
            line = game.get('total')
            
            # Get PPG data (with fuzzy matching)
            def find_team(team_name, data_dict):
                if not team_name:
                    return None
                if team_name in data_dict:
                    return data_dict[team_name]
                team_lower = team_name.lower()
                for k, v in data_dict.items():
                    if k.lower() == team_lower or team_lower in k.lower() or k.lower() in team_lower:
                        return v
                return None
            
            away_ppg_rank = find_team(away_team, ppg_data['season_ranks'])
            away_ppg_value = find_team(away_team, ppg_data['season_values'])
            away_last3_rank = find_team(away_team, ppg_data['last3_ranks'])
            away_last3_value = find_team(away_team, ppg_data['last3_values'])
            
            home_ppg_rank = find_team(home_team, ppg_data['season_ranks'])
            home_ppg_value = find_team(home_team, ppg_data['season_values'])
            home_last3_rank = find_team(home_team, ppg_data['last3_ranks'])
            home_last3_value = find_team(home_team, ppg_data['last3_values'])
            
            # Calculate combined PPG using the correct formula: (Season_A + Season_H + Last3_A + Last3_H) / 2
            combined_ppg = None
            if away_ppg_value and home_ppg_value and away_last3_value and home_last3_value:
                combined_ppg = round((away_ppg_value + home_ppg_value + away_last3_value + home_last3_value) / 2, 1)
            elif away_ppg_value and home_ppg_value:
                # Fallback to just season values if last3 not available
                combined_ppg = round(away_ppg_value + home_ppg_value, 1)
            
            edge = None
            if combined_ppg and line:
                try:
                    edge = round(combined_ppg - float(line), 1)
                except:
                    pass
            
            recommendation = ''
            if edge is not None:
                if edge >= 10:
                    recommendation = 'OVER'
                elif edge <= -9:
                    recommendation = 'UNDER'
            
            away_dots = get_ncaab_dot_color(away_ppg_rank) + get_ncaab_dot_color(away_last3_rank)
            home_dots = get_ncaab_dot_color(home_ppg_rank) + get_ncaab_dot_color(home_last3_rank)
            
            processed_games.append({
                'away': away_team,
                'home': home_team,
                'away_team': away_team,
                'home_team': home_team,
                'time': game.get('time', ''),
                'total': line,
                'opening_line': line,
                'away_ppg_rank': away_ppg_rank,
                'away_ppg_value': away_ppg_value,
                'away_last3_rank': away_last3_rank,
                'away_last3_value': away_last3_value,
                'home_ppg_rank': home_ppg_rank,
                'home_ppg_value': home_ppg_value,
                'home_last3_rank': home_last3_rank,
                'home_last3_value': home_last3_value,
                'combined_ppg': combined_ppg,
                'edge': edge,
                'recommendation': recommendation,
                'away_dots': away_dots,
                'home_dots': home_dots
            })
        
        # #3.85 / #3.90: Preserve existing plays and merge games
        existing = await db.ncaab_opportunities.find_one({"date": target_date}, {"_id": 0})
        existing_plays = existing.get('plays', []) if existing else []
        existing_games = existing.get('games', []) if existing else []
        
        # Create lookup of new games
        new_matchups = set()
        for g in processed_games:
            key = f"{g.get('away', g.get('away_team', '')).upper()}_{g.get('home', g.get('home_team', '')).upper()}"
            new_matchups.add(key)
        
        # #3.85/#3.90: Add existing games that are NOT in the new list (started games)
        added_started = 0
        for existing_game in existing_games:
            away = existing_game.get('away', existing_game.get('away_team', ''))
            home = existing_game.get('home', existing_game.get('home_team', ''))
            key = f"{away.upper()}_{home.upper()}"
            
            if key not in new_matchups:
                # This game has started or isn't in new data - keep it from cache
                existing_game['started'] = True
                processed_games.append(existing_game)
                added_started += 1
        
        if added_started > 0:
            logger.info(f"[#3.85/#3.90] Preserved {added_started} started NCAAB games in manual entry (total: {len(processed_games)})")
        
        # Preserve bet info on games
        for play in merged_plays:
            play_game = play.get('game', '')
            if ' @ ' not in play_game:
                continue
            play_away = play_game.split(' @ ')[0].lower()
            play_home = play_game.split(' @ ')[1].lower()
            
            for g in processed_games:
                g_away = (g.get('away', g.get('away_team', ''))).lower()
                g_home = (g.get('home', g.get('home_team', ''))).lower()
                if (play_away in g_away or g_away in play_away) and \
                   (play_home in g_home or g_home in play_home):
                    g['has_bet'] = True
                    g['bet_type'] = play.get('bet_type')
                    g['bet_count'] = play.get('bet_count', 1)
                    g['bet_line'] = play.get('bet_line')
                    break
        
        # Save to database
        doc = {
            "date": target_date,
            "games": processed_games,
            "plays": merged_plays,  # #3.85: Preserve existing plays + new open bets
            "last_updated": datetime.now(timezone.utc).isoformat(),
            "data_source": "manual"
        }
        
        await db.ncaab_opportunities.replace_one(
            {"date": target_date},
            doc,
            upsert=True
        )
        
        logger.info(f"[#3.85/#3.90] Manual NCAAB save: {len(processed_games)} games, {len(existing_plays)} plays preserved")
        
        return {
            "success": True,
            "date": target_date,
            "games_count": len(processed_games),
            "plays_preserved": len(existing_plays),
            "ppg_teams": len(ppg_data['season_values']),
            "games": processed_games,
            "message": f"Manually added {len(processed_games)} NCAAB games for {target_date} ({len(existing_plays)} plays preserved)"
        }
        
    except Exception as e:
        logger.error(f"Error adding manual NCAAB data: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/opportunities/ncaab/update-ppg")
async def update_ncaab_ppg_from_cbs(target_date: str = None):
    """
    Update NCAAB PPG values by running the external scraper script.
    This spawns a subprocess to avoid memory issues with Playwright.
    
    Args:
        target_date: Target date in 'YYYY-MM-DD' format. Defaults to tomorrow (Arizona time).
    """
    import subprocess
    import os
    
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # Default to tomorrow if no date specified (since button is on Tomorrow page)
        if not target_date:
            target_date = (datetime.now(arizona_tz) + timedelta(days=1)).strftime('%Y-%m-%d')
        
        logger.info(f"[NCAAB PPG Update] Starting script for {target_date}")
        
        # Run the standalone script with target_date as environment variable
        script_path = os.path.join(os.path.dirname(__file__), 'update_ncaab_ppg.py')
        
        env = os.environ.copy()
        env['TARGET_DATE'] = target_date
        
        # Run with timeout of 300 seconds
        result = subprocess.run(
            ['python3', script_path],
            capture_output=True,
            text=True,
            timeout=300,
            cwd=os.path.dirname(__file__),
            env=env
        )
        
        logger.info(f"[NCAAB PPG Update] Script output: {result.stdout[-500:] if result.stdout else 'No output'}")
        if result.stderr:
            logger.warning(f"[NCAAB PPG Update] Script stderr: {result.stderr[-500:]}")
        
        # Get the updated data from database
        existing = await db.ncaab_opportunities.find_one({"date": target_date}, {"_id": 0})
        games = existing.get('games', []) if existing else []
        plays = existing.get('plays', []) if existing else []
        
        # Count games with PPG
        games_with_ppg = sum(1 for g in games if g.get('away_ppg_value') and g.get('home_ppg_value'))
        
        return {
            "success": True,
            "date": target_date,
            "message": f"Updated PPG for {games_with_ppg} of {len(games)} NCAAB games",
            "games_count": len(games),
            "games_with_ppg": games_with_ppg,
            "plays_preserved": len(plays),
            "script_output": result.stdout[-200:] if result.stdout else ""
        }
        
    except subprocess.TimeoutExpired:
        logger.warning("[NCAAB PPG Update] Script timed out but may still be updating")
        return {
            "success": True,
            "message": "PPG update is still running. Please refresh data in a few minutes.",
            "warning": "Script timed out"
        }
    except Exception as e:
        logger.error(f"Error updating NCAAB PPG: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/ppg/upload-file")
async def upload_ppg_file(file: UploadFile):
    """
    Upload PPG Excel file to server.
    Saves to /tmp/PPG.xlsx for processing.
    """
    try:
        # Read file content
        content = await file.read()
        
        # Save to /tmp/PPG.xlsx
        with open('/tmp/PPG.xlsx', 'wb') as f:
            f.write(content)
        
        logger.info(f"[PPG Upload] Saved PPG.xlsx ({len(content)} bytes)")
        
        return {
            "success": True,
            "message": "PPG file uploaded successfully",
            "size": len(content)
        }
    except Exception as e:
        logger.error(f"Error uploading PPG file: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/ppg/upload-excel")
async def upload_ppg_excel(league: str, target_date: str = None, target_day: str = None):
    """
    Process PPG data from uploaded Excel file for a specific league.
    The Excel file should already be at /tmp/PPG.xlsx
    
    Args:
        league: 'NBA', 'NHL', or 'NCAAB'
        target_date: Target date in 'YYYY-MM-DD' format. Defaults to tomorrow.
        target_day: 'today' or 'tomorrow' - backend calculates date in Arizona timezone
    """
    import pandas as pd
    
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # If target_day is provided, calculate date in Arizona timezone
        if target_day:
            now_arizona = datetime.now(arizona_tz)
            if target_day == 'tomorrow':
                target_date = (now_arizona + timedelta(days=1)).strftime('%Y-%m-%d')
            else:  # 'today'
                target_date = now_arizona.strftime('%Y-%m-%d')
        elif not target_date:
            target_date = (datetime.now(arizona_tz) + timedelta(days=1)).strftime('%Y-%m-%d')
        
        logger.info(f"[PPG Excel] Processing {league} PPG for {target_date}")
        
        # Read Excel file
        xlsx = pd.ExcelFile('/tmp/PPG.xlsx')
        
        ppg_data = []
        
        if league == 'NBA':
            df = pd.read_excel(xlsx, sheet_name='NBA - PPG SEASON AND PPGL3', header=None)
            for i, row in df.iterrows():
                vals = row.values
                try:
                    rank = int(vals[0]) if pd.notna(vals[0]) else None
                    if rank and rank <= 30:
                        team = str(vals[1]).strip() if pd.notna(vals[1]) else None
                        season_ppg = float(vals[2]) if pd.notna(vals[2]) else None
                        last3_ppg = float(vals[3]) if pd.notna(vals[3]) else None
                        if team and season_ppg:
                            ppg_data.append({
                                'rank': rank,
                                'team': team,
                                'season_ppg': season_ppg,
                                'last3_ppg': last3_ppg
                            })
                except:
                    continue
                    
        elif league == 'NHL':
            # NHL team name normalization map
            nhl_normalize = {
                'Penguins': 'Pittsburgh', 'Golden Knights': 'Vegas', 'Maple Leafs': 'Toronto',
                'Avalanche': 'Colorado', 'Stars': 'Dallas', 'Oilers': 'Edmonton',
                'Ducks': 'Anaheim', 'Lightning': 'Tampa Bay', 'Canadiens': 'Montreal',
                'Hurricanes': 'Carolina', 'Senators': 'Ottawa', 'Bruins': 'Boston',
                'Rangers': 'NY Rangers', 'Islanders': 'NY Islanders', 'Flyers': 'Philadelphia',
                'Devils': 'New Jersey', 'Blue Jackets': 'Columbus', 'Panthers': 'Florida',
                'Sabres': 'Buffalo', 'Red Wings': 'Detroit', 'Blackhawks': 'Chicago',
                'Blues': 'St. Louis', 'Predators': 'Nashville', 'Wild': 'Minnesota',
                'Jets': 'Winnipeg', 'Kraken': 'Seattle', 'Sharks': 'San Jose',
                'Kings': 'LA Kings', 'Flames': 'Calgary', 'Canucks': 'Vancouver',
                'Capitals': 'Washington', 'Utah HC': 'Utah', 'Utah Mammoth': 'Utah', 'Mammoth': 'Utah'
            }
            
            def normalize_nhl_team(full_name):
                for key, val in nhl_normalize.items():
                    if key in full_name:
                        return val
                return full_name.replace('Montréal', 'Montreal').strip()
            
            # NHL Season GPG from "NHL - PPG SEASON" tab
            df_season = pd.read_excel(xlsx, sheet_name='NHL - PPG SEASON', header=None)
            season_gpg = {}
            for i, row in df_season.iterrows():
                vals = row.values
                try:
                    rank = int(vals[0]) if pd.notna(vals[0]) else None
                    if rank and rank <= 32:
                        team_full = str(vals[1]).strip() if pd.notna(vals[1]) else None
                        gpg = float(vals[15]) if pd.notna(vals[15]) else None
                        if team_full and gpg:
                            team = normalize_nhl_team(team_full)
                            season_gpg[team] = gpg
                except:
                    continue
            
            logger.info(f"[PPG Excel] NHL Season GPG: {len(season_gpg)} teams")
            
            # NHL Last 3 GPG from "NHL - PPGL3" tab
            df_l3 = pd.read_excel(xlsx, sheet_name='NHL - PPGL3', header=None)
            last3_gpg = {}
            last3_gpg_ordered = []  # Keep track of order to find the last one
            current_gpg = None
            for i, row in df_l3.iterrows():
                vals = row.values
                try:
                    # Row with rank has G (goals) in column 3 and GP (games played) in column 4
                    # GPG = G / GP
                    if pd.notna(vals[0]):
                        g_value = float(vals[3]) if pd.notna(vals[3]) else None
                        gp_value = float(vals[4]) if pd.notna(vals[4]) else None
                        if g_value and gp_value and gp_value > 0:
                            current_gpg = round(g_value / gp_value, 2)  # Calculate GPG
                        else:
                            current_gpg = None
                    # Team name is in column 2
                    team_name = vals[2] if pd.notna(vals[2]) else None
                    if team_name and isinstance(team_name, str) and len(team_name) > 2 and team_name != 'TEAM':
                        team = normalize_nhl_team(team_name.strip())
                        if current_gpg:
                            last3_gpg[team] = current_gpg
                            last3_gpg_ordered.append(current_gpg)
                except:
                    continue
            
            logger.info(f"[PPG Excel] NHL Last3 GPG: {len(last3_gpg)} teams")
            
            # Default Last3 GPG = the last team's value in PPGL3 (25th team)
            # This changes daily based on Excel data
            default_l3_gpg = last3_gpg_ordered[-1] if last3_gpg_ordered else 2.33
            logger.info(f"[PPG Excel] NHL default Last3 GPG for missing teams: {default_l3_gpg}")
            
            # Combine season and last3 data
            for team, season in season_gpg.items():
                l3 = last3_gpg.get(team, default_l3_gpg)  # Use default if missing
                ppg_data.append({
                    'team': team,
                    'season_gpg': season,
                    'last3_gpg': l3
                })
                    
        elif league == 'NCAAB':
            # NCAAB team name mapping (CBS Sports name -> Excel name)
            ncaab_map = {
                # A
                'App. St.': 'App State',
                # B
                # C
                'CCSU': 'C Connecticut',
                'Clev. St.': 'Cleveland St',
                'Chicago St.': 'Chicago St',
                'Colo. St.': 'Colorado St',
                'Colorado State': 'Colorado St',
                # D
                'Detroit': 'Detroit Mercy',
                # E
                'East Carolina': 'E Carolina',
                'E. Kentucky': 'E Kentucky',
                'E. Michigan': 'E Michigan',
                # F
                'FAU': 'Florida Atlantic',
                'FDU': 'F Dickinson',
                'FIU': 'Florida Intl',
                'Fla. Gulf Coast': 'Florida Gulf Coast',
                # G
                'Ga. Tech': 'Georgia Tech',
                'Ga. Southern': 'Georgia Southern',
                'George Wash.': 'G Washington',
                # I
                'IUI': None,  # Indiana University Indianapolis - not in Excel
                'Ill.-Chicago': 'Illinois Chicago',
                # J
                'J&W-Prov.': None,
                'Jax. State': 'Jacksonville St',
                # K
                'Kennesaw St.': 'Kennesaw St',
                # L
                'LIU': 'Long Island',
                'LMU': 'Loyola Mymt',
                'La. Tech': 'Louisiana Tech',
                'Loyola-Md.': 'Loyola MD',
                'Loyola Chi.': 'Loyola Chi',
                'Lindenwood': 'Lindenwood',
                # M
                'Middle Tenn.': 'Middle Tenn',
                'Missouri St.': 'Missouri St',
                'Mt St Mary\'s': 'Mount St Mary\'s',
                'Murray St.': 'Murray St',
                'Massachusetts': 'UMass',
                'MASS': 'UMass',
                'Miami Ohio': 'Miami (Ohio)',
                'MIAMI OHIO': 'Miami (Ohio)',
                'Miami (OH)': 'Miami (Ohio)',
                # N
                'N. Dak. St.': 'N Dakota St',
                'N. Iowa': 'Northern Iowa',
                'N. Kentucky': 'N Kentucky',
                'N. Mex. St.': 'New Mexico St',
                'N.J. Tech': 'NJIT',
                # O
                'Oregon St.': 'Oregon St',
                'Okla. St.': 'Oklahoma St',
                'Okla St': 'Oklahoma St',
                'Oklahoma State': 'Oklahoma St',
                # P
                'PFW': 'Purdue FW',
                # R
                # S
                'S. Illinois': 'Southern Illinois',
                'S. Carolina': 'South Carolina',
                'S Carolina': 'South Carolina',
                'SC': 'South Carolina',
                'SIUE': 'SIU Edwardsville',
                'SIU-E': 'SIU Edwardsville',
                'Saint Francis': 'St Francis PA',
                "Saint Mary's": "Saint Mary's",
                'Sam Houston': 'Sam Houston St',
                'San Fran.': 'San Francisco',
                'St. Bona.': 'St Bonaventure',
                "St. John's": "St John's",
                'St. Thomas (MN)': 'St Thomas',
                # T
                'TX A&M-CC': 'Texas A&M-CC',
                # U
                'UAB': 'UAB',
                'UConn': 'UConn',
                'UIC': 'Illinois Chicago',
                'UTEP': 'UTEP',
                'UNC-Ash.': 'NC Asheville',
                'UNCW': 'NC Wilmington',
                'UT-Rio Grande Valley': 'UT Rio Grande',
                # V
                'Va. Tech': 'Virginia Tech',
                # W
                'W. Michigan': 'Western Michigan',
                'WESTERN MICHIGAN': 'W. Michigan',
                'W Michigan': 'W. Michigan',
                'W. Carolina': 'W Carolina',
                'W. Kentucky': 'W Kentucky',
                'Wash. St.': 'Washington St',
                'Wright St.': 'Wright St',
                # Y
                'Youngstown St.': 'Youngstown St',
                # Others
                'E. Texas A&M': 'E Texas A&M',
                'East Texas A&M': 'E Texas A&M',
                'NC A&T': 'NC A&T',
            }
            
            # Try both possible sheet names
            sheet_names = xlsx.sheet_names
            ncaab_sheet = None
            for sn in sheet_names:
                if 'NCAAB' in sn.upper() and 'PPG' in sn.upper():
                    ncaab_sheet = sn
                    break
            
            if not ncaab_sheet:
                ncaab_sheet = 'NCAAB - PPG SEASON AND PPGL3'  # Fallback
            
            logger.info(f"[PPG Excel] Using NCAAB sheet: {ncaab_sheet}")
            df = pd.read_excel(xlsx, sheet_name=ncaab_sheet, header=None)
            for i, row in df.iterrows():
                vals = row.values
                try:
                    rank = int(vals[0]) if pd.notna(vals[0]) else None
                    if rank and rank <= 400:
                        team = str(vals[1]).strip() if pd.notna(vals[1]) else None
                        season_ppg = float(vals[2]) if pd.notna(vals[2]) else None
                        last3_ppg = float(vals[3]) if pd.notna(vals[3]) else None
                        if team and season_ppg:
                            ppg_data.append({
                                'rank': rank,
                                'team': team,
                                'season_ppg': season_ppg,
                                'last3_ppg': last3_ppg
                            })
                except:
                    continue
            
            # Store the mapping for use in find_ppg
            ncaab_team_mapping = ncaab_map
        else:
            raise HTTPException(status_code=400, detail=f"Unknown league: {league}")
        
        logger.info(f"[PPG Excel] Parsed {len(ppg_data)} teams for {league}")
        
        # Build lookup dictionaries
        ppg_by_team = {}
        for item in ppg_data:
            team = item['team']
            ppg_by_team[team] = item
        
        # Sort by Last3 PPG/GPG to create L3 ranks
        if league == 'NHL':
            sorted_by_l3 = sorted([t for t in ppg_data if t.get('last3_gpg')], 
                                  key=lambda x: x['last3_gpg'], reverse=True)
            sorted_by_season = sorted([t for t in ppg_data if t.get('season_gpg')], 
                                      key=lambda x: x['season_gpg'], reverse=True)
        else:
            sorted_by_l3 = sorted([t for t in ppg_data if t.get('last3_ppg')], 
                                  key=lambda x: x['last3_ppg'], reverse=True)
            sorted_by_season = sorted([t for t in ppg_data if t.get('season_ppg')], 
                                      key=lambda x: x['season_ppg'], reverse=True)
        
        last3_ranks = {t['team']: i+1 for i, t in enumerate(sorted_by_l3)}
        season_ranks = {t['team']: i+1 for i, t in enumerate(sorted_by_season)}
        
        # Get existing games
        collection_name = f"{league.lower()}_opportunities"
        existing = await db[collection_name].find_one({"date": target_date}, {"_id": 0})
        
        if not existing:
            return {
                "success": False,
                "error": f"No {league} games found for {target_date}. Run scrape-openers first."
            }
        
        existing_games = existing.get('games', [])
        existing_plays = existing.get('plays', [])
        
        # Helper to find team in PPG data
        def find_ppg(team_name):
            if not team_name:
                return None
            
            # Check direct match first
            if team_name in ppg_by_team:
                return ppg_by_team[team_name]
            
            # For NCAAB, check the mapping FIRST (before fuzzy match)
            if league == 'NCAAB':
                mapped_name = ncaab_map.get(team_name)
                if mapped_name and mapped_name in ppg_by_team:
                    logger.info(f"[PPG] Mapped '{team_name}' -> '{mapped_name}'")
                    return ppg_by_team[mapped_name]
            
            # For NHL, check team name mapping
            if league == 'NHL':
                nhl_game_map = {
                    'Mammoth': 'Utah',
                    'Utah Mammoth': 'Utah',
                    'Utah HC': 'Utah',
                }
                mapped_name = nhl_game_map.get(team_name)
                if mapped_name and mapped_name in ppg_by_team:
                    logger.info(f"[PPG] Mapped NHL '{team_name}' -> '{mapped_name}'")
                    return ppg_by_team[mapped_name]
            
            # Fuzzy match - but be careful with similar names (Michigan vs Michigan St)
            team_lower = team_name.lower().replace(".", "").replace("-", " ").replace("'", "").strip()
            
            # First pass: exact match after normalization
            for t, data in ppg_by_team.items():
                t_lower = t.lower().replace(".", "").replace("-", " ").replace("'", "").strip()
                if t_lower == team_lower:
                    return data
            
            # Second pass: fuzzy match but avoid partial matches for teams with common prefixes
            # e.g., "Michigan" should NOT match "Michigan St"
            best_match = None
            best_score = 0
            for t, data in ppg_by_team.items():
                t_lower = t.lower().replace(".", "").replace("-", " ").replace("'", "").strip()
                
                # Skip if one is a prefix of the other but they're not equal
                # (prevents "michigan" matching "michigan st")
                if t_lower != team_lower:
                    if t_lower.startswith(team_lower + " ") or team_lower.startswith(t_lower + " "):
                        continue
                
                # Check for substring match
                if t_lower in team_lower or team_lower in t_lower:
                    # Score by how similar the lengths are (prefer closer matches)
                    score = min(len(t_lower), len(team_lower)) / max(len(t_lower), len(team_lower))
                    if score > best_score:
                        best_score = score
                        best_match = data
            
            return best_match
        
        # Dot color helper
        def get_dot(rank, total):
            if rank is None:
                return '⚪'
            pct = rank / total
            if pct <= 0.25:
                return '🟢'
            elif pct <= 0.50:
                return '🔵'
            elif pct <= 0.75:
                return '🟡'
            return '🔴'
        
        total_teams = len(ppg_data)
        
        # Update games
        updated_games = []
        games_with_ppg = 0
        
        for i, game in enumerate(existing_games, 1):
            away = game.get('away_team', '')
            home = game.get('home_team', '')
            
            away_data = find_ppg(away)
            home_data = find_ppg(home)
            
            if league == 'NHL':
                away_l3 = away_data.get('last3_gpg') if away_data else None
                home_l3 = home_data.get('last3_gpg') if home_data else None
                away_season = away_data.get('season_gpg') if away_data else None
                home_season = home_data.get('season_gpg') if home_data else None
            else:
                away_l3 = away_data.get('last3_ppg') if away_data else None
                home_l3 = home_data.get('last3_ppg') if home_data else None
                away_season = away_data.get('season_ppg') if away_data else None
                home_season = home_data.get('season_ppg') if home_data else None
            
            # Calculate average of Season + Last3 for each team (keep precision until final)
            away_avg_raw = (away_season + away_l3) / 2 if away_season and away_l3 else None
            home_avg_raw = (home_season + home_l3) / 2 if home_season and home_l3 else None
            
            # Round for display
            away_avg = round(away_avg_raw, 1) if away_avg_raw else None
            home_avg = round(home_avg_raw, 1) if home_avg_raw else None
            
            away_l3_rank = last3_ranks.get(away_data['team']) if away_data and away_data.get('team') in last3_ranks else None
            home_l3_rank = last3_ranks.get(home_data['team']) if home_data and home_data.get('team') in last3_ranks else None
            away_season_rank = season_ranks.get(away_data['team']) if away_data and away_data.get('team') in season_ranks else None
            home_season_rank = season_ranks.get(home_data['team']) if home_data and home_data.get('team') in season_ranks else None
            
            # Combined = Away avg + Home avg (use raw values for precision, round final)
            combined = round(away_avg_raw + home_avg_raw, 1) if away_avg_raw and home_avg_raw else None
            if away_avg_raw and home_avg_raw:
                games_with_ppg += 1
            
            # Use bet_line for edge calculation if bet is placed, otherwise use current line
            line = game.get('total') or game.get('opening_line')
            if game.get('has_bet') and game.get('bet_line'):
                edge_line = float(game.get('bet_line'))
            else:
                edge_line = float(line) if line else None
            
            edge = round(combined - edge_line, 1) if combined and edge_line else None
            
            rec = ''
            if edge:
                if edge >= 10:
                    rec = 'OVER'
                elif edge <= -9:
                    rec = 'UNDER'
            
            updated_game = {
                **game,
                'game_num': i,
                'away_ppg_value': away_avg,  # Store the average
                'away_last3_value': away_l3,
                'away_season_ppg': away_season,
                'away_ppg_rank': away_season_rank,  # Season rank
                'away_last3_rank': away_l3_rank,    # L3 rank
                'home_ppg_value': home_avg,  # Store the average
                'home_last3_value': home_l3,
                'home_season_ppg': home_season,
                'home_ppg_rank': home_season_rank,  # Season rank
                'home_last3_rank': home_l3_rank,    # L3 rank
                'combined_ppg': combined,
                'combined_gpg': combined if league == 'NHL' else None,  # Also store as combined_gpg for NHL
                'edge': edge,
                'recommendation': rec,
                'away_dots': get_dot(away_season_rank, total_teams) + get_dot(away_l3_rank, total_teams),
                'home_dots': get_dot(home_season_rank, total_teams) + get_dot(home_l3_rank, total_teams)
            }
            updated_games.append(updated_game)
        
        # Save updated games
        await db[collection_name].update_one(
            {"date": target_date},
            {"$set": {
                "games": updated_games,
                "plays": existing_plays,
                "last_updated": datetime.now(arizona_tz).strftime('%I:%M %p'),
                "data_source": "PPG.xlsx",
                "ppg_locked": True
            }},
            upsert=True
        )
        
        logger.info(f"[PPG Excel] Updated {games_with_ppg}/{len(updated_games)} {league} games with PPG")
        
        return {
            "success": True,
            "league": league,
            "date": target_date,
            "teams_in_excel": len(ppg_data),
            "games_count": len(updated_games),
            "games_with_ppg": games_with_ppg,
            "plays_preserved": len(existing_plays)
        }
        
    except Exception as e:
        logger.error(f"[PPG Excel] Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/opportunities/nba/manual")
async def add_nba_manual_data(data: dict):
    """
    Manually add NBA games data when scraping is blocked.
    
    Expected payload:
    {
        "date": "2025-12-29",
        "ppg_data": [
            {"rank": 1, "team": "Denver Nuggets", "ppg": 125.8, "last3": 132.7},
            ...
        ],
        "games": [
            {"away": "Team A", "home": "Team B", "total": 227.5, "time": "7:00 PM"},
            ...
        ]
    }
    """
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        
        target_date = data.get('date') or datetime.now(arizona_tz).strftime('%Y-%m-%d')
        ppg_list = data.get('ppg_data', [])
        games_list = data.get('games', [])
        
        logger.info(f"Manual NBA data entry for {target_date}: {len(ppg_list)} teams, {len(games_list)} games")
        
        # Build PPG lookup dictionaries
        ppg_data = {
            'season_ranks': {},
            'season_values': {},
            'last3_ranks': {},
            'last3_values': {}
        }
        
        for item in ppg_list:
            team = item.get('team', '')
            if team:
                ppg_data['season_ranks'][team] = item.get('rank')
                ppg_data['season_values'][team] = item.get('ppg')
                ppg_data['last3_values'][team] = item.get('last3')
        
        # Create Last3 ranks based on values
        last3_sorted = sorted(ppg_data['last3_values'].items(), key=lambda x: x[1] if x[1] else 0, reverse=True)
        for i, (team, _) in enumerate(last3_sorted, 1):
            ppg_data['last3_ranks'][team] = i
        
        # NBA dot colors (30 teams, divided into 4 groups of ~8)
        def get_nba_dot_color(rank):
            if rank is None:
                return '🔵'
            if rank <= 8:
                return '🟢'
            elif rank <= 16:
                return '🔵'
            elif rank <= 24:
                return '🟡'
            else:
                return '🔴'
        
        # Process games
        processed_games = []
        
        for game in games_list:
            away_team = game.get('away', '')
            home_team = game.get('home', '')
            line = game.get('total')
            
            # Get PPG data (with fuzzy matching)
            def find_team(team_name, data_dict):
                if not team_name:
                    return None
                if team_name in data_dict:
                    return data_dict[team_name]
                team_lower = team_name.lower()
                for k, v in data_dict.items():
                    if k.lower() == team_lower or team_lower in k.lower() or k.lower() in team_lower:
                        return v
                return None
            
            away_ppg_rank = find_team(away_team, ppg_data['season_ranks'])
            away_ppg_value = find_team(away_team, ppg_data['season_values'])
            away_last3_rank = find_team(away_team, ppg_data['last3_ranks'])
            away_last3_value = find_team(away_team, ppg_data['last3_values'])
            
            home_ppg_rank = find_team(home_team, ppg_data['season_ranks'])
            home_ppg_value = find_team(home_team, ppg_data['season_values'])
            home_last3_rank = find_team(home_team, ppg_data['last3_ranks'])
            home_last3_value = find_team(home_team, ppg_data['last3_values'])
            
            # Calculate combined PPG (NBA formula: avg of all 4 values / 2)
            combined_ppg = None
            if away_ppg_value and home_ppg_value and away_last3_value and home_last3_value:
                combined_ppg = round((away_ppg_value + home_ppg_value + away_last3_value + home_last3_value) / 2, 1)
            elif away_ppg_value and home_ppg_value:
                combined_ppg = round(away_ppg_value + home_ppg_value, 1)
            
            edge = None
            if combined_ppg and line:
                try:
                    edge = round(combined_ppg - float(line), 1)
                except:
                    pass
            
            recommendation = ''
            if edge is not None:
                if edge >= 8:
                    recommendation = 'OVER'
                elif edge <= -5:
                    recommendation = 'UNDER'
            
            away_dots = get_nba_dot_color(away_ppg_rank) + get_nba_dot_color(away_last3_rank)
            home_dots = get_nba_dot_color(home_ppg_rank) + get_nba_dot_color(home_last3_rank)
            
            processed_games.append({
                'away': away_team,
                'home': home_team,
                'away_team': away_team,
                'home_team': home_team,
                'time': game.get('time', ''),
                'total': line,
                'opening_line': line,
                'away_ppg_rank': away_ppg_rank,
                'away_ppg_value': away_ppg_value,
                'away_last3_rank': away_last3_rank,
                'away_last3_value': away_last3_value,
                'home_ppg_rank': home_ppg_rank,
                'home_ppg_value': home_ppg_value,
                'home_last3_rank': home_last3_rank,
                'home_last3_value': home_last3_value,
                'combined_ppg': combined_ppg,
                'edge': edge,
                'recommendation': recommendation,
                'away_dots': away_dots,
                'home_dots': home_dots
            })
        
        # Save to database
        doc = {
            "date": target_date,
            "games": processed_games,
            "plays": [],
            "last_updated": datetime.now(timezone.utc).isoformat(),
            "data_source": "manual"
        }
        
        await db.nba_opportunities.replace_one(
            {"date": target_date},
            doc,
            upsert=True
        )
        
        return {
            "success": True,
            "date": target_date,
            "games_count": len(processed_games),
            "ppg_teams": len(ppg_data['season_values']),
            "games": processed_games,
            "message": f"Manually added {len(processed_games)} NBA games for {target_date}"
        }
        
    except Exception as e:
        logger.error(f"Error adding manual NBA data: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# Team name aliases for matching scores
NBA_TEAM_ALIASES = {
    'Bucks': 'Milwaukee', 'Hornets': 'Charlotte', 'Suns': 'Phoenix', 'Wizards': 'Washington',
    'Warriors': 'Golden State', 'Nets': 'Brooklyn', 'Magic': 'Orlando', 'Raptors': 'Toronto',
    'Nuggets': 'Denver', 'Heat': 'Miami', 'Knicks': 'New York', 'Pelicans': 'New Orleans',
    'Pacers': 'Indiana', 'Rockets': 'Houston', 'Hawks': 'Atlanta', 'Thunder': 'Okla City',
    'Timberwolves': 'Minnesota', 'Bulls': 'Chicago', 'Cavaliers': 'Cleveland', 'Spurs': 'San Antonio',
    'Mavericks': 'Dallas', 'Trail Blazers': 'Portland', 'Lakers': 'Los Angeles', 'Clippers': 'LA Clippers',
    'Celtics': 'Boston', 'Kings': 'Sacramento', 'Jazz': 'Utah', 'Grizzlies': 'Memphis',
    '76ers': 'Philadelphia', 'Pistons': 'Detroit'
}

NHL_TEAM_ALIASES = {
    'Hurricanes': 'Carolina', 'Penguins': 'Pittsburgh', 'Devils': 'New Jersey', 'Maple Leafs': 'Toronto',
    'Canadiens': 'Montreal', 'Panthers': 'Florida', 'Islanders': 'NY Islanders', 'Blackhawks': 'Chicago',
    'Flyers': 'Philadelphia', 'Canucks': 'Vancouver', 'Bruins': 'Boston', 'Sabres': 'Buffalo',
    'Red Wings': 'Detroit', 'Lightning': 'Tampa Bay', 'Senators': 'Ottawa', 'Blue Jackets': 'Columbus',
    'Rangers': 'NY Rangers', 'Capitals': 'Washington', 'Jets': 'Winnipeg', 'Wild': 'Minnesota',
    'Predators': 'Nashville', 'Blues': 'St. Louis', 'Stars': 'Dallas', 'Avalanche': 'Colorado',
    'Coyotes': 'Utah', 'Mammoth': 'Utah', 'Utah Hockey Club': 'Utah', 'Golden Knights': 'Vegas', 'Kraken': 'Seattle', 'Ducks': 'Anaheim',
    'Kings': 'Los Angeles', 'Sharks': 'San Jose', 'Flames': 'Calgary', 'Oilers': 'Edmonton'
}


@api_router.post("/scores/nba/update")
async def update_nba_scores(date: str = None):
    """
    Update NBA scores from CBS Sports for a specific date.
    Marks edge recommendations as HIT or MISS based on final scores.
    
    Args:
        date: Date in YYYY-MM-DD format. Defaults to yesterday.
    """
    try:
        from zoneinfo import ZoneInfo
        from datetime import timedelta
        
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # Default to yesterday if no date provided
        if not date:
            yesterday = datetime.now(arizona_tz) - timedelta(days=1)
            date = yesterday.strftime('%Y-%m-%d')
        
        logger.info(f"[NBA Scores] Updating scores for {date} from CBS Sports")
        
        # Scrape scores from CBS Sports (same source as Process #1)
        scraped_games = await scrape_cbssports_nba(date)
        
        logger.info(f"[NBA Scores] Scraped {len(scraped_games)} games from CBS Sports")
        
        # Scrape betting consensus from Covers.com
        consensus_data = await scrape_covers_consensus('NBA', date)
        logger.info(f"[NBA Scores] Scraped {len(consensus_data)} team consensus entries from Covers.com")
        
        # Get database games
        db_data = await db.nba_opportunities.find_one({"date": date}, {"_id": 0})
        if not db_data:
            raise HTTPException(status_code=404, detail=f"No NBA data found for {date}")
        
        db_games = db_data.get('games', [])
        logger.info(f"[NBA Scores] Found {len(db_games)} games in database")
        
        # Helper function to normalize team names
        def normalize_team(team_name):
            if not team_name:
                return ''
            team_name = team_name.strip()
            for alias, full_name in NBA_TEAM_ALIASES.items():
                if alias.lower() in team_name.lower() or team_name.lower() in alias.lower():
                    return full_name
            return team_name
        
        # Helper to get team abbreviation for Covers lookup
        def get_team_abbrev(team_name):
            """Convert team name to abbreviation for Covers lookup"""
            abbrevs = {
                'CLEVELAND': 'CLE', 'CAVALIERS': 'CLE',
                'INDIANA': 'IND', 'PACERS': 'IND',
                'MEMPHIS': 'MEM', 'GRIZZLIES': 'MEM',
                'SAN ANTONIO': 'SA', 'SPURS': 'SA',
                'LA LAKERS': 'LAL', 'LOS ANGELES LAKERS': 'LAL', 'LAKERS': 'LAL',
                'NEW ORLEANS': 'NO', 'PELICANS': 'NO',
                'DALLAS': 'DAL', 'MAVERICKS': 'DAL',
                'SACRAMENTO': 'SAC', 'KINGS': 'SAC',
                'ORLANDO': 'ORL', 'MAGIC': 'ORL',
                'WASHINGTON': 'WAS', 'WIZARDS': 'WAS',
                'MIAMI': 'MIA', 'HEAT': 'MIA',
                'MINNESOTA': 'MIN', 'TIMBERWOLVES': 'MIN',
                'BOSTON': 'BOS', 'CELTICS': 'BOS',
                'DENVER': 'DEN', 'NUGGETS': 'DEN',
                'PHOENIX': 'PHO', 'PHX': 'PHO', 'SUNS': 'PHO',
                'LA CLIPPERS': 'LAC', 'LOS ANGELES CLIPPERS': 'LAC', 'CLIPPERS': 'LAC',
                'GOLDEN STATE': 'GS', 'GSW': 'GS', 'WARRIORS': 'GS',
                'BROOKLYN': 'BKN', 'BK': 'BKN', 'NETS': 'BKN',
                'NEW YORK': 'NY', 'NYK': 'NY', 'KNICKS': 'NY',
                'PHILADELPHIA': 'PHI', '76ERS': 'PHI', 'SIXERS': 'PHI',
                'CHICAGO': 'CHI', 'BULLS': 'CHI',
                'DETROIT': 'DET', 'PISTONS': 'DET',
                'ATLANTA': 'ATL', 'HAWKS': 'ATL',
                'CHARLOTTE': 'CHA', 'HORNETS': 'CHA',
                'TORONTO': 'TOR', 'RAPTORS': 'TOR',
                'MILWAUKEE': 'MIL', 'BUCKS': 'MIL',
                'OKLAHOMA CITY': 'OKC', 'THUNDER': 'OKC',
                'PORTLAND': 'POR', 'TRAIL BLAZERS': 'POR', 'BLAZERS': 'POR',
                'UTAH': 'UTA', 'JAZZ': 'UTA',
                'HOUSTON': 'HOU', 'ROCKETS': 'HOU',
            }
            team_upper = team_name.upper() if team_name else ''
            for name, abbrev in abbrevs.items():
                if name in team_upper:
                    return abbrev
            return team_upper[:3] if team_upper else ''
        
        # Helper function to match games
        def match_game(db_game, scraped):
            db_away = normalize_team(db_game.get('away_team', ''))
            db_home = normalize_team(db_game.get('home_team', ''))
            
            for sg in scraped:
                scraped_away = normalize_team(sg.get('away_team', ''))
                scraped_home = normalize_team(sg.get('home_team', ''))
                
                if (db_away.lower() in scraped_away.lower() or scraped_away.lower() in db_away.lower()) and \
                   (db_home.lower() in scraped_home.lower() or scraped_home.lower() in db_home.lower()):
                    return sg
                if (db_away.lower() in scraped_home.lower() or scraped_home.lower() in db_away.lower()) and \
                   (db_home.lower() in scraped_away.lower() or scraped_away.lower() in db_home.lower()):
                    return sg
            return None
        
        # Update games with scores
        updated_count = 0
        hits = 0
        misses = 0
        results = []
        
        for game in db_games:
            matched = match_game(game, scraped_games)
            if matched:
                final_score = matched.get('final_score')
                game['final_score'] = final_score
                game['away_score'] = matched.get('away_score')
                game['home_score'] = matched.get('home_score')
                
                # Add consensus data
                away_abbrev = get_team_abbrev(game.get('away_team', ''))
                home_abbrev = get_team_abbrev(game.get('home_team', ''))
                
                away_consensus = consensus_data.get(away_abbrev, {})
                home_consensus = consensus_data.get(home_abbrev, {})
                
                game['away_consensus_pct'] = away_consensus.get('consensus_pct')
                game['home_consensus_pct'] = home_consensus.get('consensus_pct')
                
                # Store spreads from Covers.com for Public Record calculation
                if away_consensus.get('spread') is not None:
                    game['away_spread'] = away_consensus.get('spread')
                if home_consensus.get('spread') is not None and game.get('spread') is None:
                    game['spread'] = home_consensus.get('spread')
                
                # Determine which team has higher consensus (the "public pick")
                if away_consensus.get('consensus_pct') and home_consensus.get('consensus_pct'):
                    if away_consensus['consensus_pct'] > home_consensus['consensus_pct']:
                        game['public_pick'] = game.get('away_team')
                        game['public_pick_pct'] = away_consensus['consensus_pct']
                    else:
                        game['public_pick'] = game.get('home_team')
                        game['public_pick_pct'] = home_consensus['consensus_pct']
                
                line = game.get('total') or game.get('opening_line')
                recommendation = game.get('recommendation')
                edge = game.get('edge', 0) or 0
                
                if final_score and line:
                    # Determine result
                    if final_score > line:
                        game['result'] = 'OVER'
                    elif final_score < line:
                        game['result'] = 'UNDER'
                    else:
                        game['result'] = 'PUSH'
                    
                    # Check if recommendation hit (only for games with |edge| >= 5)
                    # If recommendation is missing but edge meets threshold, derive it from edge
                    if abs(edge) >= 8:
                        if not recommendation:
                            recommendation = 'OVER' if edge >= 5 else 'UNDER'
                            game['recommendation'] = recommendation
                        
                        if recommendation == 'OVER':
                            game['result_hit'] = final_score > line
                        elif recommendation == 'UNDER':
                            game['result_hit'] = final_score < line
                        else:
                            game['result_hit'] = None
                        
                        if game['result_hit'] == True:
                            hits += 1
                        elif game['result_hit'] == False:
                            misses += 1
                    else:
                        game['result_hit'] = None
                
                updated_count += 1
                status = "HIT" if game.get('result_hit') == True else "MISS" if game.get('result_hit') == False else "NO_REC"
                results.append({
                    "game": f"{game.get('away_team')} @ {game.get('home_team')}",
                    "final_score": final_score,
                    "line": line,
                    "edge": edge,
                    "recommendation": recommendation,
                    "result": status
                })
                logger.info(f"[NBA Scores] {game.get('away_team')} @ {game.get('home_team')}: {final_score} vs {line} | {status}")
        
        # Save to database
        await db.nba_opportunities.update_one(
            {"date": date},
            {"$set": {
                "games": db_games,
                "scores_updated": datetime.now(arizona_tz).isoformat()
            }}
        )
        
        logger.info(f"[NBA Scores] Updated {updated_count}/{len(db_games)} games")
        
        return {
            "success": True,
            "date": date,
            "message": f"Updated {updated_count} of {len(db_games)} NBA games with final scores",
            "games_updated": updated_count,
            "games_total": len(db_games),
            "edge_hits": hits,
            "edge_misses": misses,
            "hit_rate": f"{hits/(hits+misses)*100:.1f}%" if hits + misses > 0 else "N/A",
            "results": results
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating NBA scores: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/scores/nhl/update")
async def update_nhl_scores(date: str = None):
    """
    Update NHL scores from CBS Sports for a specific date.
    Marks edge recommendations as HIT or MISS based on final scores.
    
    Args:
        date: Date in YYYY-MM-DD format. Defaults to yesterday.
    """
    try:
        from zoneinfo import ZoneInfo
        from datetime import timedelta
        
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # Default to yesterday if no date provided
        if not date:
            yesterday = datetime.now(arizona_tz) - timedelta(days=1)
            date = yesterday.strftime('%Y-%m-%d')
        
        logger.info(f"[NHL Scores] Updating scores for {date} from CBS Sports")
        
        # Scrape scores from CBS Sports (same source as Process #1)
        scraped_games = await scrape_cbssports_nhl(date)
        
        logger.info(f"[NHL Scores] Scraped {len(scraped_games)} games from CBS Sports")
        
        # Scrape betting consensus from Covers.com
        consensus_data = await scrape_covers_consensus('NHL', date)
        logger.info(f"[NHL Scores] Scraped {len(consensus_data)} team consensus entries from Covers.com")
        
        # Get database games
        db_data = await db.nhl_opportunities.find_one({"date": date}, {"_id": 0})
        if not db_data:
            raise HTTPException(status_code=404, detail=f"No NHL data found for {date}")
        
        db_games = db_data.get('games', [])
        logger.info(f"[NHL Scores] Found {len(db_games)} games in database")
        
        # Create reverse mapping for NHL teams (city name -> team name)
        NHL_CITY_TO_TEAM = {v: k for k, v in NHL_TEAM_ALIASES.items()}
        
        # Helper to get team abbreviation for Covers lookup
        def get_team_abbrev(team_name):
            """Convert team name to abbreviation for Covers lookup"""
            abbrevs = {
                'COLORADO': 'COL', 'AVALANCHE': 'COL',
                'TAMPA BAY': 'TB', 'LIGHTNING': 'TB',
                'FLORIDA': 'FLA', 'PANTHERS': 'FLA',
                'TORONTO': 'TOR', 'MAPLE LEAFS': 'TOR',
                'NEW JERSEY': 'NJ', 'DEVILS': 'NJ',
                'NY ISLANDERS': 'NYI', 'ISLANDERS': 'NYI',
                'WASHINGTON': 'WAS', 'CAPITALS': 'WAS',
                'DALLAS': 'DAL', 'STARS': 'DAL',
                'CALGARY': 'CGY', 'FLAMES': 'CGY',
                'MONTREAL': 'MTL', 'CANADIENS': 'MTL',
                'ST. LOUIS': 'STL', 'BLUES': 'STL',
                'UTAH': 'UTA',
                'LOS ANGELES': 'LA', 'KINGS': 'LA',
                'BOSTON': 'BOS', 'BRUINS': 'BOS',
                'DETROIT': 'DET', 'RED WINGS': 'DET',
                'CHICAGO': 'CHI', 'BLACKHAWKS': 'CHI',
                'PITTSBURGH': 'PIT', 'PENGUINS': 'PIT',
                'PHILADELPHIA': 'PHI', 'FLYERS': 'PHI',
                'NY RANGERS': 'NYR', 'RANGERS': 'NYR',
                'CAROLINA': 'CAR', 'HURRICANES': 'CAR',
                'COLUMBUS': 'CBJ', 'BLUE JACKETS': 'CBJ',
                'OTTAWA': 'OTT', 'SENATORS': 'OTT',
                'BUFFALO': 'BUF', 'SABRES': 'BUF',
                'MINNESOTA': 'MIN', 'WILD': 'MIN',
                'WINNIPEG': 'WPG', 'JETS': 'WPG',
                'EDMONTON': 'EDM', 'OILERS': 'EDM',
                'VANCOUVER': 'VAN', 'CANUCKS': 'VAN',
                'SEATTLE': 'SEA', 'KRAKEN': 'SEA',
                'VEGAS': 'VGK', 'GOLDEN KNIGHTS': 'VGK',
                'ARIZONA': 'ARI', 'COYOTES': 'ARI',
                'SAN JOSE': 'SJ', 'SHARKS': 'SJ',
                'ANAHEIM': 'ANA', 'DUCKS': 'ANA',
                'NASHVILLE': 'NSH', 'PREDATORS': 'NSH',
            }
            team_upper = team_name.upper() if team_name else ''
            for name, abbrev in abbrevs.items():
                if name in team_upper:
                    return abbrev
            return team_upper[:3] if team_upper else ''
        
        # Helper function to normalize team names (converts to city name)
        def normalize_team(team_name):
            if not team_name:
                return ''
            team_name = team_name.strip()
            # First check if it's already a city name
            for city in NHL_TEAM_ALIASES.values():
                if city.lower() in team_name.lower() or team_name.lower() in city.lower():
                    return city
            # Then check if it's a team nickname
            for alias, city in NHL_TEAM_ALIASES.items():
                if alias.lower() in team_name.lower() or team_name.lower() in alias.lower():
                    return city
            return team_name
        
        # Helper function to match games
        def match_game(db_game, scraped):
            db_away = normalize_team(db_game.get('away_team', ''))
            db_home = normalize_team(db_game.get('home_team', ''))
            
            for sg in scraped:
                scraped_away = normalize_team(sg.get('away_team', ''))
                scraped_home = normalize_team(sg.get('home_team', ''))
                
                logger.debug(f"[NHL Match] Comparing DB: {db_away}@{db_home} vs Scraped: {scraped_away}@{scraped_home}")
                
                if (db_away.lower() == scraped_away.lower() and db_home.lower() == scraped_home.lower()):
                    return sg
                # Check reverse order
                if (db_away.lower() == scraped_home.lower() and db_home.lower() == scraped_away.lower()):
                    return sg
                # Fuzzy match
                if (db_away.lower() in scraped_away.lower() or scraped_away.lower() in db_away.lower()) and \
                   (db_home.lower() in scraped_home.lower() or scraped_home.lower() in db_home.lower()):
                    return sg
            return None
        
        # Update games with scores
        updated_count = 0
        hits = 0
        misses = 0
        results = []
        
        # NHL edge threshold is 0.5
        edge_threshold = 0.5
        
        for game in db_games:
            matched = match_game(game, scraped_games)
            if matched:
                final_score = matched.get('final_score')
                game['final_score'] = final_score
                game['away_score'] = matched.get('away_score')
                game['home_score'] = matched.get('home_score')
                
                # Add consensus data
                away_abbrev = get_team_abbrev(game.get('away_team', ''))
                home_abbrev = get_team_abbrev(game.get('home_team', ''))
                
                away_consensus = consensus_data.get(away_abbrev, {})
                home_consensus = consensus_data.get(home_abbrev, {})
                
                game['away_consensus_pct'] = away_consensus.get('consensus_pct')
                game['home_consensus_pct'] = home_consensus.get('consensus_pct')
                
                # Store spreads from Covers.com for Public Record calculation
                if away_consensus.get('spread') is not None:
                    game['away_spread'] = away_consensus.get('spread')
                if home_consensus.get('spread') is not None and game.get('spread') is None:
                    game['spread'] = home_consensus.get('spread')
                
                # Determine which team has higher consensus (the "public pick")
                if away_consensus.get('consensus_pct') and home_consensus.get('consensus_pct'):
                    if away_consensus['consensus_pct'] > home_consensus['consensus_pct']:
                        game['public_pick'] = game.get('away_team')
                        game['public_pick_pct'] = away_consensus['consensus_pct']
                    else:
                        game['public_pick'] = game.get('home_team')
                        game['public_pick_pct'] = home_consensus['consensus_pct']
                
                line = game.get('total') or game.get('opening_line')
                recommendation = game.get('recommendation')
                edge = game.get('edge', 0) or 0
                
                if final_score is not None and line:
                    # Determine result
                    if final_score > line:
                        game['result'] = 'OVER'
                    elif final_score < line:
                        game['result'] = 'UNDER'
                    else:
                        game['result'] = 'PUSH'
                    
                    # Check if recommendation hit
                    if abs(edge) >= edge_threshold and recommendation:
                        if recommendation == 'OVER':
                            game['result_hit'] = final_score > line
                        elif recommendation == 'UNDER':
                            game['result_hit'] = final_score < line
                        else:
                            game['result_hit'] = None
                        
                        if game['result_hit'] == True:
                            hits += 1
                        elif game['result_hit'] == False:
                            misses += 1
                    else:
                        game['result_hit'] = None
                
                updated_count += 1
                status = "HIT" if game.get('result_hit') == True else "MISS" if game.get('result_hit') == False else "NO_REC"
                results.append({
                    "game": f"{game.get('away_team')} @ {game.get('home_team')}",
                    "final_score": final_score,
                    "line": line,
                    "edge": edge,
                    "recommendation": recommendation,
                    "result": status
                })
                logger.info(f"[NHL Scores] {game.get('away_team')} @ {game.get('home_team')}: {final_score} vs {line} | {status}")
        
        # Save to database
        await db.nhl_opportunities.update_one(
            {"date": date},
            {"$set": {
                "games": db_games,
                "scores_updated": datetime.now(arizona_tz).isoformat()
            }}
        )
        
        logger.info(f"[NHL Scores] Updated {updated_count}/{len(db_games)} games")
        
        return {
            "success": True,
            "date": date,
            "message": f"Updated {updated_count} of {len(db_games)} NHL games with final scores",
            "games_updated": updated_count,
            "games_total": len(db_games),
            "edge_hits": hits,
            "edge_misses": misses,
            "hit_rate": f"{hits/(hits+misses)*100:.1f}%" if hits + misses > 0 else "N/A",
            "results": results
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating NHL scores: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/scores/ncaab/update")
async def update_ncaab_scores(date: str = None):
    """
    Update NCAAB scores from ScoresAndOdds.com for a specific date.
    Marks edge recommendations as HIT or MISS based on final scores.
    
    Args:
        date: Date in YYYY-MM-DD format. Defaults to yesterday.
    """
    try:
        from zoneinfo import ZoneInfo
        from datetime import timedelta
        
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # Default to yesterday if no date provided
        if not date:
            yesterday = datetime.now(arizona_tz) - timedelta(days=1)
            date = yesterday.strftime('%Y-%m-%d')
        
        logger.info(f"[NCAAB Scores] Updating scores for {date} from CBS Sports")
        
        # Scrape scores from CBS Sports (same source as Process #1)
        scraped_games = await scrape_cbssports_ncaab(date)
        
        logger.info(f"[NCAAB Scores] Scraped {len(scraped_games)} games from CBS Sports")
        
        # Scrape betting consensus from Covers.com
        consensus_data = await scrape_covers_consensus('NCAAB', date)
        logger.info(f"[NCAAB Scores] Scraped {len(consensus_data)} team consensus entries from Covers.com")
        
        # Get database games
        db_data = await db.ncaab_opportunities.find_one({"date": date}, {"_id": 0})
        if not db_data:
            raise HTTPException(status_code=404, detail=f"No NCAAB data found for {date}")
        
        db_games = db_data.get('games', [])
        logger.info(f"[NCAAB Scores] Found {len(db_games)} games in database")
        
        # Helper function to normalize NCAAB team names
        def normalize_team(team_name):
            if not team_name:
                return ''
            team_name = team_name.strip()
            # Remove common prefixes/suffixes
            team_name = team_name.replace('State', 'St.').replace('University', '')
            return team_name
        
        # Helper to get team abbreviation for Covers lookup (NCAAB uses different abbrevs)
        def get_team_abbrev(team_name):
            """Convert team name to abbreviation for Covers lookup"""
            if not team_name:
                return ''
            # NCAAB team abbreviations are often just the first few letters or specific codes
            # This is a simplified version - Covers uses various abbreviations
            team_upper = team_name.upper()
            # Common mappings
            abbrevs = {
                'DUKE': 'DUKE', 'NORTH CAROLINA': 'UNC', 'KENTUCKY': 'UK', 'KANSAS': 'KU',
                'GONZAGA': 'GONZ', 'UCLA': 'UCLA', 'VILLANOVA': 'NOVA', 'MICHIGAN': 'MICH',
                'OHIO STATE': 'OSU', 'TEXAS': 'TEX', 'LOUISVILLE': 'LOU', 'INDIANA': 'IND',
                'PURDUE': 'PUR', 'IOWA': 'IOWA', 'WISCONSIN': 'WIS', 'MINNESOTA': 'MINN',
                'MICHIGAN STATE': 'MSU', 'PENN STATE': 'PSU', 'MARYLAND': 'MD',
                'ALABAMA': 'BAMA', 'AUBURN': 'AUB', 'TENNESSEE': 'TENN', 'FLORIDA': 'FLA',
                'GEORGIA': 'UGA', 'LSU': 'LSU', 'ARKANSAS': 'ARK', 'MISSISSIPPI': 'MISS',
                'SYRACUSE': 'SYR', 'CLEMSON': 'CLEM', 'VIRGINIA': 'UVA', 'MIAMI': 'MIA',
            }
            for name, abbrev in abbrevs.items():
                if name in team_upper:
                    return abbrev
            # Default: use first 4 chars
            return team_upper.replace(' ', '')[:4]
        
        # Helper function to match games (fuzzy matching for NCAAB)
        def match_game(db_game, scraped):
            db_away = normalize_team(db_game.get('away_team', '')).lower()
            db_home = normalize_team(db_game.get('home_team', '')).lower()
            
            for sg in scraped:
                scraped_away = normalize_team(sg.get('away_team', '')).lower()
                scraped_home = normalize_team(sg.get('home_team', '')).lower()
                
                # Try to match by key words in team names
                db_away_words = set(db_away.split())
                db_home_words = set(db_home.split())
                scraped_away_words = set(scraped_away.split())
                scraped_home_words = set(scraped_home.split())
                
                # Match if significant overlap in team name words
                away_match = len(db_away_words & scraped_away_words) >= 1 or db_away in scraped_away or scraped_away in db_away
                home_match = len(db_home_words & scraped_home_words) >= 1 or db_home in scraped_home or scraped_home in db_home
                
                if away_match and home_match:
                    return sg
                    
                # Try reverse order
                away_match_rev = len(db_away_words & scraped_home_words) >= 1 or db_away in scraped_home or scraped_home in db_away
                home_match_rev = len(db_home_words & scraped_away_words) >= 1 or db_home in scraped_away or scraped_away in db_home
                
                if away_match_rev and home_match_rev:
                    return sg
            
            return None
        
        # Update games with scores
        updated_count = 0
        hits = 0
        misses = 0
        results = []
        
        # NCAAB edge threshold is 9
        edge_threshold = 9
        
        for game in db_games:
            matched = match_game(game, scraped_games)
            
            # IMPORTANT: Only update if scraped has a final score
            # Preserve existing scores if CBS doesn't have this game
            if matched and matched.get('final_score'):
                final_score = matched.get('final_score')
                game['final_score'] = final_score
                game['away_score'] = matched.get('away_score')
                game['home_score'] = matched.get('home_score')
            elif game.get('final_score'):
                # Preserve existing score that was manually added
                final_score = game.get('final_score')
                logger.info(f"[NCAAB Scores] Preserving existing score for {game.get('away_team')} @ {game.get('home_team')}: {final_score}")
            else:
                final_score = None
            
            # Add consensus data
            away_abbrev = get_team_abbrev(game.get('away_team', ''))
            home_abbrev = get_team_abbrev(game.get('home_team', ''))
            
            away_consensus = consensus_data.get(away_abbrev, {})
            home_consensus = consensus_data.get(home_abbrev, {})
            
            game['away_consensus_pct'] = away_consensus.get('consensus_pct')
            game['home_consensus_pct'] = home_consensus.get('consensus_pct')
            
            # Store spreads from Covers.com for Public Record calculation
            if away_consensus.get('spread') is not None:
                game['away_spread'] = away_consensus.get('spread')
            if home_consensus.get('spread') is not None and game.get('spread') is None:
                game['spread'] = home_consensus.get('spread')
            
            # Determine which team has higher consensus (the "public pick")
            if away_consensus.get('consensus_pct') and home_consensus.get('consensus_pct'):
                if away_consensus['consensus_pct'] > home_consensus['consensus_pct']:
                    game['public_pick'] = game.get('away_team')
                    game['public_pick_pct'] = away_consensus['consensus_pct']
                else:
                    game['public_pick'] = game.get('home_team')
                    game['public_pick_pct'] = home_consensus['consensus_pct']
            
            line = game.get('total') or game.get('opening_line')
            recommendation = game.get('recommendation')
            edge = game.get('edge', 0) or 0
            
            if final_score is not None and line:
                # Determine result
                if final_score > line:
                    game['result'] = 'OVER'
                elif final_score < line:
                    game['result'] = 'UNDER'
                else:
                    game['result'] = 'PUSH'
                
                # Check if recommendation hit (edge threshold 9 for NCAAB)
                if abs(edge) >= edge_threshold and recommendation:
                    if recommendation == 'OVER':
                        game['result_hit'] = final_score > line
                    elif recommendation == 'UNDER':
                        game['result_hit'] = final_score < line
                    else:
                        game['result_hit'] = None
                    
                    if game['result_hit'] == True:
                        hits += 1
                    elif game['result_hit'] == False:
                        misses += 1
                else:
                    game['result_hit'] = None
            
                updated_count += 1
                status = "HIT" if game.get('result_hit') == True else "MISS" if game.get('result_hit') == False else "NO_REC"
                results.append({
                    "game": f"{game.get('away_team')} @ {game.get('home_team')}",
                    "final_score": final_score,
                    "line": line,
                    "edge": edge,
                    "recommendation": recommendation,
                    "result": status
                })
                logger.info(f"[NCAAB Scores] {game.get('away_team')} @ {game.get('home_team')}: {final_score} vs {line} | {status}")
        
        # Save to database
        await db.ncaab_opportunities.update_one(
            {"date": date},
            {"$set": {
                "games": db_games,
                "scores_updated": datetime.now(arizona_tz).isoformat()
            }}
        )
        
        logger.info(f"[NCAAB Scores] Updated {updated_count}/{len(db_games)} games")
        
        return {
            "success": True,
            "date": date,
            "message": f"Updated {updated_count} of {len(db_games)} NCAAB games with final scores",
            "games_updated": updated_count,
            "games_total": len(db_games),
            "edge_hits": hits,
            "edge_misses": misses,
            "hit_rate": f"{hits/(hits+misses)*100:.1f}%" if hits + misses > 0 else "N/A",
            "results": results
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating NCAAB scores: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/bets/nba/update-results")
async def update_nba_bet_results(date: str = None):
    """
    Update NBA bet results from plays888.co History page.
    Marks games with user bets and their win/loss status.
    Shows the exact line the bet was placed at (bet_line).
    
    Args:
        date: Date in YYYY-MM-DD format. Defaults to yesterday.
    """
    try:
        from zoneinfo import ZoneInfo
        from datetime import timedelta
        import re
        
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # Default to yesterday if no date provided
        if not date:
            yesterday = datetime.now(arizona_tz) - timedelta(days=1)
            date = yesterday.strftime('%Y-%m-%d')
        
        logger.info(f"[NBA Bet Results] Updating bet results for {date}")
        
        # Get database games for that date
        db_data = await db.nba_opportunities.find_one({"date": date}, {"_id": 0})
        if not db_data:
            raise HTTPException(status_code=404, detail=f"No NBA data found for {date}")
        
        db_games = db_data.get('games', [])
        logger.info(f"[NBA Bet Results] Found {len(db_games)} games in database")
        
        # Scrape bet history from plays888.co
        settled_bets = []
        
        # Initialize Plays888 service
        service = Plays888Service()
        await service.initialize()
        
        try:
            # Login with account jac075
            login_result = await service.login("jac075", "acuna2025!")
            if not login_result.get('success'):
                raise HTTPException(status_code=401, detail="Failed to login to plays888")
            
            logger.info("[NBA Bet Results] Logged in to plays888.co")
            
            # Navigate to History page
            await service.page.goto('https://www.plays888.co/wager/History.aspx', timeout=30000)
            await service.page.wait_for_load_state('domcontentloaded')
            await service.page.wait_for_timeout(3000)
            
            # Get page text
            page_text = await service.page.inner_text('body')
            logger.info(f"[NBA Bet Results] Got history page ({len(page_text)} chars)")
            
            # Parse settled bets
            lines = page_text.split('\n')
            
            # NBA team mapping for matching
            NBA_TEAMS = {
                'ATLANTA': 'Atlanta', 'HAWKS': 'Atlanta', 'BOSTON': 'Boston', 'CELTICS': 'Boston',
                'BROOKLYN': 'Brooklyn', 'NETS': 'Brooklyn', 'CHARLOTTE': 'Charlotte', 'HORNETS': 'Charlotte',
                'CHICAGO': 'Chicago', 'BULLS': 'Chicago', 'CLEVELAND': 'Cleveland', 'CAVALIERS': 'Cleveland',
                'DALLAS': 'Dallas', 'MAVERICKS': 'Dallas', 'DENVER': 'Denver', 'NUGGETS': 'Denver',
                'DETROIT': 'Detroit', 'PISTONS': 'Detroit', 'GOLDEN STATE': 'Golden State', 'WARRIORS': 'Golden State',
                'HOUSTON': 'Houston', 'ROCKETS': 'Houston', 'INDIANA': 'Indiana', 'PACERS': 'Indiana',
                'LA CLIPPERS': 'LA Clippers', 'CLIPPERS': 'LA Clippers', 'LOS ANGELES CLIPPERS': 'LA Clippers',
                'LA LAKERS': 'Los Angeles', 'LAKERS': 'Los Angeles', 'LOS ANGELES LAKERS': 'Los Angeles',
                'MEMPHIS': 'Memphis', 'GRIZZLIES': 'Memphis', 'MIAMI': 'Miami', 'HEAT': 'Miami',
                'MILWAUKEE': 'Milwaukee', 'BUCKS': 'Milwaukee', 'MINNESOTA': 'Minnesota', 'TIMBERWOLVES': 'Minnesota',
                'NEW ORLEANS': 'New Orleans', 'PELICANS': 'New Orleans', 'NEW YORK': 'New York', 'KNICKS': 'New York',
                'OKLAHOMA CITY': 'Okla City', 'THUNDER': 'Okla City', 'OKC': 'Okla City',
                'ORLANDO': 'Orlando', 'MAGIC': 'Orlando', 'PHILADELPHIA': 'Philadelphia', '76ERS': 'Philadelphia',
                'PHOENIX': 'Phoenix', 'SUNS': 'Phoenix', 'PORTLAND': 'Portland', 'TRAIL BLAZERS': 'Portland',
                'SACRAMENTO': 'Sacramento', 'KINGS': 'Sacramento', 'SAN ANTONIO': 'San Antonio', 'SPURS': 'San Antonio',
                'TORONTO': 'Toronto', 'RAPTORS': 'Toronto', 'UTAH': 'Utah', 'JAZZ': 'Utah',
                'WASHINGTON': 'Washington', 'WIZARDS': 'Washington'
            }
            
            def normalize_team(team_str):
                if not team_str:
                    return None
                team_upper = team_str.upper().strip()
                for key, value in NBA_TEAMS.items():
                    if key in team_upper:
                        return value
                return team_str.strip()
            
            # Convert date format from YYYY-MM-DD to M/DD/YYYY for matching
            date_parts = date.split('-')
            search_date = f"{int(date_parts[1])}/{int(date_parts[2])}/{date_parts[0]}"
            search_date_alt = f"{date_parts[1]}/{date_parts[2]}/{date_parts[0]}"
            
            logger.info(f"[NBA Bet Results] Looking for bets on {search_date} or {search_date_alt}")
            
            # The History page text format example:
            # INTERNET / -1     Ticket #: 338655920
            # Dec 29 04:40 PM
            # NBA
            # STRAIGHT BET
            # [555] TOTAL o244-110
            # (DENVER NUGGETS vrs MIAMI HEAT)
            # 2200.00 / 2000.00
            # 2000.00
            # WIN
            # WIN
            # 12/29/2025 02:10 PM
            
            # Parse by looking for NBA sport entries followed by TOTAL bets
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                
                # Look for the "NBA" sport marker (on its own line)
                if line.upper() == 'NBA':
                    # Look FORWARD from NBA marker (not backwards) for bet details
                    # The structure is: NBA -> STRAIGHT BET -> [ID] TOTAL oXXX -> (TEAM vrs TEAM) -> amounts -> WIN/LOSE
                    forward_start = i
                    forward_end = min(len(lines), i + 12)  # Result is usually within 12 lines forward
                    forward_lines = lines[forward_start:forward_end]
                    forward_context = ' '.join([l.strip() for l in forward_lines])
                    
                    # Also check backward for date context
                    backward_context = ' '.join([l.strip() for l in lines[max(0, i-5):i]])
                    full_context = backward_context + ' ' + forward_context
                    
                    # Check if this bet is for the target date
                    if search_date in full_context or search_date_alt in full_context or f"Dec {int(date_parts[2])}" in full_context:
                        # Extract the bet line (e.g., "TOTAL o244-110" or "TOTAL u220-120")
                        total_match = re.search(r'TOTAL\s+([ou])(\d+\.?\d*)[½]?', forward_context, re.IGNORECASE)
                        
                        # Extract teams (e.g., "(DENVER NUGGETS vrs MIAMI HEAT)")
                        teams_match = re.search(r'\(([A-Z\s]+)\s+vrs\s+([A-Z\s]+)\)', forward_context, re.IGNORECASE)
                        
                        if total_match and teams_match:
                            bet_type = 'OVER' if total_match.group(1).lower() == 'o' else 'UNDER'
                            bet_line = float(total_match.group(2).replace('½', '.5'))
                            away_team = normalize_team(teams_match.group(1))
                            home_team = normalize_team(teams_match.group(2))
                            
                            # Extract result - look for WIN or LOSE AFTER the teams in forward context
                            # Find where teams appear and look after that
                            result = None
                            teams_pos = forward_context.upper().find('VRS')
                            if teams_pos > 0:
                                # Look for result after teams
                                after_teams = forward_context[teams_pos:]
                                result_match = re.search(r'\b(WIN|LOSE|LOSS|PUSH)\b', after_teams, re.IGNORECASE)
                                if result_match:
                                    res = result_match.group(1).upper()
                                    if res == 'WIN':
                                        result = 'won'
                                    elif res in ['LOSE', 'LOSS']:
                                        result = 'lost'
                                    elif res == 'PUSH':
                                        result = 'push'
                            
                            # Avoid duplicates - check if same teams/line already added
                            is_duplicate = any(
                                b['away_team'] == away_team and 
                                b['home_team'] == home_team and 
                                b['bet_line'] == bet_line 
                                for b in settled_bets
                            )
                            
                            if not is_duplicate:
                                settled_bets.append({
                                    'away_team': away_team,
                                    'home_team': home_team,
                                    'bet_type': bet_type,
                                    'bet_line': bet_line,
                                    'result': result
                                })
                                logger.info(f"[NBA Bet Results] Found: {away_team} @ {home_team} {bet_type} {bet_line} -> {result}")
                
                i += 1
            
        finally:
            await service.close()
        
        logger.info(f"[NBA Bet Results] Found {len(settled_bets)} settled NBA bets for {date}")
        
        # Match bets to games and update
        bets_matched = 0
        wins = 0
        losses = 0
        
        def match_teams(db_game, bet):
            db_away = db_game.get('away_team', '').lower()
            db_home = db_game.get('home_team', '').lower()
            bet_away = (bet.get('away_team') or '').lower()
            bet_home = (bet.get('home_team') or '').lower()
            
            # Match by team names
            if (bet_away in db_away or db_away in bet_away) and \
               (bet_home in db_home or db_home in bet_home):
                return True
            return False
        
        for game in db_games:
            # Find matching bet
            for bet in settled_bets:
                if match_teams(game, bet):
                    game['user_bet'] = True
                    game['bet_line'] = bet['bet_line']
                    game['bet_type'] = bet['bet_type']
                    game['has_bet'] = True
                    game['bet_result'] = bet['result']  # Store the raw result (won/lost/push)
                    
                    # Determine if bet hit
                    if bet['result'] == 'won':
                        game['user_bet_hit'] = True
                        wins += 1
                    elif bet['result'] == 'lost':
                        game['user_bet_hit'] = False
                        losses += 1
                    else:
                        game['user_bet_hit'] = None  # Push
                    
                    bets_matched += 1
                    logger.info(f"[NBA Bet Results] Matched: {game.get('away_team')} @ {game.get('home_team')} -> {bet['result']}")
                    break
        
        # FALLBACK: For games with bets but no matched result (e.g., live bets)
        # Calculate user_bet_hit directly from final_score and bet_line
        for game in db_games:
            if game.get('has_bet') and game.get('user_bet_hit') is None and game.get('final_score') and game.get('bet_line'):
                final_score = game['final_score']
                bet_line = game['bet_line']
                bet_type = game.get('bet_type', '').upper()
                
                # Check for PUSH first (final_score == bet_line)
                if final_score == bet_line:
                    game['user_bet_hit'] = None  # Push = neither hit nor miss
                    game['result'] = 'PUSH'
                    game['bet_result'] = 'push'
                    logger.info(f"[NBA Bet Results] PUSH: {game.get('away_team')} @ {game.get('home_team')} - {bet_type} {bet_line} vs {final_score}")
                elif 'OVER' in bet_type:
                    game['user_bet_hit'] = final_score > bet_line
                    game['result'] = 'OVER' if final_score > bet_line else 'UNDER'
                    if game['user_bet_hit']:
                        wins += 1
                    else:
                        losses += 1
                    logger.info(f"[NBA Bet Results] Fallback calc: {game.get('away_team')} @ {game.get('home_team')} - {bet_type} {bet_line} vs {final_score} = {'HIT' if game['user_bet_hit'] else 'MISS'}")
                elif 'UNDER' in bet_type:
                    game['user_bet_hit'] = final_score < bet_line
                    game['result'] = 'UNDER' if final_score < bet_line else 'OVER'
                    if game['user_bet_hit']:
                        wins += 1
                    else:
                        losses += 1
                    logger.info(f"[NBA Bet Results] Fallback calc: {game.get('away_team')} @ {game.get('home_team')} - {bet_type} {bet_line} vs {final_score} = {'HIT' if game['user_bet_hit'] else 'MISS'}")
        
        # Save to database with actual bet record
        await db.nba_opportunities.update_one(
            {"date": date},
            {"$set": {
                "games": db_games,
                "bet_results_updated": datetime.now(arizona_tz).isoformat(),
                "actual_bet_record": {
                    "wins": wins,
                    "losses": losses,
                    "total_bets": len(settled_bets)
                }
            }}
        )
        
        logger.info(f"[NBA Bet Results] Updated {bets_matched} games with bet results")
        
        return {
            "success": True,
            "date": date,
            "message": f"Updated {bets_matched} NBA games with bet results",
            "bets_found": len(settled_bets),
            "bets_matched": bets_matched,
            "wins": wins,
            "losses": losses,
            "win_rate": f"{wins/(wins+losses)*100:.1f}%" if wins + losses > 0 else "N/A"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating NBA bet results: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/bets/nhl/update-results")
async def update_nhl_bet_results(date: str = None):
    """
    Update NHL bet results from plays888.co History page.
    Marks games with user bets and their win/loss status.
    Shows the exact line the bet was placed at (bet_line).
    
    NHL bets can appear as:
    1. Sport marker "NHL" with "OT Included" or just TOTAL
    2. Sport marker "SOC" with "NHL - Regulation Time Only" in description
    
    Args:
        date: Date in YYYY-MM-DD format. Defaults to yesterday.
    """
    try:
        from zoneinfo import ZoneInfo
        from datetime import timedelta
        import re
        
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # Default to yesterday if no date provided
        if not date:
            yesterday = datetime.now(arizona_tz) - timedelta(days=1)
            date = yesterday.strftime('%Y-%m-%d')
        
        logger.info(f"[NHL Bet Results] Updating bet results for {date}")
        
        # Get database games for that date
        db_data = await db.nhl_opportunities.find_one({"date": date}, {"_id": 0})
        if not db_data:
            raise HTTPException(status_code=404, detail=f"No NHL data found for {date}")
        
        db_games = db_data.get('games', [])
        logger.info(f"[NHL Bet Results] Found {len(db_games)} games in database")
        
        # Scrape bet history from plays888.co
        settled_bets = []
        
        # Initialize Plays888 service
        service = Plays888Service()
        await service.initialize()
        
        try:
            # Login with account jac075
            login_result = await service.login("jac075", "acuna2025!")
            if not login_result.get('success'):
                raise HTTPException(status_code=401, detail="Failed to login to plays888")
            
            logger.info("[NHL Bet Results] Logged in to plays888.co")
            
            # Navigate to History page
            await service.page.goto('https://www.plays888.co/wager/History.aspx', timeout=30000)
            await service.page.wait_for_load_state('domcontentloaded')
            await service.page.wait_for_timeout(3000)
            
            # Get page text
            page_text = await service.page.inner_text('body')
            logger.info(f"[NHL Bet Results] Got history page ({len(page_text)} chars)")
            
            # Parse settled bets
            lines = page_text.split('\n')
            
            # NHL team mapping for matching
            NHL_TEAMS = {
                'ANAHEIM': 'Anaheim', 'DUCKS': 'Anaheim',
                'ARIZONA': 'Arizona', 'COYOTES': 'Arizona',
                'BOSTON': 'Boston', 'BRUINS': 'Boston',
                'BUFFALO': 'Buffalo', 'SABRES': 'Buffalo',
                'CALGARY': 'Calgary', 'FLAMES': 'Calgary',
                'CAROLINA': 'Carolina', 'HURRICANES': 'Carolina',
                'CHICAGO': 'Chicago', 'BLACKHAWKS': 'Chicago',
                'COLORADO': 'Colorado', 'AVALANCHE': 'Colorado',
                'COLUMBUS': 'Columbus', 'BLUE JACKETS': 'Columbus',
                'DALLAS': 'Dallas', 'STARS': 'Dallas',
                'DETROIT': 'Detroit', 'RED WINGS': 'Detroit',
                'EDMONTON': 'Edmonton', 'OILERS': 'Edmonton',
                'FLORIDA': 'Florida', 'PANTHERS': 'Florida',
                'LOS ANGELES': 'Los Angeles', 'KINGS': 'Los Angeles', 'LA KINGS': 'Los Angeles',
                'MINNESOTA': 'Minnesota', 'WILD': 'Minnesota',
                'MONTREAL': 'Montreal', 'CANADIENS': 'Montreal',
                'NASHVILLE': 'Nashville', 'PREDATORS': 'Nashville',
                'NEW JERSEY': 'New Jersey', 'DEVILS': 'New Jersey',
                'NEW YORK ISLANDERS': 'NY Islanders', 'ISLANDERS': 'NY Islanders',
                'NEW YORK RANGERS': 'NY Rangers', 'RANGERS': 'NY Rangers',
                'OTTAWA': 'Ottawa', 'SENATORS': 'Ottawa',
                'PHILADELPHIA': 'Philadelphia', 'FLYERS': 'Philadelphia',
                'PITTSBURGH': 'Pittsburgh', 'PENGUINS': 'Pittsburgh',
                'SAN JOSE': 'San Jose', 'SHARKS': 'San Jose',
                'SEATTLE': 'Seattle', 'KRAKEN': 'Seattle',
                'ST. LOUIS': 'St. Louis', 'BLUES': 'St. Louis', 'ST LOUIS': 'St. Louis',
                'TAMPA BAY': 'Tampa Bay', 'LIGHTNING': 'Tampa Bay',
                'TORONTO': 'Toronto', 'MAPLE LEAFS': 'Toronto',
                'UTAH': 'Utah',
                'VANCOUVER': 'Vancouver', 'CANUCKS': 'Vancouver',
                'VEGAS': 'Vegas', 'GOLDEN KNIGHTS': 'Vegas', 'VGK': 'Vegas',
                'WASHINGTON': 'Washington', 'CAPITALS': 'Washington',
                'WINNIPEG': 'Winnipeg', 'JETS': 'Winnipeg'
            }
            
            def normalize_team(team_str):
                if not team_str:
                    return None
                # Clean up REG.TIME suffix
                team_upper = team_str.upper().strip().replace(' REG.TIME', '').replace('REG.TIME', '')
                for key, value in NHL_TEAMS.items():
                    if key in team_upper:
                        return value
                return team_str.strip().replace(' REG.TIME', '')
            
            # Convert date format from YYYY-MM-DD to M/DD/YYYY for matching
            date_parts = date.split('-')
            search_date = f"{int(date_parts[1])}/{int(date_parts[2])}/{date_parts[0]}"
            search_date_alt = f"{date_parts[1]}/{date_parts[2]}/{date_parts[0]}"
            
            # Map month number to name for date matching
            month_names = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
                         7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
            month_name = month_names.get(int(date_parts[1]), '')
            
            logger.info(f"[NHL Bet Results] Looking for bets on {search_date} or {month_name} {int(date_parts[2])}")
            
            # NHL bets can appear in two formats:
            # 1. Sport = "NHL" with "OT Included" or regular TOTAL
            # 2. Sport = "SOC" but description contains "NHL - Regulation Time Only"
            
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                
                # Check for direct NHL marker
                is_nhl_bet = line.upper() == 'NHL'
                
                # Check for SOC marker that might be an NHL Regulation Time bet
                is_reg_time_bet = False
                if line.upper() == 'SOC':
                    # Look ahead for "NHL - Regulation Time" in description
                    look_ahead = ' '.join([l.strip() for l in lines[i:min(len(lines), i+5)]])
                    if 'NHL' in look_ahead.upper() and 'REGULATION' in look_ahead.upper():
                        is_reg_time_bet = True
                
                if is_nhl_bet or is_reg_time_bet:
                    # Look FORWARD for bet details
                    forward_start = i
                    forward_end = min(len(lines), i + 15)  # NHL bets may have more lines
                    forward_lines = lines[forward_start:forward_end]
                    forward_context = ' '.join([l.strip() for l in forward_lines])
                    
                    # Also check backward for date context
                    backward_context = ' '.join([l.strip() for l in lines[max(0, i-5):i]])
                    full_context = backward_context + ' ' + forward_context
                    
                    # Check if this bet is for the target date
                    if search_date in full_context or search_date_alt in full_context or f"{month_name} {int(date_parts[2])}" in full_context:
                        # Extract the bet line (e.g., "TOTAL o5½-125" or "TOTAL u6-110")
                        # Handle half values with ½ symbol
                        total_match = re.search(r'TOTAL\s+([ou])(\d+)[½]?\.?(\d*)', forward_context, re.IGNORECASE)
                        
                        # Extract teams - handle REG.TIME suffix
                        # Pattern: (TEAM1 [REG.TIME] vrs TEAM2 [REG.TIME])
                        teams_match = re.search(r'\(([A-Z\s\.]+?)\s+vrs\s+([A-Z\s\.]+?)\)', forward_context, re.IGNORECASE)
                        
                        if total_match and teams_match:
                            bet_type = 'OVER' if total_match.group(1).lower() == 'o' else 'UNDER'
                            # Handle half values
                            base_num = float(total_match.group(2))
                            if '½' in forward_context[total_match.start():total_match.end()+2]:
                                bet_line = base_num + 0.5
                            elif total_match.group(3):
                                bet_line = float(f"{total_match.group(2)}.{total_match.group(3)}")
                            else:
                                bet_line = base_num
                            
                            away_team = normalize_team(teams_match.group(1))
                            home_team = normalize_team(teams_match.group(2))
                            
                            # Extract result - look for WIN or LOSE AFTER the teams
                            result = None
                            teams_pos = forward_context.upper().find('VRS')
                            if teams_pos > 0:
                                after_teams = forward_context[teams_pos:]
                                result_match = re.search(r'\b(WIN|LOSE|LOSS|PUSH)\b', after_teams, re.IGNORECASE)
                                if result_match:
                                    res = result_match.group(1).upper()
                                    if res == 'WIN':
                                        result = 'won'
                                    elif res in ['LOSE', 'LOSS']:
                                        result = 'lost'
                                    elif res == 'PUSH':
                                        result = 'push'
                            
                            # Avoid duplicates
                            is_duplicate = any(
                                b['away_team'] == away_team and 
                                b['home_team'] == home_team and 
                                b['bet_line'] == bet_line 
                                for b in settled_bets
                            )
                            
                            if not is_duplicate:
                                bet_info = {
                                    'away_team': away_team,
                                    'home_team': home_team,
                                    'bet_type': bet_type,
                                    'bet_line': bet_line,
                                    'result': result,
                                    'is_reg_time': is_reg_time_bet
                                }
                                settled_bets.append(bet_info)
                                reg_marker = " (REG TIME)" if is_reg_time_bet else ""
                                logger.info(f"[NHL Bet Results] Found: {away_team} @ {home_team} {bet_type} {bet_line}{reg_marker} -> {result}")
                
                i += 1
            
        finally:
            await service.close()
        
        logger.info(f"[NHL Bet Results] Found {len(settled_bets)} settled NHL bets for {date}")
        
        # Match bets to games and update
        bets_matched = 0
        wins = 0
        losses = 0
        
        def match_teams(db_game, bet):
            db_away = db_game.get('away_team', '').lower()
            db_home = db_game.get('home_team', '').lower()
            bet_away = (bet.get('away_team') or '').lower()
            bet_home = (bet.get('home_team') or '').lower()
            
            # Match by team names (partial match for variations)
            away_match = bet_away in db_away or db_away in bet_away or \
                        any(word in db_away for word in bet_away.split() if len(word) > 3)
            home_match = bet_home in db_home or db_home in bet_home or \
                        any(word in db_home for word in bet_home.split() if len(word) > 3)
            
            return away_match and home_match
        
        for game in db_games:
            # Find matching bet
            for bet in settled_bets:
                if match_teams(game, bet):
                    game['user_bet'] = True
                    game['bet_line'] = bet['bet_line']
                    game['bet_type'] = bet['bet_type']
                    game['has_bet'] = True
                    game['is_reg_time_bet'] = bet.get('is_reg_time', False)
                    game['bet_result'] = bet['result']  # Store the raw result (won/lost/push)
                    
                    # Determine if bet hit
                    if bet['result'] == 'won':
                        game['user_bet_hit'] = True
                        wins += 1
                    elif bet['result'] == 'lost':
                        game['user_bet_hit'] = False
                        losses += 1
                    elif bet['result'] == 'push':
                        game['user_bet_hit'] = None  # Push
                        game['bet_result'] = 'push'
                    else:
                        game['user_bet_hit'] = None  # Unknown
                    
                    bets_matched += 1
                    logger.info(f"[NHL Bet Results] Matched: {game.get('away_team')} @ {game.get('home_team')} - bet_type={bet['bet_type']}, bet_line={bet['bet_line']} -> {bet['result']}")
                    break
        
        # FALLBACK: For games with bets but no matched result (e.g., live bets)
        for game in db_games:
            if game.get('has_bet') and game.get('user_bet_hit') is None and game.get('final_score') and game.get('bet_line'):
                final_score = game['final_score']
                bet_line = game['bet_line']
                bet_type = game.get('bet_type', '').upper()
                
                # Check for PUSH first (final_score == bet_line)
                if final_score == bet_line:
                    game['user_bet_hit'] = None  # Push = neither hit nor miss
                    game['result'] = 'PUSH'
                    game['bet_result'] = 'push'
                    logger.info(f"[NHL Bet Results] PUSH: {game.get('away_team')} @ {game.get('home_team')} - {bet_type} {bet_line} vs {final_score}")
                elif 'OVER' in bet_type:
                    game['user_bet_hit'] = final_score > bet_line
                    game['result'] = 'OVER' if final_score > bet_line else 'UNDER'
                    if game['user_bet_hit']:
                        wins += 1
                    else:
                        losses += 1
                    logger.info(f"[NHL Bet Results] Fallback calc: {game.get('away_team')} @ {game.get('home_team')} - {bet_type} {bet_line} vs {final_score} = {'HIT' if game['user_bet_hit'] else 'MISS'}")
                elif 'UNDER' in bet_type:
                    game['user_bet_hit'] = final_score < bet_line
                    game['result'] = 'UNDER' if final_score < bet_line else 'OVER'
                    if game['user_bet_hit']:
                        wins += 1
                    else:
                        losses += 1
                    logger.info(f"[NHL Bet Results] Fallback calc: {game.get('away_team')} @ {game.get('home_team')} - {bet_type} {bet_line} vs {final_score} = {'HIT' if game['user_bet_hit'] else 'MISS'}")
        
        # Save to database with actual bet record
        await db.nhl_opportunities.update_one(
            {"date": date},
            {"$set": {
                "games": db_games,
                "bet_results_updated": datetime.now(arizona_tz).isoformat(),
                "actual_bet_record": {
                    "wins": wins,
                    "losses": losses,
                    "total_bets": len(settled_bets)
                }
            }}
        )
        
        logger.info(f"[NHL Bet Results] Updated {bets_matched} games with bet results")
        
        return {
            "success": True,
            "date": date,
            "message": f"Updated {bets_matched} NHL games with bet results",
            "bets_found": len(settled_bets),
            "bets_matched": bets_matched,
            "wins": wins,
            "losses": losses,
            "win_rate": f"{wins/(wins+losses)*100:.1f}%" if wins + losses > 0 else "N/A"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating NHL bet results: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/bets/ncaab/update-results")
async def update_ncaab_bet_results(date: str = None):
    """
    Update NCAAB bet results from plays888.co History page.
    Marks games with user bets and their win/loss status.
    Shows the exact line the bet was placed at (bet_line).
    
    NCAAB bets appear as "CBB" (College Basketball) in the sport marker.
    Supports both TOTAL bets and SPREAD bets.
    
    Args:
        date: Date in YYYY-MM-DD format. Defaults to yesterday.
    """
    try:
        from zoneinfo import ZoneInfo
        from datetime import timedelta
        import re
        
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # Default to yesterday if no date provided
        if not date:
            yesterday = datetime.now(arizona_tz) - timedelta(days=1)
            date = yesterday.strftime('%Y-%m-%d')
        
        logger.info(f"[NCAAB Bet Results] Updating bet results for {date}")
        
        # Get database games for that date
        db_data = await db.ncaab_opportunities.find_one({"date": date}, {"_id": 0})
        if not db_data:
            raise HTTPException(status_code=404, detail=f"No NCAAB data found for {date}")
        
        db_games = db_data.get('games', [])
        logger.info(f"[NCAAB Bet Results] Found {len(db_games)} games in database")
        
        # Scrape bet history from plays888.co
        settled_bets = []
        
        # Initialize Plays888 service
        service = Plays888Service()
        await service.initialize()
        
        try:
            # Login with account jac075
            login_result = await service.login("jac075", "acuna2025!")
            if not login_result.get('success'):
                raise HTTPException(status_code=401, detail="Failed to login to plays888")
            
            logger.info("[NCAAB Bet Results] Logged in to plays888.co")
            
            # Navigate to History page
            await service.page.goto('https://www.plays888.co/wager/History.aspx', timeout=30000)
            await service.page.wait_for_load_state('domcontentloaded')
            await service.page.wait_for_timeout(3000)
            
            # Get page text
            page_text = await service.page.inner_text('body')
            logger.info(f"[NCAAB Bet Results] Got history page ({len(page_text)} chars)")
            
            # Parse settled bets
            lines = page_text.split('\n')
            
            # Convert date format from YYYY-MM-DD to M/DD/YYYY for matching
            date_parts = date.split('-')
            search_date = f"{int(date_parts[1])}/{int(date_parts[2])}/{date_parts[0]}"
            search_date_alt = f"{date_parts[1]}/{date_parts[2]}/{date_parts[0]}"
            
            # Map month number to name for date matching
            month_names = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
                         7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
            month_name = month_names.get(int(date_parts[1]), '')
            
            logger.info(f"[NCAAB Bet Results] Looking for bets on {search_date} or {month_name} {int(date_parts[2])}")
            
            # NCAAB bets appear as "CBB" (College Basketball) in the sport marker
            # Example formats:
            # CBB -> STRAIGHT BET -> [802] TOTAL u133-120 (B+½) -> (MISSOURI STATE vrs DELAWARE)
            # CBB -> STRAIGHT BET -> [803] MERRIMACK +2-110  (spread bet, single team)
            
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                
                # Look for CBB sport marker
                if line.upper() == 'CBB':
                    # Look FORWARD for bet details
                    forward_start = i
                    forward_end = min(len(lines), i + 15)
                    forward_lines = lines[forward_start:forward_end]
                    forward_context = ' '.join([l.strip() for l in forward_lines])
                    
                    # Also check backward for date context
                    backward_context = ' '.join([l.strip() for l in lines[max(0, i-5):i]])
                    full_context = backward_context + ' ' + forward_context
                    
                    # Check if this bet is for the target date
                    if search_date in full_context or search_date_alt in full_context or f"{month_name} {int(date_parts[2])}" in full_context:
                        
                        # Try to extract TOTAL bet first
                        total_match = re.search(r'TOTAL\s+([ou])(\d+)[½]?\.?(\d*)', forward_context, re.IGNORECASE)
                        
                        # Extract teams for TOTAL bets (e.g., "(MISSOURI STATE vrs DELAWARE)")
                        teams_match = re.search(r'\(([A-Z\s\.\&\-\']+?)\s+vrs\s+([A-Z\s\.\&\-\']+?)\)', forward_context, re.IGNORECASE)
                        
                        bet_info = None
                        
                        if total_match and teams_match:
                            # This is a TOTAL bet with both teams
                            bet_type = 'OVER' if total_match.group(1).lower() == 'o' else 'UNDER'
                            base_num = float(total_match.group(2))
                            if '½' in forward_context[total_match.start():total_match.end()+2]:
                                bet_line = base_num + 0.5
                            elif total_match.group(3):
                                bet_line = float(f"{total_match.group(2)}.{total_match.group(3)}")
                            else:
                                bet_line = base_num
                            
                            away_team = teams_match.group(1).strip()
                            home_team = teams_match.group(2).strip()
                            
                            # Extract result
                            result = None
                            teams_pos = forward_context.upper().find('VRS')
                            if teams_pos > 0:
                                after_teams = forward_context[teams_pos:]
                                result_match = re.search(r'\b(WIN|LOSE|LOSS|PUSH)\b', after_teams, re.IGNORECASE)
                                if result_match:
                                    res = result_match.group(1).upper()
                                    if res == 'WIN':
                                        result = 'won'
                                    elif res in ['LOSE', 'LOSS']:
                                        result = 'lost'
                                    elif res == 'PUSH':
                                        result = 'push'
                            
                            bet_info = {
                                'away_team': away_team,
                                'home_team': home_team,
                                'bet_type': bet_type,
                                'bet_line': bet_line,
                                'result': result,
                                'is_spread': False
                            }
                        
                        elif not teams_match:
                            # Check for SPREAD bet (single team, e.g., "MERRIMACK +2-110")
                            # Pattern: [ID] TEAM +/-SPREAD-ODDS
                            spread_match = re.search(r'\[[\d]+\]\s+([A-Z\s\.\&\-\']+?)\s+([+-]?\d+\.?\d*)[½]?-\d+', forward_context, re.IGNORECASE)
                            
                            if spread_match:
                                team_name = spread_match.group(1).strip()
                                spread_value = spread_match.group(2)
                                # Handle half values
                                if '½' in forward_context[spread_match.start():spread_match.end()+2]:
                                    bet_line = float(spread_value) + (0.5 if float(spread_value) >= 0 else -0.5)
                                else:
                                    bet_line = float(spread_value)
                                
                                # Extract result - look after the spread line
                                result = None
                                spread_pos = spread_match.end()
                                after_spread = forward_context[spread_pos:]
                                result_match = re.search(r'\b(WIN|LOSE|LOSS|PUSH)\b', after_spread, re.IGNORECASE)
                                if result_match:
                                    res = result_match.group(1).upper()
                                    if res == 'WIN':
                                        result = 'won'
                                    elif res in ['LOSE', 'LOSS']:
                                        result = 'lost'
                                    elif res == 'PUSH':
                                        result = 'push'
                                
                                bet_info = {
                                    'team_name': team_name,  # Single team for spread bets
                                    'away_team': None,
                                    'home_team': None,
                                    'bet_type': 'SPREAD',
                                    'bet_line': bet_line,
                                    'result': result,
                                    'is_spread': True
                                }
                        
                        if bet_info:
                            # For counting, we track all bets including duplicates
                            # But for matching to DB, we only match unique games
                            is_duplicate = False
                            if bet_info.get('is_spread'):
                                is_duplicate = any(
                                    b.get('team_name') == bet_info.get('team_name') and 
                                    b.get('bet_line') == bet_info.get('bet_line')
                                    for b in settled_bets
                                )
                            else:
                                is_duplicate = any(
                                    b.get('away_team') == bet_info.get('away_team') and 
                                    b.get('home_team') == bet_info.get('home_team') and 
                                    b.get('bet_line') == bet_info.get('bet_line') and
                                    b.get('bet_type') == bet_info.get('bet_type')
                                    for b in settled_bets
                                )
                            
                            # Always add to list - we'll count wins/losses from this list
                            # Mark duplicates so we can handle them in matching
                            bet_info['is_duplicate'] = is_duplicate
                            settled_bets.append(bet_info)
                            
                            dup_marker = " (DUP)" if is_duplicate else ""
                            if bet_info.get('is_spread'):
                                logger.info(f"[NCAAB Bet Results] Found SPREAD: {bet_info['team_name']} {bet_info['bet_line']} -> {bet_info['result']}{dup_marker}")
                            else:
                                logger.info(f"[NCAAB Bet Results] Found TOTAL: {bet_info['away_team']} @ {bet_info['home_team']} {bet_info['bet_type']} {bet_info['bet_line']} -> {bet_info['result']}{dup_marker}")
                
                i += 1
            
        finally:
            await service.close()
        
        total_bets = len(settled_bets)
        spread_bets = len([b for b in settled_bets if b.get('is_spread')])
        total_bet_count = total_bets - spread_bets
        
        # Calculate wins/losses directly from parsed bets (includes duplicates for correct count)
        all_wins = len([b for b in settled_bets if b.get('result') == 'won'])
        all_losses = len([b for b in settled_bets if b.get('result') == 'lost'])
        
        logger.info(f"[NCAAB Bet Results] Found {total_bets} settled NCAAB bets for {date} ({total_bet_count} totals, {spread_bets} spreads)")
        logger.info(f"[NCAAB Bet Results] Direct record from History: {all_wins}-{all_losses}")
        
        # Match bets to games and update (for display purposes)
        bets_matched = 0
        wins = 0
        losses = 0
        
        def normalize_team_name(name):
            """Normalize team names for matching"""
            if not name:
                return ''
            name = name.upper().strip()
            
            # Handle specific college team mappings FIRST (order matters!)
            specific_mappings = {
                # Directional schools
                'E. ILLINOIS': 'EASTERN ILLINOIS',
                'E ILLINOIS': 'EASTERN ILLINOIS',
                'W. ILLINOIS': 'WESTERN ILLINOIS',
                'W ILLINOIS': 'WESTERN ILLINOIS',
                'N. ILLINOIS': 'NORTHERN ILLINOIS',
                'S. ILLINOIS': 'SOUTHERN ILLINOIS',
                'E. KENTUCKY': 'EASTERN KENTUCKY',
                'W. KENTUCKY': 'WESTERN KENTUCKY',
                'N. KENTUCKY': 'NORTHERN KENTUCKY',
                'E. MICHIGAN': 'EASTERN MICHIGAN',
                'W. MICHIGAN': 'WESTERN MICHIGAN',
                'N. IOWA': 'NORTHERN IOWA',
                'N IOWA': 'NORTHERN IOWA',
                'N. ARIZONA': 'NORTHERN ARIZONA',
                'N. COLORADO': 'NORTHERN COLORADO',
                'N. DAKOTA': 'NORTH DAKOTA',
                'S. DAKOTA': 'SOUTH DAKOTA',
                # Tennessee schools
                'UT MARTIN': 'TENNESSEE MARTIN',
                'TENNESSEE-MARTIN': 'TENNESSEE MARTIN',
                'MIDDLE TENN': 'MIDDLE TENNESSEE',
                # Missouri schools
                'SE MISSOURI': 'SOUTHEAST MISSOURI',
                'SE MISSOURI ST': 'SOUTHEAST MISSOURI STATE',
                'SE MISSOURI ST.': 'SOUTHEAST MISSOURI STATE',
                # Florida schools
                'FGCU': 'FLORIDA GULF COAST',
                'FLA GULF COAST': 'FLORIDA GULF COAST',
                'FAU': 'FLORIDA ATLANTIC',
                'FIU': 'FLORIDA INTERNATIONAL',
                'UCF': 'CENTRAL FLORIDA',
                # Arkansas schools
                'CENT. ARKANSAS': 'CENTRAL ARKANSAS',
                'CENT ARKANSAS': 'CENTRAL ARKANSAS',
                # Other mappings
                'INDIANA ST.': 'INDIANA STATE',
                'INDIANA ST': 'INDIANA STATE',
                'MONTANA ST.': 'MONTANA STATE',
                'MONTANA ST': 'MONTANA STATE',
                'CAL POLY SLO': 'CAL POLY',
                'ILLINOIS ST.': 'ILLINOIS STATE',
                'ILLINOIS ST': 'ILLINOIS STATE',
            }
            for abbrev, full in specific_mappings.items():
                if abbrev in name:
                    name = name.replace(abbrev, full)
            
            # Remove common suffixes and normalize
            name = re.sub(r'\s+(ST|STATE|UNIV|UNIVERSITY)\.?$', ' STATE', name)
            name = name.replace('.', '').replace("'", '').replace('-', ' ')
            return name
        
        def match_teams(db_game, bet):
            """Match a bet to a database game"""
            db_away = normalize_team_name(db_game.get('away_team', ''))
            db_home = normalize_team_name(db_game.get('home_team', ''))
            
            if bet.get('is_spread'):
                # For spread bets, match single team name to either away or home
                bet_team = normalize_team_name(bet.get('team_name', ''))
                
                # Check if bet team matches away or home
                away_words = set(db_away.split())
                home_words = set(db_home.split())
                bet_words = set(bet_team.split())
                
                # Match if significant words overlap
                away_match = len(away_words & bet_words) >= min(2, len(bet_words))
                home_match = len(home_words & bet_words) >= min(2, len(bet_words))
                
                # Also try substring matching
                if not away_match and not home_match:
                    away_match = bet_team in db_away or db_away in bet_team
                    home_match = bet_team in db_home or db_home in bet_team
                
                return away_match or home_match
            else:
                # For total bets, match both teams
                bet_away = normalize_team_name(bet.get('away_team', ''))
                bet_home = normalize_team_name(bet.get('home_team', ''))
                
                # Try word overlap matching
                db_away_words = set(db_away.split())
                db_home_words = set(db_home.split())
                bet_away_words = set(bet_away.split())
                bet_home_words = set(bet_home.split())
                
                away_match = len(db_away_words & bet_away_words) >= min(2, len(bet_away_words)) or \
                            bet_away in db_away or db_away in bet_away
                home_match = len(db_home_words & bet_home_words) >= min(2, len(bet_home_words)) or \
                            bet_home in db_home or db_home in bet_home
                
                return away_match and home_match
        
        for game in db_games:
            # Find matching bet
            for bet in settled_bets:
                if match_teams(game, bet):
                    game['user_bet'] = True
                    game['bet_line'] = bet['bet_line']
                    game['bet_type'] = bet['bet_type']
                    game['has_bet'] = True
                    game['is_spread_bet'] = bet.get('is_spread', False)
                    game['bet_result'] = bet['result']  # Store the raw result (won/lost/push)
                    
                    # Determine if bet hit
                    if bet['result'] == 'won':
                        game['user_bet_hit'] = True
                        wins += 1
                    elif bet['result'] == 'lost':
                        game['user_bet_hit'] = False
                        losses += 1
                    else:
                        game['user_bet_hit'] = None  # Push
                    
                    bets_matched += 1
                    bet_type_str = f"SPREAD {bet['bet_line']}" if bet.get('is_spread') else f"{bet['bet_type']} {bet['bet_line']}"
                    logger.info(f"[NCAAB Bet Results] Matched: {game.get('away_team')} @ {game.get('home_team')} ({bet_type_str}) -> {bet['result']}")
                    break
        
        # FALLBACK: For games with bets but no matched result (e.g., live bets)
        for game in db_games:
            if game.get('has_bet') and game.get('user_bet_hit') is None and game.get('final_score') and game.get('bet_line'):
                final_score = game['final_score']
                bet_line = game['bet_line']
                bet_type = game.get('bet_type', '').upper()
                
                # Check for PUSH first (final_score == bet_line)
                if final_score == bet_line:
                    game['user_bet_hit'] = None  # Push = neither hit nor miss
                    game['result'] = 'PUSH'
                    game['bet_result'] = 'push'
                    logger.info(f"[NCAAB Bet Results] PUSH: {game.get('away_team')} @ {game.get('home_team')} - {bet_type} {bet_line} vs {final_score}")
                elif 'OVER' in bet_type:
                    game['user_bet_hit'] = final_score > bet_line
                    game['result'] = 'OVER' if final_score > bet_line else 'UNDER'
                    if game['user_bet_hit']:
                        wins += 1
                    else:
                        losses += 1
                    logger.info(f"[NCAAB Bet Results] Fallback calc: {game.get('away_team')} @ {game.get('home_team')} - {bet_type} {bet_line} vs {final_score} = {'HIT' if game['user_bet_hit'] else 'MISS'}")
                elif 'UNDER' in bet_type:
                    game['user_bet_hit'] = final_score < bet_line
                    game['result'] = 'UNDER' if final_score < bet_line else 'OVER'
                    if game['user_bet_hit']:
                        wins += 1
                    else:
                        losses += 1
                    logger.info(f"[NCAAB Bet Results] Fallback calc: {game.get('away_team')} @ {game.get('home_team')} - {bet_type} {bet_line} vs {final_score} = {'HIT' if game['user_bet_hit'] else 'MISS'}")
        
        # Save to database - also store the actual bet record from History
        await db.ncaab_opportunities.update_one(
            {"date": date},
            {"$set": {
                "games": db_games,
                "bet_results_updated": datetime.now(arizona_tz).isoformat(),
                "actual_bet_record": {
                    "wins": all_wins,
                    "losses": all_losses,
                    "total_bets": total_bets
                }
            }}
        )
        
        logger.info(f"[NCAAB Bet Results] Updated {bets_matched} games with bet results. Actual record: {all_wins}-{all_losses}")
        
        return {
            "success": True,
            "date": date,
            "message": f"Updated {bets_matched} NCAAB games with bet results. Actual record: {all_wins}-{all_losses}",
            "bets_found": total_bets,
            "total_bets": total_bet_count,
            "spread_bets": spread_bets,
            "bets_matched": bets_matched,
            "wins": all_wins,
            "losses": all_losses,
            "win_rate": f"{all_wins/(all_wins+all_losses)*100:.1f}%" if all_wins + all_losses > 0 else "N/A"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating NCAAB bet results: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/opportunities/nhl/manual")
async def add_nhl_manual_data(data: dict):
    """
    Manually add NHL games data when scraping is blocked.
    
    Expected payload:
    {
        "date": "2025-12-29",
        "gpg_data": [
            {"rank": 1, "team": "New Jersey Devils", "gpg": 3.97, "last3": 4.0},
            ...
        ],
        "games": [
            {"away": "Team A", "home": "Team B", "total": 6.5, "time": "7:00 PM"},
            ...
        ]
    }
    """
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        
        target_date = data.get('date') or datetime.now(arizona_tz).strftime('%Y-%m-%d')
        gpg_list = data.get('gpg_data', [])
        games_list = data.get('games', [])
        
        logger.info(f"Manual NHL data entry for {target_date}: {len(gpg_list)} teams, {len(games_list)} games")
        
        # Build GPG lookup dictionaries - use hardcoded defaults if no gpg_data provided
        # Default NHL GPG data from ESPN (Season) and StatMuse (Last 3) - Dec 29, 2025
        default_season_ranks = {
            'Colorado': 1, 'Dallas': 2, 'Edmonton': 3, 'Anaheim': 4, 'Carolina': 5,
            'Montreal': 6, 'Tampa Bay': 7, 'Ottawa': 8, 'Toronto': 9, 'Washington': 10,
            'Vegas': 11, 'Florida': 12, 'Pittsburgh': 13, 'Boston': 14, 'Detroit': 15,
            'Buffalo': 16, 'Minnesota': 17, 'San Jose': 18, 'Utah': 19, 'Columbus': 20,
            'Winnipeg': 21, 'Philadelphia': 22, 'Nashville': 23, 'Vancouver': 24, 'NY Islanders': 25,
            'Chicago': 26, 'New Jersey': 27, 'Calgary': 28, 'Los Angeles': 29, 'Seattle': 30,
            'NY Rangers': 31, 'St. Louis': 32
        }
        default_season_values = {
            'Colorado': 3.97, 'Dallas': 3.49, 'Edmonton': 3.38, 'Anaheim': 3.32, 'Carolina': 3.30,
            'Montreal': 3.29, 'Tampa Bay': 3.29, 'Ottawa': 3.27, 'Toronto': 3.26, 'Washington': 3.18,
            'Vegas': 3.17, 'Florida': 3.16, 'Pittsburgh': 3.14, 'Boston': 3.10, 'Detroit': 3.08,
            'Buffalo': 3.05, 'Minnesota': 3.03, 'San Jose': 3.00, 'Utah': 2.97, 'Columbus': 2.92,
            'Winnipeg': 2.92, 'Philadelphia': 2.89, 'Nashville': 2.78, 'Vancouver': 2.78, 'NY Islanders': 2.77,
            'Chicago': 2.76, 'New Jersey': 2.71, 'Calgary': 2.61, 'Los Angeles': 2.59, 'Seattle': 2.58,
            'NY Rangers': 2.55, 'St. Louis': 2.51
        }
        default_last3_ranks = {
            'Toronto': 1, 'Vegas': 2, 'Montreal': 3, 'Ottawa': 4, 'Pittsburgh': 5,
            'Tampa Bay': 6, 'Colorado': 7, 'Dallas': 8, 'Edmonton': 9, 'Carolina': 10,
            'Buffalo': 11, 'San Jose': 12, 'Columbus': 13, 'Calgary': 14, 'Seattle': 15,
            'St. Louis': 16, 'Washington': 17, 'Florida': 18, 'Detroit': 19, 'Philadelphia': 20,
            'Vancouver': 21, 'Los Angeles': 22, 'Winnipeg': 23, 'NY Rangers': 24, 'Minnesota': 25,
            'Nashville': 26, 'Chicago': 27, 'Anaheim': 28, 'NY Islanders': 29, 'Boston': 30,
            'Utah': 31, 'New Jersey': 32
        }
        default_last3_values = {
            'Toronto': 5.00, 'Vegas': 5.00, 'Montreal': 4.33, 'Ottawa': 4.33, 'Pittsburgh': 4.33,
            'Tampa Bay': 4.00, 'Colorado': 3.67, 'Dallas': 3.67, 'Edmonton': 3.67, 'Carolina': 3.67,
            'Buffalo': 3.33, 'San Jose': 3.33, 'Columbus': 3.33, 'Calgary': 3.33, 'Seattle': 3.33,
            'St. Louis': 3.33, 'Washington': 3.00, 'Florida': 3.00, 'Detroit': 3.00, 'Philadelphia': 3.00,
            'Vancouver': 3.00, 'Los Angeles': 3.00, 'Winnipeg': 2.67, 'NY Rangers': 2.67, 'Minnesota': 2.33,
            'Nashville': 2.33, 'Chicago': 2.33, 'Anaheim': 2.00, 'NY Islanders': 2.00, 'Boston': 1.67,
            'Utah': 1.67, 'New Jersey': 1.67
        }
        
        gpg_data = {
            'season_ranks': default_season_ranks.copy(),
            'season_values': default_season_values.copy(),
            'last3_ranks': default_last3_ranks.copy(),
            'last3_values': default_last3_values.copy()
        }
        
        # Override with provided gpg_data if any
        for item in gpg_list:
            team = item.get('team', '')
            if team:
                gpg_data['season_ranks'][team] = item.get('rank')
                gpg_data['season_values'][team] = item.get('gpg')
                gpg_data['last3_values'][team] = item.get('last3')
                # Recalculate last3 ranks if values changed
        
        # Recalculate Last3 ranks based on values if gpg_data was provided
        if gpg_list:
            last3_sorted = sorted(gpg_data['last3_values'].items(), key=lambda x: x[1] if x[1] else 0, reverse=True)
            for i, (team, _) in enumerate(last3_sorted, 1):
                gpg_data['last3_ranks'][team] = i
        
        # NHL dot colors (32 teams, divided into 4 groups of 8)
        def get_nhl_dot_color(rank):
            if rank is None:
                return '🔵'
            if rank <= 8:
                return '🟢'
            elif rank <= 16:
                return '🔵'
            elif rank <= 24:
                return '🟡'
            else:
                return '🔴'
        
        # Process games
        processed_games = []
        
        for game in games_list:
            away_team = game.get('away', '')
            home_team = game.get('home', '')
            line = game.get('total')
            
            # Get GPG data (with fuzzy matching)
            def find_team(team_name, data_dict):
                if not team_name:
                    return None
                if team_name in data_dict:
                    return data_dict[team_name]
                team_lower = team_name.lower()
                for k, v in data_dict.items():
                    if k.lower() == team_lower or team_lower in k.lower() or k.lower() in team_lower:
                        return v
                return None
            
            away_gpg_rank = find_team(away_team, gpg_data['season_ranks'])
            away_gpg_value = find_team(away_team, gpg_data['season_values'])
            away_last3_rank = find_team(away_team, gpg_data['last3_ranks'])
            away_last3_value = find_team(away_team, gpg_data['last3_values'])
            
            home_gpg_rank = find_team(home_team, gpg_data['season_ranks'])
            home_gpg_value = find_team(home_team, gpg_data['season_values'])
            home_last3_rank = find_team(home_team, gpg_data['last3_ranks'])
            home_last3_value = find_team(home_team, gpg_data['last3_values'])
            
            # Calculate combined GPG (NHL formula: avg of all 4 values / 2)
            combined_gpg = None
            if away_gpg_value and home_gpg_value and away_last3_value and home_last3_value:
                combined_gpg = round((away_gpg_value + home_gpg_value + away_last3_value + home_last3_value) / 2, 1)
            elif away_gpg_value and home_gpg_value:
                combined_gpg = round(away_gpg_value + home_gpg_value, 1)
            
            edge = None
            if combined_gpg and line:
                try:
                    edge = round(combined_gpg - float(line), 1)
                except:
                    pass
            
            recommendation = ''
            if edge is not None:
                if edge >= 0.6:
                    recommendation = 'OVER'
                elif edge <= -0.5:
                    recommendation = 'UNDER'
            
            away_dots = get_nhl_dot_color(away_gpg_rank) + get_nhl_dot_color(away_last3_rank)
            home_dots = get_nhl_dot_color(home_gpg_rank) + get_nhl_dot_color(home_last3_rank)
            
            processed_games.append({
                'away': away_team,
                'home': home_team,
                'away_team': away_team,
                'home_team': home_team,
                'time': game.get('time', ''),
                'total': line,
                'opening_line': line,
                'away_gpg_rank': away_gpg_rank,
                'away_gpg_value': away_gpg_value,
                'away_last3_rank': away_last3_rank,
                'away_last3_value': away_last3_value,
                'home_gpg_rank': home_gpg_rank,
                'home_gpg_value': home_gpg_value,
                'home_last3_rank': home_last3_rank,
                'home_last3_value': home_last3_value,
                'combined_gpg': combined_gpg,
                'edge': edge,
                'recommendation': recommendation,
                'away_dots': away_dots,
                'home_dots': home_dots
            })
        
        # Save to database
        doc = {
            "date": target_date,
            "games": processed_games,
            "plays": [],
            "last_updated": datetime.now(timezone.utc).isoformat(),
            "data_source": "manual"
        }
        
        await db.nhl_opportunities.replace_one(
            {"date": target_date},
            doc,
            upsert=True
        )
        
        return {
            "success": True,
            "date": target_date,
            "games_count": len(processed_games),
            "gpg_teams": len(gpg_data['season_values']),
            "games": processed_games,
            "message": f"Manually added {len(processed_games)} NHL games for {target_date}"
        }
        
    except Exception as e:
        logger.error(f"Error adding manual NHL data: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ================= NFL OPPORTUNITIES =================

@api_router.get("/opportunities/nfl/weeks")
async def get_nfl_weeks():
    """Get list of available NFL weeks with their date ranges"""
    try:
        # Get all NFL data
        cursor = db.nfl_opportunities.find({}, {"_id": 0})
        all_docs = await cursor.to_list(length=100)
        
        # Build week info
        week_info = {}
        for doc in all_docs:
            date = doc.get('date')
            games = doc.get('games', [])
            for game in games:
                week = game.get('week')
                if week and isinstance(week, int):
                    if week not in week_info:
                        week_info[week] = {
                            'week': week,
                            'dates': [],
                            'game_count': 0
                        }
                    if date not in week_info[week]['dates']:
                        week_info[week]['dates'].append(date)
                    week_info[week]['game_count'] += 1
        
        # Sort and format
        weeks = []
        for week in sorted(week_info.keys()):
            info = week_info[week]
            dates = sorted(info['dates'])
            weeks.append({
                'week': week,
                'dates': dates,
                'date_range': f"{dates[0]} to {dates[-1]}" if len(dates) > 1 else dates[0],
                'game_count': info['game_count']
            })
        
        return {
            "success": True,
            "weeks": weeks,
            "total_weeks": len(weeks)
        }
    except Exception as e:
        logger.error(f"Error getting NFL weeks: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/opportunities/nfl/week/{week_num}")
async def get_nfl_opportunities_by_week(week_num: int):
    """Get NFL betting opportunities for a specific week"""
    try:
        # Get all NFL data
        cursor = db.nfl_opportunities.find({}, {"_id": 0})
        all_docs = await cursor.to_list(length=100)
        
        # Filter games for the requested week
        week_games = []
        dates_in_week = set()
        
        for doc in all_docs:
            date = doc.get('date')
            games = doc.get('games', [])
            
            for game in games:
                if game.get('week') == week_num:
                    game_copy = dict(game)
                    game_copy['date'] = date  # Include date in game object
                    week_games.append(game_copy)
                    dates_in_week.add(date)
        
        # Get compound record
        record = await db.compound_records.find_one({"league": "NFL"}, {"_id": 0})
        compound_record = {
            "hits": record.get('hits', 0) if record else 0,
            "misses": record.get('misses', 0) if record else 0
        }
        
        dates_sorted = sorted(list(dates_in_week))
        
        return {
            "success": True,
            "week": week_num,
            "dates": dates_sorted,
            "date_range": f"{dates_sorted[0]} to {dates_sorted[-1]}" if len(dates_sorted) > 1 else (dates_sorted[0] if dates_sorted else "N/A"),
            "last_updated": datetime.now().strftime('%H:%M %p'),
            "games": week_games,
            "plays": [],
            "compound_record": compound_record,
            "data_source": "historical"
        }
    except Exception as e:
        logger.error(f"Error getting NFL opportunities by week: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/opportunities/nfl")
async def get_nfl_opportunities(day: str = "today"):
    """Get NFL betting opportunities. day parameter: 'yesterday', 'today', 'tomorrow', or a specific date 'YYYY-MM-DD'"""
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        
        # Check if day is a specific date format (YYYY-MM-DD)
        if len(day) == 10 and day[4] == '-' and day[7] == '-':
            target_date = day
        elif day == "tomorrow":
            target_date = (datetime.now(arizona_tz) + timedelta(days=1)).strftime('%Y-%m-%d')
        elif day == "yesterday":
            target_date = (datetime.now(arizona_tz) - timedelta(days=1)).strftime('%Y-%m-%d')
        else:
            target_date = datetime.now(arizona_tz).strftime('%Y-%m-%d')
        
        # Get cached NFL opportunities
        cached = await db.nfl_opportunities.find_one({"date": target_date}, {"_id": 0})
        
        # Get compound record - NFL should start at 0-0
        record = await db.compound_records.find_one({"league": "NFL"}, {"_id": 0})
        compound_record = {
            "hits": record.get('hits', 0) if record else 0,
            "misses": record.get('misses', 0) if record else 0
        }
        
        if cached and cached.get('games'):
            return {
                "success": True,
                "date": target_date,
                "last_updated": cached.get('last_updated'),
                "games": cached.get('games', []),
                "plays": cached.get('plays', []),
                "compound_record": compound_record,
                "data_source": cached.get('data_source', 'plays888.co')
            }
        
        return {
            "success": True,
            "date": target_date,
            "message": "No NFL opportunities data yet. Click refresh to load games.",
            "games": [],
            "plays": [],
            "compound_record": compound_record,
            "data_source": None
        }
    except Exception as e:
        logger.error(f"Error getting NFL opportunities: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/opportunities/nfl/refresh")
async def refresh_nfl_opportunities(day: str = "today", use_live_lines: bool = False):
    """Manually refresh NFL opportunities data. 
    day parameter: 'yesterday', 'today' or 'tomorrow'
    use_live_lines: if True, fetch O/U lines from plays888.co (always used for NFL)
    
    #3 PROCESS: After 5am Arizona time, TODAY's data automatically uses Plays888 for live lines
    """
    try:
        from zoneinfo import ZoneInfo
        arizona_tz = ZoneInfo('America/Phoenix')
        now_arizona = datetime.now(arizona_tz)
        current_hour = now_arizona.hour
        
        # #3 PROCESS: Automatically use Plays888 for TODAY after 5am Arizona time
        if day == "today" and current_hour >= 5:
            use_live_lines = True
            logger.info(f"[#3 Process] NFL: After 5am Arizona ({now_arizona.strftime('%I:%M %p')}), using Plays888 for live lines")
        
        if day == "tomorrow":
            target_date = (now_arizona + timedelta(days=1)).strftime('%Y-%m-%d')
        elif day == "yesterday":
            target_date = (now_arizona - timedelta(days=1)).strftime('%Y-%m-%d')
        else:
            target_date = now_arizona.strftime('%Y-%m-%d')
        
        # NFL PPG Rankings (2025 Season) - From teamrankings.com/nfl/stat/points-per-game (Dec 27, 2025)
        ppg_season = {
            'LA Rams': 1, 'Seattle': 2, 'Buffalo': 3, 'Detroit': 4, 'Dallas': 5,
            'Indianapolis': 6, 'Jacksonville': 7, 'New England': 8, 'San Francisco': 9, 'Chicago': 10,
            'Pittsburgh': 11, 'Green Bay': 12, 'Cincinnati': 13, 'Baltimore': 14, 'Denver': 15,
            'LA Chargers': 16, 'Philadelphia': 17, 'Tampa Bay': 18, 'Houston': 19, 'Kansas City': 20,
            'Arizona': 21, 'Miami': 22, 'NY Giants': 23, 'Washington': 24, 'Minnesota': 25,
            'Atlanta': 26, 'Carolina': 27, 'NY Jets': 28, 'New Orleans': 29, 'Tennessee': 30,
            'Cleveland': 31, 'Las Vegas': 32
        }
        
        # NFL PPG Last 3 Games Rankings (sorted by Last 3 PPG descending)
        ppg_last3 = {
            'LA Rams': 1, 'Jacksonville': 2, 'San Francisco': 3, 'Buffalo': 4, 'Seattle': 5,
            'New England': 6, 'Pittsburgh': 7, 'Houston': 8, 'Tennessee': 9, 'Cincinnati': 10,
            'Philadelphia': 11, 'Denver': 12, 'Minnesota': 13, 'Chicago': 14, 'Dallas': 15,
            'LA Chargers': 16, 'New Orleans': 17, 'Carolina': 18, 'Baltimore': 19, 'Green Bay': 20,
            'Miami': 21, 'Washington': 22, 'Detroit': 23, 'Tampa Bay': 24, 'Atlanta': 25,
            'Indianapolis': 26, 'Arizona': 27, 'Cleveland': 28, 'NY Giants': 29, 'Las Vegas': 30,
            'NY Jets': 31, 'Kansas City': 32
        }
        
        # Actual PPG values (2025 Season) from teamrankings.com
        ppg_season_values = {
            'LA Rams': 30.5, 'Seattle': 29.5, 'Buffalo': 28.9, 'Detroit': 28.9, 'Dallas': 28.4,
            'Indianapolis': 27.9, 'Jacksonville': 27.3, 'New England': 27.3, 'San Francisco': 26.1, 'Chicago': 25.8,
            'Pittsburgh': 24.3, 'Green Bay': 24.3, 'Cincinnati': 23.9, 'Baltimore': 23.9, 'Denver': 23.9,
            'LA Chargers': 23.3, 'Philadelphia': 23.3, 'Tampa Bay': 23.1, 'Houston': 23.1, 'Kansas City': 21.9,
            'Arizona': 21.4, 'Miami': 21.1, 'NY Giants': 20.9, 'Washington': 20.8, 'Minnesota': 20.5,
            'Atlanta': 20.5, 'Carolina': 19.1, 'NY Jets': 18.8, 'New Orleans': 17.0, 'Tennessee': 16.7,
            'Cleveland': 16.4, 'Las Vegas': 14.5
        }
        
        # PPG Last 3 Games values from teamrankings.com
        ppg_last3_values = {
            'LA Rams': 41.0, 'Jacksonville': 39.3, 'San Francisco': 37.0, 'Buffalo': 32.3, 'Seattle': 31.0,
            'New England': 30.7, 'Pittsburgh': 28.0, 'Houston': 27.7, 'Tennessee': 27.0, 'Cincinnati': 26.3,
            'Philadelphia': 26.3, 'Denver': 24.7, 'Minnesota': 24.3, 'Chicago': 24.7, 'Dallas': 24.3,
            'LA Chargers': 24.0, 'New Orleans': 24.3, 'Carolina': 23.7, 'Baltimore': 23.3, 'Green Bay': 23.3,
            'Miami': 23.3, 'Washington': 23.3, 'Detroit': 22.7, 'Tampa Bay': 22.7, 'Atlanta': 21.3,
            'Indianapolis': 20.7, 'Arizona': 18.7, 'Cleveland': 17.3, 'NY Giants': 16.3,
            'Las Vegas': 12.7, 'NY Jets': 12.0, 'Kansas City': 11.7
        }
        
        games_raw = []
        data_source = "plays888.co"
        open_bets = []
        
        # Check if yesterday - use hardcoded Christmas Day games
        if day == "yesterday":
            # Dec 25 (Christmas Day) NFL games - Week 16
            # Actual results from scoresandodds.com
            games_raw = [
                {"time": "1:00 PM", "away": "Dallas", "home": "Washington", "total": 50.5, "final_score": 53, "user_bet": False},  # DAL 30 + WAS 23 = 53 > 50.5 = OVER
                {"time": "4:30 PM", "away": "Detroit", "home": "Minnesota", "total": 45.5, "final_score": 33, "user_bet": True, "bet_type": "OVER"},  # DET 10 + MIN 23 = 33 < 45.5 = MISS
                {"time": "8:15 PM", "away": "Denver", "home": "Kansas City", "total": 37.5, "final_score": 33, "user_bet": False},  # DEN 20 + KC 13 = 33 < 37.5 = UNDER
            ]
            data_source = "scoresandodds.com"
        else:
            # Fetch from plays888.co for today/tomorrow
            try:
                # Get connection credentials
                conn = await db.connections.find_one({}, {"_id": 0}, sort=[("created_at", -1)])
                if conn and conn.get("is_connected"):
                    username = conn["username"]
                    password = decrypt_password(conn["password_encrypted"])
                    
                    # Create new scraper instance
                    scraper = Plays888Service()
                    await scraper.login(username, password)
                    live_games = await scraper.scrape_totals("NFL")
                    
                    # Also fetch open bets for ENANO account
                    await scraper.close()
                    scraper = Plays888Service()
                    enano_conn = await db.connections.find_one({"username": "jac075"}, {"_id": 0})
                    if enano_conn:
                        await scraper.login("jac075", decrypt_password(enano_conn["password_encrypted"]))
                        open_bets = await scraper.scrape_open_bets()
                        await scraper.close()
                    
                    if live_games:
                        # Convert plays888 data to our format
                        for game in live_games:
                            away_name = convert_plays888_team_name(game.get('away', ''))
                            home_name = convert_plays888_team_name(game.get('home', ''))
                            total = game.get('total', 0)
                            time_str = game.get('time', 'TBD')
                            
                            games_raw.append({
                                "away": away_name,
                                "home": home_name,
                                "total": total,
                                "time": time_str
                            })
                        
                        logger.info(f"Got {len(live_games)} NFL games from plays888.co")
                        data_source = "plays888.co"
            except Exception as e:
                logger.error(f"Error fetching live NFL lines: {e}")
        
        # #3.85/#3.90: MERGE live games with cached games to keep ALL games visible ALL DAY
        # Games that have started disappear from Plays888, but we want to show them
        # For NFL, we need to check BOTH today's date AND the week's main date (yesterday for Week 17)
        if day == "today":
            all_cached_games = []
            
            # Check today's cache
            cached_today = await db.nfl_opportunities.find_one({"date": target_date}, {"_id": 0})
            if cached_today and cached_today.get('games'):
                all_cached_games.extend(cached_today['games'])
            
            # Also check yesterday's cache (NFL weeks span multiple days)
            yesterday = (now_arizona - timedelta(days=1)).strftime('%Y-%m-%d')
            cached_yesterday = await db.nfl_opportunities.find_one({"date": yesterday}, {"_id": 0})
            if cached_yesterday and cached_yesterday.get('games'):
                all_cached_games.extend(cached_yesterday['games'])
            
            logger.info(f"[#3.90] Found {len(all_cached_games)} total cached NFL games from today + yesterday")
            
            if all_cached_games:
                # Create lookup of current games from Plays888
                live_matchups = set()
                for g in games_raw:
                    key = f"{g['away'].upper()}_{g['home'].upper()}"
                    live_matchups.add(key)
                
                # Add cached games that are NOT in the live list (started games)
                added_started = 0
                seen_matchups = set(live_matchups)  # Track what we've added
                
                for cached_game in all_cached_games:
                    away = cached_game.get('away_team', cached_game.get('away', ''))
                    home = cached_game.get('home_team', cached_game.get('home', ''))
                    key = f"{away.upper()}_{home.upper()}"
                    
                    if key not in seen_matchups:
                        # This game has started or wasn't in live list - add it from cache
                        games_raw.append({
                            "time": cached_game.get('time', 'Started'),
                            "away": away,
                            "home": home,
                            "total": cached_game.get('total'),
                            "started": True  # Mark as started
                        })
                        seen_matchups.add(key)
                        added_started += 1
                        logger.debug(f"[#3.90] Added from cache: {away} @ {home}")
                
                if added_started > 0:
                    logger.info(f"[#3.90] Added {added_started} started NFL games from cache (total: {len(games_raw)} games)")
                    data_source = "plays888.co + cached"
        
        # Fallback: If no games from Plays888, check database cache
        if not games_raw and day in ["today", "tomorrow"]:
            cached = await db.nfl_opportunities.find_one({"date": target_date}, {"_id": 0})
            if cached and cached.get('games'):
                for g in cached['games']:
                    games_raw.append({
                        "time": g.get('time', ''),
                        "away": g.get('away_team', g.get('away', '')),
                        "home": g.get('home_team', g.get('home', '')),
                        "total": g.get('total')
                    })
                data_source = "cached (from last night's scrape)"
                logger.info(f"Using {len(games_raw)} cached NFL games from database for {target_date}")
        
        # Process open bets for NFL
        nfl_open_bets = {}
        for bet in open_bets:
            if bet.get('sport') == 'NFL':
                away_team = convert_plays888_team_name(bet.get('away_team', ''))
                home_team = convert_plays888_team_name(bet.get('home_team', ''))
                game_key = f"{away_team}_{home_team}"
                
                if game_key not in nfl_open_bets:
                    nfl_open_bets[game_key] = {
                        "bet_type": bet.get('bet_type'),
                        "bet_line": bet.get('total_line'),
                        "bet_risk": bet.get('risk', 0),
                        "bet_count": 1,
                        "away_team": away_team,
                        "home_team": home_team
                    }
                else:
                    # Check for hedged bets (opposite directions)
                    existing_type = nfl_open_bets[game_key].get('bet_type')
                    new_type = bet.get('bet_type')
                    if existing_type and new_type and existing_type != new_type:
                        nfl_open_bets[game_key]['hedged'] = True
                    nfl_open_bets[game_key]['bet_count'] += 1
        
        # #3.85: Add games from open bets that aren't in the schedule (started games with bets)
        if day == "today":
            existing_matchups = set()
            for g in games_raw:
                key = f"{g['away'].upper()}_{g['home'].upper()}"
                existing_matchups.add(key)
            
            added_from_bets = 0
            for game_key, bet_data in nfl_open_bets.items():
                away = bet_data.get('away_team', '')
                home = bet_data.get('home_team', '')
                check_key = f"{away.upper()}_{home.upper()}"
                
                if check_key not in existing_matchups and away and home:
                    # This bet's game isn't in the schedule - add it
                    bet_line = bet_data.get('bet_line', 45.0)
                    games_raw.append({
                        "time": "Started",
                        "away": away,
                        "home": home,
                        "total": bet_line,
                        "started": True,
                        "from_bet": True
                    })
                    existing_matchups.add(check_key)
                    added_from_bets += 1
                    logger.info(f"[#3.85] Added bet game to NFL schedule: {away} @ {home} (line: {bet_line})")
            
            if added_from_bets > 0:
                logger.info(f"[#3.85] Added {added_from_bets} games from NFL open bets (total: {len(games_raw)} games)")
        
        games = []
        plays = []
        
        for i, game_data in enumerate(games_raw, 1):
            away_team = game_data.get('away', 'TBD')
            home_team = game_data.get('home', 'TBD')
            total = game_data.get('total', 0)
            time_str = game_data.get('time', 'TBD')
            
            # Get rankings
            away_ppg_rank = ppg_season.get(away_team, 16)
            home_ppg_rank = ppg_season.get(home_team, 16)
            away_last3_rank = ppg_last3.get(away_team, 16)
            home_last3_rank = ppg_last3.get(home_team, 16)
            
            # Calculate combined PPG
            away_ppg = ppg_season_values.get(away_team, 22.0)
            home_ppg = ppg_season_values.get(home_team, 22.0)
            combined_ppg = round((away_ppg + home_ppg), 1)
            
            # Calculate combined game average (Season + Last 3)
            away_last3_ppg = ppg_last3_values.get(away_team, 22.0)
            home_last3_ppg = ppg_last3_values.get(home_team, 22.0)
            game_avg = round((away_ppg + home_ppg + away_last3_ppg + home_last3_ppg) / 2, 1)
            
            # Calculate edge (positive = OVER, negative = UNDER)
            edge = round(game_avg - total, 1) if total else None
            
            # Determine recommendation based on edge
            recommendation = None
            color = None
            if edge is not None:
                if edge >= 7:  # NFL threshold is +7
                    recommendation = "OVER"
                    color = "green"
                elif edge <= -7:
                    recommendation = "UNDER"
                    color = "red"
            
            # Check for open bet on this game (for today/tomorrow) or use hardcoded user_bet (for yesterday)
            game_key = f"{away_team}_{home_team}"
            bet_data = nfl_open_bets.get(game_key, {})
            
            # For yesterday, use hardcoded user_bet from games_raw
            if day == "yesterday":
                has_bet = game_data.get('user_bet', False)
                user_bet_type = game_data.get('bet_type', '')
                final_score = game_data.get('final_score')
            else:
                has_bet = bool(bet_data) and not bet_data.get('hedged', False)
                user_bet_type = bet_data.get('bet_type') if has_bet else None
                final_score = None
            
            # Calculate dots based on rankings
            def get_dot_color(rank: int) -> str:
                if rank <= 8:
                    return "🟢"  # Green: Top tier
                elif rank <= 16:
                    return "🔵"  # Blue: Upper middle
                elif rank <= 24:
                    return "🟡"  # Yellow: Lower middle
                else:
                    return "🔴"  # Red: Bottom tier
            
            dots = f"{get_dot_color(away_ppg_rank)}{get_dot_color(away_last3_rank)}{get_dot_color(home_ppg_rank)}{get_dot_color(home_last3_rank)}"
            away_dots = f"{get_dot_color(away_ppg_rank)}{get_dot_color(away_last3_rank)}"
            home_dots = f"{get_dot_color(home_ppg_rank)}{get_dot_color(home_last3_rank)}"
            
            game = {
                "game_num": i,
                "time": time_str,
                "away_team": away_team,
                "home_team": home_team,
                "away_ppg_rank": away_ppg_rank,
                "home_ppg_rank": home_ppg_rank,
                "away_last3_rank": away_last3_rank,
                "home_last3_rank": home_last3_rank,
                "dots": dots,
                "away_dots": away_dots,
                "home_dots": home_dots,
                "total": total,
                "combined_ppg": game_avg,  # For NFL, use game_avg (Season + Last 3 / 2)
                "game_avg": game_avg,
                "edge": edge,
                "recommendation": recommendation,
                "color": color,
                "has_bet": has_bet,
                "user_bet": has_bet,  # Alias for compatibility
                "bet_type": user_bet_type,
                "bet_count": bet_data.get('bet_count', 0) if bet_data else 0
            }
            
            # Add final score and result for yesterday's games
            if final_score is not None:
                game["final_score"] = final_score
                # Calculate result_hit based on system recommendation
                if recommendation:
                    if recommendation == "OVER":
                        game["result_hit"] = final_score > total
                    else:
                        game["result_hit"] = final_score < total
                else:
                    game["result_hit"] = None
                    
                # Calculate user_bet_hit based on user's actual bet
                # Use bet_line if available, otherwise use total
                bet_line_for_eval = bet_data.get('bet_line') if bet_data else None
                line_for_eval = bet_line_for_eval if bet_line_for_eval else total
                
                if has_bet and user_bet_type:
                    # Check for PUSH first
                    if final_score == line_for_eval:
                        game["user_bet_hit"] = None  # Push
                        game["result"] = "PUSH"
                        game["bet_result"] = "push"
                    elif user_bet_type.upper() == "OVER":
                        game["user_bet_hit"] = final_score > line_for_eval
                    elif user_bet_type.upper() == "UNDER":
                        game["user_bet_hit"] = final_score < line_for_eval
                    else:
                        game["user_bet_hit"] = None
                else:
                    game["user_bet_hit"] = None
            
            games.append(game)
            
            # Add to plays if meets threshold (edge >= 7)
            if recommendation and edge is not None and abs(edge) >= 7 and not bet_data.get('hedged', False):
                play = {
                    "game": f"{away_team} @ {home_team}",
                    "total": total,
                    "combined_ppg": game_avg,  # Use game_avg which includes Season + Last 3
                    "edge": edge,
                    "recommendation": recommendation,
                    "has_bet": has_bet,
                    "bet_type": bet_data.get('bet_type') if has_bet else None,
                    "bet_count": bet_data.get('bet_count', 0) if has_bet else 0
                }
                
                # If user has bet, show bet-time values
                if has_bet:
                    bet_line = bet_data.get('bet_line', total)
                    bet_edge = round(game_avg - bet_line, 1) if bet_line else edge
                    play['bet_line'] = bet_line
                    play['bet_edge'] = bet_edge
                    play['live_edge'] = edge
                
                plays.append(play)
        
        # Save to database
        await db.nfl_opportunities.update_one(
            {"date": target_date},
            {"$set": {
                "date": target_date,
                "last_updated": datetime.now(arizona_tz).strftime('%I:%M %p'),
                "games": games,
                "plays": plays,
                "data_source": data_source
            }},
            upsert=True
        )
        
        # Get compound record
        record = await db.compound_records.find_one({"league": "NFL"}, {"_id": 0})
        compound_record = {
            "hits": record.get('hits', 0) if record else 0,
            "misses": record.get('misses', 0) if record else 0
        }
        
        return {
            "success": True,
            "message": f"NFL opportunities refreshed (source: {data_source})",
            "date": target_date,
            "last_updated": datetime.now(arizona_tz).strftime('%I:%M %p'),
            "games": games,
            "plays": plays,
            "compound_record": compound_record,
            "data_source": data_source
        }
    except Exception as e:
        logger.error(f"Error refreshing NFL opportunities: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/process/8pm")
async def trigger_8pm_process():
    """
    Manually trigger the 8pm process for all leagues.
    This runs:
    #1 - Scrape tomorrow's games opening lines from ScoresAndOdds
    #2 - Fill PPG data and 4-dot analysis for all games
    """
    try:
        result = await execute_8pm_process()
        return result
    except Exception as e:
        logger.error(f"Error in 8pm process: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/process/5am")
async def trigger_5am_process():
    """
    Manually trigger the 5am morning process for all leagues.
    This runs:
    #4 - Get yesterday's scores from ScoresAndOdds + mark edge HITs/MISSes
    #5 - Get bet results from Plays888 History
    #6 - Update betting and edge records
    """
    try:
        result = await morning_data_refresh()
        return result
    except Exception as e:
        logger.error(f"Error in 5am process: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ================= RANKING PPG FEATURE =================

class RankingPPGUpdate(BaseModel):
    league: str
    date: str
    game_num: int
    ranking_type: str  # "high" or "low"

@api_router.post("/opportunities/ranking-ppg")
async def set_ranking_ppg(data: RankingPPGUpdate):
    """Set ranking PPG selection (High/Low) for a game"""
    try:
        league = data.league.upper()
        collection_map = {
            'NBA': 'nba_opportunities',
            'NHL': 'nhl_opportunities',
            'NCAAB': 'ncaab_opportunities'
        }
        
        if league not in collection_map:
            raise HTTPException(status_code=400, detail=f"Invalid league: {league}")
        
        collection = db[collection_map[league]]
        
        # Get the document for this date
        doc = await collection.find_one({"date": data.date})
        if not doc:
            raise HTTPException(status_code=404, detail=f"No data found for {league} on {data.date}")
        
        games = doc.get('games', [])
        game_found = False
        
        for game in games:
            if game.get('game_num') == data.game_num:
                game['ranking_ppg'] = data.ranking_type  # "high" or "low"
                game_found = True
                break
        
        if not game_found:
            raise HTTPException(status_code=404, detail=f"Game #{data.game_num} not found")
        
        # Update the document
        await collection.update_one(
            {"date": data.date},
            {"$set": {"games": games}}
        )
        
        logger.info(f"Set ranking PPG to '{data.ranking_type}' for {league} game #{data.game_num} on {data.date}")
        
        return {
            "success": True,
            "message": f"Ranking PPG set to '{data.ranking_type}' for game #{data.game_num}",
            "league": league,
            "date": data.date,
            "game_num": data.game_num,
            "ranking_type": data.ranking_type
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error setting ranking PPG: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ================= BET CANCELLED FLAG =================

@api_router.get("/records/public-by-threshold/{league}")
async def get_public_records_by_threshold(league: str, threshold: int = 57):
    """
    Calculate Public Record dynamically based on consensus threshold.
    Threshold can be 57-70%.
    """
    try:
        league_upper = league.upper()
        if league_upper not in ['NBA', 'NHL', 'NCAAB', 'NFL']:
            raise HTTPException(status_code=400, detail="Invalid league")
        
        # For NBA, NHL and NCAAB, return stored record values
        if league_upper in ['NBA', 'NHL', 'NCAAB']:
            stored_record = await db.public_records.find_one({"league": league_upper})
            if stored_record:
                hits = stored_record.get('hits', 0)
                misses = stored_record.get('misses', 0)
                total = hits + misses
                win_pct = (hits / total * 100) if total > 0 else 0
                return {
                    "league": league_upper,
                    "threshold": f"{threshold}%",
                    "record": f"{hits}-{misses}",
                    "hits": hits,
                    "misses": misses,
                    "total_games": total,
                    "win_pct": round(win_pct, 1)
                }
        
        collection_name = f"{league_upper.lower()}_opportunities"
        
        # Get all documents (only used for NFL now)
        cursor = db[collection_name].find({})
        docs = await cursor.to_list(length=1000)
        
        total_hits = 0
        total_misses = 0
        
        for doc in docs:
            games = doc.get('games', [])
            
            for g in games:
                away_pct = g.get('away_consensus_pct') or 0
                home_pct = g.get('home_consensus_pct') or 0
                
                is_away_public = away_pct >= home_pct
                public_pct = away_pct if is_away_public else home_pct
                
                # Apply threshold filter
                if public_pct < threshold:
                    continue
                
                if is_away_public:
                    public_spread = g.get('away_spread')
                else:
                    public_spread = g.get('spread')
                
                away_score = g.get('away_score')
                home_score = g.get('home_score')
                
                if away_score is None or home_score is None or public_spread is None:
                    continue
                
                try:
                    if is_away_public:
                        covered = float(away_score) + float(public_spread) > float(home_score)
                        push = float(away_score) + float(public_spread) == float(home_score)
                    else:
                        covered = float(home_score) + float(public_spread) > float(away_score)
                        push = float(home_score) + float(public_spread) == float(away_score)
                    
                    if push:
                        continue
                    
                    if covered:
                        total_hits += 1
                    else:
                        total_misses += 1
                except (ValueError, TypeError):
                    continue
        
        total_games = total_hits + total_misses
        win_pct = (total_hits / total_games * 100) if total_games > 0 else 0
        
        return {
            "league": league_upper,
            "threshold": f"{threshold}%",
            "record": f"{total_hits}-{total_misses}",
            "hits": total_hits,
            "misses": total_misses,
            "total_games": total_games,
            "win_pct": round(win_pct, 1)
        }
    except Exception as e:
        logger.error(f"Error calculating public record by threshold: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/records/public-compound/{league}")
async def get_public_compound_records(league: str):
    """
    Calculate compound Public Records for fade-the-public analysis.
    Returns records for threshold ranges: 57-58, 59-60, 61-62, etc.
    """
    try:
        league_upper = league.upper()
        if league_upper not in ['NBA', 'NHL', 'NCAAB', 'NFL']:
            raise HTTPException(status_code=400, detail="Invalid league")
        
        collection_name = f"{league_upper.lower()}_opportunities"
        
        # Get all documents
        cursor = db[collection_name].find({})
        docs = await cursor.to_list(length=1000)
        
        # Define threshold pairs
        threshold_pairs = [
            (57, 58), (59, 60), (61, 62), (63, 64),
            (65, 66), (67, 68), (69, 70), (71, 72),
            (73, 74), (75, 76), (77, 78), (79, 80),
        ]
        
        results = []
        
        for low, high in threshold_pairs:
            public_wins = 0
            public_losses = 0
            
            for doc in docs:
                games = doc.get('games', [])
                
                for g in games:
                    away_pct = g.get('away_consensus_pct') or 0
                    home_pct = g.get('home_consensus_pct') or 0
                    
                    # Determine public side - must be within threshold range
                    public_side = None
                    if low <= away_pct <= high:
                        public_side = 'away'
                    elif low <= home_pct <= high:
                        public_side = 'home'
                    else:
                        continue
                    
                    # Get spread and scores
                    spread = g.get('spread')  # Home team spread
                    away_score = g.get('away_score')
                    home_score = g.get('home_score')
                    
                    if away_score is None or home_score is None or spread is None:
                        continue
                    
                    try:
                        # Calculate result based on home spread
                        home_result = float(home_score) + float(spread) - float(away_score)
                        
                        if home_result == 0:  # Push
                            continue
                        
                        if public_side == 'home':
                            if home_result > 0:
                                public_wins += 1
                            else:
                                public_losses += 1
                        else:  # away
                            if home_result < 0:
                                public_wins += 1
                            else:
                                public_losses += 1
                    except (ValueError, TypeError):
                        continue
            
            total = public_wins + public_losses
            if total > 0:
                public_win_pct = round(public_wins / total * 100, 1)
                fade_win_pct = round(public_losses / total * 100, 1)
                
                results.append({
                    "range": f"{low}-{high}%",
                    "low": low,
                    "high": high,
                    "public_record": f"{public_wins}-{public_losses}",
                    "fade_record": f"{public_losses}-{public_wins}",
                    "public_wins": public_wins,
                    "public_losses": public_losses,
                    "fade_wins": public_losses,
                    "fade_losses": public_wins,
                    "public_win_pct": public_win_pct,
                    "fade_win_pct": fade_win_pct,
                    "total_games": total
                })
        
        # Sort by threshold range (highest percentage first: 77-78%, 75-76%, etc.)
        results.sort(key=lambda x: x['high'], reverse=True)
        
        return {
            "league": league_upper,
            "compound_records": results
        }
    except Exception as e:
        logger.error(f"Error calculating compound public records: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class BetCancelledUpdate(BaseModel):
    league: str
    date: str
    game_num: int
    cancelled: bool = True

@api_router.post("/opportunities/bet-cancelled")
async def set_bet_cancelled(data: BetCancelledUpdate):
    """Mark a game's bet as cancelled (e.g., when user places both OVER and UNDER)"""
    try:
        league = data.league.upper()
        collection_map = {
            'NBA': 'nba_opportunities',
            'NHL': 'nhl_opportunities',
            'NCAAB': 'ncaab_opportunities'
        }
        
        if league not in collection_map:
            raise HTTPException(status_code=400, detail=f"Invalid league: {league}")
        
        collection = db[collection_map[league]]
        
        # Get the document for this date
        doc = await collection.find_one({"date": data.date})
        if not doc:
            raise HTTPException(status_code=404, detail=f"No data found for {league} on {data.date}")
        
        games = doc.get('games', [])
        game_found = False
        
        for game in games:
            if game.get('game_num') == data.game_num:
                if data.cancelled:
                    # Mark as cancelled and clear bet indicators
                    game['bet_cancelled'] = True
                    game['has_bet'] = False
                    game['user_bet'] = False
                    game['bet_type'] = None
                    game['bet_line'] = None
                    game['bet_types'] = []
                    game['bet_lines'] = []
                    game['bet_count'] = 0
                else:
                    # Uncancel - remove the flag (bet will be re-added on next refresh)
                    game['bet_cancelled'] = False
                game_found = True
                break
        
        if not game_found:
            raise HTTPException(status_code=404, detail=f"Game #{data.game_num} not found")
        
        # Update the document
        await collection.update_one(
            {"date": data.date},
            {"$set": {"games": games}}
        )
        
        action = "cancelled" if data.cancelled else "uncancelled"
        logger.info(f"Bet {action} for {league} game #{data.game_num} on {data.date}")
        
        return {
            "success": True,
            "message": f"Bet {action} for game #{data.game_num}",
            "league": league,
            "date": data.date,
            "game_num": data.game_num,
            "cancelled": data.cancelled
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error setting bet cancelled: {e}")
        raise HTTPException(status_code=500, detail=str(e))



@api_router.delete("/opportunities/ranking-ppg")
async def clear_ranking_ppg(league: str, date: str, game_num: int):
    """Clear ranking PPG selection for a game"""
    try:
        league = league.upper()
        collection_map = {
            'NBA': 'nba_opportunities',
            'NHL': 'nhl_opportunities',
            'NCAAB': 'ncaab_opportunities'
        }
        
        if league not in collection_map:
            raise HTTPException(status_code=400, detail=f"Invalid league: {league}")
        
        collection = db[collection_map[league]]
        
        # Get the document for this date
        doc = await collection.find_one({"date": date})
        if not doc:
            raise HTTPException(status_code=404, detail=f"No data found for {league} on {date}")
        
        games = doc.get('games', [])
        game_found = False
        
        for game in games:
            if game.get('game_num') == game_num:
                if 'ranking_ppg' in game:
                    del game['ranking_ppg']
                game_found = True
                break
        
        if not game_found:
            raise HTTPException(status_code=404, detail=f"Game #{game_num} not found")
        
        # Update the document
        await collection.update_one(
            {"date": date},
            {"$set": {"games": games}}
        )
        
        logger.info(f"Cleared ranking PPG for {league} game #{game_num} on {date}")
        
        return {
            "success": True,
            "message": f"Ranking PPG cleared for game #{game_num}"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error clearing ranking PPG: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/records/ranking-ppg-summary")
async def get_ranking_ppg_summary():
    """Get Ranking PPG records summary for all leagues from stored records"""
    try:
        summary = {}
        
        for league in ['NBA', 'NHL', 'NCAAB', 'NFL']:
            # Read from stored ranking_ppg_records collection
            record = await db.ranking_ppg_records.find_one({"league": league})
            
            if record:
                # Handle both old format (high_hits) and new format (high: {hits, misses})
                if 'high' in record and isinstance(record['high'], dict):
                    high_hits = record['high'].get('hits', 0)
                    high_misses = record['high'].get('misses', 0)
                else:
                    high_hits = record.get('high_hits', 0)
                    high_misses = record.get('high_misses', 0)
                    
                if 'low' in record and isinstance(record['low'], dict):
                    low_hits = record['low'].get('hits', 0)
                    low_misses = record['low'].get('misses', 0)
                else:
                    low_hits = record.get('low_hits', 0)
                    low_misses = record.get('low_misses', 0)
            else:
                high_hits = high_misses = low_hits = low_misses = 0
            
            summary[league] = {
                "high_record": f"{high_hits}-{high_misses}",
                "low_record": f"{low_hits}-{low_misses}",
                "high_hits": high_hits,
                "high_misses": high_misses,
                "low_hits": low_hits,
                "low_misses": low_misses
            }
        
        return summary
    except Exception as e:
        logger.error(f"Error getting ranking PPG summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/records/public-summary")
async def get_public_records_summary():
    """Get Public Consensus spread records summary for all leagues from stored records"""
    try:
        summary = {}
        
        for league in ['NBA', 'NHL', 'NCAAB']:
            # Read from stored public_records collection
            record = await db.public_records.find_one({"league": league})
            
            if record:
                hits = record.get('hits', 0)
                misses = record.get('misses', 0)
            else:
                hits = misses = 0
            
            summary[league] = {
                "record": f"{hits}-{misses}",
                "hits": hits,
                "misses": misses
            }
        
        return summary
    except Exception as e:
        logger.error(f"Error getting public records summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/records/public-detail/{league}")
async def get_public_records_detail(league: str):
    """
    Get detailed Public Record breakdown by date and game for verification.
    Uses 57%+ consensus threshold and Covers.com spread data.
    """
    try:
        league_upper = league.upper()
        if league_upper not in ['NBA', 'NHL', 'NCAAB', 'NFL']:
            raise HTTPException(status_code=400, detail="Invalid league")
        
        record = await db.public_records.find_one({"league": league_upper})
        
        if not record:
            return {
                "league": league_upper,
                "total_record": "0-0",
                "hits": 0,
                "misses": 0,
                "games": [],
                "by_date": {}
            }
        
        games = record.get('games', [])
        
        # Group by date
        by_date = {}
        for g in games:
            date = g.get('date', 'unknown')
            if date not in by_date:
                by_date[date] = {"hits": 0, "misses": 0, "games": []}
            
            if g.get('result') == 'HIT':
                by_date[date]["hits"] += 1
            else:
                by_date[date]["misses"] += 1
            
            by_date[date]["games"].append({
                "game": g.get('game'),
                "public_pick": g.get('public_pick'),
                "consensus_pct": g.get('consensus_pct'),
                "spread": g.get('spread'),
                "result": g.get('result')
            })
        
        return {
            "league": league_upper,
            "total_record": f"{record.get('hits', 0)}-{record.get('misses', 0)}",
            "hits": record.get('hits', 0),
            "misses": record.get('misses', 0),
            "threshold": "57%",
            "spread_source": "CBS Sports Live Line",
            "by_date": by_date
        }
    except Exception as e:
        logger.error(f"Error getting public records detail: {e}")
        raise HTTPException(status_code=500, detail=str(e))




@api_router.get("/process/status")
async def get_process_status():
    """Get status of scheduled processes and tomorrow's data."""
    from zoneinfo import ZoneInfo
    arizona_tz = ZoneInfo('America/Phoenix')
    now_arizona = datetime.now(arizona_tz)
    tomorrow = (now_arizona + timedelta(days=1)).strftime('%Y-%m-%d')
    
    status = {
        "current_time_arizona": now_arizona.strftime('%Y-%m-%d %I:%M %p'),
        "tomorrow_date": tomorrow,
        "leagues": {}
    }
    
    for league in ['nba', 'nhl', 'nfl']:
        coll_name = f"{league}_opportunities"
        doc = await db[coll_name].find_one({"date": tomorrow}, {"_id": 0})
        
        if doc:
            games = doc.get('games', [])
            ppg_populated = doc.get('ppg_populated', False)
            sample_game = games[0] if games else {}
            
            status["leagues"][league.upper()] = {
                "games_count": len(games),
                "ppg_populated": ppg_populated,
                "ppg_updated_at": doc.get('ppg_updated_at'),
                "has_dots": 'dots' in sample_game if sample_game else False,
                "has_combined_ppg": 'combined_ppg' in sample_game if sample_game else False
            }
        else:
            status["leagues"][league.upper()] = {
                "games_count": 0,
                "ppg_populated": False,
                "ppg_updated_at": None,
                "has_dots": False,
                "has_combined_ppg": False
            }
    
    return status


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    global monitoring_enabled
    monitoring_enabled = False
    if scheduler.running:
        scheduler.shutdown()
    client.close()
    await plays888_service.close()
